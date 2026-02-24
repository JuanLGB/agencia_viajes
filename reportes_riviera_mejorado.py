from database import conectar
from datetime import datetime
import os

def reporte_riviera_consola(usuario_actual):
    """Muestra reporte detallado de Riviera Maya en consola"""
    conexion = conectar()
    cursor = conexion.cursor()
    
    print("\n" + "="*70)
    print("🏖️  REPORTE DETALLADO - RIVIERA MAYA")
    print("="*70)
    
    # Filtro por vendedora si no es admin
    if usuario_actual["rol"] == "VENDEDORA":
        print(f"Vendedora: {usuario_actual['nombre']}")
        filtro_usuario = f"AND usuario_id = {usuario_actual['id_vendedora']}"
    else:
        filtro_usuario = ""
    
    # ===== OBTENER VENTAS ACTIVAS =====
    cursor.execute(f"""
        SELECT v.id, v.cliente, v.destino, v.fecha_inicio, v.fecha_fin,
               v.noches, v.adultos, v.menores, v.tipo_habitacion,
               v.precio_total, v.pagado, v.saldo, v.ganancia,
               v.estado, vd.nombre, v.tipo_venta, v.es_bloqueo
        FROM ventas v
        JOIN vendedoras vd ON v.vendedora_id = vd.id
        WHERE v.estado != 'CERRADO' {filtro_usuario}
        ORDER BY v.fecha_inicio
    """)
    
    ventas = cursor.fetchall()
    
    if not ventas:
        print("\n⚠️  No hay ventas activas.")
        conexion.close()
        return
    
    print(f"\n\n{'='*70}")
    print(f"📋 VENTAS ACTIVAS ({len(ventas)})")
    print("="*70)
    
    # Totales
    total_vendido = 0
    total_cobrado = 0
    total_saldo = 0
    total_ganancia = 0
    total_personas = 0
    ventas_bloqueo = 0
    ventas_generales = 0
    
    for venta in ventas:
        venta_id = venta[0]
        tipo_icon = "📦" if venta[16] else "🏖️"
        estado_icon = "✅" if venta[13] == "LIQUIDADO" else "⏳"
        
        print(f"\n{tipo_icon} {estado_icon} ID: {venta_id} | {venta[1]}")
        print(f"   Destino: {venta[2]}")
        print(f"   Fechas: {venta[3]} al {venta[4]} ({venta[5]} noches)")
        print(f"   Pasajeros: {venta[6]} adultos + {venta[7]} menores = {venta[6] + venta[7]} total")
        print(f"   Habitación: {venta[8]}")
        print(f"   Vendedora: {venta[14]}")
        print(f"   Tipo: {venta[15]}")
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre, tipo
            FROM pasajeros
            WHERE venta_id = ?
            ORDER BY tipo DESC
        """, (venta_id,))
        
        pasajeros = cursor.fetchall()
        
        if pasajeros:
            print(f"\n   👥 PASAJEROS:")
            for p in pasajeros:
                print(f"      • {p[0]} ({p[1]})")
        
        print(f"\n   💵 PAGOS:")
        print(f"      Total: ${venta[9]:,.2f}")
        print(f"      Pagado: ${venta[10]:,.2f}")
        print(f"      Saldo: ${venta[11]:,.2f}")
        print(f"      Ganancia: ${venta[12]:,.2f}")
        
        # Historial de abonos
        cursor.execute("""
            SELECT fecha, monto
            FROM abonos
            WHERE venta_id = ?
            ORDER BY fecha
        """, (venta_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            print(f"\n   📅 Historial de pagos:")
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
                print(f"      {fecha}: ${abono[1]:,.2f}")
        
        print("-" * 70)
        
        # Acumular totales
        total_vendido += venta[9]
        total_cobrado += venta[10]
        total_saldo += venta[11]
        total_ganancia += venta[12]
        total_personas += venta[6] + venta[7]
        
        if venta[16]:  # es_bloqueo
            ventas_bloqueo += 1
        else:
            ventas_generales += 1
    
    # ===== RESUMEN GENERAL =====
    print(f"\n{'='*70}")
    print("📊 RESUMEN GENERAL")
    print("="*70)
    
    print(f"\n📋 VENTAS:")
    print(f"   Total ventas: {len(ventas)}")
    print(f"   📦 Bloqueos: {ventas_bloqueo}")
    print(f"   🏖️  Generales: {ventas_generales}")
    print(f"   👥 Total pasajeros: {total_personas}")
    
    print(f"\n💰 FINANCIERO:")
    print(f"   Total vendido: ${total_vendido:,.2f}")
    print(f"   Total cobrado: ${total_cobrado:,.2f}")
    print(f"   Total saldo: ${total_saldo:,.2f}")
    porcentaje_cobrado = (total_cobrado / total_vendido * 100) if total_vendido > 0 else 0
    print(f"   % Cobrado: {porcentaje_cobrado:.1f}%")
    
    print(f"\n📈 GANANCIA:")
    print(f"   Ganancia total: ${total_ganancia:,.2f}")
    if total_vendido > 0:
        margen = (total_ganancia / total_vendido * 100)
        print(f"   Margen promedio: {margen:.1f}%")
    
    # ===== ALERTAS =====
    print(f"\n⚠️  ALERTAS:")
    alertas = []
    
    # Ventas con saldo alto
    cursor.execute(f"""
        SELECT COUNT(*) FROM ventas
        WHERE saldo > 20000 AND estado != 'CERRADO' {filtro_usuario}
    """)
    
    ventas_saldo_alto = cursor.fetchone()[0]
    if ventas_saldo_alto > 0:
        alertas.append(f"   • {ventas_saldo_alto} venta(s) con saldo > $20,000")
    
    # % Cobrado bajo
    if porcentaje_cobrado < 70 and len(ventas) > 0:
        alertas.append(f"   • ⚠️  Solo {porcentaje_cobrado:.0f}% cobrado del total")
    
    if alertas:
        for alerta in alertas:
            print(alerta)
    else:
        print("   ✅ Sin alertas")
    
    print("\n" + "="*70)
    
    conexion.close()


def exportar_reporte_riviera_excel(usuario_actual):
    """Exporta reporte de Riviera Maya a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n❌ Error: Se requiere openpyxl para exportar a Excel.")
        print("Instala con: pip install openpyxl")
        return
    
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Filtro por vendedora si no es admin
    if usuario_actual["rol"] == "VENDEDORA":
        filtro_usuario = f"AND usuario_id = {usuario_actual['id_vendedora']}"
        nombre_reporte = f"Riviera_{usuario_actual['nombre'].replace(' ', '_')}"
    else:
        filtro_usuario = ""
        nombre_reporte = "Riviera_General"
    
    # Obtener ventas
    cursor.execute(f"""
        SELECT v.id, v.cliente, v.destino, v.fecha_inicio, v.fecha_fin,
               v.noches, v.adultos, v.menores, v.tipo_habitacion,
               v.precio_total, v.pagado, v.saldo, v.ganancia,
               v.estado, vd.nombre, v.tipo_venta, v.es_bloqueo
        FROM ventas v
        JOIN vendedoras vd ON v.vendedora_id = vd.id
        WHERE v.estado != 'CERRADO' {filtro_usuario}
        ORDER BY v.fecha_inicio
    """)
    
    ventas = cursor.fetchall()
    
    if not ventas:
        print("\n⚠️  No hay ventas activas para exportar.")
        conexion.close()
        return
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: VENTAS Y PAGOS =====
    ws_ventas = wb.create_sheet("Ventas y Pagos")
    
    ws_ventas['A1'] = "REPORTE RIVIERA MAYA - VENTAS ACTIVAS"
    ws_ventas['A1'].font = Font(bold=True, size=16)
    ws_ventas.merge_cells('A1:M1')
    
    # Encabezados
    headers = ["ID", "Cliente", "Destino", "Fecha Inicio", "Fecha Fin", "Noches",
               "Adultos", "Menores", "Total Pax", "Habitación", "Vendedora",
               "Total", "Pagado", "Saldo", "Estado"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_ventas.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    row = 4
    total_vendido = 0
    total_cobrado = 0
    total_saldo = 0
    
    for venta in ventas:
        ws_ventas.cell(row=row, column=1, value=venta[0])
        ws_ventas.cell(row=row, column=2, value=venta[1])
        ws_ventas.cell(row=row, column=3, value=venta[2])
        ws_ventas.cell(row=row, column=4, value=venta[3])
        ws_ventas.cell(row=row, column=5, value=venta[4])
        ws_ventas.cell(row=row, column=6, value=venta[5])
        ws_ventas.cell(row=row, column=7, value=venta[6])
        ws_ventas.cell(row=row, column=8, value=venta[7])
        ws_ventas.cell(row=row, column=9, value=venta[6] + venta[7])
        ws_ventas.cell(row=row, column=10, value=venta[8])
        ws_ventas.cell(row=row, column=11, value=venta[14])
        ws_ventas.cell(row=row, column=12, value=venta[9])
        ws_ventas.cell(row=row, column=13, value=venta[10])
        ws_ventas.cell(row=row, column=14, value=venta[11])
        ws_ventas.cell(row=row, column=15, value=venta[13])
        
        # Formato moneda
        for col in [12, 13, 14]:
            ws_ventas.cell(row=row, column=col).number_format = '$#,##0.00'
        
        # Borders
        for col in range(1, 16):
            ws_ventas.cell(row=row, column=col).border = border
        
        total_vendido += venta[9]
        total_cobrado += venta[10]
        total_saldo += venta[11]
        
        row += 1
    
    # Totales
    row += 1
    ws_ventas.cell(row=row, column=11, value="TOTALES:").font = Font(bold=True)
    ws_ventas.cell(row=row, column=12, value=total_vendido).font = Font(bold=True)
    ws_ventas.cell(row=row, column=13, value=total_cobrado).font = Font(bold=True)
    ws_ventas.cell(row=row, column=14, value=total_saldo).font = Font(bold=True)
    
    for col in [12, 13, 14]:
        ws_ventas.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_ventas.column_dimensions['A'].width = 8
    ws_ventas.column_dimensions['B'].width = 30
    ws_ventas.column_dimensions['C'].width = 25
    for col in ['D', 'E']:
        ws_ventas.column_dimensions[col].width = 12
    for col in ['F', 'G', 'H', 'I']:
        ws_ventas.column_dimensions[col].width = 10
    ws_ventas.column_dimensions['J'].width = 15
    ws_ventas.column_dimensions['K'].width = 25
    for col in ['L', 'M', 'N']:
        ws_ventas.column_dimensions[col].width = 15
    ws_ventas.column_dimensions['O'].width = 12
    
    # ===== HOJA 2: PASAJEROS =====
    ws_pasajeros = wb.create_sheet("Pasajeros")
    
    ws_pasajeros['A1'] = "LISTA DE PASAJEROS POR VENTA"
    ws_pasajeros['A1'].font = Font(bold=True, size=14)
    ws_pasajeros.merge_cells('A1:E1')
    
    # Encabezados
    headers_pas = ["ID Venta", "Cliente", "Pasajero", "Tipo", "Habitación"]
    
    for col, header in enumerate(headers_pas, start=1):
        cell = ws_pasajeros.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for venta in ventas:
        venta_id = venta[0]
        cliente = venta[1]
        habitacion = venta[8]
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre, tipo
            FROM pasajeros
            WHERE venta_id = ?
            ORDER BY tipo DESC
        """, (venta_id,))
        
        pasajeros = cursor.fetchall()
        
        primera_fila = True
        for pasajero in pasajeros:
            ws_pasajeros.cell(row=row, column=1, value=venta_id if primera_fila else "")
            ws_pasajeros.cell(row=row, column=2, value=cliente if primera_fila else "")
            ws_pasajeros.cell(row=row, column=3, value=pasajero[0])
            ws_pasajeros.cell(row=row, column=4, value=pasajero[1])
            ws_pasajeros.cell(row=row, column=5, value=habitacion if primera_fila else "")
            
            for col in range(1, 6):
                ws_pasajeros.cell(row=row, column=col).border = border
            
            primera_fila = False
            row += 1
        
        row += 1  # Línea separadora
    
    # Ajustar anchos
    ws_pasajeros.column_dimensions['A'].width = 10
    ws_pasajeros.column_dimensions['B'].width = 30
    ws_pasajeros.column_dimensions['C'].width = 35
    ws_pasajeros.column_dimensions['D'].width = 12
    ws_pasajeros.column_dimensions['E'].width = 15
    
    # ===== HOJA 3: HISTORIAL DE PAGOS =====
    ws_pagos = wb.create_sheet("Historial de Pagos")
    
    ws_pagos['A1'] = "HISTORIAL COMPLETO DE PAGOS"
    ws_pagos['A1'].font = Font(bold=True, size=14)
    ws_pagos.merge_cells('A1:E1')
    
    # Encabezados
    headers_pagos = ["ID Venta", "Cliente", "Fecha Pago", "Monto", "Saldo Actual"]
    
    for col, header in enumerate(headers_pagos, start=1):
        cell = ws_pagos.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for venta in ventas:
        venta_id = venta[0]
        cliente = venta[1]
        saldo_actual = venta[11]
        
        # Obtener abonos
        cursor.execute("""
            SELECT fecha, monto
            FROM abonos
            WHERE venta_id = ?
            ORDER BY fecha
        """, (venta_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            primera_fila = True
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
                
                ws_pagos.cell(row=row, column=1, value=venta_id if primera_fila else "")
                ws_pagos.cell(row=row, column=2, value=cliente if primera_fila else "")
                ws_pagos.cell(row=row, column=3, value=fecha)
                ws_pagos.cell(row=row, column=4, value=abono[1])
                ws_pagos.cell(row=row, column=5, value=saldo_actual if primera_fila else "")
                
                # Formato
                ws_pagos.cell(row=row, column=4).number_format = '$#,##0.00'
                ws_pagos.cell(row=row, column=5).number_format = '$#,##0.00'
                
                for col in range(1, 6):
                    ws_pagos.cell(row=row, column=col).border = border
                
                primera_fila = False
                row += 1
    
    # Ajustar anchos
    ws_pagos.column_dimensions['A'].width = 10
    ws_pagos.column_dimensions['B'].width = 30
    ws_pagos.column_dimensions['C'].width = 20
    ws_pagos.column_dimensions['D'].width = 15
    ws_pagos.column_dimensions['E'].width = 15
    
    # Guardar archivo
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"Reporte_{nombre_reporte}_{fecha_actual}.xlsx"
    
    wb.save(nombre_archivo)
    
    print(f"\n✅ Reporte exportado exitosamente:")
    print(f"📁 {nombre_archivo}")
    
    # Mostrar resumen
    print(f"\n📊 RESUMEN:")
    print(f"   Ventas: {len(ventas)}")
    print(f"   Total vendido: ${total_vendido:,.2f}")
    print(f"   Total cobrado: ${total_cobrado:,.2f}")
    print(f"   Saldo pendiente: ${total_saldo:,.2f}")
    
    conexion.close()
    
    return nombre_archivo


def menu_reportes_riviera(usuario_actual):
    """Menú principal de reportes Riviera Maya"""
    while True:
        print("\n" + "="*70)
        print("📊 REPORTES - RIVIERA MAYA")
        print("="*70)
        print("\n1. Ver reporte en consola")
        print("2. Exportar reporte a Excel")
        print("3. Volver")
        
        opcion = input("\nSelecciona una opción: ").strip()
        
        if opcion == "1":
            reporte_riviera_consola(usuario_actual)
        elif opcion == "2":
            archivo = exportar_reporte_riviera_excel(usuario_actual)
            if archivo:
                print(f"\n💡 Archivo guardado en la carpeta actual")
        elif opcion == "3":
            break
        else:
            print("❌ Opción inválida.")


if __name__ == "__main__":
    # Para pruebas - normalmente se llamaría desde main.py
    usuario_test = {"rol": "ADMIN", "nombre": "Admin"}
    menu_reportes_riviera(usuario_test)
