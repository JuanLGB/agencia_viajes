from database import conectar
from datetime import datetime, timedelta
import os

def reporte_movimientos_semanal_consola():
    """Muestra reporte de todos los abonos de la semana en consola"""
    print("\n" + "="*70)
    print("📊 REPORTE SEMANAL DE MOVIMIENTOS")
    print("="*70)
    
    # Solicitar rango de fechas
    print("\n📅 RANGO DE FECHAS")
    print("1. Última semana (últimos 7 días)")
    print("2. Últimas 2 semanas (últimos 14 días)")
    print("3. Último mes (últimos 30 días)")
    print("4. Rango personalizado")
    
    opcion = input("\nSelecciona opción: ").strip()
    
    hoy = datetime.now()
    
    if opcion == "1":
        fecha_inicio = (hoy - timedelta(days=7)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "ÚLTIMOS 7 DÍAS"
    elif opcion == "2":
        fecha_inicio = (hoy - timedelta(days=14)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "ÚLTIMOS 14 DÍAS"
    elif opcion == "3":
        fecha_inicio = (hoy - timedelta(days=30)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "ÚLTIMOS 30 DÍAS"
    elif opcion == "4":
        fecha_inicio = input("Fecha inicio (DD-MM-YYYY): ").strip()
        fecha_fin = input("Fecha fin (DD-MM-YYYY): ").strip()
        try:
            inicio_dt = datetime.strptime(fecha_inicio, "%d-%m-%Y")
            fin_dt = datetime.strptime(fecha_fin, "%d-%m-%Y")
            fecha_inicio = inicio_dt.strftime("%Y-%m-%d")
            fecha_fin = fin_dt.strftime("%Y-%m-%d")
            titulo_periodo = f"{inicio_dt.strftime('%d-%m-%Y')} al {fin_dt.strftime('%d-%m-%Y')}"
        except:
            print("❌ Fechas inválidas.")
            return
    else:
        print("❌ Opción inválida.")
        return
    
    conexion = conectar()
    cursor = conexion.cursor()
    
    print("\n" + "="*70)
    print(f"💰 MOVIMIENTOS - {titulo_periodo}")
    print("="*70)
    
    # ===== 1. ABONOS RIVIERA MAYA =====
    print(f"\n{'='*70}")
    print("🏖️  ABONOS - RIVIERA MAYA")
    print("="*70)
    
    cursor.execute("""
        SELECT a.fecha, v.cliente, v.destino, a.monto, vd.nombre, v.tipo_venta, v.id
        FROM abonos a
        JOIN ventas v ON a.venta_id = v.id
        JOIN vendedoras vd ON v.vendedora_id = vd.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_riviera = cursor.fetchall()
    total_riviera = 0
    
    if abonos_riviera:
        for abono in abonos_riviera:
            fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
            tipo_icon = "📦" if abono[5] == "Bloqueo" else "🏖️"
            print(f"\n{tipo_icon} {fecha}")
            print(f"   Viaje ID: {abono[6]}")
            print(f"   Cliente: {abono[1]}")
            print(f"   Destino: {abono[2]}")
            print(f"   Vendedora: {abono[4]}")
            print(f"   Tipo: {abono[5]}")
            print(f"   💵 Monto: ${abono[3]:,.2f}")
            total_riviera += abono[3]
            print("-" * 70)
        
        print(f"\n💰 TOTAL RIVIERA MAYA: ${total_riviera:,.2f}")
    else:
        print("\n⚠️  Sin movimientos en este periodo")
    
    # ===== 2. ABONOS VIAJES NACIONALES =====
    print(f"\n\n{'='*70}")
    print("🎫 ABONOS - VIAJES NACIONALES")
    print("="*70)
    
    cursor.execute("""
        SELECT a.fecha, c.nombre_cliente, vn.nombre_viaje, vn.destino, a.monto, 
               vd.nombre, vn.id, c.id
        FROM abonos_nacionales a
        JOIN clientes_nacionales c ON a.cliente_id = c.id
        JOIN viajes_nacionales vn ON c.viaje_id = vn.id
        JOIN vendedoras vd ON a.vendedora_id = vd.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_nacionales = cursor.fetchall()
    total_nacionales = 0
    
    if abonos_nacionales:
        for abono in abonos_nacionales:
            fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
            print(f"\n🎫 {fecha}")
            print(f"   Viaje: {abono[2]} (ID: {abono[6]})")
            print(f"   Destino: {abono[3]}")
            print(f"   Cliente: {abono[1]} (ID: {abono[7]})")
            print(f"   Vendedora: {abono[5]}")
            print(f"   💵 Monto: ${abono[4]:,.2f}")
            total_nacionales += abono[4]
            print("-" * 70)
        
        print(f"\n💰 TOTAL NACIONALES: ${total_nacionales:,.2f}")
    else:
        print("\n⚠️  Sin movimientos en este periodo")
    
    # ===== 3. ABONOS VIAJES INTERNACIONALES =====
    print(f"\n\n{'='*70}")
    print("🌎 ABONOS - VIAJES INTERNACIONALES")
    print("="*70)
    
    cursor.execute("""
        SELECT a.fecha, c.nombre_cliente, vi.destino, a.moneda, a.monto_original,
               a.tipo_cambio, a.monto_usd, vi.id, c.id
        FROM abonos_internacionales a
        JOIN clientes_internacionales c ON a.cliente_id = c.id
        JOIN viajes_internacionales vi ON c.viaje_id = vi.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_internacionales = cursor.fetchall()
    total_internacionales_usd = 0
    
    if abonos_internacionales:
        for abono in abonos_internacionales:
            fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
            print(f"\n🌎 {fecha}")
            print(f"   Viaje ID: {abono[7]}")
            print(f"   Destino: {abono[2]}")
            print(f"   Cliente: {abono[1]} (ID: {abono[8]})")
            
            if abono[3] == "USD":
                print(f"   💵 Monto: ${abono[6]:,.2f} USD")
            else:
                print(f"   💵 Monto: ${abono[4]:,.2f} MXN (TC: {abono[5]:.2f}) = ${abono[6]:,.2f} USD")
            
            total_internacionales_usd += abono[6]
            print("-" * 70)
        
        print(f"\n💰 TOTAL INTERNACIONALES: ${total_internacionales_usd:,.2f} USD")
    else:
        print("\n⚠️  Sin movimientos en este periodo")
    
    # ===== RESUMEN GENERAL =====
    print(f"\n\n{'='*70}")
    print("📊 RESUMEN GENERAL")
    print("="*70)
    
    total_abonos = len(abonos_riviera) + len(abonos_nacionales) + len(abonos_internacionales)
    
    print(f"\n📋 MOVIMIENTOS:")
    print(f"   Riviera Maya: {len(abonos_riviera)} abonos")
    print(f"   Viajes Nacionales: {len(abonos_nacionales)} abonos")
    print(f"   Viajes Internacionales: {len(abonos_internacionales)} abonos")
    print(f"   TOTAL ABONOS: {total_abonos}")
    
    print(f"\n💰 INGRESOS (MXN):")
    print(f"   Riviera Maya: ${total_riviera:,.2f}")
    print(f"   Viajes Nacionales: ${total_nacionales:,.2f}")
    print(f"   Total MXN: ${total_riviera + total_nacionales:,.2f}")
    
    if total_internacionales_usd > 0:
        print(f"\n💵 INGRESOS (USD):")
        print(f"   Viajes Internacionales: ${total_internacionales_usd:,.2f} USD")
    
    print("\n" + "="*70)
    
    conexion.close()


def exportar_movimientos_excel():
    """Exporta reporte de movimientos a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n❌ Error: Se requiere openpyxl para exportar a Excel.")
        print("Instala con: pip install openpyxl")
        return
    
    print("\n" + "="*70)
    print("📊 EXPORTAR REPORTE DE MOVIMIENTOS")
    print("="*70)
    
    # Solicitar rango de fechas
    print("\n📅 RANGO DE FECHAS")
    print("1. Última semana (últimos 7 días)")
    print("2. Últimas 2 semanas (últimos 14 días)")
    print("3. Último mes (últimos 30 días)")
    print("4. Rango personalizado")
    
    opcion = input("\nSelecciona opción: ").strip()
    
    hoy = datetime.now()
    
    if opcion == "1":
        fecha_inicio = (hoy - timedelta(days=7)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "Ultimos_7_dias"
    elif opcion == "2":
        fecha_inicio = (hoy - timedelta(days=14)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "Ultimos_14_dias"
    elif opcion == "3":
        fecha_inicio = (hoy - timedelta(days=30)).strftime("%Y-%m-%d")
        fecha_fin = hoy.strftime("%Y-%m-%d")
        titulo_periodo = "Ultimo_mes"
    elif opcion == "4":
        fecha_inicio = input("Fecha inicio (DD-MM-YYYY): ").strip()
        fecha_fin = input("Fecha fin (DD-MM-YYYY): ").strip()
        try:
            inicio_dt = datetime.strptime(fecha_inicio, "%d-%m-%Y")
            fin_dt = datetime.strptime(fecha_fin, "%d-%m-%Y")
            fecha_inicio = inicio_dt.strftime("%Y-%m-%d")
            fecha_fin = fin_dt.strftime("%Y-%m-%d")
            titulo_periodo = f"{inicio_dt.strftime('%d%m%Y')}_{fin_dt.strftime('%d%m%Y')}"
        except:
            print("❌ Fechas inválidas.")
            return
    else:
        print("❌ Opción inválida.")
        return
    
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: RIVIERA MAYA =====
    ws_riviera = wb.create_sheet("Riviera Maya")
    
    ws_riviera['A1'] = "ABONOS - RIVIERA MAYA"
    ws_riviera['A1'].font = Font(bold=True, size=16)
    ws_riviera.merge_cells('A1:H1')
    
    # Encabezados
    headers_riviera = ["Fecha", "ID Viaje", "Cliente", "Destino", "Tipo", "Vendedora", "Monto"]
    
    for col, header in enumerate(headers_riviera, start=1):
        cell = ws_riviera.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener datos
    cursor.execute("""
        SELECT a.fecha, v.cliente, v.destino, a.monto, vd.nombre, v.tipo_venta, v.id
        FROM abonos a
        JOIN ventas v ON a.venta_id = v.id
        JOIN vendedoras vd ON v.vendedora_id = vd.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_riviera = cursor.fetchall()
    
    row = 4
    total_riviera = 0
    
    for abono in abonos_riviera:
        fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
        
        ws_riviera.cell(row=row, column=1, value=fecha)
        ws_riviera.cell(row=row, column=2, value=abono[6])
        ws_riviera.cell(row=row, column=3, value=abono[1])
        ws_riviera.cell(row=row, column=4, value=abono[2])
        ws_riviera.cell(row=row, column=5, value=abono[5])
        ws_riviera.cell(row=row, column=6, value=abono[4])
        ws_riviera.cell(row=row, column=7, value=abono[3])
        
        ws_riviera.cell(row=row, column=7).number_format = '$#,##0.00'
        
        for col in range(1, 8):
            ws_riviera.cell(row=row, column=col).border = border
        
        total_riviera += abono[3]
        row += 1
    
    # Total
    row += 1
    ws_riviera.cell(row=row, column=6, value="TOTAL:").font = Font(bold=True)
    ws_riviera.cell(row=row, column=7, value=total_riviera).font = Font(bold=True)
    ws_riviera.cell(row=row, column=7).number_format = '$#,##0.00'
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_riviera.column_dimensions[col].width = 18
    
    # ===== HOJA 2: VIAJES NACIONALES =====
    ws_nacionales = wb.create_sheet("Viajes Nacionales")
    
    ws_nacionales['A1'] = "ABONOS - VIAJES NACIONALES"
    ws_nacionales['A1'].font = Font(bold=True, size=16)
    ws_nacionales.merge_cells('A1:H1')
    
    # Encabezados
    headers_nac = ["Fecha", "ID Viaje", "Viaje", "Destino", "ID Cliente", "Cliente", "Vendedora", "Monto"]
    
    for col, header in enumerate(headers_nac, start=1):
        cell = ws_nacionales.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener datos
    cursor.execute("""
        SELECT a.fecha, c.nombre_cliente, vn.nombre_viaje, vn.destino, a.monto, 
               vd.nombre, vn.id, c.id
        FROM abonos_nacionales a
        JOIN clientes_nacionales c ON a.cliente_id = c.id
        JOIN viajes_nacionales vn ON c.viaje_id = vn.id
        JOIN vendedoras vd ON a.vendedora_id = vd.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_nacionales = cursor.fetchall()
    
    row = 4
    total_nacionales = 0
    
    for abono in abonos_nacionales:
        fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
        
        ws_nacionales.cell(row=row, column=1, value=fecha)
        ws_nacionales.cell(row=row, column=2, value=abono[6])
        ws_nacionales.cell(row=row, column=3, value=abono[2])
        ws_nacionales.cell(row=row, column=4, value=abono[3])
        ws_nacionales.cell(row=row, column=5, value=abono[7])
        ws_nacionales.cell(row=row, column=6, value=abono[1])
        ws_nacionales.cell(row=row, column=7, value=abono[5])
        ws_nacionales.cell(row=row, column=8, value=abono[4])
        
        ws_nacionales.cell(row=row, column=8).number_format = '$#,##0.00'
        
        for col in range(1, 9):
            ws_nacionales.cell(row=row, column=col).border = border
        
        total_nacionales += abono[4]
        row += 1
    
    # Total
    row += 1
    ws_nacionales.cell(row=row, column=7, value="TOTAL:").font = Font(bold=True)
    ws_nacionales.cell(row=row, column=8, value=total_nacionales).font = Font(bold=True)
    ws_nacionales.cell(row=row, column=8).number_format = '$#,##0.00'
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_nacionales.column_dimensions[col].width = 18
    
    # ===== HOJA 3: VIAJES INTERNACIONALES =====
    ws_internacionales = wb.create_sheet("Viajes Internacionales")
    
    ws_internacionales['A1'] = "ABONOS - VIAJES INTERNACIONALES"
    ws_internacionales['A1'].font = Font(bold=True, size=16)
    ws_internacionales.merge_cells('A1:I1')
    
    # Encabezados
    headers_int = ["Fecha", "ID Viaje", "Destino", "ID Cliente", "Cliente", "Moneda", "Monto Original", "T.C.", "Monto USD"]
    
    for col, header in enumerate(headers_int, start=1):
        cell = ws_internacionales.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener datos
    cursor.execute("""
        SELECT a.fecha, c.nombre_cliente, vi.destino, a.moneda, a.monto_original,
               a.tipo_cambio, a.monto_usd, vi.id, c.id
        FROM abonos_internacionales a
        JOIN clientes_internacionales c ON a.cliente_id = c.id
        JOIN viajes_internacionales vi ON c.viaje_id = vi.id
        WHERE DATE(a.fecha) BETWEEN ? AND ?
        ORDER BY a.fecha DESC
    """, (fecha_inicio, fecha_fin))
    
    abonos_internacionales = cursor.fetchall()
    
    row = 4
    total_internacionales_usd = 0
    
    for abono in abonos_internacionales:
        fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
        
        ws_internacionales.cell(row=row, column=1, value=fecha)
        ws_internacionales.cell(row=row, column=2, value=abono[7])
        ws_internacionales.cell(row=row, column=3, value=abono[2])
        ws_internacionales.cell(row=row, column=4, value=abono[8])
        ws_internacionales.cell(row=row, column=5, value=abono[1])
        ws_internacionales.cell(row=row, column=6, value=abono[3])
        ws_internacionales.cell(row=row, column=7, value=abono[4])
        ws_internacionales.cell(row=row, column=8, value=abono[5])
        ws_internacionales.cell(row=row, column=9, value=abono[6])
        
        ws_internacionales.cell(row=row, column=7).number_format = '$#,##0.00'
        ws_internacionales.cell(row=row, column=8).number_format = '#,##0.00'
        ws_internacionales.cell(row=row, column=9).number_format = '$#,##0.00'
        
        for col in range(1, 10):
            ws_internacionales.cell(row=row, column=col).border = border
        
        total_internacionales_usd += abono[6]
        row += 1
    
    # Total
    row += 1
    ws_internacionales.cell(row=row, column=8, value="TOTAL:").font = Font(bold=True)
    ws_internacionales.cell(row=row, column=9, value=total_internacionales_usd).font = Font(bold=True)
    ws_internacionales.cell(row=row, column=9).number_format = '$#,##0.00'
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_internacionales.column_dimensions[col].width = 16
    
    # ===== HOJA 4: RESUMEN =====
    ws_resumen = wb.create_sheet("Resumen")
    
    ws_resumen['A1'] = "RESUMEN DE MOVIMIENTOS"
    ws_resumen['A1'].font = Font(bold=True, size=16)
    ws_resumen.merge_cells('A1:C1')
    
    row = 3
    ws_resumen[f'A{row}'] = "PERIODO:"
    ws_resumen[f'B{row}'] = f"{titulo_periodo.replace('_', ' ')}"
    ws_resumen[f'A{row}'].font = Font(bold=True)
    
    row += 2
    ws_resumen[f'A{row}'] = "Categoría"
    ws_resumen[f'B{row}'] = "Cantidad Abonos"
    ws_resumen[f'C{row}'] = "Total"
    for col in ['A', 'B', 'C']:
        cell = ws_resumen[f'{col}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row += 1
    resumen_data = [
        ["Riviera Maya", len(abonos_riviera), total_riviera],
        ["Viajes Nacionales", len(abonos_nacionales), total_nacionales],
        ["", "", ""],
        ["TOTAL MXN", len(abonos_riviera) + len(abonos_nacionales), total_riviera + total_nacionales],
        ["", "", ""],
        ["Viajes Internacionales (USD)", len(abonos_internacionales), total_internacionales_usd]
    ]
    
    for data in resumen_data:
        ws_resumen[f'A{row}'] = data[0]
        ws_resumen[f'B{row}'] = data[1]
        ws_resumen[f'C{row}'] = data[2]
        
        if "TOTAL" in str(data[0]) or row == row + 3:
            ws_resumen[f'A{row}'].font = Font(bold=True)
            ws_resumen[f'B{row}'].font = Font(bold=True)
            ws_resumen[f'C{row}'].font = Font(bold=True)
        
        if isinstance(data[2], (int, float)) and data[2] > 0:
            ws_resumen[f'C{row}'].number_format = '$#,##0.00'
        
        for col in ['A', 'B', 'C']:
            ws_resumen[f'{col}{row}'].border = border
        
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 20
    
    # Guardar archivo
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"Movimientos_{titulo_periodo}_{fecha_actual}.xlsx"
    
    wb.save(nombre_archivo)
    
    print(f"\n✅ Reporte exportado exitosamente:")
    print(f"📁 {nombre_archivo}")
    
    # Mostrar resumen
    print(f"\n📊 RESUMEN:")
    print(f"   Total abonos: {len(abonos_riviera) + len(abonos_nacionales) + len(abonos_internacionales)}")
    print(f"   Riviera Maya: ${total_riviera:,.2f}")
    print(f"   Viajes Nacionales: ${total_nacionales:,.2f}")
    print(f"   Viajes Internacionales: ${total_internacionales_usd:,.2f} USD")
    
    conexion.close()
    
    return nombre_archivo


def menu_reportes_movimientos():
    """Menú principal de reportes de movimientos"""
    while True:
        print("\n" + "="*70)
        print("📊 REPORTES DE MOVIMIENTOS")
        print("="*70)
        print("\n1. Ver reporte en consola")
        print("2. Exportar reporte a Excel")
        print("3. Volver")
        
        opcion = input("\nSelecciona una opción: ").strip()
        
        if opcion == "1":
            reporte_movimientos_semanal_consola()
        elif opcion == "2":
            archivo = exportar_movimientos_excel()
            if archivo:
                print(f"\n💡 Archivo guardado en la carpeta actual")
        elif opcion == "3":
            break
        else:
            print("❌ Opción inválida.")


if __name__ == "__main__":
    menu_reportes_movimientos()
