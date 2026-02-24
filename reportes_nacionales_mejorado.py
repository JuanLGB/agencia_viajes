from database import conectar
from datetime import datetime
import os

def reporte_nacionales_consola():
    """Muestra reporte detallado de viajes nacionales en consola"""
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Mostrar viajes disponibles
    cursor.execute("""
        SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso, estado
        FROM viajes_nacionales
        ORDER BY fecha_salida DESC
    """)
    
    viajes = cursor.fetchall()
    
    if not viajes:
        print("\n❌ No hay viajes nacionales registrados.")
        conexion.close()
        return
    
    print("\n" + "="*70)
    print("🎫 VIAJES NACIONALES DISPONIBLES")
    print("="*70)
    
    for v in viajes:
        estado_icon = "✅" if v[5] == "ACTIVO" else "🔒"
        print(f"{estado_icon} {v[0]}. {v[1]} - {v[2]} ({v[3]} al {v[4]}) - {v[5]}")
    
    try:
        viaje_id = int(input("\nSelecciona ID del viaje para ver reporte: "))
    except ValueError:
        print("❌ ID inválido.")
        conexion.close()
        return
    
    # Obtener información del viaje
    cursor.execute("""
        SELECT nombre_viaje, destino, fecha_salida, fecha_regreso, dias, noches,
               cupos_totales, cupos_vendidos, cupos_disponibles,
               precio_persona_doble, precio_persona_triple, estado
        FROM viajes_nacionales
        WHERE id = ?
    """, (viaje_id,))
    
    viaje = cursor.fetchone()
    
    if not viaje:
        print("❌ Viaje no encontrado.")
        conexion.close()
        return
    
    # ===== ENCABEZADO =====
    print("\n" + "="*70)
    print(f"🎫 REPORTE DETALLADO - {viaje[0].upper()}")
    print("="*70)
    
    print(f"\n📅 INFORMACIÓN DEL VIAJE:")
    print(f"   Nombre: {viaje[0]}")
    print(f"   Destino: {viaje[1]}")
    print(f"   Fechas: {viaje[2]} al {viaje[3]}")
    print(f"   Duración: {viaje[4]} días / {viaje[5]} noches")
    print(f"   Estado: {viaje[11]}")
    
    print(f"\n👥 CUPOS:")
    print(f"   Total: {viaje[6]} personas")
    print(f"   Vendidos: {viaje[7]} ({(viaje[7]/viaje[6]*100):.1f}%)")
    print(f"   Disponibles: {viaje[8]}")
    
    print(f"\n💵 PRECIOS BASE:")
    print(f"   Por persona (doble): ${viaje[9]:,.2f}")
    print(f"   Por persona (triple): ${viaje[10]:,.2f}")
    
    # ===== CLIENTES Y PASAJEROS =====
    cursor.execute("""
        SELECT c.id, c.nombre_cliente, v.nombre, c.adultos, c.menores,
               c.habitaciones_doble, c.habitaciones_triple,
               c.total_pagar, c.total_abonado, c.saldo, c.estado
        FROM clientes_nacionales c
        JOIN vendedoras v ON c.vendedora_id = v.id
        WHERE c.viaje_id = ?
        ORDER BY c.nombre_cliente
    """, (viaje_id,))
    
    clientes = cursor.fetchall()
    
    if not clientes:
        print("\n\n⚠️  No hay clientes registrados en este viaje.")
        conexion.close()
        return
    
    print(f"\n\n{'='*70}")
    print(f"👥 CLIENTES REGISTRADOS ({len(clientes)})")
    print("="*70)
    
    # Totales
    total_vendido = 0
    total_abonado = 0
    total_saldo = 0
    total_hab_dobles = 0
    total_hab_triples = 0
    
    for cliente in clientes:
        cliente_id = cliente[0]
        estado_icon = "✅" if cliente[10] == "LIQUIDADO" else "⏳"
        
        print(f"\n{estado_icon} {cliente[1]} (Vendedora: {cliente[2]})")
        print(f"   Pasajeros: {cliente[3]} adultos + {cliente[4]} menores = {cliente[3] + cliente[4]} total")
        print(f"   Habitaciones: {cliente[5]} doble(s) + {cliente[6]} triple(s)")
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre_completo, tipo, habitacion_asignada
            FROM pasajeros_nacionales
            WHERE cliente_id = ?
            ORDER BY habitacion_asignada, tipo DESC
        """, (cliente_id,))
        
        pasajeros = cursor.fetchall()
        
        if pasajeros:
            print(f"\n   📋 DISTRIBUCIÓN DE HABITACIONES:")
            
            # Agrupar por habitación
            habitaciones = {}
            for p in pasajeros:
                hab = p[2] if p[2] else "Sin asignar"
                if hab not in habitaciones:
                    habitaciones[hab] = []
                habitaciones[hab].append(f"{p[0]} ({p[1]})")
            
            for hab, personas in sorted(habitaciones.items()):
                print(f"      {hab}:")
                for persona in personas:
                    print(f"         • {persona}")
        
        print(f"\n   💵 PAGOS:")
        print(f"      Total: ${cliente[7]:,.2f}")
        print(f"      Abonado: ${cliente[8]:,.2f}")
        print(f"      Saldo: ${cliente[9]:,.2f}")
        
        # Historial de abonos
        cursor.execute("""
            SELECT a.fecha, a.monto, v.nombre
            FROM abonos_nacionales a
            JOIN vendedoras v ON a.vendedora_id = v.id
            WHERE a.cliente_id = ?
            ORDER BY a.fecha
        """, (cliente_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            print(f"\n   📅 Historial de pagos:")
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
                print(f"      {fecha}: ${abono[1]:,.2f} (por {abono[2]})")
        
        print("-" * 70)
        
        # Acumular totales
        total_vendido += cliente[7]
        total_abonado += cliente[8]
        total_saldo += cliente[9]
        total_hab_dobles += cliente[5]
        total_hab_triples += cliente[6]
    
    # ===== RESUMEN GENERAL =====
    print(f"\n{'='*70}")
    print("📊 RESUMEN GENERAL")
    print("="*70)
    
    print(f"\n🏨 HABITACIONES:")
    print(f"   Dobles: {total_hab_dobles}")
    print(f"   Triples: {total_hab_triples}")
    print(f"   Total: {total_hab_dobles + total_hab_triples}")
    
    print(f"\n💰 FINANCIERO:")
    print(f"   Total vendido: ${total_vendido:,.2f}")
    print(f"   Total abonado: ${total_abonado:,.2f}")
    print(f"   Total saldo: ${total_saldo:,.2f}")
    porcentaje_cobrado = (total_abonado / total_vendido * 100) if total_vendido > 0 else 0
    print(f"   % Cobrado: {porcentaje_cobrado:.1f}%")
    
    # ===== ALERTAS =====
    print(f"\n⚠️  ALERTAS:")
    alertas = []
    
    # Clientes con saldo alto
    cursor.execute("""
        SELECT COUNT(*) FROM clientes_nacionales
        WHERE viaje_id = ? AND saldo > 5000
    """, (viaje_id,))
    
    clientes_saldo_alto = cursor.fetchone()[0]
    if clientes_saldo_alto > 0:
        alertas.append(f"   • {clientes_saldo_alto} cliente(s) con saldo > $5,000")
    
    # % Cobrado bajo
    if porcentaje_cobrado < 50 and len(clientes) > 0:
        alertas.append(f"   • ⚠️  Solo {porcentaje_cobrado:.0f}% cobrado del total")
    
    # Cupos disponibles
    if viaje[8] > 0 and viaje[11] == "ACTIVO":
        alertas.append(f"   • {viaje[8]} cupos aún disponibles para venta")
    
    if alertas:
        for alerta in alertas:
            print(alerta)
    else:
        print("   ✅ Sin alertas")
    
    print("\n" + "="*70)
    
    conexion.close()


def exportar_reporte_nacionales_excel():
    """Exporta reporte de viaje nacional a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n❌ Error: Se requiere openpyxl para exportar a Excel.")
        print("Instala con: pip install openpyxl")
        return
    
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Mostrar viajes disponibles
    cursor.execute("""
        SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso, estado
        FROM viajes_nacionales
        ORDER BY fecha_salida DESC
    """)
    
    viajes = cursor.fetchall()
    
    if not viajes:
        print("\n❌ No hay viajes nacionales registrados.")
        conexion.close()
        return
    
    print("\n" + "="*70)
    print("🎫 VIAJES NACIONALES DISPONIBLES")
    print("="*70)
    
    for v in viajes:
        estado_icon = "✅" if v[5] == "ACTIVO" else "🔒"
        print(f"{estado_icon} {v[0]}. {v[1]} - {v[2]} ({v[3]} al {v[4]}) - {v[5]}")
    
    try:
        viaje_id = int(input("\nSelecciona ID del viaje para exportar: "))
    except ValueError:
        print("❌ ID inválido.")
        conexion.close()
        return
    
    # Obtener información del viaje
    cursor.execute("""
        SELECT nombre_viaje, destino, fecha_salida, fecha_regreso, dias, noches,
               cupos_totales, cupos_vendidos, cupos_disponibles,
               precio_persona_doble, precio_persona_triple, estado
        FROM viajes_nacionales
        WHERE id = ?
    """, (viaje_id,))
    
    viaje = cursor.fetchone()
    
    if not viaje:
        print("❌ Viaje no encontrado.")
        conexion.close()
        return
    
    nombre_viaje = viaje[0]
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: INFORMACIÓN GENERAL =====
    ws_info = wb.create_sheet("Información General")
    
    ws_info['A1'] = f"REPORTE - {nombre_viaje.upper()}"
    ws_info['A1'].font = Font(bold=True, size=16)
    ws_info.merge_cells('A1:D1')
    
    row = 3
    ws_info[f'A{row}'] = "INFORMACIÓN DEL VIAJE"
    ws_info[f'A{row}'].font = title_font
    
    row += 1
    info_data = [
        ["Nombre:", nombre_viaje],
        ["Destino:", viaje[1]],
        ["Fecha salida:", viaje[2]],
        ["Fecha regreso:", viaje[3]],
        ["Duración:", f"{viaje[4]} días / {viaje[5]} noches"],
        ["Estado:", viaje[11]],
        ["", ""],
        ["CUPOS", ""],
        ["Total:", viaje[6]],
        ["Vendidos:", viaje[7]],
        ["Disponibles:", viaje[8]],
        ["% Ocupación:", f"{(viaje[7]/viaje[6]*100):.1f}%"],
        ["", ""],
        ["PRECIOS BASE", ""],
        ["Por persona (doble):", f"${viaje[9]:,.2f}"],
        ["Por persona (triple):", f"${viaje[10]:,.2f}"]
    ]
    
    for data in info_data:
        ws_info[f'A{row}'] = data[0]
        ws_info[f'B{row}'] = data[1]
        if data[0] in ["CUPOS", "PRECIOS BASE", "INFORMACIÓN DEL VIAJE"]:
            ws_info[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Ajustar anchos
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 20
    
    # ===== HOJA 2: CLIENTES Y PAGOS =====
    ws_clientes = wb.create_sheet("Clientes y Pagos")
    
    ws_clientes['A1'] = "CLIENTES Y ESTADO DE PAGOS"
    ws_clientes['A1'].font = Font(bold=True, size=14)
    ws_clientes.merge_cells('A1:K1')
    
    # Encabezados
    headers = ["ID", "Cliente", "Vendedora", "Adultos", "Menores", "Total Pax", 
               "Dobles", "Triples", "Total", "Abonado", "Saldo", "Estado"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_clientes.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener clientes
    cursor.execute("""
        SELECT c.id, c.nombre_cliente, v.nombre, c.adultos, c.menores,
               c.habitaciones_doble, c.habitaciones_triple,
               c.total_pagar, c.total_abonado, c.saldo, c.estado
        FROM clientes_nacionales c
        JOIN vendedoras v ON c.vendedora_id = v.id
        WHERE c.viaje_id = ?
        ORDER BY c.nombre_cliente
    """, (viaje_id,))
    
    clientes = cursor.fetchall()
    
    row = 4
    total_vendido = 0
    total_abonado = 0
    total_saldo = 0
    
    for cliente in clientes:
        ws_clientes.cell(row=row, column=1, value=cliente[0])
        ws_clientes.cell(row=row, column=2, value=cliente[1])
        ws_clientes.cell(row=row, column=3, value=cliente[2])
        ws_clientes.cell(row=row, column=4, value=cliente[3])
        ws_clientes.cell(row=row, column=5, value=cliente[4])
        ws_clientes.cell(row=row, column=6, value=cliente[3] + cliente[4])
        ws_clientes.cell(row=row, column=7, value=cliente[5])
        ws_clientes.cell(row=row, column=8, value=cliente[6])
        ws_clientes.cell(row=row, column=9, value=cliente[7])
        ws_clientes.cell(row=row, column=10, value=cliente[8])
        ws_clientes.cell(row=row, column=11, value=cliente[9])
        ws_clientes.cell(row=row, column=12, value=cliente[10])
        
        # Formato moneda
        for col in [9, 10, 11]:
            ws_clientes.cell(row=row, column=col).number_format = '$#,##0.00'
        
        # Borders
        for col in range(1, 13):
            ws_clientes.cell(row=row, column=col).border = border
        
        total_vendido += cliente[7]
        total_abonado += cliente[8]
        total_saldo += cliente[9]
        
        row += 1
    
    # Totales
    row += 1
    ws_clientes.cell(row=row, column=8, value="TOTALES:").font = Font(bold=True)
    ws_clientes.cell(row=row, column=9, value=total_vendido).font = Font(bold=True)
    ws_clientes.cell(row=row, column=10, value=total_abonado).font = Font(bold=True)
    ws_clientes.cell(row=row, column=11, value=total_saldo).font = Font(bold=True)
    
    for col in [9, 10, 11]:
        ws_clientes.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_clientes.column_dimensions['A'].width = 8
    ws_clientes.column_dimensions['B'].width = 30
    ws_clientes.column_dimensions['C'].width = 25
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws_clientes.column_dimensions[col].width = 10
    for col in ['I', 'J', 'K']:
        ws_clientes.column_dimensions[col].width = 15
    ws_clientes.column_dimensions['L'].width = 12
    
    # ===== HOJA 3: DISTRIBUCIÓN DE HABITACIONES =====
    ws_hab = wb.create_sheet("Habitaciones")
    
    ws_hab['A1'] = "DISTRIBUCIÓN DE HABITACIONES Y PASAJEROS"
    ws_hab['A1'].font = Font(bold=True, size=14)
    ws_hab.merge_cells('A1:E1')
    
    # Encabezados
    headers_hab = ["Cliente", "Habitación", "Pasajero", "Tipo", "Total Pax"]
    
    for col, header in enumerate(headers_hab, start=1):
        cell = ws_hab.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_cliente = cliente[1]
        total_pax = cliente[3] + cliente[4]
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre_completo, tipo, habitacion_asignada
            FROM pasajeros_nacionales
            WHERE cliente_id = ?
            ORDER BY habitacion_asignada, tipo DESC
        """, (cliente_id,))
        
        pasajeros = cursor.fetchall()
        
        primera_fila = True
        for pasajero in pasajeros:
            ws_hab.cell(row=row, column=1, value=nombre_cliente if primera_fila else "")
            ws_hab.cell(row=row, column=2, value=pasajero[2] if pasajero[2] else "Sin asignar")
            ws_hab.cell(row=row, column=3, value=pasajero[0])
            ws_hab.cell(row=row, column=4, value=pasajero[1])
            ws_hab.cell(row=row, column=5, value=total_pax if primera_fila else "")
            
            for col in range(1, 6):
                ws_hab.cell(row=row, column=col).border = border
            
            primera_fila = False
            row += 1
        
        row += 1
    
    # Ajustar anchos
    ws_hab.column_dimensions['A'].width = 30
    ws_hab.column_dimensions['B'].width = 20
    ws_hab.column_dimensions['C'].width = 35
    ws_hab.column_dimensions['D'].width = 12
    ws_hab.column_dimensions['E'].width = 12
    
    # ===== HOJA 4: HISTORIAL DE PAGOS =====
    ws_pagos = wb.create_sheet("Historial de Pagos")
    
    ws_pagos['A1'] = "HISTORIAL DETALLADO DE PAGOS"
    ws_pagos['A1'].font = Font(bold=True, size=14)
    ws_pagos.merge_cells('A1:E1')
    
    # Encabezados
    headers_pagos = ["Cliente", "Fecha", "Monto", "Vendedora", "Saldo Actual"]
    
    for col, header in enumerate(headers_pagos, start=1):
        cell = ws_pagos.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_cliente = cliente[1]
        saldo_actual = cliente[9]
        
        # Obtener abonos
        cursor.execute("""
            SELECT a.fecha, a.monto, v.nombre
            FROM abonos_nacionales a
            JOIN vendedoras v ON a.vendedora_id = v.id
            WHERE a.cliente_id = ?
            ORDER BY a.fecha
        """, (cliente_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            primera_fila = True
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
                
                ws_pagos.cell(row=row, column=1, value=nombre_cliente if primera_fila else "")
                ws_pagos.cell(row=row, column=2, value=fecha)
                ws_pagos.cell(row=row, column=3, value=abono[1])
                ws_pagos.cell(row=row, column=4, value=abono[2])
                ws_pagos.cell(row=row, column=5, value=saldo_actual if primera_fila else "")
                
                # Formato
                ws_pagos.cell(row=row, column=3).number_format = '$#,##0.00'
                ws_pagos.cell(row=row, column=5).number_format = '$#,##0.00'
                
                for col in range(1, 6):
                    ws_pagos.cell(row=row, column=col).border = border
                
                primera_fila = False
                row += 1
    
    # Ajustar anchos
    ws_pagos.column_dimensions['A'].width = 30
    ws_pagos.column_dimensions['B'].width = 20
    ws_pagos.column_dimensions['C'].width = 15
    ws_pagos.column_dimensions['D'].width = 25
    ws_pagos.column_dimensions['E'].width = 15
    
    # Guardar archivo
    nombre_limpio = nombre_viaje.replace(" ", "_").replace("/", "-")
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"Reporte_{nombre_limpio}_{fecha_actual}.xlsx"
    
    wb.save(nombre_archivo)
    
    print(f"\n✅ Reporte exportado exitosamente:")
    print(f"📁 {nombre_archivo}")
    
    # Mostrar resumen
    print(f"\n📊 RESUMEN:")
    print(f"   Clientes: {len(clientes)}")
    print(f"   Total vendido: ${total_vendido:,.2f}")
    print(f"   Total abonado: ${total_abonado:,.2f}")
    print(f"   Saldo pendiente: ${total_saldo:,.2f}")
    
    conexion.close()
    
    return nombre_archivo


def menu_reportes_nacionales():
    """Menú principal de reportes viajes nacionales"""
    while True:
        print("\n" + "="*70)
        print("📊 REPORTES - VIAJES NACIONALES")
        print("="*70)
        print("\n1. Ver reporte en consola")
        print("2. Exportar reporte a Excel")
        print("3. Volver")
        
        opcion = input("\nSelecciona una opción: ").strip()
        
        if opcion == "1":
            reporte_nacionales_consola()
        elif opcion == "2":
            archivo = exportar_reporte_nacionales_excel()
            if archivo:
                print(f"\n💡 Archivo guardado en la carpeta actual")
        elif opcion == "3":
            break
        else:
            print("❌ Opción inválida.")


if __name__ == "__main__":
    menu_reportes_nacionales()
