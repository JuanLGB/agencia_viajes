from database import conectar
from datetime import datetime
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference


def crear_reporte_excel_mensual(mes, anio):
    """Genera reporte financiero mensual en Excel"""
    
    from reportes_financieros import (
        obtener_ingresos_riviera,
        obtener_ingresos_nacionales,
        obtener_ingresos_internacionales,
        obtener_gastos_operativos
    )
    
    mes_nombre = calendar.month_name[mes]
    
    # Obtener datos
    riviera = obtener_ingresos_riviera(mes, anio)
    nacionales = obtener_ingresos_nacionales(mes, anio)
    internacionales_usd = obtener_ingresos_internacionales(mes, anio)
    tipo_cambio = 17.0
    internacionales_mxn = internacionales_usd * tipo_cambio
    
    total_ingresos = riviera['total'] + nacionales + internacionales_mxn
    
    gastos_data = obtener_gastos_operativos(mes, anio)
    total_gastos = gastos_data['total']
    
    utilidad_neta = total_ingresos - total_gastos
    margen = (utilidad_neta / total_ingresos * 100) if total_ingresos > 0 else 0
    
    # Crear workbook
    wb = Workbook()
    
    # ==== HOJA 1: RESUMEN EJECUTIVO ====
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    
    # Estilos
    titulo_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    titulo_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    
    subtitulo_font = Font(name='Arial', size=12, bold=True)
    subtitulo_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    total_font = Font(name='Arial', size=11, bold=True)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Título principal
    ws1['A1'] = f'REPORTE FINANCIERO - {mes_nombre.upper()} {anio}'
    ws1['A1'].font = titulo_font
    ws1['A1'].fill = titulo_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.merge_cells('A1:D1')
    ws1.row_dimensions[1].height = 30
    
    # Fecha de generación
    ws1['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(size=9, italic=True)
    ws1.merge_cells('A2:D2')
    
    # INGRESOS
    ws1['A4'] = '💰 INGRESOS'
    ws1['A4'].font = subtitulo_font
    ws1['A4'].fill = subtitulo_fill
    ws1.merge_cells('A4:D4')
    
    row = 5
    ws1[f'A{row}'] = 'Concepto'
    ws1[f'B{row}'] = 'Monto (MXN)'
    for col in ['A', 'B']:
        ws1[f'{col}{row}'].font = header_font
        ws1[f'{col}{row}'].fill = header_fill
        ws1[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    ws1[f'A{row}'] = 'Riviera Maya - General'
    ws1[f'B{row}'] = riviera['general']
    ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = 'Riviera Maya - Bloqueos'
    ws1[f'B{row}'] = riviera['bloqueos']
    ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = 'Riviera Maya - Grupos'
    ws1[f'B{row}'] = riviera['grupos']
    ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = 'Viajes Nacionales'
    ws1[f'B{row}'] = nacionales
    ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = f'Viajes Internacionales (${internacionales_usd:,.2f} USD)'
    ws1[f'B{row}'] = internacionales_mxn
    ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = 'TOTAL INGRESOS'
    ws1[f'B{row}'] = total_ingresos
    ws1[f'B{row}'].number_format = '$#,##0.00'
    ws1[f'A{row}'].font = total_font
    ws1[f'B{row}'].font = total_font
    ws1[f'A{row}'].fill = total_fill
    ws1[f'B{row}'].fill = total_fill
    
    # GASTOS
    row += 2
    ws1[f'A{row}'] = '💸 GASTOS OPERATIVOS'
    ws1[f'A{row}'].font = subtitulo_font
    ws1[f'A{row}'].fill = subtitulo_fill
    ws1.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws1[f'A{row}'] = 'Categoría'
    ws1[f'B{row}'] = 'Monto (MXN)'
    for col in ['A', 'B']:
        ws1[f'{col}{row}'].font = header_font
        ws1[f'{col}{row}'].fill = header_fill
        ws1[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    if gastos_data['categorias']:
        for cat in gastos_data['categorias']:
            row += 1
            ws1[f'A{row}'] = cat[0]
            ws1[f'B{row}'] = cat[1]
            ws1[f'B{row}'].number_format = '$#,##0.00'
    
    row += 1
    ws1[f'A{row}'] = 'TOTAL GASTOS'
    ws1[f'B{row}'] = total_gastos
    ws1[f'B{row}'].number_format = '$#,##0.00'
    ws1[f'A{row}'].font = total_font
    ws1[f'B{row}'].font = total_font
    ws1[f'A{row}'].fill = total_fill
    ws1[f'B{row}'].fill = total_fill
    
    # UTILIDAD
    row += 2
    ws1[f'A{row}'] = '💵 UTILIDAD NETA'
    ws1[f'B{row}'] = utilidad_neta
    ws1[f'B{row}'].number_format = '$#,##0.00'
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="006600")
    ws1[f'B{row}'].font = Font(size=14, bold=True, color="006600")
    ws1[f'A{row}'].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    ws1[f'B{row}'].fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    
    row += 1
    ws1[f'A{row}'] = 'Margen de Utilidad'
    ws1[f'B{row}'] = margen / 100
    ws1[f'B{row}'].number_format = '0.00%'
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'].font = Font(bold=True)
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20
    
    # Guardar archivo
    nombre_archivo = f'Reporte_Financiero_{mes_nombre}_{anio}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    try:
        wb.save(nombre_archivo)
        print(f"\n✅ Reporte Excel generado: {nombre_archivo}")
        return nombre_archivo
    except Exception as e:
        print(f"\n❌ Error al generar Excel: {e}")
        return None


def crear_reporte_excel_anual(anio):
    """Genera reporte financiero anual en Excel"""
    
    from reportes_financieros import (
        obtener_ingresos_riviera,
        obtener_ingresos_nacionales,
        obtener_ingresos_internacionales,
        obtener_gastos_operativos
    )
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumen Anual"
    
    # Estilos
    titulo_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    titulo_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Título
    ws1['A1'] = f'REPORTE FINANCIERO ANUAL - {anio}'
    ws1['A1'].font = titulo_font
    ws1['A1'].fill = titulo_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.merge_cells('A1:E1')
    ws1.row_dimensions[1].height = 30
    
    # Fecha
    ws1['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(size=9, italic=True)
    ws1.merge_cells('A2:E2')
    
    # Headers
    ws1['A4'] = 'Mes'
    ws1['B4'] = 'Ingresos (MXN)'
    ws1['C4'] = 'Gastos (MXN)'
    ws1['D4'] = 'Utilidad (MXN)'
    ws1['E4'] = 'Margen %'
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws1[f'{col}4'].font = header_font
        ws1[f'{col}4'].fill = header_fill
        ws1[f'{col}4'].alignment = Alignment(horizontal='center')
    
    # Datos por mes
    row = 5
    total_ingresos = 0
    total_gastos = 0
    
    for mes in range(1, 13):
        riviera = obtener_ingresos_riviera(mes, anio)
        nacionales = obtener_ingresos_nacionales(mes, anio)
        internacionales_usd = obtener_ingresos_internacionales(mes, anio)
        internacionales_mxn = internacionales_usd * 17.0
        
        ingresos = riviera['total'] + nacionales + internacionales_mxn
        gastos_data = obtener_gastos_operativos(mes, anio)
        gastos = gastos_data['total']
        utilidad = ingresos - gastos
        margen = (utilidad / ingresos * 100) if ingresos > 0 else 0
        
        ws1[f'A{row}'] = calendar.month_name[mes]
        ws1[f'B{row}'] = ingresos
        ws1[f'C{row}'] = gastos
        ws1[f'D{row}'] = utilidad
        ws1[f'E{row}'] = margen / 100
        
        ws1[f'B{row}'].number_format = '$#,##0.00'
        ws1[f'C{row}'].number_format = '$#,##0.00'
        ws1[f'D{row}'].number_format = '$#,##0.00'
        ws1[f'E{row}'].number_format = '0.00%'
        
        total_ingresos += ingresos
        total_gastos += gastos
        row += 1
    
    # Totales
    utilidad_anual = total_ingresos - total_gastos
    margen_anual = (utilidad_anual / total_ingresos * 100) if total_ingresos > 0 else 0
    
    ws1[f'A{row}'] = 'TOTAL'
    ws1[f'B{row}'] = total_ingresos
    ws1[f'C{row}'] = total_gastos
    ws1[f'D{row}'] = utilidad_anual
    ws1[f'E{row}'] = margen_anual / 100
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws1[f'{col}{row}'].font = Font(bold=True)
        ws1[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    ws1[f'B{row}'].number_format = '$#,##0.00'
    ws1[f'C{row}'].number_format = '$#,##0.00'
    ws1[f'D{row}'].number_format = '$#,##0.00'
    ws1[f'E{row}'].number_format = '0.00%'
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 12
    
    # Guardar
    nombre_archivo = f'Reporte_Financiero_Anual_{anio}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    try:
        wb.save(nombre_archivo)
        print(f"\n✅ Reporte Excel generado: {nombre_archivo}")
        return nombre_archivo
    except Exception as e:
        print(f"\n❌ Error al generar Excel: {e}")
        return None


def menu_reportes_excel():
    """Menú de reportes Excel"""
    print("\n" + "="*60)
    print("📊 EXPORTAR REPORTES A EXCEL")
    print("="*60)
    print("\n1. Reporte mensual")
    print("2. Reporte anual")
    print("3. Volver")
    
    opcion = input("\nSelecciona una opción: ").strip()
    
    if opcion == "1":
        try:
            mes = int(input("\nMes (1-12): "))
            anio = int(input("Año: "))
            crear_reporte_excel_mensual(mes, anio)
        except ValueError:
            print("❌ Valores inválidos.")
    
    elif opcion == "2":
        try:
            anio = int(input("\nAño: "))
            crear_reporte_excel_anual(anio)
        except ValueError:
            print("❌ Año inválido.")
    
    elif opcion == "3":
        return
    else:
        print("❌ Opción inválida.")


if __name__ == "__main__":
    menu_reportes_excel()
