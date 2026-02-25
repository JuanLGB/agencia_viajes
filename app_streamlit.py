import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go
import urllib.request
import re
import json
import io
import os

# Módulo de transferencias
try:
    from transferencias import mostrar_pagina_transferencias
    TRANSFERENCIAS_DISPONIBLE = True
except ImportError:
    TRANSFERENCIAS_DISPONIBLE = False

# ── Generador de recibos Turismar ─────────────────────────────────────────
try:
    from generar_recibo import generar_recibo_pdf
    RECIBOS_DISPONIBLES = True
except ImportError:
    RECIBOS_DISPONIBLES = False

# ── Generador de cupones Turismar ─────────────────────────────────────────
try:
    from generar_cupon import generar_cupon_pdf
    CUPONES_DISPONIBLES = True
except ImportError:
    CUPONES_DISPONIBLES = False

# Rutas a los assets del recibo (misma carpeta que app_streamlit.py)
_DIR   = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else "."
LOGO_PATH = os.path.join(_DIR, "logo_turismar_clean.png")
NINA_PATH  = os.path.join(_DIR, "nina_turismar.png")
SELLO_PATH = os.path.join(_DIR, "sello_abono.jpg")

# Contador de recibo global (usa SQLite para persistencia)
def _siguiente_num_recibo() -> int:
    """Obtiene y avanza el correlativo de recibos."""
    conn = sqlite3.connect(DB_NAME, check_same_thread=False)
    cur  = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS config_recibos (
            clave TEXT PRIMARY KEY,
            valor INTEGER DEFAULT 1
        )
    """)
    cur.execute("INSERT OR IGNORE INTO config_recibos (clave, valor) VALUES ('num_recibo', 1)")
    cur.execute("SELECT valor FROM config_recibos WHERE clave='num_recibo'")
    num = cur.fetchone()[0]
    cur.execute("UPDATE config_recibos SET valor=? WHERE clave='num_recibo'", (num + 1,))
    conn.commit()
    conn.close()
    return num

def _boton_recibo(
    numero, fecha_str, cliente, monto,
    concepto, forma_pago, agente,
    key_suffix="",
    total_viaje=None, pagado_acumulado=None, nuevo_saldo=None,
):
    """
    Muestra el botón de descarga del recibo PDF.
    Llama a generar_recibo_pdf y crea el download_button de Streamlit.
    """
    if not RECIBOS_DISPONIBLES:
        st.caption("⚠️ generar_recibo.py no encontrado junto a app_streamlit.py")
        return
    try:
        pdf_bytes = generar_recibo_pdf(
            numero           = numero,
            fecha            = fecha_str,
            recibide         = cliente,
            cantidad         = monto,
            concepto         = concepto,
            forma_pago       = forma_pago,
            agente           = agente,
            logo_path        = LOGO_PATH,
            nina_path        = NINA_PATH,
            sello_path       = SELLO_PATH,
            total_viaje      = total_viaje,
            pagado_acumulado = pagado_acumulado,
            nuevo_saldo      = nuevo_saldo,
        )
        st.download_button(
            label    = f"🧾 Descargar Recibo No. {str(numero).zfill(3)}",
            data     = pdf_bytes,
            file_name= f"recibo_{str(numero).zfill(3)}_{cliente[:20].replace(' ','_')}.pdf",
            mime     = "application/pdf",
            key      = f"rec_{numero}_{key_suffix}",
            use_container_width=True,
        )
    except Exception as e:
        st.warning(f"⚠️ No se pudo generar el recibo: {e}")

# Configuración de la página
st.set_page_config(
    page_title="Sistema Agencia de Viajes - Riviera Maya",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilos CSS personalizados
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&family=DM+Serif+Display&display=swap');

    html, body, [class*="css"] { font-family: 'Outfit', sans-serif; }
    .main { padding-top: 1.5rem; }
    .block-container { max-width: 1200px; padding: 1rem 2rem 3rem; }

    h1 {
        font-family: 'DM Serif Display', serif;
        color: #0D1B2A;
        font-size: 2rem;
        border-bottom: 3px solid #0077B6;
        padding-bottom: 0.4rem;
        margin-bottom: 1.2rem;
    }
    h2, h3 { color: #0D1B2A; font-weight: 600; }

    [data-testid="metric-container"] {
        background: linear-gradient(135deg, #f8fbff 0%, #eef5ff 100%);
        border: 1px solid #cce3ff;
        border-radius: 12px;
        padding: 16px 20px;
        box-shadow: 0 2px 8px rgba(0,119,182,0.08);
    }
    [data-testid="metric-container"] label {
        font-size: 0.78rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        color: #0077B6 !important;
    }
    [data-testid="metric-container"] [data-testid="metric-value"] {
        font-family: 'DM Serif Display', serif;
        font-size: 1.6rem;
        color: #0D1B2A;
    }

    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0077B6 0%, #00B4D8 100%);
        color: white;
        border: none;
        border-radius: 10px;
        font-weight: 600;
        font-size: 0.95rem;
        padding: 0.6rem 1.4rem;
        box-shadow: 0 4px 14px rgba(0,119,182,0.35);
        transition: all 0.2s ease;
    }
    .stButton > button[kind="primary"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0,119,182,0.45);
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0D1B2A 0%, #1a2e45 100%);
    }
    [data-testid="stSidebar"] .stMarkdown,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p { color: #c8d8e8 !important; }
    [data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.08);
        color: #e8f4fd;
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 8px;
        width: 100%;
        text-align: left;
        font-weight: 500;
        transition: all 0.2s;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(0,180,216,0.25);
        border-color: #00B4D8;
        color: white;
    }

    hr { border-color: #E5E7EB; margin: 1.5rem 0; }

    .success-box {
        padding: 1rem 1.2rem; border-radius: 10px;
        background: linear-gradient(135deg, #d4edda, #c3e6cb);
        border-left: 4px solid #28a745; color: #155724; font-weight: 500;
    }
    .warning-box {
        padding: 1rem 1.2rem; border-radius: 10px;
        background: linear-gradient(135deg, #fff3cd, #ffeaa7);
        border-left: 4px solid #FFB703; color: #856404; font-weight: 500;
    }
    .error-box {
        padding: 1rem 1.2rem; border-radius: 10px;
        background: linear-gradient(135deg, #f8d7da, #f5c6cb);
        border-left: 4px solid #FF6B6B; color: #721c24; font-weight: 500;
    }
    .card-bloqueo {
        background: linear-gradient(135deg, #fff8e1, #fff3cd);
        border-left: 4px solid #FFB703;
        border-radius: 10px; padding: 14px 20px; margin: 8px 0;
    }
    .card-grupo {
        background: linear-gradient(135deg, #f0fff4, #e6f9ef);
        border-left: 4px solid #2D6A4F;
        border-radius: 10px; padding: 14px 20px; margin: 8px 0;
    }
    .card-general {
        background: linear-gradient(135deg, #f0f9ff, #e8f4fd);
        border-left: 4px solid #0077B6;
        border-radius: 10px; padding: 14px 20px; margin: 8px 0;
    }
    </style>
""", unsafe_allow_html=True)

# Base de datos
DB_NAME = "agencia.db"

def obtener_tipo_cambio_megatravel():
    """
    Obtiene el tipo de cambio USD/MXN desde Megatravel.
    Si es sábado o domingo usa tipo de cambio del viernes.
    Retorna (tipo_cambio, fuente) o (None, None) si falla.
    """
    try:
        from datetime import datetime
        dia = datetime.now().weekday()  # 5=Sab, 6=Dom
        url = "https://www.megatravel.com.mx/"
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=8) as resp:
            html = resp.read().decode("utf-8")
            # Patron principal
            match = re.search(r"Tipo de\s+Cambio\s+([\d.]+)", html)
            if not match:
                # Patron alternativo: buscar número de 2 dígitos punto 2 dígitos cerca de "Cambio"
                match = re.search(r"[Cc]ambio[^\d]*(\d{2}\.\d{2,4})", html)
            if match:
                tc = float(match.group(1))
                if tc < 10 or tc > 30:   # Rango razonable MXN/USD
                    return None, None
                fuente = "Megatravel"
                if dia == 5:
                    fuente += " (viernes — hoy es sábado)"
                elif dia == 6:
                    fuente += " (viernes — hoy es domingo)"
                return tc, fuente
    except Exception:
        pass
    return None, None


def obtener_tipo_cambio_api():
    """API de respaldo exchangerate-api"""
    try:
        url = "https://api.exchangerate-api.com/v4/latest/USD"
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            tc = data["rates"]["MXN"]
            return tc, "exchangerate-api.com"
    except Exception:
        pass
    return None, None


def obtener_tipo_cambio():
    """
    Intenta obtener el tipo de cambio USD/MXN.
    Orden: Megatravel → API respaldo → None
    Retorna (tipo_cambio_float, fuente_str) o (None, None)
    """
    tc, fuente = obtener_tipo_cambio_megatravel()
    if tc:
        return tc, fuente
    tc, fuente = obtener_tipo_cambio_api()
    if tc:
        return tc, fuente
    return None, None



def conectar_db():
    """Conecta a la base de datos"""
    return sqlite3.connect(DB_NAME, check_same_thread=False)


def inicializar_base_datos():
    """Crea todas las tablas si no existen (incluyendo grupos y bloqueos)"""
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS vendedoras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            usuario TEXT,
            activa INTEGER DEFAULT 1,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS operadores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            contacto TEXT,
            notas TEXT,
            activo INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS hoteles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            activo INTEGER DEFAULT 1,
            veces_usado INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente TEXT,
            celular_responsable TEXT,
            tipo_venta TEXT DEFAULT 'General',
            destino TEXT,
            fecha_inicio TEXT,
            fecha_fin TEXT,
            noches INTEGER DEFAULT 0,
            adultos INTEGER DEFAULT 0,
            menores INTEGER DEFAULT 0,
            tipo_habitacion TEXT,
            precio_adulto REAL DEFAULT 0,
            precio_menor REAL DEFAULT 0,
            precio_total REAL DEFAULT 0,
            porcentaje_ganancia REAL DEFAULT 0,
            ganancia REAL DEFAULT 0,
            costo_mayorista REAL DEFAULT 0,
            pagado REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            estado TEXT DEFAULT 'ACTIVO',
            vendedora_id INTEGER,
            usuario_id INTEGER,
            es_bloqueo INTEGER DEFAULT 0,
            bloqueo_id INTEGER,
            es_grupo INTEGER DEFAULT 0,
            grupo_id INTEGER,
            operador TEXT,
            comision_vendedora REAL DEFAULT 0,
            comision_pagada INTEGER DEFAULT 0,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS abonos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER,
            monto REAL,
            fecha TEXT,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS pasajeros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER,
            nombre TEXT,
            tipo TEXT DEFAULT 'ADULTO'
        );
        CREATE TABLE IF NOT EXISTS bloqueos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hotel TEXT,
            operador TEXT,
            celular_responsable TEXT,
            fecha_inicio TEXT,
            fecha_fin TEXT,
            noches INTEGER DEFAULT 0,
            habitaciones_totales INTEGER DEFAULT 0,
            habitaciones_vendidas INTEGER DEFAULT 0,
            habitaciones_disponibles INTEGER DEFAULT 0,
            precio_noche_doble REAL DEFAULT 0,
            precio_noche_triple REAL DEFAULT 0,
            precio_noche_cuadruple REAL DEFAULT 0,
            precio_menor_doble REAL DEFAULT 0,
            precio_menor_triple REAL DEFAULT 0,
            precio_menor_cuadruple REAL DEFAULT 0,
            costo_real REAL DEFAULT 0,
            estado TEXT DEFAULT 'ACTIVO',
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS grupos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_grupo TEXT NOT NULL,
            operador TEXT,
            hotel TEXT,
            fecha_inicio TEXT,
            fecha_fin TEXT,
            noches INTEGER DEFAULT 0,
            habitaciones_totales INTEGER DEFAULT 0,
            habitaciones_vendidas INTEGER DEFAULT 0,
            habitaciones_disponibles INTEGER DEFAULT 0,
            precio_noche_doble REAL DEFAULT 0,
            precio_noche_triple REAL DEFAULT 0,
            precio_noche_cuadruple REAL DEFAULT 0,
            precio_menor_doble REAL DEFAULT 0,
            precio_menor_triple REAL DEFAULT 0,
            precio_menor_cuadruple REAL DEFAULT 0,
            costo_real REAL DEFAULT 0,
            responsable TEXT DEFAULT 'Mamá',
            celular_responsable TEXT,
            estado TEXT DEFAULT 'ACTIVO',
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS viajes_nacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cotizacion_id INTEGER,
            nombre_viaje TEXT NOT NULL,
            destino TEXT,
            fecha_salida TEXT,
            fecha_regreso TEXT,
            dias INTEGER DEFAULT 0,
            noches INTEGER DEFAULT 0,
            cupos_totales INTEGER DEFAULT 0,
            cupos_vendidos INTEGER DEFAULT 0,
            cupos_disponibles INTEGER DEFAULT 0,
            precio_persona_doble REAL DEFAULT 0,
            precio_persona_triple REAL DEFAULT 0,
            estado TEXT DEFAULT 'ACTIVO',
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS clientes_nacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            viaje_id INTEGER,
            vendedora_id INTEGER,
            nombre_cliente TEXT,
            celular_responsable TEXT,
            adultos INTEGER DEFAULT 0,
            menores INTEGER DEFAULT 0,
            habitaciones_doble INTEGER DEFAULT 0,
            habitaciones_triple INTEGER DEFAULT 0,
            total_pagar REAL DEFAULT 0,
            total_abonado REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            ganancia REAL DEFAULT 0,
            estado TEXT DEFAULT 'ADEUDO',
            comision_pagada INTEGER DEFAULT 0,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS abonos_nacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            monto REAL,
            fecha TEXT,
            vendedora_id INTEGER,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS pasajeros_nacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            nombre_completo TEXT,
            tipo TEXT DEFAULT 'ADULTO',
            habitacion_asignada TEXT
        );
        CREATE TABLE IF NOT EXISTS cotizaciones_nacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_viaje TEXT NOT NULL,
            destino TEXT,
            fecha_salida TEXT,
            fecha_regreso TEXT,
            dias INTEGER DEFAULT 0,
            noches INTEGER DEFAULT 0,
            personas_proyectadas INTEGER DEFAULT 20,
            costo_vuelo_real REAL DEFAULT 0,
            precio_vuelo_venta REAL DEFAULT 0,
            costo_traslados_total REAL DEFAULT 0,
            precio_traslados_persona REAL DEFAULT 0,
            costo_entradas_real REAL DEFAULT 0,
            precio_entradas_venta REAL DEFAULT 0,
            precio_persona_doble REAL DEFAULT 0,
            precio_persona_triple REAL DEFAULT 0,
            inversion_total REAL DEFAULT 0,
            ganancia_proyectada REAL DEFAULT 0,
            estado TEXT DEFAULT 'BORRADOR',
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS hoteles_cotizacion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cotizacion_id INTEGER,
            nombre_hotel TEXT,
            noches INTEGER DEFAULT 0,
            costo_doble_real REAL DEFAULT 0,
            precio_doble_venta REAL DEFAULT 0,
            costo_triple_real REAL DEFAULT 0,
            precio_triple_venta REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS viajes_internacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            destino TEXT,
            fecha_salida TEXT,
            fecha_regreso TEXT,
            dias INTEGER DEFAULT 0,
            noches INTEGER DEFAULT 0,
            cupos_totales INTEGER DEFAULT 0,
            cupos_vendidos INTEGER DEFAULT 0,
            cupos_disponibles INTEGER DEFAULT 0,
            precio_adulto_doble_usd REAL DEFAULT 0,
            precio_adulto_triple_usd REAL DEFAULT 0,
            precio_menor_doble_usd REAL DEFAULT 0,
            precio_menor_triple_usd REAL DEFAULT 0,
            porcentaje_ganancia REAL DEFAULT 0,
            estado TEXT DEFAULT 'ACTIVO',
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS clientes_internacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            viaje_id INTEGER,
            vendedora_id INTEGER,
            nombre_cliente TEXT,
            adultos INTEGER DEFAULT 0,
            menores INTEGER DEFAULT 0,
            habitaciones_doble INTEGER DEFAULT 0,
            habitaciones_triple INTEGER DEFAULT 0,
            total_usd REAL DEFAULT 0,
            abonado_usd REAL DEFAULT 0,
            saldo_usd REAL DEFAULT 0,
            ganancia_usd REAL DEFAULT 0,
            estado TEXT DEFAULT 'ADEUDO',
            estado_comision TEXT DEFAULT 'PENDIENTE',
            comision_pagada INTEGER DEFAULT 0,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS abonos_internacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            monto_usd REAL,
            tipo_cambio REAL DEFAULT 0,
            monto_mxn REAL DEFAULT 0,
            fecha TEXT,
            vendedora_id INTEGER,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS pasajeros_internacionales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente_id INTEGER,
            nombre_completo TEXT,
            tipo TEXT DEFAULT 'ADULTO',
            habitacion_asignada TEXT
        );
        CREATE TABLE IF NOT EXISTS pasaportes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendedora_id INTEGER,
            cliente TEXT NOT NULL,
            celular TEXT,
            tipo TEXT DEFAULT 'Nuevo',
            fecha_cita TEXT,
            fecha_entrega_est TEXT,
            costo_oficial REAL DEFAULT 0,
            cargo_servicio REAL DEFAULT 0,
            total REAL DEFAULT 0,
            pagado REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            estado TEXT DEFAULT 'En trámite',
            notas TEXT,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS visas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendedora_id INTEGER,
            cliente TEXT NOT NULL,
            celular TEXT,
            pais_destino TEXT,
            tipo_visa TEXT DEFAULT 'Turista',
            es_familiar INTEGER DEFAULT 0,
            num_integrantes INTEGER DEFAULT 1,
            fecha_cita TEXT,
            fecha_entrega_est TEXT,
            costo_oficial REAL DEFAULT 0,
            cargo_servicio REAL DEFAULT 0,
            total REAL DEFAULT 0,
            pagado REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            estado TEXT DEFAULT 'En trámite',
            notas TEXT,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS vuelos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendedora_id INTEGER,
            pasajero TEXT NOT NULL,
            celular TEXT,
            tipo TEXT DEFAULT 'Nacional',
            aerolinea TEXT,
            origen TEXT,
            destino TEXT,
            fecha_vuelo TEXT,
            hora_vuelo TEXT,
            num_pasajeros INTEGER DEFAULT 1,
            costo_compra REAL DEFAULT 0,
            cargo_servicio REAL DEFAULT 0,
            total REAL DEFAULT 0,
            pagado REAL DEFAULT 0,
            saldo REAL DEFAULT 0,
            estado TEXT DEFAULT 'Cotizado',
            notas TEXT,
            fecha_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS abonos_pasaportes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pasaporte_id INTEGER,
            monto REAL,
            fecha TEXT,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS abonos_visas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            visa_id INTEGER,
            monto REAL,
            fecha TEXT,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS abonos_vuelos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vuelo_id INTEGER,
            monto REAL,
            fecha TEXT,
            metodo_pago TEXT DEFAULT 'Efectivo'
        );
        CREATE TABLE IF NOT EXISTS config_comisiones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT UNIQUE,
            porcentaje_comision REAL DEFAULT 10
        );
        CREATE TABLE IF NOT EXISTS historial_comisiones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendedora TEXT,
            fecha TEXT,
            tipo TEXT,
            referencia_id INTEGER,
            monto REAL,
            metodo_pago TEXT,
            nota TEXT
        );
    """)
    cursor.execute("INSERT OR IGNORE INTO config_comisiones (tipo, porcentaje_comision) VALUES ('riviera', 10)")
    cursor.execute("INSERT OR IGNORE INTO config_comisiones (tipo, porcentaje_comision) VALUES ('nacionales', 10)")
    cursor.execute("INSERT OR IGNORE INTO config_comisiones (tipo, porcentaje_comision) VALUES ('internacionales', 10)")

    migraciones = [
        "ALTER TABLE abonos ADD COLUMN metodo_pago TEXT DEFAULT 'Efectivo'",
        "ALTER TABLE clientes_nacionales ADD COLUMN ganancia REAL DEFAULT 0",
        "ALTER TABLE clientes_internacionales ADD COLUMN ganancia_usd REAL DEFAULT 0",
        "ALTER TABLE clientes_internacionales ADD COLUMN vendedora_id INTEGER DEFAULT 0",
        "ALTER TABLE abonos_nacionales ADD COLUMN metodo_pago TEXT DEFAULT 'Efectivo'",
        "ALTER TABLE abonos_internacionales ADD COLUMN metodo_pago TEXT DEFAULT 'Efectivo'",
        "ALTER TABLE abonos_internacionales ADD COLUMN monto_mxn REAL DEFAULT 0",
        "ALTER TABLE abonos_internacionales ADD COLUMN tipo_cambio REAL DEFAULT 0",
        "ALTER TABLE historial_comisiones ADD COLUMN vendedora_id INTEGER DEFAULT 0",
        "ALTER TABLE ventas ADD COLUMN celular_responsable TEXT",
        # ── Localización y datos para cupón ──
        "ALTER TABLE ventas ADD COLUMN no_localizador TEXT",
        "ALTER TABLE ventas ADD COLUMN clave_confirmacion TEXT",
        "ALTER TABLE ventas ADD COLUMN plan_alimento TEXT DEFAULT 'Todo incluido'",
        "ALTER TABLE ventas ADD COLUMN requerimientos_especiales TEXT",
        "ALTER TABLE ventas ADD COLUMN edades_menores TEXT",
        # ── Hoteles: ampliar catálogo ──
        "ALTER TABLE hoteles ADD COLUMN direccion TEXT",
        "ALTER TABLE hoteles ADD COLUMN telefono TEXT",
        "ALTER TABLE hoteles ADD COLUMN estrellas INTEGER DEFAULT 4",
        # ── Confirmación de reserva hotel ──
        "ALTER TABLE ventas ADD COLUMN reserva_confirmada INTEGER DEFAULT 0",
        "ALTER TABLE ventas ADD COLUMN reserva_confirmada_por TEXT",
        "ALTER TABLE ventas ADD COLUMN reserva_confirmada_fecha TEXT",
    ]
    for sql in migraciones:
        try:
            cursor.execute(sql)
        except Exception:
            pass

    conn.commit()
    conn.close()


def verificar_login(usuario, password):
    """Verifica las credenciales del usuario"""
    import json
    import os
    
    # Cargar usuarios desde JSON
    ruta_usuarios = "usuarios.json"
    
    if not os.path.exists(ruta_usuarios):
        return None
    
    try:
        with open(ruta_usuarios, "r", encoding="utf-8") as archivo:
            usuarios = json.load(archivo)
        
        for u in usuarios:
            if u["usuario"] == usuario and u["password"] == password:
                return {
                    "id_vendedora": u.get("id_vendedora", 1),
                    "nombre": u.get("nombre", usuario),
                    "rol": u["rol"],
                    "usuario": usuario
                }
    except:
        pass
    
    return None

# Función para obtener métricas del dashboard
def obtener_metricas_dashboard(usuario_actual):
    """Obtiene métricas generales para el dashboard"""
    conn = conectar_db()
    cursor = conn.cursor()
    
    if usuario_actual["rol"] == "VENDEDORA":
        filtro = f"AND usuario_id = {usuario_actual['id_vendedora']}"
    else:
        filtro = ""
    
    # Ventas activas Riviera
    cursor.execute(f"""
        SELECT 
            COUNT(*) as total_ventas,
            SUM(precio_total) as total_vendido,
            SUM(pagado) as total_cobrado,
            SUM(saldo) as total_saldo,
            SUM(ganancia) as ganancia_total
        FROM ventas
        WHERE estado != 'CERRADO' {filtro}
    """)
    
    riviera = cursor.fetchone()
    
    # Viajes nacionales
    cursor.execute("""
        SELECT COUNT(*) FROM viajes_nacionales WHERE estado = 'ACTIVO'
    """)
    nacionales = cursor.fetchone()[0]
    
    # Viajes internacionales
    cursor.execute("""
        SELECT COUNT(*) FROM viajes_internacionales WHERE estado = 'ACTIVO'
    """)
    internacionales = cursor.fetchone()[0]
    
    # Comisiones pendientes
    if usuario_actual["rol"] == "VENDEDORA":
        cursor.execute(f"""
            SELECT SUM(comision_vendedora)
            FROM ventas
            WHERE usuario_id = {usuario_actual['id_vendedora']} AND estado = 'LIQUIDADO'
        """)
    else:
        cursor.execute("""
            SELECT SUM(comision_vendedora)
            FROM ventas
            WHERE estado = 'LIQUIDADO'
        """)
    
    comisiones = cursor.fetchone()[0] or 0
    
    conn.close()
    
    return {
        "riviera_ventas": riviera[0] or 0,
        "riviera_vendido": riviera[1] or 0,
        "riviera_cobrado": riviera[2] or 0,
        "riviera_saldo": riviera[3] or 0,
        "riviera_ganancia": riviera[4] or 0,
        "nacionales": nacionales,
        "internacionales": internacionales,
        "comisiones": comisiones
    }


def obtener_operadores():
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, nombre FROM operadores WHERE activo = 1 ORDER BY nombre")
        ops = cursor.fetchall()
    except:
        ops = []
    conn.close()
    return ops


def obtener_vendedoras():
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre FROM vendedoras WHERE activa = 1 ORDER BY nombre")
    vendedoras = cursor.fetchall()
    conn.close()
    return vendedoras


def obtener_hoteles():
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("SELECT nombre FROM hoteles WHERE activo = 1 ORDER BY veces_usado DESC, nombre")
    hoteles = [row[0] for row in cursor.fetchall()]
    conn.close()
    return hoteles


def obtener_hoteles_completos():
    """Devuelve lista de dicts con todos los datos del hotel para el cupón."""
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT nombre,
                   COALESCE(direccion,'') AS direccion,
                   COALESCE(telefono,'') AS telefono,
                   COALESCE(estrellas, 4) AS estrellas
            FROM hoteles WHERE activo = 1
            ORDER BY veces_usado DESC, nombre
        """)
        cols = ['nombre','direccion','telefono','estrellas']
        rows = cursor.fetchall()
    except Exception:
        cursor.execute("SELECT nombre FROM hoteles WHERE activo = 1 ORDER BY nombre")
        rows  = [(r[0],'','',4) for r in cursor.fetchall()]
        cols  = ['nombre','direccion','telefono','estrellas']
    conn.close()
    return [dict(zip(cols, r)) for r in rows]


def obtener_bloqueos_disponibles():
    conn = conectar_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, hotel, fecha_inicio, fecha_fin, noches,
               habitaciones_disponibles,
               precio_noche_doble, precio_noche_triple, precio_noche_cuadruple,
               precio_menor_doble, precio_menor_triple, precio_menor_cuadruple
        FROM bloqueos
        WHERE estado = 'ACTIVO' AND habitaciones_disponibles > 0
        ORDER BY fecha_inicio
    """)
    cols = ['id','hotel','fecha_inicio','fecha_fin','noches','disponibles',
            'precio_doble','precio_triple','precio_cuadruple',
            'menor_doble','menor_triple','menor_cuadruple']
    rows = cursor.fetchall()
    conn.close()
    return [dict(zip(cols, r)) for r in rows]


def obtener_grupos_disponibles():
    """Retorna grupos activos con habitaciones disponibles"""
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, nombre_grupo, operador, hotel, fecha_inicio, fecha_fin, noches,
                   habitaciones_totales, habitaciones_vendidas, habitaciones_disponibles,
                   precio_noche_doble, precio_noche_triple, precio_noche_cuadruple,
                   precio_menor_doble, precio_menor_triple, precio_menor_cuadruple,
                   costo_real, responsable, celular_responsable, estado
            FROM grupos
            WHERE estado = 'ACTIVO' AND habitaciones_disponibles > 0
            ORDER BY fecha_inicio
        """)
        cols = ['id','nombre_grupo','operador','hotel','fecha_inicio','fecha_fin','noches',
                'habitaciones_totales','habitaciones_vendidas','habitaciones_disponibles',
                'precio_noche_doble','precio_noche_triple','precio_noche_cuadruple',
                'precio_menor_doble','precio_menor_triple','precio_menor_cuadruple',
                'costo_real','responsable','celular_responsable','estado']
        rows = cursor.fetchall()
        conn.close()
        return [dict(zip(cols, r)) for r in rows]
    except Exception:
        conn.close()
        return []


def formulario_nueva_venta():
    """Formulario completo para registrar nueva venta Riviera Maya — con Grupos y Bloqueos"""
    st.subheader("➕ Nueva Venta - Riviera Maya")
    usuario = st.session_state.usuario_actual

    # ── Tipo de venta ─────────────────────────────────────────────────────────
    st.markdown("#### 🏷️ Tipo de Venta")

    col_t1, col_t2, col_t3 = st.columns(3)
    with col_t1:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#e8f4fd,#d0e8f9);border-left:4px solid #0077B6;
        border-radius:10px;padding:14px 18px;text-align:center;">
        <div style="font-size:1.8rem">🏨</div>
        <div style="font-weight:700;color:#0077B6;font-size:0.95rem">VENTA GENERAL</div>
        <div style="font-size:0.78rem;color:#4a6fa5;margin-top:4px">Hotel libre, fechas y precios propios</div>
        </div>""", unsafe_allow_html=True)
    with col_t2:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#fff8e1,#fff0b3);border-left:4px solid #FFB703;
        border-radius:10px;padding:14px 18px;text-align:center;">
        <div style="font-size:1.8rem">📦</div>
        <div style="font-weight:700;color:#856404;font-size:0.95rem">BLOQUEO</div>
        <div style="font-size:0.78rem;color:#9a7520;margin-top:4px">Habitaciones pre-compradas al mayorista</div>
        </div>""", unsafe_allow_html=True)
    with col_t3:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#f0fff4,#d8f5e5);border-left:4px solid #2D6A4F;
        border-radius:10px;padding:14px 18px;text-align:center;">
        <div style="font-size:1.8rem">👥</div>
        <div style="font-weight:700;color:#2D6A4F;font-size:0.95rem">GRUPO</div>
        <div style="font-size:0.78rem;color:#3d8b65;margin-top:4px">Venta dentro de un grupo registrado</div>
        </div>""", unsafe_allow_html=True)

    tipo_venta = st.radio(
        "Selecciona el tipo:",
        ["🏨 Venta General", "📦 Bloqueo", "👥 Grupo"],
        horizontal=True,
        label_visibility="collapsed",
        key="rv_tipo_venta"
    )
    es_general = "General" in tipo_venta
    es_bloqueo = "Bloqueo" in tipo_venta
    es_grupo   = "Grupo"   in tipo_venta

    st.divider()

    # ── Selección de Bloqueo ──────────────────────────────────────────────────
    bloqueo_sel = None
    if es_bloqueo:
        bloqueos = obtener_bloqueos_disponibles()
        if not bloqueos:
            st.error("❌ No hay bloqueos disponibles actualmente. Registra uno en Configuración → Bloqueos.")
            return
        st.markdown("#### 📦 Seleccionar Bloqueo")
        opciones_bloqueo = {
            f"ID {b['id']} | 🏨 {b['hotel']} | {b['fecha_inicio']} → {b['fecha_fin']} | {b['noches']}n | Disp: {b['disponibles']} hab": b
            for b in bloqueos
        }
        bloqueo_label = st.selectbox("Bloqueo disponible:", list(opciones_bloqueo.keys()), key="rv_bloqueo_sel")
        bloqueo_sel   = opciones_bloqueo[bloqueo_label]
        st.markdown(f"""
        <div class="card-bloqueo">
        🏨 <b>{bloqueo_sel['hotel']}</b> &nbsp;|&nbsp;
        📅 {bloqueo_sel['fecha_inicio']} → {bloqueo_sel['fecha_fin']} &nbsp;|&nbsp;
        🌙 {bloqueo_sel['noches']} noches &nbsp;|&nbsp;
        🏠 {bloqueo_sel['disponibles']} habitaciones disponibles
        </div>""", unsafe_allow_html=True)
        st.divider()

    # ── Selección de Grupo ────────────────────────────────────────────────────
    grupo_sel = None
    if es_grupo:
        grupos = obtener_grupos_disponibles()
        if not grupos:
            st.error("❌ No hay grupos activos con habitaciones disponibles. Regístralos en Configuración → Grupos.")
            return
        st.markdown("#### 👥 Seleccionar Grupo")
        opciones_grupo = {
            f"ID {g['id']} | 👥 {g['nombre_grupo']} | 🏨 {g['hotel']} | {g['fecha_inicio']} → {g['fecha_fin']} | Disp: {g['habitaciones_disponibles']} hab": g
            for g in grupos
        }
        grupo_label = st.selectbox("Grupo disponible:", list(opciones_grupo.keys()), key="rv_grupo_sel")
        grupo_sel   = opciones_grupo[grupo_label]

        col_g1, col_g2, col_g3, col_g4 = st.columns(4)
        col_g1.metric("Hotel", grupo_sel['hotel'])
        col_g2.metric("Noches", grupo_sel['noches'])
        col_g3.metric("Hab. disponibles", grupo_sel['habitaciones_disponibles'])
        col_g4.metric("Operador", grupo_sel['operador'] or "—")

        st.markdown(f"""
        <div class="card-grupo">
        👤 <b>Responsable:</b> {grupo_sel['responsable']} &nbsp;|&nbsp;
        📱 {grupo_sel['celular_responsable']} &nbsp;|&nbsp;
        📅 {grupo_sel['fecha_inicio']} → {grupo_sel['fecha_fin']}
        </div>""", unsafe_allow_html=True)
        st.divider()

    # ── Datos del cliente ─────────────────────────────────────────────────────
    st.markdown("#### 👤 Datos del Cliente")
    col1, col2 = st.columns(2)
    with col1:
        cliente = st.text_input("Nombre del cliente *", key="rv_cliente")
    with col2:
        celular = st.text_input("Celular del cliente *", key="rv_celular")

    # ── Hotel y Operador (solo General) ──────────────────────────────────────
    if es_general:
        st.markdown("#### 🏨 Hotel y Operador")
        col1, col2 = st.columns(2)
        with col1:
            hoteles = obtener_hoteles()
            destino = st.selectbox("Hotel / Destino *", hoteles, key="rv_destino") if hoteles else st.text_input("Hotel / Destino *", key="rv_destino_txt")
        with col2:
            operadores = obtener_operadores()
            if operadores:
                ops_dict = {"Sin operador": None}
                ops_dict.update({op[1]: op[1] for op in operadores})
                operador = ops_dict[st.selectbox("Operador mayorista", list(ops_dict.keys()), key="rv_operador")]
            else:
                operador = st.text_input("Operador mayorista (opcional)", key="rv_operador_txt")
    elif es_bloqueo:
        destino  = bloqueo_sel['hotel']
        operador = None
    else:  # grupo
        destino  = grupo_sel['hotel']
        operador = grupo_sel['operador']

    # ── Fechas ────────────────────────────────────────────────────────────────
    if es_general:
        st.markdown("#### 📅 Fechas")
        col1, col2 = st.columns(2)
        with col1:
            fecha_inicio = st.date_input("Fecha de salida *", key="rv_fi")
        with col2:
            fecha_fin = st.date_input("Fecha de regreso *", key="rv_ff")
        noches = (fecha_fin - fecha_inicio).days
        if noches > 0:
            st.success(f"✅ {noches} noches")
        elif noches == 0:
            st.warning("⚠️ 0 noches — verifica las fechas")
        else:
            st.error("❌ Fecha de regreso anterior a la de salida")
    elif es_bloqueo:
        from datetime import datetime as dt
        fecha_inicio = dt.strptime(bloqueo_sel['fecha_inicio'], "%d-%m-%Y").date()
        fecha_fin    = dt.strptime(bloqueo_sel['fecha_fin'],    "%d-%m-%Y").date()
        noches = bloqueo_sel['noches']
        st.info(f"📅 {bloqueo_sel['fecha_inicio']} → {bloqueo_sel['fecha_fin']} | 🌙 {noches} noches")
    else:  # grupo
        from datetime import datetime as dt
        fecha_inicio = dt.strptime(grupo_sel['fecha_inicio'], "%d-%m-%Y").date()
        fecha_fin    = dt.strptime(grupo_sel['fecha_fin'],    "%d-%m-%Y").date()
        noches = grupo_sel['noches']
        st.info(f"📅 {grupo_sel['fecha_inicio']} → {grupo_sel['fecha_fin']} | 🌙 {noches} noches")

    # ── Pasajeros y habitación ────────────────────────────────────────────────
    st.markdown("#### 🛏️ Habitación y Pasajeros")
    col1, col2 = st.columns(2)
    with col1:
        adultos = st.selectbox("Número de adultos *", [2, 3, 4], key="rv_adultos")
    with col2:
        max_menores = {2: 2, 3: 1, 4: 0}[adultos]
        menores = st.selectbox("Número de menores", list(range(max_menores + 1)), key="rv_menores")
    tipo_habitacion = {2: "DOBLE", 3: "TRIPLE", 4: "CUÁDRUPLE"}[adultos]
    st.info(f"🛏️ Habitación: **{tipo_habitacion}**  —  {adultos} adultos{f' + {menores} menor(es)' if menores else ''}")

    # ── Precios ───────────────────────────────────────────────────────────────
    st.markdown("#### 💰 Precios")

    if es_bloqueo:
        key_hab = {2: 'precio_doble', 3: 'precio_triple', 4: 'precio_cuadruple'}[adultos]
        key_men = {2: 'menor_doble',  3: 'menor_triple',  4: 'menor_cuadruple'}[adultos]
        precio_adulto = round((bloqueo_sel[key_hab] / adultos) * noches, 2)
        precio_menor  = round(bloqueo_sel[key_men] * noches, 2)
        col1, col2 = st.columns(2)
        col1.metric("Precio por adulto (total estancia)", f"${precio_adulto:,.2f}")
        col2.metric("Precio por menor (total estancia)",  f"${precio_menor:,.2f}")

    elif es_grupo:
        key_noche_ad = {2: 'precio_noche_doble', 3: 'precio_noche_triple', 4: 'precio_noche_cuadruple'}[adultos]
        key_noche_mn = {2: 'precio_menor_doble', 3: 'precio_menor_triple', 4: 'precio_menor_cuadruple'}[adultos]
        precio_adulto = round((grupo_sel[key_noche_ad] / adultos) * noches, 2)
        precio_menor  = round(grupo_sel[key_noche_mn] * noches, 2)
        col1, col2 = st.columns(2)
        col1.metric("Precio por adulto (total estancia)", f"${precio_adulto:,.2f}")
        col2.metric("Precio por menor (total estancia)",  f"${precio_menor:,.2f}")
        st.caption(f"💡 Basado en precio/noche del grupo × {noches} noches ÷ {adultos} adultos")

    else:  # general
        col1, col2 = st.columns(2)
        with col1:
            precio_adulto = st.number_input("Precio por adulto *", min_value=0.0, step=100.0, format="%.2f", key="rv_precio_ad")
        with col2:
            precio_menor  = st.number_input("Precio por menor",   min_value=0.0, step=100.0, format="%.2f", key="rv_precio_mn")

    total_adultos = adultos * precio_adulto
    total_menores = menores * precio_menor
    precio_total  = total_adultos + total_menores

    if precio_total > 0:
        st.markdown("**📊 Desglose del total:**")
        col1, col2, col3 = st.columns(3)
        col1.metric(f"{adultos} adultos × ${precio_adulto:,.2f}", f"${total_adultos:,.2f}")
        if menores > 0:
            col2.metric(f"{menores} menor(es) × ${precio_menor:,.2f}", f"${total_menores:,.2f}")
        col3.metric("💵 TOTAL", f"${precio_total:,.2f}")

    # ── Ganancia ──────────────────────────────────────────────────────────────
    st.markdown("#### 📈 Ganancia")
    porcentaje_ganancia = st.slider("Porcentaje de ganancia (%)", 1, 50, 15, key="rv_pct_gan")
    ganancia        = round(precio_total * (porcentaje_ganancia / 100), 2)
    costo_mayorista = round(precio_total - ganancia, 2)
    col1, col2, col3 = st.columns(3)
    col1.metric("Ganancia", f"${ganancia:,.2f}")
    col2.metric("Costo mayorista", f"${costo_mayorista:,.2f}")
    col3.metric("Margen", f"{porcentaje_ganancia}%")

    # ── Abono inicial ─────────────────────────────────────────────────────────
    st.markdown("#### 💳 Abono Inicial")
    col1, col2 = st.columns(2)
    with col1:
        abono_inicial = st.number_input(
            "Abono inicial ($)", min_value=0.0,
            max_value=float(precio_total) if precio_total > 0 else 999999.0,
            step=100.0, format="%.2f", key="rv_abono"
        )
    with col2:
        metodo_pago = st.selectbox("Método de pago *",
            ["Efectivo", "Transferencia", "Tarjeta de crédito",
             "Tarjeta de débito", "Depósito bancario", "Otro"], key="rv_metodo_pago")

    saldo_inicial = precio_total - abono_inicial
    liquidado     = saldo_inicial <= 0.01

    col1, col2 = st.columns(2)
    col1.metric("Saldo restante", f"${saldo_inicial:,.2f}")
    if liquidado and precio_total > 0:
        st.success("🎉 ¡Esta venta quedará LIQUIDADA con el abono inicial!")
    elif abono_inicial > 0:
        pct_ab = (abono_inicial / precio_total * 100) if precio_total > 0 else 0
        st.info(f"📊 Adelanto: {pct_ab:.1f}% del total")

    # ── Vendedora ─────────────────────────────────────────────────────────────
    st.markdown("#### 🧑‍💼 Vendedora")
    if usuario["rol"] == "ADMIN":
        vendedoras = obtener_vendedoras()
        if not vendedoras:
            st.error("❌ No hay vendedoras registradas.")
            return
        vend_dict    = {v[1]: v[0] for v in vendedoras}
        id_vendedora = vend_dict[st.selectbox("Vendedora *", list(vend_dict.keys()), key="rv_vendedora")]
    else:
        id_vendedora = usuario["id_vendedora"]
        st.info(f"Vendedora: **{usuario['nombre']}**")

    # ── Pasajeros ─────────────────────────────────────────────────────────────
    st.markdown("#### 🧳 Nombres de Pasajeros")
    nombres_pasajeros = []
    cols_pas = st.columns(2)
    for i in range(adultos):
        with cols_pas[i % 2]:
            n = st.text_input(f"Adulto {i+1}", key=f"rv_adulto_{i}")
            nombres_pasajeros.append({"nombre": n, "tipo": "ADULTO"})
    for i in range(menores):
        with cols_pas[i % 2]:
            n = st.text_input(f"Menor {i+1}", key=f"rv_menor_{i}")
            nombres_pasajeros.append({"nombre": n, "tipo": "MENOR"})

    # ── Localización y datos para cupón ──────────────────────────────────
    st.markdown("#### 🔑 Localización y Datos del Cupón")
    col1, col2 = st.columns(2)
    with col1:
        no_localizador = st.text_input(
            "No. de Localizador",
            placeholder="Ej: ABC123456",
            help="Número de referencia que da el mayorista al reservar",
            key="rv_localizador"
        )
    with col2:
        clave_confirmacion = st.text_input(
            "Clave de Confirmación",
            placeholder="Ej: 69-7543060",
            help="Se obtiene cuando el viaje está pagado al 100%. Puedes dejarla vacía y agregarla después.",
            key="rv_clave_confirm"
        )
    if clave_confirmacion:
        st.success("✅ Con clave de confirmación — se podrá generar el cupón al guardar.")
    elif no_localizador:
        st.info("💡 La clave de confirmación puede agregarse después cuando el cliente liquide.")

    col1, col2 = st.columns(2)
    with col1:
        plan_alimento = st.selectbox(
            "Plan de alimento",
            ["Todo incluido", "Solo desayuno", "Media pensión", "Solo alojamiento", "Otro"],
            key="rv_plan_alimento"
        )
    with col2:
        edades_menores_txt = st.text_input(
            "Edades de menores",
            placeholder="Ej: 5 y 8 años",
            key="rv_edades_menores"
        ) if menores > 0 else ""

    requerimientos_esp = st.text_area(
        "Requerimientos especiales (opcional)",
        placeholder="Ej: Luna de miel, solicito habitación planta baja, etc.",
        height=70,
        key="rv_requerimientos"
    )

    st.divider()

    # ── Guardar ───────────────────────────────────────────────────────────────
    if st.button("💾 Registrar Venta", type="primary", use_container_width=True, key="rv_guardar"):
        errores = []
        if not cliente.strip():   errores.append("El nombre del cliente es obligatorio.")
        if not celular.strip():   errores.append("El celular del cliente es obligatorio.")
        if noches <= 0:           errores.append("Las fechas no son válidas.")
        if precio_total <= 0:     errores.append("El precio total debe ser mayor a 0.")
        if any(p["nombre"].strip() == "" for p in nombres_pasajeros):
            errores.append("Todos los nombres de pasajeros son obligatorios.")

        if errores:
            for e in errores:
                st.error(f"❌ {e}")
            return

        try:
            conn   = conectar_db()
            cursor = conn.cursor()
            fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            fecha_i_str    = fecha_inicio.strftime("%d-%m-%Y")
            fecha_f_str    = fecha_fin.strftime("%d-%m-%Y")

            if es_bloqueo:
                tipo_venta_db = "Bloqueo"
                bloqueo_id_db = bloqueo_sel['id']
                grupo_id_db   = None
                es_bloqueo_db = 1
                es_grupo_db   = 0
            elif es_grupo:
                tipo_venta_db = "Grupo"
                bloqueo_id_db = None
                grupo_id_db   = grupo_sel['id']
                es_bloqueo_db = 0
                es_grupo_db   = 1
            else:
                tipo_venta_db = "General"
                bloqueo_id_db = None
                grupo_id_db   = None
                es_bloqueo_db = 0
                es_grupo_db   = 0

            estado_inicial = "LIQUIDADO" if liquidado else "ACTIVO"

            cursor.execute("""
                INSERT INTO ventas (
                    cliente, celular_responsable, tipo_venta, destino,
                    fecha_inicio, fecha_fin, noches,
                    adultos, menores, tipo_habitacion,
                    precio_adulto, precio_menor, precio_total,
                    porcentaje_ganancia, ganancia, costo_mayorista,
                    pagado, saldo, estado, vendedora_id, usuario_id,
                    es_bloqueo, bloqueo_id, es_grupo, grupo_id,
                    operador, no_localizador, clave_confirmacion,
                    plan_alimento, edades_menores, requerimientos_especiales,
                    fecha_registro
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                cliente.strip(), celular.strip(), tipo_venta_db, destino,
                fecha_i_str, fecha_f_str, noches,
                adultos, menores, tipo_habitacion,
                precio_adulto, precio_menor, precio_total,
                porcentaje_ganancia, ganancia, costo_mayorista,
                abono_inicial, saldo_inicial, estado_inicial,
                id_vendedora, usuario.get("id_vendedora", id_vendedora),
                es_bloqueo_db, bloqueo_id_db, es_grupo_db, grupo_id_db,
                operador,
                no_localizador.strip() or None,
                clave_confirmacion.strip() or None,
                plan_alimento,
                edades_menores_txt.strip() if menores > 0 else None,
                requerimientos_esp.strip() or None,
                fecha_registro
            ))
            venta_id = cursor.lastrowid

            # Abono inicial si hay monto
            if abono_inicial > 0:
                try:
                    cursor.execute(
                        "INSERT INTO abonos (venta_id, monto, fecha, metodo_pago) VALUES (?,?,?,?)",
                        (venta_id, abono_inicial, fecha_registro, metodo_pago)
                    )
                except Exception:
                    cursor.execute(
                        "INSERT INTO abonos (venta_id, monto, fecha) VALUES (?,?,?)",
                        (venta_id, abono_inicial, fecha_registro)
                    )

            # Pasajeros
            for p in nombres_pasajeros:
                cursor.execute(
                    "INSERT INTO pasajeros (venta_id, nombre, tipo) VALUES (?,?,?)",
                    (venta_id, p["nombre"].strip(), p["tipo"])
                )

            # Descontar inventario
            if es_bloqueo:
                cursor.execute("""
                    UPDATE bloqueos
                    SET habitaciones_vendidas    = habitaciones_vendidas + 1,
                        habitaciones_disponibles = habitaciones_disponibles - 1
                    WHERE id = ?
                """, (bloqueo_id_db,))
                cursor.execute("""
                    UPDATE bloqueos SET estado = 'AGOTADO'
                    WHERE id = ? AND habitaciones_disponibles = 0
                """, (bloqueo_id_db,))

            elif es_grupo:
                cursor.execute("""
                    UPDATE grupos
                    SET habitaciones_vendidas    = habitaciones_vendidas + 1,
                        habitaciones_disponibles = habitaciones_disponibles - 1
                    WHERE id = ? AND habitaciones_disponibles >= 1
                """, (grupo_id_db,))
                cursor.execute("""
                    UPDATE grupos SET estado = 'AGOTADO'
                    WHERE id = ? AND habitaciones_disponibles = 0
                """, (grupo_id_db,))

            # Si liquidado: calcular comisión
            if liquidado:
                cursor.execute("""
                    UPDATE ventas SET comision_vendedora = ROUND(ganancia * 0.10, 2)
                    WHERE id = ?
                """, (venta_id,))

            conn.commit()
            conn.close()
            tipo_icon = {"General": "🏨", "Bloqueo": "📦", "Grupo": "👥"}[tipo_venta_db]
            st.success(f"✅ Venta {tipo_icon} **{tipo_venta_db}** registrada correctamente (ID: {venta_id})")
            if liquidado:
                st.success("🎉 ¡Venta LIQUIDADA desde el inicio!")
                st.balloons()
            else:
                st.info(f"💳 Abono inicial registrado: ${abono_inicial:,.2f} | Saldo: ${saldo_inicial:,.2f}")

            # ── Recibo del abono inicial ───────────────────────────────────
            if abono_inicial > 0:
                num_rec_rv  = _siguiente_num_recibo()
                fecha_rec_rv = datetime.now().strftime("%d-%B-%Y").replace(
                    "January","enero").replace("February","febrero").replace(
                    "March","marzo").replace("April","abril").replace(
                    "May","mayo").replace("June","junio").replace(
                    "July","julio").replace("August","agosto").replace(
                    "September","septiembre").replace("October","octubre").replace(
                    "November","noviembre").replace("December","diciembre")
                concepto_rv = f"Abono inicial - {destino}"
                vend_rv     = usuario.get("nombre","Agente")
                st.divider()
                st.markdown("#### 🧾 Recibo del Abono Inicial")
                _boton_recibo(
                    numero     = num_rec_rv,
                    fecha_str  = fecha_rec_rv,
                    cliente    = cliente.strip(),
                    monto      = abono_inicial,
                    concepto   = concepto_rv,
                    forma_pago = metodo_pago,
                    agente     = vend_rv,
                    key_suffix = f"rv_new_{venta_id}",
                )

        except Exception as e:
            st.error(f"❌ Error al guardar: {e}")
            st.caption("💡 Verifica que las columnas `es_grupo`, `grupo_id`, `pagado` existan en la tabla `ventas`.")




def alertas_otros():
    """Alertas de citas próximas de Pasaportes y Visas (7 días)"""
    from datetime import date, timedelta
    hoy     = date.today()
    limite  = hoy + timedelta(days=7)

    usuario  = st.session_state.usuario_actual
    es_admin = usuario.get("rol") == "ADMIN"
    id_vend  = usuario.get("id_vendedora", 0) or 0
    filtro_p = "" if es_admin else f"AND vendedora_id = {id_vend}"

    conn = conectar_db()
    try:
        df_citas = pd.read_sql_query(f"""
            SELECT 'Pasaporte' AS tipo, id, cliente, celular, fecha_cita, estado
            FROM pasaportes
            WHERE fecha_cita IS NOT NULL
              AND date(fecha_cita) BETWEEN date('{hoy.isoformat()}') AND date('{limite.isoformat()}')
              AND estado NOT IN ('Entregado')
              {filtro_p}
            UNION ALL
            SELECT 'Visa' AS tipo, id, cliente, celular, fecha_cita, estado
            FROM visas
            WHERE fecha_cita IS NOT NULL
              AND date(fecha_cita) BETWEEN date('{hoy.isoformat()}') AND date('{limite.isoformat()}')
              AND estado NOT IN ('Entregada')
              {filtro_p}
            ORDER BY fecha_cita ASC
        """, conn)
    except:
        df_citas = pd.DataFrame()
    conn.close()

    if df_citas.empty:
        return

    st.markdown(
        """<div style="background:linear-gradient(135deg,#f3e5f5,#ede7f6);
        border-left:5px solid #7B1FA2;border-radius:12px;padding:16px 20px;margin-bottom:16px;">
        <div style="font-size:1.1rem;font-weight:700;color:#4A148C;">
        🗓️ Citas Próximas — Pasaportes y Visas (próximos 7 días)</div>
        <div style="font-size:0.85rem;color:#6A1B9A;margin-top:4px;">
        Recuerda confirmar y preparar la documentación con anticipación.</div>
        </div>""", unsafe_allow_html=True)

    for _, row in df_citas.iterrows():
        try:
            fecha_cita_dt = date.fromisoformat(str(row['fecha_cita']))
        except:
            continue
        dias_rest = (fecha_cita_dt - hoy).days
        dias_txt  = "¡HOY!" if dias_rest == 0 else ("¡Mañana!" if dias_rest == 1 else f"En {dias_rest} días")
        icono     = "🛂" if row['tipo'] == "Pasaporte" else "🌎"
        color     = "#7B1FA2" if dias_rest <= 2 else "#4A148C"

        col1, col2, col3 = st.columns([3, 2, 2])
        with col1:
            st.markdown(f"""<div style="background:#fdf3ff;border:1px solid #ce93d8;
            border-radius:8px;padding:10px 14px;">
            <b>{icono} {row['cliente']}</b><br>
            <span style="font-size:0.82rem;color:#555;">📋 {row['tipo']} | 📞 {row['celular'] or '—'}</span><br>
            <span style="font-size:0.82rem;color:#555;">Estado: {row['estado']}</span>
            </div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""<div style="background:#fdf3ff;border:1px solid #ce93d8;
            border-radius:8px;padding:10px 14px;text-align:center;">
            <div style="font-size:0.78rem;color:#888;font-weight:600;">FECHA CITA</div>
            <div style="font-weight:700;">{row['fecha_cita']}</div>
            <div style="font-size:1.1rem;font-weight:800;color:{color};">{dias_txt}</div>
            </div>""", unsafe_allow_html=True)
        with col3:
            st.markdown(f"""<div style="background:#fdf3ff;border:1px solid #ce93d8;
            border-radius:8px;padding:10px 14px;text-align:center;">
            <div style="font-size:0.85rem;color:#555;">Ver detalle en<br><b>🗂️ Otros → {row['tipo']}s</b></div>
            </div>""", unsafe_allow_html=True)
        st.markdown("<div style='margin-bottom:8px;'></div>", unsafe_allow_html=True)

    st.caption(f"Total: **{len(df_citas)}** cita(s) en los próximos 7 días.")
    st.divider()


def alertas_riviera():
    """
    Alertas para Riviera Maya:
      🔴 Adeudos a 45 días del viaje
      🟡 Viajes en los próximos 10 días
    """
    from datetime import date, timedelta
    hoy = date.today()
    limite_adeudo   = hoy + timedelta(days=45)
    limite_proximos = hoy + timedelta(days=10)
    limite_adeudo   = hoy + timedelta(days=45)
    limite_proximos = hoy + timedelta(days=10)

    usuario  = st.session_state.usuario_actual
    es_admin = usuario.get("rol") == "ADMIN"
    id_vend  = usuario.get("id_vendedora")
    filtro_vend = "" if es_admin else f"AND v.usuario_id = {id_vend}"

    conn = conectar_db()

    try:
        df_adeudos = pd.read_sql_query(f"""
            SELECT v.id, v.cliente, v.celular_responsable, v.destino,
                   v.fecha_inicio, v.precio_total, v.pagado, v.saldo,
                   COALESCE(vd.nombre,'—') AS vendedora, v.tipo_venta
            FROM ventas v
            LEFT JOIN vendedoras vd ON v.vendedora_id = vd.id
            WHERE v.estado NOT IN ('CERRADO','LIQUIDADO')
              AND COALESCE(v.saldo, 0) > 0
              AND v.fecha_inicio IS NOT NULL AND v.fecha_inicio != ''
              AND date(substr(v.fecha_inicio,7,4)||'-'||substr(v.fecha_inicio,4,2)||'-'||substr(v.fecha_inicio,1,2))
                  BETWEEN date('{hoy.isoformat()}') AND date('{limite_adeudo.isoformat()}')
              {filtro_vend}
            ORDER BY v.fecha_inicio ASC
        """, conn)
    except Exception:
        df_adeudos = pd.DataFrame()

    try:
        df_proximos = pd.read_sql_query(f"""
            SELECT v.id, v.cliente, v.celular_responsable, v.destino,
                   v.fecha_inicio, v.fecha_fin, v.noches,
                   v.precio_total, v.pagado, v.saldo, v.estado,
                   v.tipo_habitacion, v.adultos, v.menores, v.operador, v.tipo_venta,
                   COALESCE(vd.nombre,'—') AS vendedora,
                   COALESCE(v.reserva_confirmada, 0) AS reserva_confirmada,
                   v.reserva_confirmada_por, v.reserva_confirmada_fecha
            FROM ventas v
            LEFT JOIN vendedoras vd ON v.vendedora_id = vd.id
            WHERE v.estado != 'CERRADO'
              AND COALESCE(v.reserva_confirmada, 0) = 0
              AND v.fecha_inicio IS NOT NULL AND v.fecha_inicio != ''
              AND date(substr(v.fecha_inicio,7,4)||'-'||substr(v.fecha_inicio,4,2)||'-'||substr(v.fecha_inicio,1,2))
                  BETWEEN date('{hoy.isoformat()}') AND date('{limite_proximos.isoformat()}')
              {filtro_vend}
            ORDER BY v.fecha_inicio ASC
        """, conn)
    except Exception:
        df_proximos = pd.DataFrame()

    conn.close()

    if df_adeudos.empty and df_proximos.empty:
        return

    st.markdown("## 🚨 Alertas Importantes")

    # ── Bloque 1: Adeudos ────────────────────────────────────────────────────
    if not df_adeudos.empty:
        st.markdown(
            """<div style="background:linear-gradient(135deg,#fff0f0,#ffe0e0);
            border-left:5px solid #FF4444;border-radius:12px;padding:16px 20px;margin-bottom:16px;">
            <div style="font-size:1.1rem;font-weight:700;color:#CC0000;">
            🔴 Adeudos Críticos — Viajes en menos de 45 días</div>
            <div style="font-size:0.85rem;color:#990000;margin-top:4px;">
            ⚠️ El mayorista exige liquidación <b>30 días antes</b>. Estos clientes aún tienen saldo pendiente.</div>
            </div>""", unsafe_allow_html=True)

        for _, row in df_adeudos.iterrows():
            try:
                partes = str(row['fecha_inicio']).split('-')
                if len(partes) == 3 and len(partes[2]) == 4:
                    fecha_dt = date(int(partes[2]), int(partes[1]), int(partes[0]))
                else:
                    fecha_dt = date.fromisoformat(str(row['fecha_inicio']))
            except Exception:
                continue

            dias_rest  = (fecha_dt - hoy).days
            precio     = float(row['precio_total']) if row['precio_total'] else 0
            pagado_val = float(row['pagado'])       if row['pagado'] is not None else 0
            saldo_val  = float(row['saldo'])        if row['saldo']  is not None else 0
            pct        = (pagado_val / precio * 100) if precio > 0 else 0
            color_urg  = "#CC0000" if dias_rest <= 30 else "#E65100"
            icono_urg  = "🚨" if dias_rest <= 30 else "⚠️"

            col1, col2, col3, col4 = st.columns([3,2,2,2])
            with col1:
                st.markdown(f"""<div style="background:#fff5f5;border:1px solid #ffcccc;
                border-radius:8px;padding:10px 14px;">
                <b>👤 {row['cliente']}</b><br>
                <span style="font-size:0.82rem;color:#555;">📍 {row['destino']}</span><br>
                <span style="font-size:0.82rem;color:#555;">👩‍💼 {row['vendedora']} | 🏷️ {row.get('tipo_venta') or 'General'}</span><br>
                <span style="font-size:0.82rem;color:#555;">📞 {row.get('celular_responsable') or '—'}</span>
                </div>""", unsafe_allow_html=True)
            with col2:
                st.markdown(f"""<div style="background:#fff5f5;border:1px solid #ffcccc;
                border-radius:8px;padding:10px 14px;text-align:center;">
                <div style="font-size:0.78rem;color:#888;font-weight:600;">FECHA VIAJE</div>
                <div style="font-weight:700;">{row['fecha_inicio']}</div>
                <div style="font-size:1.1rem;font-weight:800;color:{color_urg};">{icono_urg} {dias_rest} días</div>
                </div>""", unsafe_allow_html=True)
            with col3:
                st.markdown(f"""<div style="background:#fff5f5;border:1px solid #ffcccc;
                border-radius:8px;padding:10px 14px;text-align:center;">
                <div style="font-size:0.78rem;color:#888;font-weight:600;">SALDO PENDIENTE</div>
                <div style="font-size:1.15rem;font-weight:800;color:#CC0000;">${saldo_val:,.2f}</div>
                <div style="font-size:0.8rem;color:#888;">de ${precio:,.2f}</div>
                </div>""", unsafe_allow_html=True)
            with col4:
                st.markdown(f"""<div style="background:#fff5f5;border:1px solid #ffcccc;
                border-radius:8px;padding:10px 14px;text-align:center;">
                <div style="font-size:0.78rem;color:#888;font-weight:600;">PAGADO</div>
                <div style="font-size:1.1rem;font-weight:700;color:#2E7D32;">{pct:.0f}%</div>
                <div style="font-size:0.8rem;color:#888;">${pagado_val:,.2f}</div>
                </div>""", unsafe_allow_html=True)
            st.markdown("<div style='margin-bottom:8px;'></div>", unsafe_allow_html=True)

        st.caption(f"Total: **{len(df_adeudos)}** cliente(s) con adeudo en los próximos 45 días.")

    # ── Bloque 2: Próximos viajes ─────────────────────────────────────────────
    if not df_proximos.empty:
        st.markdown(
            """<div style="background:linear-gradient(135deg,#fffde7,#fff8c4);
            border-left:5px solid #F9A825;border-radius:12px;padding:16px 20px;margin-bottom:16px;">
            <div style="font-size:1.1rem;font-weight:700;color:#E65100;">
            🟡 Viajes en los Próximos 10 Días — Acción Requerida</div>
            <div style="font-size:0.85rem;color:#BF360C;margin-top:4px;">
            📋 Confirma con el <b>mayorista y el hotel</b> que la reserva existe y los pasajeros no tendrán problemas.<br>
            ✅ Cuando hayas verificado la reserva, marca el viaje como confirmado para quitarlo de esta lista.</div>
            </div>""", unsafe_allow_html=True)

        for _, row in df_proximos.iterrows():
            try:
                partes = str(row['fecha_inicio']).split('-')
                if len(partes) == 3 and len(partes[2]) == 4:
                    fecha_dt = date(int(partes[2]), int(partes[1]), int(partes[0]))
                else:
                    fecha_dt = date.fromisoformat(str(row['fecha_inicio']))
            except Exception:
                continue

            dias_rest  = (fecha_dt - hoy).days
            saldo_val  = float(row['saldo']) if row['saldo'] is not None else 0
            estado_col = {"LIQUIDADO":"#2E7D32","ACTIVO":"#E65100","ADEUDO":"#CC0000"}.get(str(row.get('estado','')),"#555")
            dias_txt   = "¡HOY!" if dias_rest == 0 else ("¡Mañana!" if dias_rest == 1 else f"En {dias_rest} días")
            viaje_key  = str(row['id'])

            col1, col2, col3, col_btn = st.columns([3, 2, 3, 1.5])
            with col1:
                st.markdown(f"""<div style="background:#fffef0;border:1px solid #ffe082;
                border-radius:8px;padding:10px 14px;">
                <b>👤 {row['cliente']}</b><br>
                <span style="font-size:0.82rem;color:#555;">📍 {row['destino']}</span><br>
                <span style="font-size:0.82rem;color:#555;">👩‍💼 {row['vendedora']} | 🏷️ {row.get('tipo_venta') or 'General'}</span><br>
                <span style="font-size:0.82rem;color:#555;">📞 {row.get('celular_responsable') or '—'}</span>
                </div>""", unsafe_allow_html=True)
            with col2:
                st.markdown(f"""<div style="background:#fffef0;border:1px solid #ffe082;
                border-radius:8px;padding:10px 14px;text-align:center;">
                <div style="font-size:0.78rem;color:#888;font-weight:600;">SALIDA</div>
                <div style="font-weight:700;">{row['fecha_inicio']}</div>
                <div style="font-size:1.1rem;font-weight:800;color:#E65100;">{dias_txt}</div>
                <div style="font-size:0.82rem;font-weight:600;color:{estado_col};">{row.get('estado','—')}</div>
                </div>""", unsafe_allow_html=True)
            with col3:
                hab    = row.get('tipo_habitacion') or '—'
                adt    = int(row['adultos'])  if row['adultos']  is not None else 0
                men    = int(row['menores'])  if row['menores']  is not None else 0
                op     = row.get('operador') or '—'
                noches = int(row['noches']) if row['noches'] is not None else 0
                st.markdown(f"""<div style="background:#fffef0;border:1px solid #ffe082;
                border-radius:8px;padding:10px 14px;">
                <span style="font-size:0.82rem;color:#555;">🛏️ <b>{hab}</b> | 👥 {adt} adultos{f' + {men} menor(es)' if men else ''}</span><br>
                <span style="font-size:0.82rem;color:#555;">🏢 Operador: <b>{op}</b></span><br>
                <span style="font-size:0.82rem;color:#555;">📅 Regreso: {row.get('fecha_fin','—')} ({noches} noches)</span><br>
                <span style="font-size:0.85rem;font-weight:700;color:{'#CC0000' if saldo_val > 0 else '#2E7D32'};">
                {'⚠️ Saldo: $'+f'{saldo_val:,.2f}' if saldo_val > 0 else '✅ Liquidado'}</span>
                </div>""", unsafe_allow_html=True)
            with col_btn:
                st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)
                if st.button(
                    "✅ Reserva\nConfirmada",
                    key=f"confirmar_10d_{viaje_key}",
                    use_container_width=True,
                    help="Confirma que verificaste la reserva con el hotel/mayorista. Se quitará permanentemente de esta alerta."
                ):
                    try:
                        conn_upd = conectar_db()
                        nombre_quien = usuario.get("nombre", "—")
                        fecha_confirmacion = datetime.now().strftime("%Y-%m-%d %H:%M")
                        conn_upd.execute("""
                            UPDATE ventas
                            SET reserva_confirmada = 1,
                                reserva_confirmada_por = ?,
                                reserva_confirmada_fecha = ?
                            WHERE id = ?
                        """, (nombre_quien, fecha_confirmacion, int(row['id'])))
                        conn_upd.commit()
                        conn_upd.close()
                        st.toast(f"✅ Reserva de {row['cliente']} confirmada por {nombre_quien}", icon="✅")
                    except Exception as e:
                        st.error(f"Error al confirmar: {e}")
                    st.rerun()

            st.markdown("<div style='margin-bottom:8px;'></div>", unsafe_allow_html=True)

        st.caption(f"Total: **{len(df_proximos)}** viaje(s) pendiente(s) de confirmación en los próximos 10 días.")

    st.divider()


def mostrar_dashboard():
    """Muestra el dashboard principal"""
    st.title("🏖️ Dashboard - Agencia Riviera Maya")
    
    usuario = st.session_state.usuario_actual
    
    st.markdown(f"### Bienvenido, **{usuario['nombre']}** 👋")
    st.markdown(f"**Rol:** {usuario['rol']}")
    
    st.divider()

    # Alertas al tope del dashboard
    alertas_riviera()
    alertas_otros()

    # Obtener métricas
    metricas = obtener_metricas_dashboard(usuario)
    
    # Métricas principales
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            label="💰 Total Vendido (Riviera)",
            value=f"${metricas['riviera_vendido']:,.2f}",
            delta=f"{metricas['riviera_ventas']} ventas activas"
        )
    
    with col2:
        porcentaje = (metricas['riviera_cobrado'] / metricas['riviera_vendido'] * 100) if metricas['riviera_vendido'] > 0 else 0
        st.metric(
            label="✅ Total Cobrado",
            value=f"${metricas['riviera_cobrado']:,.2f}",
            delta=f"{porcentaje:.1f}% cobrado"
        )
    
    with col3:
        st.metric(
            label="⏳ Saldo Pendiente",
            value=f"${metricas['riviera_saldo']:,.2f}",
            delta="Por cobrar"
        )
    
    with col4:
        st.metric(
            label="📈 Ganancia",
            value=f"${metricas['riviera_ganancia']:,.2f}",
            delta=f"Margen: {(metricas['riviera_ganancia']/metricas['riviera_vendido']*100):.1f}%" if metricas['riviera_vendido'] > 0 else "0%"
        )
    
    st.divider()
    
    # Segunda fila de métricas
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            label="🎫 Viajes Nacionales",
            value=metricas['nacionales'],
            delta="Activos"
        )
    
    with col2:
        st.metric(
            label="🌎 Viajes Internacionales",
            value=metricas['internacionales'],
            delta="Activos"
        )
    
    with col3:
        st.metric(
            label="💵 Comisiones Pendientes",
            value=f"${metricas['comisiones']:,.2f}",
            delta="Por pagar"
        )
    
    # Gráficas
    st.divider()
    st.subheader("📊 Análisis Visual")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Gráfica de ventas por estado
        conn = conectar_db()
        df_estados = pd.read_sql_query("""
            SELECT estado, COUNT(*) as cantidad, SUM(saldo) as saldo_total
            FROM ventas
            WHERE estado != 'CERRADO'
            GROUP BY estado
        """, conn)
        
        if not df_estados.empty:
            fig = px.pie(df_estados, values='cantidad', names='estado', 
                        title='Ventas por Estado')
            st.plotly_chart(fig, use_container_width=True)
        
        conn.close()
    
    with col2:
        # Gráfica de cobrado vs pendiente
        labels = ['Cobrado', 'Pendiente']
        values = [metricas['riviera_cobrado'], metricas['riviera_saldo']]
        
        fig = go.Figure(data=[go.Pie(labels=labels, values=values, hole=.3)])
        fig.update_layout(title_text="Estado de Cobranza")
        st.plotly_chart(fig, use_container_width=True)

def pagina_ventas_riviera():
    """Página de ventas de Riviera Maya"""
    st.title("\U0001f3d6\ufe0f Ventas - Riviera Maya")

    tabs = st.tabs(["\U0001f4cb Ver Ventas", "\u2795 Nueva Venta", "\U0001f4b0 Registrar Pago"])

    # TAB 0 — VER VENTAS
    with tabs[0]:
        alertas_riviera()
        st.subheader("Ventas Activas")
        usuario = st.session_state.usuario_actual
        conn = conectar_db()

        if usuario["rol"] == "VENDEDORA":
            query = f"""
                SELECT v.id, v.cliente, v.celular_responsable, v.destino,
                       v.operador, v.tipo_habitacion, v.adultos, v.menores,
                       v.fecha_inicio, v.fecha_fin, v.noches,
                       v.precio_total, v.pagado, v.saldo,
                       v.porcentaje_ganancia, v.ganancia,
                       v.estado, vd.nombre AS vendedora, v.tipo_venta
                FROM ventas v
                JOIN vendedoras vd ON v.vendedora_id = vd.id
                WHERE v.estado != 'CERRADO'
                ORDER BY v.id DESC
            """
        else:
            query = """
                SELECT v.id, v.cliente, v.celular_responsable, v.destino,
                       v.operador, v.tipo_habitacion, v.adultos, v.menores,
                       v.fecha_inicio, v.fecha_fin, v.noches,
                       v.precio_total, v.pagado, v.saldo,
                       v.porcentaje_ganancia, v.ganancia,
                       v.estado, vd.nombre AS vendedora, v.tipo_venta
                FROM ventas v
                JOIN vendedoras vd ON v.vendedora_id = vd.id
                WHERE v.estado != 'CERRADO'
                ORDER BY v.id DESC
            """

        df = pd.read_sql_query(query, conn)
        conn.close()

        if df.empty:
            st.info("No hay ventas activas.")
        else:
            df.columns = ['ID','Cliente','Celular','Destino','Operador',
                          'Habitación','Adultos','Menores',
                          'Salida','Regreso','Noches',
                          'Total','Pagado','Saldo',
                          '% Ganancia','Ganancia',
                          'Estado','Vendedora','Tipo']

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Ventas", len(df))
            col2.metric("Total vendido",   f"${df['Total'].sum():,.2f}")
            col3.metric("Total cobrado",   f"${df['Pagado'].sum():,.2f}")
            col4.metric("Saldo pendiente", f"${df['Saldo'].sum():,.2f}")

            st.divider()

            col1, col2, col3 = st.columns(3)
            with col1:
                filtro_estado = st.selectbox("Filtrar por estado",
                    ["Todos"] + sorted(df['Estado'].unique().tolist()))
            with col2:
                filtro_vendedora_rv = st.selectbox("Filtrar por vendedora",
                    ["Todas"] + sorted(df['Vendedora'].unique().tolist()),
                    key="rv_filtro_vend")
            with col3:
                filtro_buscar = st.text_input("\U0001f50d Buscar cliente o destino")

            df_f = df.copy()
            if filtro_estado != "Todos":
                df_f = df_f[df_f['Estado'] == filtro_estado]
            if filtro_vendedora_rv != "Todas":
                df_f = df_f[df_f['Vendedora'] == filtro_vendedora_rv]
            if filtro_buscar:
                mask = (
                    df_f['Cliente'].str.contains(filtro_buscar, case=False, na=False) |
                    df_f['Destino'].str.contains(filtro_buscar, case=False, na=False)
                )
                df_f = df_f[mask]

            df_show = df_f.copy()
            for col in ['Total','Pagado','Saldo','Ganancia']:
                df_show[col] = df_show[col].apply(lambda x: f"${x:,.2f}")

            st.dataframe(df_show, use_container_width=True, hide_index=True)

            # Detalle expandible
            st.divider()
            st.markdown("#### \U0001f50e Ver detalle de venta")
            venta_ids = df_f['ID'].tolist()
            if venta_ids:
                id_sel = st.selectbox("Selecciona ID", venta_ids)
                row = df_f[df_f['ID'] == id_sel].iloc[0]

                with st.expander(f"\U0001f4c4 {row['Cliente']} | {row['Destino']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Cliente:** {row['Cliente']}")
                    c1.write(f"**Celular:** {row['Celular']}")
                    c1.write(f"**Operador:** {row['Operador'] or '—'}")
                    c2.write(f"**Destino:** {row['Destino']}")
                    c2.write(f"**Habitación:** {row['Habitación']}")
                    c2.write(f"**Pasajeros:** {row['Adultos']} adultos + {row['Menores']} menores")
                    c3.write(f"**Salida:** {row['Salida']}")
                    c3.write(f"**Regreso:** {row['Regreso']}")
                    c3.write(f"**Noches:** {row['Noches']}")

                    st.divider()
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total",   f"${row['Total']:,.2f}")
                    c2.metric("Pagado",  f"${row['Pagado']:,.2f}")
                    c3.metric("Saldo",   f"${row['Saldo']:,.2f}")
                    c4.metric("Ganancia",f"${row['Ganancia']:,.2f} ({row['% Ganancia']}%)")

                    conn2 = conectar_db()
                    df_pas = pd.read_sql_query(
                        "SELECT nombre, tipo FROM pasajeros WHERE venta_id = ?",
                        conn2, params=(id_sel,))
                    df_abonos = pd.read_sql_query(
                        "SELECT fecha, monto FROM abonos WHERE venta_id = ? ORDER BY fecha",
                        conn2, params=(id_sel,))
                    conn2.close()

                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**👥 Pasajeros:**")
                        if not df_pas.empty:
                            st.dataframe(df_pas, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin pasajeros registrados")
                    with c2:
                        st.markdown("**💳 Historial de abonos:**")
                        if not df_abonos.empty:
                            df_abonos['monto'] = df_abonos['monto'].apply(lambda x: f"${x:,.2f}")
                            st.dataframe(df_abonos, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin abonos registrados")

                    # ── Localización y cupón ───────────────────────────────
                    st.divider()
                    st.markdown("#### 🔑 Localización y Cupón")

                    conn3 = conectar_db()
                    try:
                        row_extra = conn3.execute("""
                            SELECT no_localizador, clave_confirmacion,
                                   plan_alimento, edades_menores,
                                   requerimientos_especiales
                            FROM ventas WHERE id = ?
                        """, (id_sel,)).fetchone()
                    except Exception:
                        row_extra = None
                    conn3.close()

                    loc_actual   = row_extra[0] if row_extra else ""
                    clave_actual = row_extra[1] if row_extra else ""
                    plan_actual  = row_extra[2] if row_extra else "Todo incluido"
                    edad_actual  = row_extra[3] if row_extra else ""
                    req_actual   = row_extra[4] if row_extra else ""

                    col_loc1, col_loc2 = st.columns(2)
                    with col_loc1:
                        ed_localizador = st.text_input(
                            "No. de Localizador",
                            value=loc_actual or "",
                            key=f"ed_loc_{id_sel}"
                        )
                    with col_loc2:
                        ed_clave = st.text_input(
                            "Clave de Confirmación",
                            value=clave_actual or "",
                            key=f"ed_clave_{id_sel}"
                        )

                    col_loc3, col_loc4 = st.columns(2)
                    with col_loc3:
                        planes    = ["Todo incluido","Solo desayuno","Media pensión","Solo alojamiento","Otro"]
                        plan_idx  = planes.index(plan_actual) if plan_actual in planes else 0
                        ed_plan   = st.selectbox("Plan de alimento", planes,
                                                  index=plan_idx, key=f"ed_plan_{id_sel}")
                    with col_loc4:
                        ed_edades = st.text_input("Edades de menores",
                                                   value=edad_actual or "",
                                                   key=f"ed_edades_{id_sel}")

                    ed_req = st.text_area("Requerimientos especiales",
                                          value=req_actual or "",
                                          height=60, key=f"ed_req_{id_sel}")

                    col_b1, col_b2 = st.columns(2)
                    with col_b1:
                        if st.button("💾 Guardar cambios", key=f"save_loc_{id_sel}",
                                     use_container_width=True):
                            try:
                                conn_upd = conectar_db()
                                conn_upd.execute("""
                                    UPDATE ventas
                                    SET no_localizador         = ?,
                                        clave_confirmacion     = ?,
                                        plan_alimento          = ?,
                                        edades_menores         = ?,
                                        requerimientos_especiales = ?
                                    WHERE id = ?
                                """, (
                                    ed_localizador.strip() or None,
                                    ed_clave.strip() or None,
                                    ed_plan,
                                    ed_edades.strip() or None,
                                    ed_req.strip() or None,
                                    id_sel
                                ))
                                conn_upd.commit()
                                conn_upd.close()
                                st.success("✅ Datos actualizados correctamente.")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error al guardar: {e}")

                    with col_b2:
                        if ed_clave.strip() and CUPONES_DISPONIBLES:
                            hoteles_full = obtener_hoteles_completos()
                            hotel_data   = next(
                                (h for h in hoteles_full if h['nombre'] == row['Destino']), {}
                            )
                            try:
                                pdf_cupon = generar_cupon_pdf(
                                    titular         = row['Cliente'],
                                    clave_confirm   = ed_clave.strip(),
                                    hotel_nombre    = row['Destino'],
                                    hotel_direccion = hotel_data.get('direccion', ''),
                                    hotel_telefono  = hotel_data.get('telefono', ''),
                                    hotel_estrellas = hotel_data.get('estrellas', 4),
                                    tipo_habitacion = row['Habitación'],
                                    plan_alimento   = ed_plan,
                                    fecha_entrada   = row['Salida'],
                                    fecha_salida    = row['Regreso'],
                                    adultos         = int(row['Adultos']),
                                    menores         = int(row['Menores']),
                                    edades_menores  = ed_edades or "",
                                    requerimientos  = ed_req or "",
                                    logo_path       = LOGO_PATH,
                                )
                                st.download_button(
                                    label    = "🎫 Descargar Cupón de Acceso",
                                    data     = pdf_cupon,
                                    file_name= f"cupon_{row['Cliente'].replace(' ','_')[:20]}.pdf",
                                    mime     = "application/pdf",
                                    key      = f"cupon_{id_sel}",
                                    use_container_width=True,
                                    type     = "primary"
                                )
                            except Exception as e:
                                st.error(f"❌ Error al generar cupón: {e}")
                        elif not ed_clave.strip():
                            st.warning("⚠️ Agrega la Clave de Confirmación para generar el cupón.",
                                       icon="🔑")
                        elif not CUPONES_DISPONIBLES:
                            st.caption("⚠️ generar_cupon.py no encontrado junto a app_streamlit.py")

            csv = df_show.to_csv(index=False).encode('utf-8')
            st.download_button("\U0001f4e5 Descargar CSV", csv,
                file_name=f"ventas_riviera_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # TAB 1 — NUEVA VENTA
    with tabs[1]:
        formulario_nueva_venta()

    # TAB 2 — REGISTRAR PAGO
    with tabs[2]:
        st.subheader("\U0001f4b0 Registrar Pago")
        usuario = st.session_state.usuario_actual
        conn = conectar_db()

        if usuario["rol"] == "VENDEDORA":
            query_ventas = f"""
                SELECT v.id, v.cliente, v.destino, v.precio_total, v.pagado, v.saldo
                FROM ventas v
                WHERE v.estado NOT IN ('CERRADO','LIQUIDADO')
                  AND v.usuario_id = {usuario['id_vendedora']}
                ORDER BY v.id DESC
            """
        else:
            query_ventas = """
                SELECT v.id, v.cliente, v.destino, v.precio_total, v.pagado, v.saldo
                FROM ventas v
                WHERE v.estado NOT IN ('CERRADO','LIQUIDADO')
                ORDER BY v.id DESC
            """

        df_ventas = pd.read_sql_query(query_ventas, conn)
        conn.close()

        if df_ventas.empty:
            st.info("No hay ventas con saldo pendiente.")
        else:
            opciones_venta = {
                f"ID {row['id']} | {row['cliente']} — {row['destino']} | Saldo: ${row['saldo']:,.2f}": row
                for _, row in df_ventas.iterrows()
            }
            venta_label = st.selectbox("Selecciona la venta:", list(opciones_venta.keys()))
            venta_sel = opciones_venta[venta_label]

            col1, col2, col3 = st.columns(3)
            col1.metric("Total",           f"${venta_sel['precio_total']:,.2f}")
            col2.metric("Ya pagado",        f"${venta_sel['pagado']:,.2f}")
            col3.metric("Saldo pendiente",  f"${venta_sel['saldo']:,.2f}")

            st.divider()
            st.markdown("#### \U0001f4b3 Datos del pago")

            col1, col2 = st.columns(2)
            with col1:
                monto_pago = st.number_input(
                    "Monto del abono *",
                    min_value=0.01,
                    max_value=float(venta_sel['saldo']),
                    step=100.0, format="%.2f"
                )
            with col2:
                metodo_pago = st.selectbox("Método de pago *",
                    ["Efectivo", "Transferencia", "Tarjeta de crédito",
                     "Tarjeta de débito", "Depósito bancario", "Otro"])

            nuevo_pagado = venta_sel['pagado'] + monto_pago
            nuevo_saldo  = venta_sel['saldo']  - monto_pago
            liquidado    = nuevo_saldo <= 0.01

            col1, col2 = st.columns(2)
            col1.metric("Nuevo total pagado", f"${nuevo_pagado:,.2f}", delta=f"+${monto_pago:,.2f}")
            col2.metric("Nuevo saldo",        f"${nuevo_saldo:,.2f}", delta=f"-${monto_pago:,.2f}")

            if liquidado:
                st.success("\U0001f389 ¡Con este pago la venta quedará **LIQUIDADA**!")
            else:
                pct = (nuevo_pagado / venta_sel['precio_total']) * 100
                st.info(f"\U0001f4ca Avance: **{pct:.1f}%** cobrado")

            st.divider()

            if st.button("\U0001f4be Registrar Pago", type="primary", use_container_width=True):
                try:
                    conn = conectar_db()
                    cursor = conn.cursor()
                    fecha_pago = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    # Intentar con metodo_pago
                    try:
                        cursor.execute(
                            "INSERT INTO abonos (venta_id, monto, fecha, metodo_pago) VALUES (?,?,?,?)",
                            (int(venta_sel['id']), monto_pago, fecha_pago, metodo_pago)
                        )
                    except:
                        cursor.execute(
                            "INSERT INTO abonos (venta_id, monto, fecha) VALUES (?,?,?)",
                            (int(venta_sel['id']), monto_pago, fecha_pago)
                        )

                    nuevo_estado = "LIQUIDADO" if liquidado else "ACTIVO"
                    if liquidado:
                        # Al liquidar: calcular comisión = 10% de la ganancia
                        cursor.execute("""
                            UPDATE ventas
                            SET pagado = pagado + ?, saldo = saldo - ?, estado = ?,
                                comision_vendedora = ROUND(ganancia * 0.10, 2)
                            WHERE id = ?
                        """, (monto_pago, monto_pago, nuevo_estado, int(venta_sel['id'])))
                    else:
                        cursor.execute("""
                            UPDATE ventas
                            SET pagado = pagado + ?, saldo = saldo - ?, estado = ?
                            WHERE id = ?
                        """, (monto_pago, monto_pago, nuevo_estado, int(venta_sel['id'])))

                    conn.commit()
                    conn.close()

                    # ── Generar recibo automático ──────────────────────────
                    num_rec = _siguiente_num_recibo()
                    fecha_rec = datetime.now().strftime("%d-%B-%Y").replace(
                        "January","enero").replace("February","febrero").replace(
                        "March","marzo").replace("April","abril").replace(
                        "May","mayo").replace("June","junio").replace(
                        "July","julio").replace("August","agosto").replace(
                        "September","septiembre").replace("October","octubre").replace(
                        "November","noviembre").replace("December","diciembre")
                    concepto_rec = f"{'Liquidación' if liquidado else 'Abono'} - {venta_sel['destino']}"
                    vend_nombre  = usuario.get("nombre", "Agente")

                    if liquidado:
                        st.success("✅ ¡Venta **LIQUIDADA**! 🎉")
                        st.balloons()
                    else:
                        st.success(f"✅ Abono de ${monto_pago:,.2f} ({metodo_pago}) registrado.")

                    st.divider()
                    st.markdown("#### 🧾 Recibo de Pago")
                    _boton_recibo(
                        numero     = num_rec,
                        fecha_str  = fecha_rec,
                        cliente    = venta_sel['cliente'],
                        monto      = monto_pago,
                        concepto   = concepto_rec,
                        forma_pago = metodo_pago,
                        agente     = vend_nombre,
                        key_suffix = f"rv_{venta_sel['id']}",
                    )

                except Exception as e:
                    st.error(f"\u274c Error: {e}")


def obtener_viajes_nacionales_activos():
    """Obtiene lista de viajes nacionales activos"""
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso,
                   dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles,
                   precio_persona_doble, precio_persona_triple, estado
            FROM viajes_nacionales
            WHERE estado = 'ACTIVO'
            ORDER BY fecha_salida
        """)
        cols = ['id','nombre_viaje','destino','fecha_salida','fecha_regreso',
                'dias','noches','cupos_totales','cupos_vendidos','cupos_disponibles',
                'precio_doble','precio_triple','estado']
        rows = cursor.fetchall()
        conn.close()
        return [dict(zip(cols, r)) for r in rows]
    except Exception as e:
        conn.close()
        return []


def obtener_cotizaciones_nacionales():
    """Obtiene cotizaciones disponibles para cargar en viaje"""
    conn = conectar_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso,
                   dias, noches, personas_proyectadas,
                   precio_persona_doble, precio_persona_triple
            FROM cotizaciones_nacionales
            ORDER BY id DESC
        """)
        cols = ['id','nombre_viaje','destino','fecha_salida','fecha_regreso',
                'dias','noches','personas_proyectadas','precio_doble','precio_triple']
        rows = cursor.fetchall()
        conn.close()
        return [dict(zip(cols, r)) for r in rows]
    except:
        conn.close()
        return []


def formulario_registro_cliente_nacional():
    """Formulario para registrar un cliente en un viaje nacional"""
    st.subheader("➕ Registrar Cliente en Viaje Nacional")
    usuario = st.session_state.usuario_actual

    viajes = obtener_viajes_nacionales_activos()
    if not viajes:
        st.warning("⚠️ No hay viajes nacionales activos con cupos. Pide al administrador que registre viajes primero.")
        return

    # ── Selección de viaje ──────────────────────────────────────────────────
    st.markdown("#### 🎫 Selecciona el Viaje")
    opciones_viaje = {
        f"{v['nombre_viaje']} | {v['destino']} | {v['fecha_salida']} → {v['fecha_regreso']} | Disp: {v['cupos_disponibles']}": v
        for v in viajes if v['cupos_disponibles'] > 0
    }
    if not opciones_viaje:
        st.error("❌ No hay viajes con cupos disponibles.")
        return

    viaje_label = st.selectbox("Viaje disponible:", list(opciones_viaje.keys()), key="cn_viaje")
    viaje_sel = opciones_viaje[viaje_label]

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Cupos disponibles", viaje_sel['cupos_disponibles'])
    col2.metric("Días / Noches", f"{viaje_sel['dias']}d / {viaje_sel['noches']}n")
    col3.metric("Precio doble p/p", f"${viaje_sel['precio_doble']:,.2f}")
    col4.metric("Precio triple p/p", f"${viaje_sel['precio_triple']:,.2f}")

    st.divider()

    # ── Datos del cliente ──────────────────────────────────────────────────
    st.markdown("#### 👤 Datos del Cliente")
    col1, col2 = st.columns(2)
    with col1:
        cliente = st.text_input("Nombre del cliente *", key="cn_cliente")
    with col2:
        celular = st.text_input("Celular del cliente *", key="cn_celular")

    # Vendedora
    if usuario["rol"] == "ADMIN":
        vendedoras = obtener_vendedoras()
        if not vendedoras:
            st.error("❌ No hay vendedoras registradas.")
            return
        vend_dict = {v[1]: v[0] for v in vendedoras}
        id_vendedora = vend_dict[st.selectbox("Vendedora *", list(vend_dict.keys()), key="cn_vend")]
    else:
        id_vendedora = usuario["id_vendedora"]
        st.info(f"Vendedora: **{usuario['nombre']}**")

    st.divider()

    # ── Distribución de habitaciones ────────────────────────────────────────
    st.markdown("#### 🛏️ Distribución de Habitaciones")
    col1, col2 = st.columns(2)
    with col1:
        hab_dobles = st.number_input(
            f"Habitaciones dobles (2 pax × ${viaje_sel['precio_doble']:,.2f})",
            min_value=0, value=1, step=1, key="cn_dobles"
        )
    with col2:
        hab_triples = st.number_input(
            f"Habitaciones triples (3 pax × ${viaje_sel['precio_triple']:,.2f})",
            min_value=0, value=0, step=1, key="cn_triples"
        )

    adultos_hab = (hab_dobles * 2) + (hab_triples * 3)

    col1, col2 = st.columns(2)
    with col1:
        menores = st.number_input("Número de menores (sin cupo extra)", min_value=0, value=0, step=1, key="cn_menores")
    with col2:
        st.metric("Total adultos por habitación", adultos_hab)

    total_personas = adultos_hab + menores

    if total_personas == 0:
        st.warning("⚠️ Selecciona al menos una habitación.")
        return

    if adultos_hab > viaje_sel['cupos_disponibles']:
        st.error(f"❌ Solo hay {viaje_sel['cupos_disponibles']} cupos disponibles. Reduce las habitaciones.")
        return

    # ── Cálculo de precios ──────────────────────────────────────────────────
    st.markdown("#### 💰 Resumen de Precios")

    total_dobles  = hab_dobles  * 2 * viaje_sel['precio_doble']
    total_triples = hab_triples * 3 * viaje_sel['precio_triple']
    total_pagar   = total_dobles + total_triples

    col1, col2, col3 = st.columns(3)
    if hab_dobles > 0:
        col1.metric(
            f"{hab_dobles} hab. doble(s) × 2 × ${viaje_sel['precio_doble']:,.2f}",
            f"${total_dobles:,.2f}"
        )
    if hab_triples > 0:
        col2.metric(
            f"{hab_triples} hab. triple(s) × 3 × ${viaje_sel['precio_triple']:,.2f}",
            f"${total_triples:,.2f}"
        )
    col3.metric("💵 TOTAL A PAGAR", f"${total_pagar:,.2f}")

    # ── Abono inicial ──────────────────────────────────────────────────────
    st.markdown("#### 💳 Abono Inicial")
    col1, col2 = st.columns(2)
    with col1:
        abono_inicial = st.number_input(
            "Abono inicial *",
            min_value=0.0, max_value=float(total_pagar),
            step=100.0, format="%.2f", key="cn_abono"
        )
    with col2:
        metodo_abono = st.selectbox("Método de pago *",
            ["Efectivo","Transferencia","Tarjeta de crédito",
             "Tarjeta de débito","Depósito bancario","Otro"], key="cn_metodo")

    saldo = total_pagar - abono_inicial
    liquidado = saldo <= 0.01

    col1, col2 = st.columns(2)
    col1.metric("Abono inicial", f"${abono_inicial:,.2f}")
    col2.metric("Saldo restante", f"${saldo:,.2f}")
    if liquidado:
        st.success("🎉 ¡Este cliente liquidará el viaje con el abono inicial!")

    st.divider()

    # ── Nombres de pasajeros por habitación ─────────────────────────────────
    st.markdown("#### 🧳 Pasajeros por Habitación")
    pasajeros = []

    for i in range(hab_dobles):
        st.markdown(f"**🛏️ Habitación Doble {i+1}**")
        cols_d = st.columns(2)
        for j in range(2):
            with cols_d[j]:
                nombre = st.text_input(f"Pasajero {j+1}", key=f"cn_d{i}_{j}", placeholder="Nombre completo")
                pasajeros.append({"nombre": nombre, "tipo": "ADULTO", "habitacion": f"Doble {i+1}"})

    for i in range(hab_triples):
        st.markdown(f"**🛏️ Habitación Triple {i+1}**")
        cols_t = st.columns(3)
        for j in range(3):
            with cols_t[j]:
                nombre = st.text_input(f"Pasajero {j+1}", key=f"cn_t{i}_{j}", placeholder="Nombre completo")
                pasajeros.append({"nombre": nombre, "tipo": "ADULTO", "habitacion": f"Triple {i+1}"})

    if menores > 0:
        st.markdown(f"**👶 Menores ({menores})**")
        cols_m = st.columns(2)
        for i in range(menores):
            with cols_m[i % 2]:
                nombre_m = st.text_input(f"Menor {i+1}", key=f"cn_m{i}", placeholder="Nombre completo")
                hab_m = st.selectbox(
                    f"Habitación del menor {i+1}",
                    [f"Doble {j+1}" for j in range(hab_dobles)] +
                    [f"Triple {j+1}" for j in range(hab_triples)],
                    key=f"cn_mh{i}"
                )
                pasajeros.append({"nombre": nombre_m, "tipo": "MENOR", "habitacion": hab_m})

    st.divider()

    if st.button("💾 Registrar Cliente en Viaje", type="primary", use_container_width=True):
        errores = []
        if not cliente.strip():
            errores.append("El nombre del cliente es obligatorio.")
        if not celular.strip():
            errores.append("El celular es obligatorio.")
        if adultos_hab == 0:
            errores.append("Debes seleccionar al menos una habitación.")
        if total_pagar <= 0:
            errores.append("El total a pagar debe ser mayor a 0.")
        if any(p["nombre"].strip() == "" for p in pasajeros):
            errores.append("Todos los nombres de pasajeros son obligatorios.")

        if errores:
            for e in errores:
                st.error(f"❌ {e}")
            return

        try:
            conn = conectar_db()
            cursor = conn.cursor()
            fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            estado_cliente = "LIQUIDADO" if liquidado else "ADEUDO"

            cursor.execute("""
                INSERT INTO clientes_nacionales (
                    viaje_id, vendedora_id, nombre_cliente, celular_responsable,
                    adultos, menores,
                    habitaciones_doble, habitaciones_triple,
                    total_pagar, total_abonado, saldo, estado, fecha_registro
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                viaje_sel['id'], id_vendedora,
                cliente.strip(), celular.strip(),
                adultos_hab, menores,
                hab_dobles, hab_triples,
                total_pagar, abono_inicial, saldo,
                estado_cliente, fecha_registro
            ))
            cliente_id = cursor.lastrowid

            # Abono inicial
            try:
                cursor.execute(
                    "INSERT INTO abonos_nacionales (cliente_id, monto, fecha, vendedora_id, metodo_pago) VALUES (?,?,?,?,?)",
                    (cliente_id, abono_inicial, fecha_registro, id_vendedora, metodo_abono)
                )
            except:
                cursor.execute(
                    "INSERT INTO abonos_nacionales (cliente_id, monto, fecha, vendedora_id) VALUES (?,?,?,?)",
                    (cliente_id, abono_inicial, fecha_registro, id_vendedora)
                )

            # Pasajeros
            for p in pasajeros:
                cursor.execute(
                    "INSERT INTO pasajeros_nacionales (cliente_id, nombre_completo, tipo, habitacion_asignada) VALUES (?,?,?,?)",
                    (cliente_id, p["nombre"].strip(), p["tipo"], p["habitacion"])
                )

            # Actualizar cupos
            cursor.execute("""
                UPDATE viajes_nacionales
                SET cupos_vendidos = cupos_vendidos + ?,
                    cupos_disponibles = cupos_disponibles - ?
                WHERE id = ?
            """, (adultos_hab, adultos_hab, viaje_sel['id']))

            cursor.execute("""
                UPDATE viajes_nacionales SET estado = 'AGOTADO'
                WHERE id = ? AND cupos_disponibles <= 0
            """, (viaje_sel['id'],))

            conn.commit()
            conn.close()

            st.success(f"✅ **{cliente}** registrado en **{viaje_sel['nombre_viaje']}** — {total_personas} pasajero(s) | Saldo: ${saldo:,.2f}")
            if liquidado:
                st.balloons()

            # ── Recibo del abono inicial ───────────────────────────────────
            if abono_inicial > 0:
                num_rec_cn  = _siguiente_num_recibo()
                fecha_rec_cn = datetime.now().strftime("%d-%B-%Y").replace(
                    "January","enero").replace("February","febrero").replace(
                    "March","marzo").replace("April","abril").replace(
                    "May","mayo").replace("June","junio").replace(
                    "July","julio").replace("August","agosto").replace(
                    "September","septiembre").replace("October","octubre").replace(
                    "November","noviembre").replace("December","diciembre")
                concepto_cn = f"Abono inicial — {viaje_sel['nombre_viaje']} ({viaje_sel.get('destino','')})"
                vend_cn     = usuario.get("nombre", "Agente")
                st.divider()
                st.markdown("#### 🧾 Recibo del Abono Inicial")
                _boton_recibo(
                    numero     = num_rec_cn,
                    fecha_str  = fecha_rec_cn,
                    cliente    = cliente.strip(),
                    monto      = abono_inicial,
                    concepto   = concepto_cn,
                    forma_pago = metodo_abono,
                    agente     = vend_cn,
                    key_suffix = f"cn_new_{cliente_id}",
                    total_viaje      = total_pagar,
                    pagado_acumulado = abono_inicial,
                    nuevo_saldo      = saldo,
                )
            else:
                st.rerun()

        except Exception as e:
            st.error(f"❌ Error al guardar: {e}")
            st.caption("💡 Verifica que las tablas `clientes_nacionales`, `pasajeros_nacionales` y `abonos_nacionales` existan en la BD.")


def pagina_viajes_nacionales():
    """Página completa de viajes nacionales"""
    st.title("🎫 Viajes Nacionales")
    usuario = st.session_state.usuario_actual

    tabs_labels = ["📋 Ver Viajes", "👥 Ver Clientes", "➕ Registrar Cliente", "💰 Registrar Pago"]
    if usuario["rol"] == "ADMIN":
        tabs_labels.append("📝 Cotizador")
        tabs_labels.append("⚙️ Gestionar Viajes")

    tabs = st.tabs(tabs_labels)

    # ── TAB 0: VER VIAJES ──────────────────────────────────────────────────
    with tabs[0]:
        st.subheader("Viajes Nacionales")

        conn = conectar_db()
        try:
            df_viajes = pd.read_sql_query("""
                SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso,
                       dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles,
                       precio_persona_doble, precio_persona_triple, estado
                FROM viajes_nacionales
                ORDER BY fecha_salida
            """, conn)
        except:
            df_viajes = pd.DataFrame()
        conn.close()

        if df_viajes.empty:
            st.info("ℹ️ No hay viajes nacionales registrados aún.")
        else:
            activos = df_viajes[df_viajes['estado'] == 'ACTIVO']
            col1, col2, col3 = st.columns(3)
            col1.metric("Viajes activos", len(activos))
            col2.metric("Cupos disponibles", int(activos['cupos_disponibles'].sum()) if not activos.empty else 0)
            col3.metric("Total viajes", len(df_viajes))

            st.divider()

            filtro_estado = st.selectbox("Filtrar por estado:",
                ["Todos"] + sorted(df_viajes['estado'].unique().tolist()), key="vn_filtro")
            df_f = df_viajes if filtro_estado == "Todos" else df_viajes[df_viajes['estado'] == filtro_estado]

            df_show = df_f.copy()
            df_show['precio_persona_doble']  = df_show['precio_persona_doble'].apply(lambda x: f"${x:,.2f}")
            df_show['precio_persona_triple'] = df_show['precio_persona_triple'].apply(lambda x: f"${x:,.2f}")
            df_show.columns = ['ID','Viaje','Destino','Salida','Regreso',
                                'Días','Noches','Cupos Total','Vendidos','Disponibles',
                                'P. Doble p/p','P. Triple p/p','Estado']
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            # Detalle
            st.divider()
            st.markdown("#### 🔍 Detalle de un viaje")
            viaje_ids = df_f['id'].tolist()
            if viaje_ids:
                id_sel = st.selectbox("Selecciona ID:", viaje_ids, key="vn_det_id")
                row = df_f[df_f['id'] == id_sel].iloc[0]
                with st.expander(f"📄 {row['nombre_viaje']} — {row['destino']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Viaje:** {row['nombre_viaje']}")
                    c1.write(f"**Destino:** {row['destino']}")
                    c1.write(f"**Salida:** {row['fecha_salida']}  →  **Regreso:** {row['fecha_regreso']}")
                    c2.metric("Días / Noches", f"{row['dias']}d / {row['noches']}n")
                    c2.metric("Cupos totales", row['cupos_totales'])
                    c3.metric("Cupos disponibles", row['cupos_disponibles'])
                    c3.metric("Estado", row['estado'])

                    st.markdown(f"**💰 Precio doble p/p:** ${row['precio_persona_doble']:,.2f}  |  **Precio triple p/p:** ${row['precio_persona_triple']:,.2f}")

                    if row['cupos_totales'] > 0:
                        pct = row['cupos_vendidos'] / row['cupos_totales']
                        st.markdown(f"**Ocupación: {pct*100:.0f}%**")
                        st.progress(pct)

    # ── TAB 1: VER CLIENTES ────────────────────────────────────────────────
    with tabs[1]:
        st.subheader("Clientes de Viajes Nacionales")

        conn = conectar_db()
        try:
            if usuario["rol"] == "VENDEDORA":
                query_cl = f"""
                    SELECT cn.id, cn.nombre_cliente, cn.celular_responsable,
                           vj.nombre_viaje, vj.destino, vj.fecha_salida, vj.fecha_regreso,
                           cn.adultos, cn.menores,
                           cn.habitaciones_doble, cn.habitaciones_triple,
                           cn.total_pagar, cn.total_abonado, cn.saldo,
                           cn.estado, vd.nombre AS vendedora
                    FROM clientes_nacionales cn
                    JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                    JOIN vendedoras vd ON cn.vendedora_id = vd.id
                    WHERE cn.estado != 'CERRADO'
                    ORDER BY cn.id DESC
                """
            else:
                query_cl = """
                    SELECT cn.id, cn.nombre_cliente, cn.celular_responsable,
                           vj.nombre_viaje, vj.destino, vj.fecha_salida, vj.fecha_regreso,
                           cn.adultos, cn.menores,
                           cn.habitaciones_doble, cn.habitaciones_triple,
                           cn.total_pagar, cn.total_abonado, cn.saldo,
                           cn.estado, vd.nombre AS vendedora
                    FROM clientes_nacionales cn
                    JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                    JOIN vendedoras vd ON cn.vendedora_id = vd.id
                    WHERE cn.estado != 'CERRADO'
                    ORDER BY cn.id DESC
                """
            df_cl = pd.read_sql_query(query_cl, conn)
        except Exception as e:
            df_cl = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_cl.empty:
            st.info("No hay clientes de viajes nacionales registrados.")
        else:
            df_cl.columns = ['ID','Cliente','Celular','Viaje','Destino','Salida','Regreso',
                              'Adultos','Menores','Hab. Doble','Hab. Triple',
                              'Total','Abonado','Saldo','Estado','Vendedora']

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Clientes", len(df_cl))
            col2.metric("Total a cobrar",  f"${df_cl['Total'].sum():,.2f}")
            col3.metric("Total cobrado",   f"${df_cl['Abonado'].sum():,.2f}")
            col4.metric("Saldo pendiente", f"${df_cl['Saldo'].sum():,.2f}")

            st.divider()

            col1, col2, col3 = st.columns(3)
            with col1:
                filtro_vj = st.selectbox("Filtrar por viaje",
                    ["Todos"] + sorted(df_cl['Viaje'].unique().tolist()), key="cl_vj")
            with col2:
                filtro_vend_nac = st.selectbox("Filtrar por vendedora",
                    ["Todas"] + sorted(df_cl['Vendedora'].unique().tolist()), key="cl_vend")
            with col3:
                filtro_cl_buscar = st.text_input("🔍 Buscar cliente", key="cl_buscar")

            df_clf = df_cl.copy()
            if filtro_vj != "Todos":
                df_clf = df_clf[df_clf['Viaje'] == filtro_vj]
            if filtro_vend_nac != "Todas":
                df_clf = df_clf[df_clf['Vendedora'] == filtro_vend_nac]
            if filtro_cl_buscar:
                df_clf = df_clf[df_clf['Cliente'].str.contains(filtro_cl_buscar, case=False, na=False)]

            df_show_cl = df_clf.copy()
            for col in ['Total','Abonado','Saldo']:
                df_show_cl[col] = df_show_cl[col].apply(lambda x: f"${x:,.2f}")
            st.dataframe(df_show_cl, use_container_width=True, hide_index=True)

            # Detalle expandible
            st.divider()
            st.markdown("#### 🔍 Ver detalle de cliente")
            cl_ids = df_clf['ID'].tolist()
            if cl_ids:
                id_cl_sel = st.selectbox("Selecciona ID cliente:", cl_ids, key="cl_det")
                row_cl = df_clf[df_clf['ID'] == id_cl_sel].iloc[0]
                with st.expander(f"📄 {row_cl['Cliente']} | {row_cl['Viaje']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Cliente:** {row_cl['Cliente']}")
                    c1.write(f"**Celular:** {row_cl['Celular']}")
                    c1.write(f"**Vendedora:** {row_cl['Vendedora']}")
                    c2.write(f"**Viaje:** {row_cl['Viaje']}")
                    c2.write(f"**Destino:** {row_cl['Destino']}")
                    c2.write(f"**Salida:** {row_cl['Salida']} → {row_cl['Regreso']}")
                    c3.write(f"**Adultos:** {row_cl['Adultos']}  |  **Menores:** {row_cl['Menores']}")
                    c3.write(f"**Hab. Dobles:** {row_cl['Hab. Doble']}  |  **Hab. Triples:** {row_cl['Hab. Triple']}")
                    c3.write(f"**Estado:** {row_cl['Estado']}")

                    st.divider()
                    c1, c2, c3 = st.columns(3)
                    total_val   = df_clf[df_clf['ID']==id_cl_sel]['Total'].values[0]
                    abonado_val = df_clf[df_clf['ID']==id_cl_sel]['Abonado'].values[0]
                    saldo_val   = df_clf[df_clf['ID']==id_cl_sel]['Saldo'].values[0]
                    c1.metric("Total",   f"${total_val:,.2f}")
                    c2.metric("Abonado", f"${abonado_val:,.2f}")
                    c3.metric("Saldo",   f"${saldo_val:,.2f}")

                    conn2 = conectar_db()
                    try:
                        df_pas_n = pd.read_sql_query(
                            "SELECT nombre_completo, tipo, habitacion_asignada FROM pasajeros_nacionales WHERE cliente_id = ?",
                            conn2, params=(id_cl_sel,))
                        df_abonos_n = pd.read_sql_query(
                            "SELECT fecha, monto FROM abonos_nacionales WHERE cliente_id = ? ORDER BY fecha",
                            conn2, params=(id_cl_sel,))
                    except:
                        df_pas_n = pd.DataFrame()
                        df_abonos_n = pd.DataFrame()
                    conn2.close()

                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**🧳 Pasajeros:**")
                        if not df_pas_n.empty:
                            df_pas_n.columns = ['Nombre','Tipo','Habitación']
                            st.dataframe(df_pas_n, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin pasajeros registrados")
                    with c2:
                        st.markdown("**💳 Historial de abonos:**")
                        if not df_abonos_n.empty:
                            df_abonos_n['monto'] = df_abonos_n['monto'].apply(lambda x: f"${x:,.2f}")
                            df_abonos_n.columns = ['Fecha','Monto']
                            st.dataframe(df_abonos_n, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin abonos registrados")

            csv_cl = df_show_cl.to_csv(index=False).encode('utf-8')
            st.download_button("📥 Descargar CSV", csv_cl,
                file_name=f"clientes_nacionales_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # ── TAB 2: REGISTRAR CLIENTE ───────────────────────────────────────────
    with tabs[2]:
        formulario_registro_cliente_nacional()

    # ── TAB 3: REGISTRAR PAGO ─────────────────────────────────────────────
    with tabs[3]:
        st.subheader("💰 Registrar Pago - Viajes Nacionales")

        conn = conectar_db()
        try:
            if usuario["rol"] == "VENDEDORA":
                query_pago = f"""
                    SELECT cn.id, cn.nombre_cliente, vj.nombre_viaje, vj.destino,
                           cn.total_pagar, cn.total_abonado, cn.saldo
                    FROM clientes_nacionales cn
                    JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                    WHERE cn.estado NOT IN ('CERRADO','LIQUIDADO')
                    ORDER BY cn.id DESC
                """
            else:
                query_pago = """
                    SELECT cn.id, cn.nombre_cliente, vj.nombre_viaje, vj.destino,
                           cn.total_pagar, cn.total_abonado, cn.saldo
                    FROM clientes_nacionales cn
                    JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                    WHERE cn.estado NOT IN ('CERRADO','LIQUIDADO')
                    ORDER BY cn.id DESC
                """
            df_pago = pd.read_sql_query(query_pago, conn)
        except Exception as e:
            df_pago = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_pago.empty:
            st.info("No hay clientes con saldo pendiente.")
        else:
            df_pago.columns = ['id','cliente','viaje','destino','total_pagar','total_abonado','saldo']
            opciones_pago = {
                f"ID {row['id']} | {row['cliente']} — {row['viaje']} | Saldo: ${row['saldo']:,.2f}": row
                for _, row in df_pago.iterrows()
            }
            pago_label = st.selectbox("Selecciona el cliente:", list(opciones_pago.keys()), key="vn_pago_sel")
            pago_sel = opciones_pago[pago_label]

            col1, col2, col3 = st.columns(3)
            col1.metric("Total",          f"${pago_sel['total_pagar']:,.2f}")
            col2.metric("Ya abonado",     f"${pago_sel['total_abonado']:,.2f}")
            col3.metric("Saldo pendiente",f"${pago_sel['saldo']:,.2f}")

            st.divider()
            st.markdown("#### 💳 Datos del Abono")
            col1, col2 = st.columns(2)
            with col1:
                monto_pago_n = st.number_input("Monto del abono *",
                    min_value=0.01, max_value=float(pago_sel['saldo']),
                    step=100.0, format="%.2f", key="vn_monto_pago")
            with col2:
                metodo_pago_n = st.selectbox("Método de pago *",
                    ["Efectivo","Transferencia","Tarjeta de crédito",
                     "Tarjeta de débito","Depósito bancario","Otro"], key="vn_metodo_pago")

            nuevo_abonado_n = pago_sel['total_abonado'] + monto_pago_n
            nuevo_saldo_n   = pago_sel['saldo'] - monto_pago_n
            liquidado_n     = nuevo_saldo_n <= 0.01

            col1, col2 = st.columns(2)
            col1.metric("Nuevo total abonado", f"${nuevo_abonado_n:,.2f}", delta=f"+${monto_pago_n:,.2f}")
            col2.metric("Nuevo saldo",         f"${nuevo_saldo_n:,.2f}",  delta=f"-${monto_pago_n:,.2f}")

            if liquidado_n:
                st.success("🎉 ¡Con este pago el cliente quedará **LIQUIDADO**!")
            else:
                pct_n = (nuevo_abonado_n / pago_sel['total_pagar']) * 100
                st.info(f"📊 Avance: **{pct_n:.1f}%** cobrado")

            st.divider()
            if st.button("💾 Registrar Abono Nacional", type="primary", use_container_width=True):
                try:
                    conn = conectar_db()
                    cursor = conn.cursor()
                    fecha_pago_n = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    try:
                        cursor.execute(
                            "INSERT INTO abonos_nacionales (cliente_id, monto, fecha, vendedora_id, metodo_pago) VALUES (?,?,?,?,?)",
                            (int(pago_sel['id']), monto_pago_n, fecha_pago_n, usuario['id_vendedora'], metodo_pago_n)
                        )
                    except:
                        cursor.execute(
                            "INSERT INTO abonos_nacionales (cliente_id, monto, fecha, vendedora_id) VALUES (?,?,?,?)",
                            (int(pago_sel['id']), monto_pago_n, fecha_pago_n, usuario['id_vendedora'])
                        )

                    nuevo_estado_n = "LIQUIDADO" if liquidado_n else "ADEUDO"
                    cursor.execute("""
                        UPDATE clientes_nacionales
                        SET total_abonado = total_abonado + ?,
                            saldo = saldo - ?,
                            estado = ?
                        WHERE id = ?
                    """, (monto_pago_n, monto_pago_n, nuevo_estado_n, int(pago_sel['id'])))

                    conn.commit()
                    conn.close()

                    # ── Recibo ─────────────────────────────────────────────
                    num_rec_n   = _siguiente_num_recibo()
                    fecha_rec_n = datetime.now().strftime("%d-%B-%Y").replace(
                        "January","enero").replace("February","febrero").replace(
                        "March","marzo").replace("April","abril").replace(
                        "May","mayo").replace("June","junio").replace(
                        "July","julio").replace("August","agosto").replace(
                        "September","septiembre").replace("October","octubre").replace(
                        "November","noviembre").replace("December","diciembre")
                    vend_n      = usuario.get("nombre", "Agente")
                    concepto_n  = f"{'Liquidación' if liquidado_n else 'Abono'} - {pago_sel.get('viaje','Viaje Nacional')}"

                    if liquidado_n:
                        st.success(f"✅ ¡**{pago_sel['cliente']}** ha LIQUIDADO su viaje! 🎉")
                        st.balloons()
                    else:
                        st.success(f"✅ Abono de ${monto_pago_n:,.2f} ({metodo_pago_n}) registrado. Saldo: ${nuevo_saldo_n:,.2f}")

                    st.divider()
                    st.markdown("#### 🧾 Recibo de Pago")
                    _boton_recibo(
                        numero     = num_rec_n,
                        fecha_str  = fecha_rec_n,
                        cliente    = pago_sel['cliente'],
                        monto      = monto_pago_n,
                        concepto   = concepto_n,
                        forma_pago = metodo_pago_n,
                        agente     = vend_n,
                        key_suffix = f"nac_{pago_sel['id']}",
                    )

                except Exception as e:
                    st.error(f"❌ Error: {e}")

    # ── TAB 4: COTIZADOR (solo ADMIN) ──────────────────────────────────────
    if usuario["rol"] == "ADMIN":
        with tabs[4]:
            st.subheader("📝 Cotizador de Viajes Nacionales")
            subtab_cot = st.tabs(["➕ Nueva Cotización", "📋 Ver Cotizaciones"])

            # ── Nueva Cotización ────────────────────────────────────────────
            with subtab_cot[0]:
                st.markdown("#### Datos Generales del Viaje")
                col1, col2 = st.columns(2)
                with col1:
                    cot_nombre  = st.text_input("Nombre del viaje *", key="cot_nombre", placeholder="Ej: Cascadas de Agua Azul")
                    cot_destino = st.text_input("Destino *", key="cot_destino", placeholder="Ej: Chiapas, México")
                with col2:
                    cot_salida  = st.date_input("Fecha de salida *", key="cot_salida")
                    cot_regreso = st.date_input("Fecha de regreso *", key="cot_regreso")

                if cot_regreso > cot_salida:
                    cot_dias   = (cot_regreso - cot_salida).days + 1
                    cot_noches = cot_dias - 1
                    st.success(f"✅ {cot_dias} días / {cot_noches} noches")
                else:
                    cot_dias, cot_noches = 0, 0
                    st.error("❌ La fecha de regreso debe ser posterior a la salida")

                cot_personas = st.number_input("Personas proyectadas *", min_value=1, value=20, step=1, key="cot_personas")

                st.divider()

                # ── VUELOS ──────────────────────────────────────────────────
                st.markdown("#### ✈️ Vuelos")
                col1, col2 = st.columns(2)
                with col1:
                    cot_vuelo_costo  = st.number_input("Costo real por persona *", min_value=0.0, step=100.0, format="%.2f", key="cot_vc")
                with col2:
                    cot_vuelo_venta  = st.number_input("Precio venta por persona *", min_value=0.0, step=100.0, format="%.2f", key="cot_vv")

                if cot_vuelo_venta > 0:
                    gan_vuelo_pax   = cot_vuelo_venta - cot_vuelo_costo
                    gan_vuelo_total = gan_vuelo_pax * cot_personas
                    col1, col2 = st.columns(2)
                    col1.metric("Ganancia por persona", f"${gan_vuelo_pax:,.2f}")
                    col2.metric("Ganancia total vuelos", f"${gan_vuelo_total:,.2f}")
                else:
                    gan_vuelo_pax   = 0.0
                    gan_vuelo_total = 0.0

                st.divider()

                # ── HOSPEDAJE ───────────────────────────────────────────────
                st.markdown("#### 🏨 Hospedaje")
                num_hoteles = st.radio("¿Cuántos hoteles?", [1, 2, 3, 4], horizontal=True, key="cot_num_hoteles")

                hoteles_data = []
                costo_total_doble  = 0.0
                precio_total_doble = 0.0
                costo_total_triple  = 0.0
                precio_total_triple = 0.0

                for i in range(num_hoteles):
                    st.markdown(f"**🏨 Hotel {i+1}**")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        h_nombre = st.text_input(f"Nombre del hotel {i+1}", key=f"cot_h_nom{i}", placeholder="Ej: Hotel Palenque")
                    with col2:
                        h_noches = st.number_input(f"Noches en hotel {i+1}", min_value=1, value=cot_noches if cot_noches > 0 else 1, key=f"cot_h_noch{i}")
                    with col3:
                        st.write("")  # spacer

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**Base Doble** (precio por noche)")
                        h_costo_doble  = st.number_input("Costo real/noche", min_value=0.0, step=50.0, format="%.2f", key=f"cot_h_cd{i}")
                        h_precio_doble = st.number_input("Precio venta/noche", min_value=0.0, step=50.0, format="%.2f", key=f"cot_h_pd{i}")
                    with col2:
                        st.markdown("**Base Triple** (precio por noche)")
                        h_costo_triple  = st.number_input("Costo real/noche ", min_value=0.0, step=50.0, format="%.2f", key=f"cot_h_ct{i}")
                        h_precio_triple = st.number_input("Precio venta/noche ", min_value=0.0, step=50.0, format="%.2f", key=f"cot_h_pt{i}")

                    h_costo_doble_total  = h_costo_doble  * h_noches
                    h_precio_doble_total = h_precio_doble * h_noches
                    h_costo_triple_total  = h_costo_triple  * h_noches
                    h_precio_triple_total = h_precio_triple * h_noches

                    if h_precio_doble > 0 or h_precio_triple > 0:
                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric(f"Total doble ({h_noches}n)", f"${h_precio_doble_total:,.2f}")
                        col2.metric("Gan. doble/noche", f"${(h_precio_doble - h_costo_doble):,.2f}")
                        col3.metric(f"Total triple ({h_noches}n)", f"${h_precio_triple_total:,.2f}")
                        col4.metric("Gan. triple/noche", f"${(h_precio_triple - h_costo_triple):,.2f}")

                    hoteles_data.append({
                        "nombre": h_nombre,
                        "noches": h_noches,
                        "costo_doble": h_costo_doble_total,
                        "precio_doble": h_precio_doble_total,
                        "costo_triple": h_costo_triple_total,
                        "precio_triple": h_precio_triple_total
                    })

                    costo_total_doble  += h_costo_doble_total
                    precio_total_doble += h_precio_doble_total
                    costo_total_triple  += h_costo_triple_total
                    precio_total_triple += h_precio_triple_total

                st.divider()

                # ── TRASLADOS ───────────────────────────────────────────────
                st.markdown("#### 🚌 Traslados / Transporte")
                col1, col2 = st.columns(2)
                with col1:
                    cot_traslado_costo_total = st.number_input("Costo TOTAL del transporte *", min_value=0.0, step=500.0, format="%.2f", key="cot_tc")
                with col2:
                    cot_traslado_venta_pax   = st.number_input("Precio venta por persona *",   min_value=0.0, step=100.0, format="%.2f", key="cot_tv")

                if cot_personas > 0 and cot_traslado_costo_total > 0:
                    cot_traslado_costo_pax = cot_traslado_costo_total / cot_personas
                    gan_traslado_pax       = cot_traslado_venta_pax - cot_traslado_costo_pax
                    gan_traslado_total     = gan_traslado_pax * cot_personas
                    col1, col2, col3 = st.columns(3)
                    col1.metric("Costo real por persona", f"${cot_traslado_costo_pax:,.2f}")
                    col2.metric("Ganancia por persona",   f"${gan_traslado_pax:,.2f}")
                    col3.metric("Ganancia total traslados", f"${gan_traslado_total:,.2f}")
                else:
                    cot_traslado_costo_pax = 0.0
                    gan_traslado_pax       = 0.0
                    gan_traslado_total     = 0.0

                st.divider()

                # ── ENTRADAS / TOURS ────────────────────────────────────────
                st.markdown("#### 🎫 Entradas / Tours / Atracciones")
                col1, col2 = st.columns(2)
                with col1:
                    cot_entradas_costo = st.number_input("Costo real por persona *", min_value=0.0, step=50.0, format="%.2f", key="cot_ec")
                with col2:
                    cot_entradas_venta = st.number_input("Precio venta por persona *", min_value=0.0, step=50.0, format="%.2f", key="cot_ev")

                if cot_entradas_venta > 0:
                    gan_entradas_pax   = cot_entradas_venta - cot_entradas_costo
                    gan_entradas_total = gan_entradas_pax * cot_personas
                    col1, col2 = st.columns(2)
                    col1.metric("Ganancia por persona", f"${gan_entradas_pax:,.2f}")
                    col2.metric("Ganancia total entradas", f"${gan_entradas_total:,.2f}")
                else:
                    gan_entradas_pax   = 0.0
                    gan_entradas_total = 0.0

                st.divider()

                # ── RESUMEN FINAL ───────────────────────────────────────────
                precio_pax_doble  = cot_vuelo_venta + precio_total_doble  + cot_traslado_venta_pax + cot_entradas_venta
                precio_pax_triple = cot_vuelo_venta + precio_total_triple + cot_traslado_venta_pax + cot_entradas_venta

                inversion_vuelos    = cot_vuelo_costo * cot_personas
                inversion_hoteles   = costo_total_doble * (cot_personas / 2) if cot_personas > 0 else 0
                inversion_traslados = cot_traslado_costo_total
                inversion_entradas  = cot_entradas_costo * cot_personas
                inversion_total     = inversion_vuelos + inversion_hoteles + inversion_traslados + inversion_entradas

                ganancia_proyectada = (
                    gan_vuelo_total + gan_traslado_total + gan_entradas_total +
                    ((precio_total_doble - costo_total_doble) * (cot_personas / 2))
                )

                if precio_pax_doble > 0 or precio_pax_triple > 0:
                    st.markdown("### 📊 Resumen de Cotización")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**🎯 Precio por persona — Base DOBLE**")
                        desglose_doble = {
                            "✈️ Vuelo":      cot_vuelo_venta,
                            "🏨 Hotel":      precio_total_doble,
                            "🚌 Traslados":  cot_traslado_venta_pax,
                            "🎫 Entradas":   cot_entradas_venta,
                        }
                        for concepto, valor in desglose_doble.items():
                            col_a, col_b = st.columns([2,1])
                            col_a.write(concepto)
                            col_b.write(f"${valor:,.2f}")
                        st.markdown(f"**TOTAL DOBLE: ${precio_pax_doble:,.2f}**")

                    with col2:
                        st.markdown("**🎯 Precio por persona — Base TRIPLE**")
                        desglose_triple = {
                            "✈️ Vuelo":      cot_vuelo_venta,
                            "🏨 Hotel":      precio_total_triple,
                            "🚌 Traslados":  cot_traslado_venta_pax,
                            "🎫 Entradas":   cot_entradas_venta,
                        }
                        for concepto, valor in desglose_triple.items():
                            col_a, col_b = st.columns([2,1])
                            col_a.write(concepto)
                            col_b.write(f"${valor:,.2f}")
                        st.markdown(f"**TOTAL TRIPLE: ${precio_pax_triple:,.2f}**")

                    st.divider()
                    col1, col2 = st.columns(2)
                    col1.metric("💸 Inversión total estimada", f"${inversion_total:,.2f}")
                    col2.metric("📈 Ganancia proyectada",      f"${ganancia_proyectada:,.2f}")

                    st.divider()
                    if st.button("💾 Guardar Cotización", type="primary", use_container_width=True, key="cot_guardar"):
                        errores_cot = []
                        if not cot_nombre.strip():  errores_cot.append("El nombre del viaje es obligatorio.")
                        if not cot_destino.strip(): errores_cot.append("El destino es obligatorio.")
                        if cot_dias <= 0:           errores_cot.append("Las fechas no son válidas.")
                        if precio_pax_doble <= 0:   errores_cot.append("El precio base doble debe ser mayor a 0.")

                        if errores_cot:
                            for e in errores_cot:
                                st.error(f"❌ {e}")
                        else:
                            try:
                                conn = conectar_db()
                                cursor = conn.cursor()
                                fecha_reg_cot = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                                cursor.execute("""
                                    INSERT INTO cotizaciones_nacionales (
                                        nombre_viaje, destino, fecha_salida, fecha_regreso, dias, noches,
                                        personas_proyectadas, costo_vuelo_real, precio_vuelo_venta,
                                        costo_traslados_total, precio_traslados_persona,
                                        costo_entradas_real, precio_entradas_venta,
                                        precio_persona_doble, precio_persona_triple,
                                        inversion_total, ganancia_proyectada, estado, fecha_registro
                                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'BORRADOR',?)
                                """, (
                                    cot_nombre.strip(), cot_destino.strip(),
                                    cot_salida.strftime("%d-%m-%Y"), cot_regreso.strftime("%d-%m-%Y"),
                                    cot_dias, cot_noches,
                                    cot_personas, cot_vuelo_costo, cot_vuelo_venta,
                                    cot_traslado_costo_total, cot_traslado_venta_pax,
                                    cot_entradas_costo, cot_entradas_venta,
                                    precio_pax_doble, precio_pax_triple,
                                    inversion_total, ganancia_proyectada,
                                    fecha_reg_cot
                                ))
                                cot_id = cursor.lastrowid

                                for hotel in hoteles_data:
                                    if hotel["nombre"].strip():
                                        cursor.execute("""
                                            INSERT INTO hoteles_cotizacion (
                                                cotizacion_id, nombre_hotel, noches,
                                                costo_doble_real, precio_doble_venta,
                                                costo_triple_real, precio_triple_venta
                                            ) VALUES (?,?,?,?,?,?,?)
                                        """, (
                                            cot_id, hotel["nombre"].strip(), hotel["noches"],
                                            hotel["costo_doble"],  hotel["precio_doble"],
                                            hotel["costo_triple"], hotel["precio_triple"]
                                        ))

                                conn.commit()
                                conn.close()
                                st.success(f"✅ Cotización **{cot_nombre}** guardada con ID: {cot_id}")
                                st.info("💡 Ahora puedes crear el viaje desde esta cotización en **Gestionar Viajes → Desde Cotización**")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error al guardar: {e}")
                                st.caption("💡 Verifica que existan las tablas `cotizaciones_nacionales` y `hoteles_cotizacion` en la BD.")

            # ── Ver Cotizaciones ────────────────────────────────────────────
            with subtab_cot[1]:
                st.markdown("#### 📋 Cotizaciones Registradas")
                conn = conectar_db()
                try:
                    df_cots = pd.read_sql_query("""
                        SELECT id, nombre_viaje, destino, fecha_salida, fecha_regreso,
                               dias, noches, personas_proyectadas,
                               precio_persona_doble, precio_persona_triple,
                               inversion_total, ganancia_proyectada, estado
                        FROM cotizaciones_nacionales
                        ORDER BY id DESC
                    """, conn)
                except:
                    df_cots = pd.DataFrame()
                conn.close()

                if df_cots.empty:
                    st.info("No hay cotizaciones registradas aún.")
                else:
                    df_cots_show = df_cots.copy()
                    for col in ['precio_persona_doble','precio_persona_triple','inversion_total','ganancia_proyectada']:
                        df_cots_show[col] = df_cots_show[col].apply(lambda x: f"${x:,.2f}")
                    df_cots_show.columns = ['ID','Viaje','Destino','Salida','Regreso',
                                            'Días','Noches','Personas',
                                            'P. Doble p/p','P. Triple p/p',
                                            'Inversión','Ganancia Proy.','Estado']
                    st.dataframe(df_cots_show, use_container_width=True, hide_index=True)

                    # Detalle
                    st.divider()
                    cot_ids = df_cots['id'].tolist()
                    id_cot_det = st.selectbox("Ver detalle de cotización:", cot_ids, key="cot_det_id")
                    row_cot = df_cots[df_cots['id'] == id_cot_det].iloc[0]

                    with st.expander(f"📄 {row_cot['nombre_viaje']} — {row_cot['destino']}", expanded=True):
                        c1, c2, c3 = st.columns(3)
                        c1.write(f"**Viaje:** {row_cot['nombre_viaje']}")
                        c1.write(f"**Destino:** {row_cot['destino']}")
                        c1.write(f"**Fechas:** {row_cot['fecha_salida']} → {row_cot['fecha_regreso']}")
                        c2.metric("Días / Noches", f"{row_cot['dias']}d / {row_cot['noches']}n")
                        c2.metric("Personas proyectadas", row_cot['personas_proyectadas'])
                        c3.metric("Precio doble p/p",  f"${row_cot['precio_persona_doble']:,.2f}")
                        c3.metric("Precio triple p/p", f"${row_cot['precio_persona_triple']:,.2f}")

                        st.divider()
                        col1, col2 = st.columns(2)
                        col1.metric("💸 Inversión total", f"${row_cot['inversion_total']:,.2f}")
                        col2.metric("📈 Ganancia proyectada", f"${row_cot['ganancia_proyectada']:,.2f}")

                        # Hoteles de la cotización
                        conn2 = conectar_db()
                        try:
                            df_hots = pd.read_sql_query(
                                "SELECT nombre_hotel, noches, costo_doble_real, precio_doble_venta, costo_triple_real, precio_triple_venta FROM hoteles_cotizacion WHERE cotizacion_id = ?",
                                conn2, params=(id_cot_det,))
                        except:
                            df_hots = pd.DataFrame()
                        conn2.close()

                        if not df_hots.empty:
                            st.markdown("**🏨 Hoteles incluidos:**")
                            df_hots.columns = ['Hotel','Noches','Costo Doble','P. Doble','Costo Triple','P. Triple']
                            for col in ['Costo Doble','P. Doble','Costo Triple','P. Triple']:
                                df_hots[col] = df_hots[col].apply(lambda x: f"${x:,.2f}")
                            st.dataframe(df_hots, hide_index=True, use_container_width=True)

    # ── TAB 5: GESTIONAR VIAJES (solo ADMIN) ──────────────────────────────
    if usuario["rol"] == "ADMIN":
        with tabs[5]:
            st.subheader("⚙️ Gestionar Viajes Nacionales")
            subtab_g = st.tabs(["➕ Nuevo Viaje", "📋 Desde Cotización", "✏️ Editar / Cerrar Viaje"])

            # ── Nuevo viaje manual ──────────────────────────────────────────
            with subtab_g[0]:
                st.markdown("#### ✈️ Registrar Nuevo Viaje Nacional (Manual)")
                col1, col2 = st.columns(2)
                with col1:
                    g_nombre  = st.text_input("Nombre del viaje *", key="g_nombre", placeholder="Ej: Cascadas de Agua Azul")
                    g_destino = st.text_input("Destino *", key="g_destino", placeholder="Ej: Chiapas, México")
                    g_salida  = st.date_input("Fecha de salida *", key="g_salida")
                    g_regreso = st.date_input("Fecha de regreso *", key="g_regreso")
                with col2:
                    g_cupos    = st.number_input("Cupos totales *", min_value=1, value=20, key="g_cupos")
                    g_p_doble  = st.number_input("Precio por persona (base doble) *", min_value=0.0, step=100.0, format="%.2f", key="g_p_doble")
                    g_p_triple = st.number_input("Precio por persona (base triple)",  min_value=0.0, step=100.0, format="%.2f", key="g_p_triple")

                if g_regreso > g_salida:
                    dias_g   = (g_regreso - g_salida).days + 1
                    noches_g = dias_g - 1
                    st.info(f"📅 {dias_g} días / {noches_g} noches")

                if st.button("💾 Registrar Viaje", type="primary", use_container_width=True, key="g_registrar"):
                    errores_g = []
                    if not g_nombre.strip():  errores_g.append("El nombre del viaje es obligatorio.")
                    if not g_destino.strip(): errores_g.append("El destino es obligatorio.")
                    if g_regreso <= g_salida: errores_g.append("La fecha de regreso debe ser posterior a la salida.")
                    if g_p_doble <= 0:        errores_g.append("El precio base doble debe ser mayor a 0.")

                    if errores_g:
                        for e in errores_g:
                            st.error(f"❌ {e}")
                    else:
                        try:
                            dias_g   = (g_regreso - g_salida).days + 1
                            noches_g = dias_g - 1
                            conn = conectar_db()
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO viajes_nacionales (
                                    nombre_viaje, destino, fecha_salida, fecha_regreso,
                                    dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles,
                                    precio_persona_doble, precio_persona_triple, estado, fecha_registro
                                ) VALUES (?,?,?,?,?,?,?,0,?,?,?,'ACTIVO',?)
                            """, (
                                g_nombre.strip(), g_destino.strip(),
                                g_salida.strftime("%d-%m-%Y"), g_regreso.strftime("%d-%m-%Y"),
                                dias_g, noches_g, g_cupos, g_cupos,
                                g_p_doble, g_p_triple,
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            ))
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Viaje **{g_nombre}** registrado con {g_cupos} cupos.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")

            # ── Cargar desde cotización ─────────────────────────────────────
            with subtab_g[1]:
                st.markdown("#### 📋 Crear Viaje desde Cotización")
                cotizaciones = obtener_cotizaciones_nacionales()
                if not cotizaciones:
                    st.info("No hay cotizaciones disponibles en la tabla `cotizaciones`.")
                else:
                    opts_cot = {
                        f"ID {c['id']} | {c['nombre_viaje']} | {c['destino']} | {c['fecha_salida']} → {c['fecha_regreso']}": c
                        for c in cotizaciones
                    }
                    cot_label = st.selectbox("Selecciona cotización:", list(opts_cot.keys()), key="cot_sel")
                    cot_sel = opts_cot[cot_label]

                    col1, col2 = st.columns(2)
                    col1.write(f"**Nombre:** {cot_sel['nombre_viaje']}")
                    col1.write(f"**Destino:** {cot_sel['destino']}")
                    col1.write(f"**Fechas:** {cot_sel['fecha_salida']} → {cot_sel['fecha_regreso']}")
                    col2.metric("Personas proyectadas", cot_sel['personas_proyectadas'])
                    col2.metric("Precio doble p/p",   f"${cot_sel['precio_doble']:,.2f}")
                    col2.metric("Precio triple p/p",  f"${cot_sel['precio_triple']:,.2f}")

                    cupos_override = st.number_input("Cupos a registrar (editable)",
                        min_value=1, value=int(cot_sel['personas_proyectadas']), key="cot_cupos")

                    if st.button("💾 Crear Viaje desde Cotización", type="primary", key="cot_crear"):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO viajes_nacionales (
                                    cotizacion_id, nombre_viaje, destino, fecha_salida, fecha_regreso,
                                    dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles,
                                    precio_persona_doble, precio_persona_triple, estado, fecha_registro
                                ) VALUES (?,?,?,?,?,?,?,?,0,?,?,?,'ACTIVO',?)
                            """, (
                                cot_sel['id'], cot_sel['nombre_viaje'], cot_sel['destino'],
                                cot_sel['fecha_salida'], cot_sel['fecha_regreso'],
                                cot_sel['dias'], cot_sel['noches'],
                                cupos_override, cupos_override,
                                cot_sel['precio_doble'], cot_sel['precio_triple'],
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            ))
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Viaje **{cot_sel['nombre_viaje']}** creado desde cotización con {cupos_override} cupos.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")

            # ── Editar / Cerrar viaje ───────────────────────────────────────
            with subtab_g[2]:
                st.markdown("#### ✏️ Editar o Cerrar un Viaje")
                conn = conectar_db()
                try:
                    df_edit = pd.read_sql_query(
                        "SELECT id, nombre_viaje, destino, cupos_totales, cupos_vendidos, cupos_disponibles, precio_persona_doble, precio_persona_triple, estado FROM viajes_nacionales ORDER BY id DESC",
                        conn)
                except:
                    df_edit = pd.DataFrame()
                conn.close()

                if df_edit.empty:
                    st.info("No hay viajes para editar.")
                else:
                    viaje_edit_opts = {
                        f"ID {row['id']} | {row['nombre_viaje']} ({row['estado']})": row
                        for _, row in df_edit.iterrows()
                    }
                    edit_label = st.selectbox("Selecciona viaje:", list(viaje_edit_opts.keys()), key="edit_viaje")
                    edit_row = viaje_edit_opts[edit_label]

                    col1, col2 = st.columns(2)
                    with col1:
                        edit_cupos   = st.number_input("Cupos totales",
                            min_value=int(edit_row['cupos_vendidos']),
                            value=int(edit_row['cupos_totales']), key="edit_cupos")
                        edit_p_doble = st.number_input("Precio doble p/p",
                            min_value=0.0, value=float(edit_row['precio_persona_doble']),
                            step=100.0, format="%.2f", key="edit_pd")
                    with col2:
                        edit_p_triple = st.number_input("Precio triple p/p",
                            min_value=0.0, value=float(edit_row['precio_persona_triple']),
                            step=100.0, format="%.2f", key="edit_pt")
                        edit_estado = st.selectbox("Estado",
                            ["ACTIVO","AGOTADO","CERRADO"],
                            index=["ACTIVO","AGOTADO","CERRADO"].index(edit_row['estado']),
                            key="edit_estado")

                    if st.button("💾 Guardar Cambios", type="primary", key="edit_guardar"):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            nuevos_disp = edit_cupos - int(edit_row['cupos_vendidos'])
                            cursor.execute("""
                                UPDATE viajes_nacionales
                                SET cupos_totales = ?, cupos_disponibles = ?,
                                    precio_persona_doble = ?, precio_persona_triple = ?,
                                    estado = ?
                                WHERE id = ?
                            """, (edit_cupos, max(0, nuevos_disp),
                                  edit_p_doble, edit_p_triple,
                                  edit_estado, int(edit_row['id'])))
                            conn.commit()
                            conn.close()
                            st.success("✅ Viaje actualizado correctamente.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")

def pagina_viajes_internacionales():
    """Página completa de viajes internacionales"""
    st.title("🌎 Viajes Internacionales")
    usuario = st.session_state.usuario_actual

    tabs_labels = ["📋 Ver Viajes", "👥 Ver Clientes", "➕ Registrar Cliente", "💰 Registrar Pago"]
    if usuario["rol"] == "ADMIN":
        tabs_labels.append("⚙️ Gestionar Viajes")

    tabs = st.tabs(tabs_labels)

    # ── TAB 0: VER VIAJES ──────────────────────────────────────────────────
    with tabs[0]:
        st.subheader("Viajes Internacionales")
        conn = conectar_db()
        try:
            df_vi = pd.read_sql_query("""
                SELECT id, destino, fecha_salida, fecha_regreso, dias, noches,
                       cupos_totales, cupos_vendidos, cupos_disponibles,
                       precio_adulto_doble_usd, precio_adulto_triple_usd,
                       precio_menor_doble_usd, precio_menor_triple_usd,
                       porcentaje_ganancia, estado
                FROM viajes_internacionales
                ORDER BY fecha_salida
            """, conn)
        except:
            df_vi = pd.DataFrame()
        conn.close()

        if df_vi.empty:
            st.info("ℹ️ No hay viajes internacionales registrados aún.")
        else:
            activos = df_vi[df_vi["estado"] == "ACTIVO"]
            col1, col2, col3 = st.columns(3)
            col1.metric("Viajes activos", len(activos))
            col2.metric("Cupos disponibles", int(activos["cupos_disponibles"].sum()) if not activos.empty else 0)
            col3.metric("Total viajes", len(df_vi))

            st.divider()
            filtro_vi = st.selectbox("Filtrar por estado:",
                ["Todos"] + sorted(df_vi["estado"].unique().tolist()), key="vi_filtro")
            df_vif = df_vi if filtro_vi == "Todos" else df_vi[df_vi["estado"] == filtro_vi]

            df_vi_show = df_vif.copy()
            for col in ["precio_adulto_doble_usd","precio_adulto_triple_usd",
                        "precio_menor_doble_usd","precio_menor_triple_usd"]:
                df_vi_show[col] = df_vi_show[col].apply(lambda x: f"${x:,.2f}")
            df_vi_show.columns = ["ID","Destino","Salida","Regreso","Días","Noches",
                                   "Cupos Total","Vendidos","Disponibles",
                                   "Ad. Doble","Ad. Triple","Men. Doble","Men. Triple",
                                   "% Gan.","Estado"]
            st.dataframe(df_vi_show, use_container_width=True, hide_index=True)

            # Detalle
            st.divider()
            st.markdown("#### 🔍 Detalle de un viaje")
            vi_ids = df_vif["id"].tolist()
            if vi_ids:
                id_vi_sel = st.selectbox("Selecciona ID:", vi_ids, key="vi_det_id")
                row_vi = df_vif[df_vif["id"] == id_vi_sel].iloc[0]
                with st.expander(f"📄 {row_vi['destino']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Destino:** {row_vi['destino']}")
                    c1.write(f"**Salida:** {row_vi['fecha_salida']}  →  **Regreso:** {row_vi['fecha_regreso']}")
                    c1.write(f"**Duración:** {row_vi['dias']} días / {row_vi['noches']} noches")
                    c2.metric("Cupos totales",     row_vi["cupos_totales"])
                    c2.metric("Cupos vendidos",    row_vi["cupos_vendidos"])
                    c3.metric("Cupos disponibles", row_vi["cupos_disponibles"])
                    c3.metric("% Ganancia",        f"{row_vi['porcentaje_ganancia']}%")

                    st.markdown("**💵 Precios (USD):**")
                    pc1, pc2, pc3, pc4 = st.columns(4)
                    pc1.metric("Adulto Doble",  f"${row_vi['precio_adulto_doble_usd']:,.2f}")
                    pc2.metric("Adulto Triple", f"${row_vi['precio_adulto_triple_usd']:,.2f}")
                    pc3.metric("Menor Doble",   f"${row_vi['precio_menor_doble_usd']:,.2f}")
                    pc4.metric("Menor Triple",  f"${row_vi['precio_menor_triple_usd']:,.2f}")

                    if row_vi["cupos_totales"] > 0:
                        pct_vi = row_vi["cupos_vendidos"] / row_vi["cupos_totales"]
                        st.markdown(f"**Ocupación: {pct_vi*100:.0f}%**")
                        st.progress(pct_vi)

    # ── TAB 1: VER CLIENTES ────────────────────────────────────────────────
    with tabs[1]:
        st.subheader("Clientes de Viajes Internacionales")
        conn = conectar_db()
        try:
            if usuario["rol"] == "VENDEDORA":
                query_ci = f"""
                    SELECT ci.id, ci.nombre_cliente,
                           vi.destino, vi.fecha_salida, vi.fecha_regreso,
                           ci.adultos, ci.menores,
                           ci.habitaciones_doble, ci.habitaciones_triple,
                           ci.total_usd, ci.abonado_usd, ci.saldo_usd,
                           ci.ganancia_usd, ci.estado
                    FROM clientes_internacionales ci
                    JOIN viajes_internacionales vi ON ci.viaje_id = vi.id
                    WHERE ci.estado != 'CERRADO'
                    ORDER BY ci.id DESC
                """
            else:
                query_ci = """
                    SELECT ci.id, ci.nombre_cliente,
                           vi.destino, vi.fecha_salida, vi.fecha_regreso,
                           ci.adultos, ci.menores,
                           ci.habitaciones_doble, ci.habitaciones_triple,
                           ci.total_usd, ci.abonado_usd, ci.saldo_usd,
                           ci.ganancia_usd, ci.estado
                    FROM clientes_internacionales ci
                    JOIN viajes_internacionales vi ON ci.viaje_id = vi.id
                    WHERE ci.estado != 'CERRADO'
                    ORDER BY ci.id DESC
                """
            df_ci = pd.read_sql_query(query_ci, conn)
        except Exception as e:
            df_ci = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_ci.empty:
            st.info("No hay clientes de viajes internacionales registrados.")
        else:
            df_ci.columns = ["ID","Cliente","Destino","Salida","Regreso",
                              "Adultos","Menores","Hab.Doble","Hab.Triple",
                              "Total USD","Abonado USD","Saldo USD","Ganancia USD","Estado"]

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Clientes",           len(df_ci))
            col2.metric("Total a cobrar",     f"${df_ci['Total USD'].sum():,.2f} USD")
            col3.metric("Total cobrado",      f"${df_ci['Abonado USD'].sum():,.2f} USD")
            col4.metric("Saldo pendiente",    f"${df_ci['Saldo USD'].sum():,.2f} USD")

            st.divider()
            col1, col2 = st.columns(2)
            with col1:
                filtro_ci_dest = st.selectbox("Filtrar por destino",
                    ["Todos"] + sorted(df_ci["Destino"].unique().tolist()), key="ci_dest")
            with col2:
                filtro_ci_buscar = st.text_input("🔍 Buscar cliente", key="ci_buscar")

            df_cif = df_ci.copy()
            if filtro_ci_dest != "Todos":
                df_cif = df_cif[df_cif["Destino"] == filtro_ci_dest]
            if filtro_ci_buscar:
                df_cif = df_cif[df_cif["Cliente"].str.contains(filtro_ci_buscar, case=False, na=False)]

            df_ci_show = df_cif.copy()
            for col in ["Total USD","Abonado USD","Saldo USD","Ganancia USD"]:
                df_ci_show[col] = df_ci_show[col].apply(lambda x: f"${x:,.2f}")
            st.dataframe(df_ci_show, use_container_width=True, hide_index=True)

            # Detalle expandible
            st.divider()
            st.markdown("#### 🔍 Ver detalle de cliente")
            ci_ids = df_cif["ID"].tolist()
            if ci_ids:
                id_ci_sel = st.selectbox("Selecciona ID cliente:", ci_ids, key="ci_det")
                row_ci = df_cif[df_cif["ID"] == id_ci_sel].iloc[0]
                with st.expander(f"📄 {row_ci['Cliente']} | {row_ci['Destino']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Cliente:** {row_ci['Cliente']}")
                    c1.write(f"**Destino:** {row_ci['Destino']}")
                    c1.write(f"**Salida:** {row_ci['Salida']} → {row_ci['Regreso']}")
                    c2.write(f"**Adultos:** {row_ci['Adultos']}  |  **Menores:** {row_ci['Menores']}")
                    c2.write(f"**Hab. Dobles:** {row_ci['Hab.Doble']}  |  **Hab. Triples:** {row_ci['Hab.Triple']}")
                    c3.write(f"**Estado:** {row_ci['Estado']}")

                    st.divider()
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total",    f"${df_cif[df_cif['ID']==id_ci_sel]['Total USD'].values[0]:,.2f} USD")
                    c2.metric("Abonado",  f"${df_cif[df_cif['ID']==id_ci_sel]['Abonado USD'].values[0]:,.2f} USD")
                    c3.metric("Saldo",    f"${df_cif[df_cif['ID']==id_ci_sel]['Saldo USD'].values[0]:,.2f} USD")
                    c4.metric("Ganancia", f"${df_cif[df_cif['ID']==id_ci_sel]['Ganancia USD'].values[0]:,.2f} USD")

                    conn2 = conectar_db()
                    try:
                        df_pas_i = pd.read_sql_query(
                            "SELECT nombre_completo, tipo, habitacion_asignada FROM pasajeros_internacionales WHERE cliente_id = ?",
                            conn2, params=(id_ci_sel,))
                        df_abonos_i = pd.read_sql_query(
                            "SELECT fecha, moneda, monto_original, tipo_cambio, monto_usd FROM abonos_internacionales WHERE cliente_id = ? ORDER BY fecha",
                            conn2, params=(id_ci_sel,))
                    except:
                        df_pas_i = pd.DataFrame()
                        df_abonos_i = pd.DataFrame()
                    conn2.close()

                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**🧳 Pasajeros:**")
                        if not df_pas_i.empty:
                            df_pas_i.columns = ["Nombre","Tipo","Habitación"]
                            st.dataframe(df_pas_i, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin pasajeros registrados")
                    with c2:
                        st.markdown("**💳 Historial de abonos:**")
                        if not df_abonos_i.empty:
                            df_abonos_i.columns = ["Fecha","Moneda","Monto Original","T.C.","Monto USD"]
                            df_abonos_i["Monto Original"] = df_abonos_i.apply(
                                lambda r: f"${r['Monto Original']:,.2f} {r['Moneda']}", axis=1)
                            df_abonos_i["Monto USD"] = df_abonos_i["Monto USD"].apply(lambda x: f"${x:,.2f} USD")
                            df_abonos_i["T.C."] = df_abonos_i.apply(
                                lambda r: f"{r['T.C.']:.2f}" if r["Moneda"] == "MXN" else "—", axis=1)
                            st.dataframe(df_abonos_i[["Fecha","Monto Original","T.C.","Monto USD"]],
                                         hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin abonos registrados")

            csv_ci = df_ci_show.to_csv(index=False).encode("utf-8")
            st.download_button("📥 Descargar CSV", csv_ci,
                file_name=f"clientes_internacionales_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # ── TAB 2: REGISTRAR CLIENTE ───────────────────────────────────────────
    with tabs[2]:
        st.subheader("➕ Registrar Cliente en Viaje Internacional")

        conn = conectar_db()
        try:
            viajes_int = pd.read_sql_query("""
                SELECT id, destino, fecha_salida, fecha_regreso, cupos_disponibles,
                       precio_adulto_doble_usd, precio_adulto_triple_usd,
                       precio_menor_doble_usd, precio_menor_triple_usd,
                       porcentaje_ganancia
                FROM viajes_internacionales
                WHERE estado = 'ACTIVO' AND cupos_disponibles > 0
                ORDER BY fecha_salida
            """, conn)
        except:
            viajes_int = pd.DataFrame()
        conn.close()

        if viajes_int.empty:
            st.warning("⚠️ No hay viajes internacionales activos con cupos disponibles.")
        else:
            # Selección de viaje
            st.markdown("#### 🌎 Selecciona el Viaje")
            viajes_int_dict = {
                f"{row['destino']} | {row['fecha_salida']} → {row['fecha_regreso']} | Disp: {row['cupos_disponibles']}": row
                for _, row in viajes_int.iterrows()
            }
            vi_label = st.selectbox("Viaje disponible:", list(viajes_int_dict.keys()), key="ri_viaje")
            vi_sel = viajes_int_dict[vi_label]

            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Cupos disp.", vi_sel["cupos_disponibles"])
            col2.metric("Ad. Doble",   f"${vi_sel['precio_adulto_doble_usd']:,.2f}")
            col3.metric("Ad. Triple",  f"${vi_sel['precio_adulto_triple_usd']:,.2f}")
            col4.metric("Men. Doble",  f"${vi_sel['precio_menor_doble_usd']:,.2f}")
            col5.metric("Men. Triple", f"${vi_sel['precio_menor_triple_usd']:,.2f}")

            st.divider()

            # Datos del cliente
            st.markdown("#### 👤 Datos del Cliente")
            nombre_cliente_i = st.text_input("Nombre del cliente *", key="ri_nombre")

            st.divider()

            # Habitaciones
            st.markdown("#### 🛏️ Distribución de Habitaciones")
            col1, col2 = st.columns(2)
            with col1:
                ri_hab_dobles = st.number_input(
                    f"Habitaciones dobles (2 adultos × ${vi_sel['precio_adulto_doble_usd']:,.2f} USD)",
                    min_value=0, value=1, step=1, key="ri_dobles")
            with col2:
                ri_hab_triples = st.number_input(
                    f"Habitaciones triples (3 adultos × ${vi_sel['precio_adulto_triple_usd']:,.2f} USD)",
                    min_value=0, value=0, step=1, key="ri_triples")

            adultos_doble_n  = ri_hab_dobles * 2
            adultos_triple_n = ri_hab_triples * 3
            ri_adultos       = adultos_doble_n + adultos_triple_n

            if ri_adultos > vi_sel["cupos_disponibles"]:
                st.error(f"❌ Solo hay {vi_sel['cupos_disponibles']} cupos. Reduce las habitaciones.")
                st.stop()

            ri_menores = st.number_input("Número de menores", min_value=0, value=0, step=1, key="ri_menores")

            # Distribución de menores si hay
            ri_menores_doble  = 0
            ri_menores_triple = 0
            if ri_menores > 0:
                st.markdown("**👶 Distribución de menores por tipo de habitación:**")
                col1, col2 = st.columns(2)
                with col1:
                    ri_menores_doble = st.number_input(
                        f"Menores en hab. dobles (× ${vi_sel['precio_menor_doble_usd']:,.2f} USD)",
                        min_value=0, max_value=ri_menores, value=ri_menores, step=1, key="ri_md")
                with col2:
                    ri_menores_triple = st.number_input(
                        f"Menores en hab. triples (× ${vi_sel['precio_menor_triple_usd']:,.2f} USD)",
                        min_value=0, max_value=ri_menores, value=0, step=1, key="ri_mt")

                if ri_menores_doble + ri_menores_triple != ri_menores:
                    st.error(f"❌ La suma de menores ({ri_menores_doble + ri_menores_triple}) debe ser {ri_menores}.")

            st.divider()

            # Cálculo de precios USD
            st.markdown("#### 💵 Resumen de Precios (USD)")
            total_ad_doble  = adultos_doble_n  * vi_sel["precio_adulto_doble_usd"]
            total_ad_triple = adultos_triple_n * vi_sel["precio_adulto_triple_usd"]
            total_men_doble  = ri_menores_doble  * vi_sel["precio_menor_doble_usd"]
            total_men_triple = ri_menores_triple * vi_sel["precio_menor_triple_usd"]
            total_usd_ri = total_ad_doble + total_ad_triple + total_men_doble + total_men_triple
            ganancia_usd_ri = total_usd_ri * (vi_sel["porcentaje_ganancia"] / 100)

            if ri_hab_dobles > 0:
                col1, col2 = st.columns(2)
                col1.metric(f"{adultos_doble_n} adultos doble × ${vi_sel['precio_adulto_doble_usd']:,.2f}",
                            f"${total_ad_doble:,.2f}")
                if ri_menores_doble > 0:
                    col2.metric(f"{ri_menores_doble} menores doble × ${vi_sel['precio_menor_doble_usd']:,.2f}",
                                f"${total_men_doble:,.2f}")
            if ri_hab_triples > 0:
                col1, col2 = st.columns(2)
                col1.metric(f"{adultos_triple_n} adultos triple × ${vi_sel['precio_adulto_triple_usd']:,.2f}",
                            f"${total_ad_triple:,.2f}")
                if ri_menores_triple > 0:
                    col2.metric(f"{ri_menores_triple} menores triple × ${vi_sel['precio_menor_triple_usd']:,.2f}",
                                f"${total_men_triple:,.2f}")

            col1, col2 = st.columns(2)
            col1.metric("💵 TOTAL (USD)", f"${total_usd_ri:,.2f}")
            col2.metric(f"📈 Ganancia ({vi_sel['porcentaje_ganancia']}%)", f"${ganancia_usd_ri:,.2f} USD")

            st.divider()

            # Abono inicial con moneda
            st.markdown("#### 💳 Abono Inicial")
            col1, col2 = st.columns(2)
            with col1:
                ri_moneda = st.radio("Moneda del abono:", ["USD", "MXN"], horizontal=True, key="ri_moneda")
            with col2:
                ri_metodo = st.selectbox("Método de pago *",
                    ["Efectivo","Transferencia","Tarjeta de crédito",
                     "Tarjeta de débito","Depósito bancario","Otro"], key="ri_metodo")

            ri_tipo_cambio = 1.0
            ri_monto_original = 0.0
            ri_abono_usd = 0.0

            if ri_moneda == "USD":
                ri_abono_usd = st.number_input("Monto del abono (USD) *",
                    min_value=0.0, max_value=float(total_usd_ri),
                    step=50.0, format="%.2f", key="ri_abono_usd")
                ri_monto_original = ri_abono_usd
                ri_tipo_cambio = 1.0
            else:
                # Obtener tipo de cambio automático
                if "tc_megatravel" not in st.session_state:
                    st.session_state.tc_megatravel = None
                    st.session_state.tc_fuente = None

                col_tc1, col_tc2 = st.columns([2, 1])
                with col_tc2:
                    if st.button("🔄 Obtener TC de Megatravel", key="ri_btn_tc"):
                        with st.spinner("Consultando Megatravel..."):
                            tc_auto, fuente_auto = obtener_tipo_cambio()
                            if tc_auto:
                                st.session_state.tc_megatravel = tc_auto
                                st.session_state.tc_fuente = fuente_auto
                            else:
                                st.session_state.tc_megatravel = None
                                st.session_state.tc_fuente = "No disponible"

                with col_tc1:
                    tc_default = st.session_state.tc_megatravel if st.session_state.tc_megatravel else 17.0
                    ri_tipo_cambio = st.number_input("Tipo de cambio (MXN/USD) *",
                        min_value=1.0, value=float(tc_default), step=0.01, format="%.2f", key="ri_tc")

                if st.session_state.tc_fuente:
                    if st.session_state.tc_megatravel:
                        st.success(f"💱 TC obtenido de: **{st.session_state.tc_fuente}** — ${st.session_state.tc_megatravel:.2f} MXN/USD")
                    else:
                        st.warning(f"⚠️ {st.session_state.tc_fuente}. Ingresa el tipo de cambio manualmente.")

                ri_monto_original = st.number_input("Monto del abono (MXN) *",
                    min_value=0.0, step=100.0, format="%.2f", key="ri_abono_mxn")
                ri_abono_usd = ri_monto_original / ri_tipo_cambio if ri_tipo_cambio > 0 else 0
                if ri_monto_original > 0:
                    st.info(f"💱 ${ri_monto_original:,.2f} MXN ÷ {ri_tipo_cambio:.2f} = **${ri_abono_usd:,.2f} USD**")

            ri_saldo_usd = total_usd_ri - ri_abono_usd
            ri_liquidado = ri_saldo_usd <= 0.01

            col1, col2 = st.columns(2)
            col1.metric("Abono inicial", f"${ri_abono_usd:,.2f} USD")
            col2.metric("Saldo restante", f"${ri_saldo_usd:,.2f} USD")
            if ri_liquidado:
                st.success("🎉 ¡Este cliente liquidará el viaje con el abono inicial!")

            st.divider()

            # Pasajeros por habitación
            st.markdown("#### 🧳 Pasajeros por Habitación")
            pasajeros_i = []

            for i in range(ri_hab_dobles):
                st.markdown(f"**🛏️ Habitación Doble {i+1}**")
                cols_d = st.columns(2)
                for j in range(2):
                    with cols_d[j]:
                        n = st.text_input(f"Pasajero {j+1}", key=f"ri_d{i}_{j}")
                        pasajeros_i.append({"nombre": n, "tipo": "ADULTO", "habitacion": f"Doble {i+1}"})

            for i in range(ri_hab_triples):
                st.markdown(f"**🛏️ Habitación Triple {i+1}**")
                cols_t = st.columns(3)
                for j in range(3):
                    with cols_t[j]:
                        n = st.text_input(f"Pasajero {j+1}", key=f"ri_t{i}_{j}")
                        pasajeros_i.append({"nombre": n, "tipo": "ADULTO", "habitacion": f"Triple {i+1}"})

            if ri_menores > 0:
                st.markdown(f"**👶 Menores ({ri_menores})**")
                hab_opciones_i = ([f"Doble {j+1}" for j in range(ri_hab_dobles)] +
                                  [f"Triple {j+1}" for j in range(ri_hab_triples)])
                cols_m = st.columns(2)
                for i in range(ri_menores):
                    with cols_m[i % 2]:
                        n_m = st.text_input(f"Menor {i+1}", key=f"ri_m{i}")
                        h_m = st.selectbox(f"Habitación menor {i+1}",
                            hab_opciones_i, key=f"ri_mh{i}")
                        pasajeros_i.append({"nombre": n_m, "tipo": "MENOR", "habitacion": h_m})

            st.divider()

            if st.button("💾 Registrar Cliente Internacional", type="primary", use_container_width=True):
                errores_ri = []
                if not nombre_cliente_i.strip():
                    errores_ri.append("El nombre del cliente es obligatorio.")
                if ri_adultos == 0:
                    errores_ri.append("Debes seleccionar al menos una habitación.")
                if total_usd_ri <= 0:
                    errores_ri.append("El total debe ser mayor a 0.")
                if ri_menores > 0 and ri_menores_doble + ri_menores_triple != ri_menores:
                    errores_ri.append("La distribución de menores no cuadra con el total.")
                if any(p["nombre"].strip() == "" for p in pasajeros_i):
                    errores_ri.append("Todos los nombres de pasajeros son obligatorios.")

                if errores_ri:
                    for e in errores_ri:
                        st.error(f"❌ {e}")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        fecha_reg_i = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        estado_ci = "LIQUIDADO" if ri_liquidado else "ADEUDO"

                        cursor.execute("""
                            INSERT INTO clientes_internacionales (
                                viaje_id, nombre_cliente, adultos, menores,
                                habitaciones_doble, habitaciones_triple,
                                total_usd, abonado_usd, saldo_usd, ganancia_usd,
                                estado, fecha_registro
                            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (
                            int(vi_sel["id"]), nombre_cliente_i.strip(),
                            ri_adultos, ri_menores,
                            ri_hab_dobles, ri_hab_triples,
                            total_usd_ri, ri_abono_usd, ri_saldo_usd, ganancia_usd_ri,
                            estado_ci, fecha_reg_i
                        ))
                        ci_id = cursor.lastrowid

                        # Abono inicial
                        cursor.execute("""
                            INSERT INTO abonos_internacionales
                                (cliente_id, fecha, moneda, monto_original, tipo_cambio, monto_usd)
                            VALUES (?,?,?,?,?,?)
                        """, (ci_id, fecha_reg_i, ri_moneda,
                              ri_monto_original, ri_tipo_cambio, ri_abono_usd))

                        # Pasajeros
                        for p in pasajeros_i:
                            cursor.execute("""
                                INSERT INTO pasajeros_internacionales
                                    (cliente_id, nombre_completo, tipo, habitacion_asignada)
                                VALUES (?,?,?,?)
                            """, (ci_id, p["nombre"].strip(), p["tipo"], p["habitacion"]))

                        # Actualizar cupos
                        cursor.execute("""
                            UPDATE viajes_internacionales
                            SET cupos_vendidos = cupos_vendidos + ?,
                                cupos_disponibles = cupos_disponibles - ?
                            WHERE id = ?
                        """, (ri_adultos, ri_adultos, int(vi_sel["id"])))

                        cursor.execute("""
                            UPDATE viajes_internacionales SET estado = 'AGOTADO'
                            WHERE id = ? AND cupos_disponibles <= 0
                        """, (int(vi_sel["id"]),))

                        conn.commit()
                        conn.close()
                        st.success(f"✅ **{nombre_cliente_i}** registrado en **{vi_sel['destino']}** — Total: ${total_usd_ri:,.2f} USD | Saldo: ${ri_saldo_usd:,.2f} USD")
                        if ri_liquidado:
                            st.balloons()
                        st.rerun()

                    except Exception as e:
                        st.error(f"❌ Error al guardar: {e}")

    # ── TAB 3: REGISTRAR PAGO ─────────────────────────────────────────────
    with tabs[3]:
        st.subheader("💰 Registrar Pago - Viajes Internacionales")

        conn = conectar_db()
        try:
            query_pi = """
                SELECT ci.id, ci.nombre_cliente, vi.destino,
                       ci.total_usd, ci.abonado_usd, ci.saldo_usd
                FROM clientes_internacionales ci
                JOIN viajes_internacionales vi ON ci.viaje_id = vi.id
                WHERE ci.estado NOT IN ('CERRADO','LIQUIDADO')
                ORDER BY ci.id DESC
            """
            df_pi = pd.read_sql_query(query_pi, conn)
        except Exception as e:
            df_pi = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_pi.empty:
            st.info("No hay clientes internacionales con saldo pendiente.")
        else:
            df_pi.columns = ["id","cliente","destino","total_usd","abonado_usd","saldo_usd"]
            opts_pi = {
                f"ID {row['id']} | {row['cliente']} — {row['destino']} | Saldo: ${row['saldo_usd']:,.2f} USD": row
                for _, row in df_pi.iterrows()
            }
            pi_label = st.selectbox("Selecciona el cliente:", list(opts_pi.keys()), key="pi_sel")
            pi_sel = opts_pi[pi_label]

            col1, col2, col3 = st.columns(3)
            col1.metric("Total",           f"${pi_sel['total_usd']:,.2f} USD")
            col2.metric("Ya abonado",      f"${pi_sel['abonado_usd']:,.2f} USD")
            col3.metric("Saldo pendiente", f"${pi_sel['saldo_usd']:,.2f} USD")

            st.divider()
            st.markdown("#### 💳 Datos del Abono")

            col1, col2 = st.columns(2)
            with col1:
                pi_moneda = st.radio("Moneda:", ["USD", "MXN"], horizontal=True, key="pi_moneda")
            with col2:
                pi_metodo = st.selectbox("Método de pago",
                    ["Efectivo","Transferencia","Tarjeta de crédito",
                     "Tarjeta de débito","Depósito bancario","Otro"], key="pi_metodo")

            pi_tipo_cambio   = 1.0
            pi_monto_original = 0.0
            pi_monto_usd     = 0.0

            if pi_moneda == "USD":
                pi_monto_usd = st.number_input("Monto del abono (USD) *",
                    min_value=0.01, max_value=float(pi_sel["saldo_usd"]),
                    step=50.0, format="%.2f", key="pi_monto_usd")
                pi_monto_original = pi_monto_usd
                pi_tipo_cambio = 1.0
            else:
                # Reusar TC de session_state si ya fue cargado, o cargar nuevo
                if "tc_megatravel" not in st.session_state:
                    st.session_state.tc_megatravel = None
                    st.session_state.tc_fuente = None

                col_tc1, col_tc2 = st.columns([2, 1])
                with col_tc2:
                    if st.button("🔄 Obtener TC de Megatravel", key="pi_btn_tc"):
                        with st.spinner("Consultando Megatravel..."):
                            tc_auto, fuente_auto = obtener_tipo_cambio()
                            if tc_auto:
                                st.session_state.tc_megatravel = tc_auto
                                st.session_state.tc_fuente = fuente_auto
                            else:
                                st.session_state.tc_megatravel = None
                                st.session_state.tc_fuente = "No disponible"

                with col_tc1:
                    tc_default = st.session_state.tc_megatravel if st.session_state.tc_megatravel else 17.0
                    pi_tipo_cambio = st.number_input("Tipo de cambio (MXN/USD) *",
                        min_value=1.0, value=float(tc_default), step=0.01, format="%.2f", key="pi_tc")

                if st.session_state.tc_fuente:
                    if st.session_state.tc_megatravel:
                        st.success(f"💱 TC obtenido de: **{st.session_state.tc_fuente}** — ${st.session_state.tc_megatravel:.2f} MXN/USD")
                    else:
                        st.warning(f"⚠️ {st.session_state.tc_fuente}. Ingresa el tipo de cambio manualmente.")

                pi_monto_original = st.number_input("Monto del abono (MXN) *",
                    min_value=0.0, step=100.0, format="%.2f", key="pi_monto_mxn")
                pi_monto_usd = pi_monto_original / pi_tipo_cambio if pi_tipo_cambio > 0 else 0
                if pi_monto_original > 0:
                    st.info(f"💱 ${pi_monto_original:,.2f} MXN ÷ {pi_tipo_cambio:.2f} = **${pi_monto_usd:,.2f} USD**")

            if pi_monto_usd > 0:
                nuevo_abonado_i = pi_sel["abonado_usd"] + pi_monto_usd
                nuevo_saldo_i   = pi_sel["saldo_usd"]   - pi_monto_usd
                liquidado_i     = nuevo_saldo_i <= 0.01

                col1, col2 = st.columns(2)
                col1.metric("Nuevo total abonado", f"${nuevo_abonado_i:,.2f} USD", delta=f"+${pi_monto_usd:,.2f}")
                col2.metric("Nuevo saldo",         f"${nuevo_saldo_i:,.2f} USD",   delta=f"-${pi_monto_usd:,.2f}")

                if liquidado_i:
                    st.success("🎉 ¡Con este pago el cliente quedará **LIQUIDADO**!")
                else:
                    pct_i = (nuevo_abonado_i / pi_sel["total_usd"]) * 100
                    st.info(f"📊 Avance: **{pct_i:.1f}%** cobrado")

                st.divider()
                if st.button("💾 Registrar Abono Internacional", type="primary", use_container_width=True):
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        fecha_pi = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                        cursor.execute("""
                            INSERT INTO abonos_internacionales
                                (cliente_id, fecha, moneda, monto_original, tipo_cambio, monto_usd)
                            VALUES (?,?,?,?,?,?)
                        """, (int(pi_sel["id"]), fecha_pi, pi_moneda,
                              pi_monto_original, pi_tipo_cambio, pi_monto_usd))

                        nuevo_estado_i = "LIQUIDADO" if liquidado_i else "ADEUDO"
                        cursor.execute("""
                            UPDATE clientes_internacionales
                            SET abonado_usd = abonado_usd + ?,
                                saldo_usd   = saldo_usd   - ?,
                                estado = ?
                            WHERE id = ?
                        """, (pi_monto_usd, pi_monto_usd, nuevo_estado_i, int(pi_sel["id"])))

                        conn.commit()
                        conn.close()

                        # ── Recibo ─────────────────────────────────────────
                        num_rec_i   = _siguiente_num_recibo()
                        fecha_rec_i = datetime.now().strftime("%d-%B-%Y").replace(
                            "January","enero").replace("February","febrero").replace(
                            "March","marzo").replace("April","abril").replace(
                            "May","mayo").replace("June","junio").replace(
                            "July","julio").replace("August","agosto").replace(
                            "September","septiembre").replace("October","octubre").replace(
                            "November","noviembre").replace("December","diciembre")
                        vend_i     = usuario.get("nombre", "Agente")
                        concepto_i = f"{'Liquidación' if liquidado_i else 'Abono'} - {pi_sel.get('destino','Viaje Internacional')}"

                        if liquidado_i:
                            st.success(f"✅ ¡**{pi_sel['cliente']}** ha LIQUIDADO su viaje! 🎉")
                            st.balloons()
                        else:
                            st.success(f"✅ Abono de ${pi_monto_usd:,.2f} USD registrado. Saldo: ${nuevo_saldo_i:,.2f} USD")

                        st.divider()
                        st.markdown("#### 🧾 Recibo de Pago")
                        _boton_recibo(
                            numero     = num_rec_i,
                            fecha_str  = fecha_rec_i,
                            cliente    = pi_sel['cliente'],
                            monto      = pi_monto_usd,
                            concepto   = concepto_i,
                            forma_pago = pi_metodo,
                            agente     = vend_i,
                            key_suffix = f"int_{pi_sel['id']}",
                        )

                    except Exception as e:
                        st.error(f"❌ Error: {e}")

    # ── TAB 4: GESTIONAR VIAJES (solo ADMIN) ──────────────────────────────
    if usuario["rol"] == "ADMIN":
        with tabs[4]:
            st.subheader("⚙️ Gestionar Viajes Internacionales")
            subtab_gi = st.tabs(["➕ Nuevo Viaje", "✏️ Editar / Cerrar Viaje"])

            with subtab_gi[0]:
                st.markdown("#### 🌎 Registrar Nuevo Viaje Internacional")
                col1, col2 = st.columns(2)
                with col1:
                    gi_destino = st.text_input("Destino *", key="gi_dest", placeholder="Ej: Europa 10 días")
                    gi_salida  = st.date_input("Fecha de salida *", key="gi_sal")
                    gi_regreso = st.date_input("Fecha de regreso *", key="gi_reg")
                    gi_cupos   = st.number_input("Pasajeros proyectados *", min_value=1, value=20, key="gi_cupos")
                with col2:
                    st.markdown("**💵 Precios en USD:**")
                    gi_p_ad_doble  = st.number_input("Adulto base doble",  min_value=0.0, step=50.0, format="%.2f", key="gi_pad")
                    gi_p_ad_triple = st.number_input("Adulto base triple", min_value=0.0, step=50.0, format="%.2f", key="gi_pat")
                    gi_p_men_doble  = st.number_input("Menor base doble",  min_value=0.0, step=50.0, format="%.2f", key="gi_pmd")
                    gi_p_men_triple = st.number_input("Menor base triple", min_value=0.0, step=50.0, format="%.2f", key="gi_pmt")
                    gi_ganancia = st.slider("Porcentaje de ganancia (%)", 1, 50, 15, key="gi_gan")

                if gi_regreso > gi_salida:
                    gi_dias   = (gi_regreso - gi_salida).days + 1
                    gi_noches = gi_dias - 1
                    st.info(f"📅 {gi_dias} días / {gi_noches} noches")

                if st.button("💾 Registrar Viaje Internacional", type="primary", use_container_width=True, key="gi_reg_btn"):
                    errores_gi = []
                    if not gi_destino.strip():  errores_gi.append("El destino es obligatorio.")
                    if gi_regreso <= gi_salida: errores_gi.append("La fecha de regreso debe ser posterior.")
                    if gi_p_ad_doble <= 0:      errores_gi.append("El precio adulto doble debe ser mayor a 0.")

                    if errores_gi:
                        for e in errores_gi:
                            st.error(f"❌ {e}")
                    else:
                        try:
                            gi_dias   = (gi_regreso - gi_salida).days + 1
                            gi_noches = gi_dias - 1
                            conn = conectar_db()
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO viajes_internacionales (
                                    destino, fecha_salida, fecha_regreso, dias, noches,
                                    cupos_totales, cupos_vendidos, cupos_disponibles,
                                    precio_adulto_doble_usd, precio_adulto_triple_usd,
                                    precio_menor_doble_usd, precio_menor_triple_usd,
                                    porcentaje_ganancia, estado, fecha_registro
                                ) VALUES (?,?,?,?,?,?,0,?,?,?,?,?,?,'ACTIVO',?)
                            """, (
                                gi_destino.strip(),
                                gi_salida.strftime("%d-%m-%Y"), gi_regreso.strftime("%d-%m-%Y"),
                                gi_dias, gi_noches,
                                gi_cupos, gi_cupos,
                                gi_p_ad_doble, gi_p_ad_triple,
                                gi_p_men_doble, gi_p_men_triple,
                                gi_ganancia,
                                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            ))
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Viaje **{gi_destino}** registrado con {gi_cupos} cupos.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")

            with subtab_gi[1]:
                st.markdown("#### ✏️ Editar o Cerrar un Viaje")
                conn = conectar_db()
                try:
                    df_gi_edit = pd.read_sql_query("""
                        SELECT id, destino, cupos_totales, cupos_vendidos, cupos_disponibles,
                               precio_adulto_doble_usd, precio_adulto_triple_usd,
                               precio_menor_doble_usd, precio_menor_triple_usd,
                               porcentaje_ganancia, estado
                        FROM viajes_internacionales ORDER BY id DESC
                    """, conn)
                except:
                    df_gi_edit = pd.DataFrame()
                conn.close()

                if df_gi_edit.empty:
                    st.info("No hay viajes para editar.")
                else:
                    gi_edit_opts = {
                        f"ID {row['id']} | {row['destino']} ({row['estado']})": row
                        for _, row in df_gi_edit.iterrows()
                    }
                    gi_edit_label = st.selectbox("Selecciona viaje:", list(gi_edit_opts.keys()), key="gi_edit_sel")
                    gi_edit_row = gi_edit_opts[gi_edit_label]

                    col1, col2 = st.columns(2)
                    with col1:
                        gi_e_cupos   = st.number_input("Cupos totales",
                            min_value=int(gi_edit_row["cupos_vendidos"]),
                            value=int(gi_edit_row["cupos_totales"]), key="gi_e_cupos")
                        gi_e_pad     = st.number_input("Adulto doble (USD)",
                            min_value=0.0, value=float(gi_edit_row["precio_adulto_doble_usd"]),
                            step=50.0, format="%.2f", key="gi_e_pad")
                        gi_e_pat     = st.number_input("Adulto triple (USD)",
                            min_value=0.0, value=float(gi_edit_row["precio_adulto_triple_usd"]),
                            step=50.0, format="%.2f", key="gi_e_pat")
                    with col2:
                        gi_e_pmd     = st.number_input("Menor doble (USD)",
                            min_value=0.0, value=float(gi_edit_row["precio_menor_doble_usd"]),
                            step=50.0, format="%.2f", key="gi_e_pmd")
                        gi_e_pmt     = st.number_input("Menor triple (USD)",
                            min_value=0.0, value=float(gi_edit_row["precio_menor_triple_usd"]),
                            step=50.0, format="%.2f", key="gi_e_pmt")
                        gi_e_estado  = st.selectbox("Estado",
                            ["ACTIVO","AGOTADO","CERRADO"],
                            index=["ACTIVO","AGOTADO","CERRADO"].index(gi_edit_row["estado"]),
                            key="gi_e_estado")

                    if st.button("💾 Guardar Cambios", type="primary", key="gi_e_guardar"):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            nuevos_disp_i = gi_e_cupos - int(gi_edit_row["cupos_vendidos"])
                            cursor.execute("""
                                UPDATE viajes_internacionales
                                SET cupos_totales = ?, cupos_disponibles = ?,
                                    precio_adulto_doble_usd = ?, precio_adulto_triple_usd = ?,
                                    precio_menor_doble_usd  = ?, precio_menor_triple_usd  = ?,
                                    estado = ?
                                WHERE id = ?
                            """, (gi_e_cupos, max(0, nuevos_disp_i),
                                  gi_e_pad, gi_e_pat, gi_e_pmd, gi_e_pmt,
                                  gi_e_estado, int(gi_edit_row["id"])))
                            conn.commit()
                            conn.close()
                            st.success("✅ Viaje actualizado correctamente.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")



# ═══════════════════════════════════════════════════════════════════════════
#  MÓDULO DE REPORTES — TURISMAR
# ═══════════════════════════════════════════════════════════════════════════

def _excel_bytes_from_wb(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def _estilo_header(ws, row, cols, bg="1F4E79", fg="FFFFFF"):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill  = PatternFill("solid", fgColor=bg)
    font  = Font(bold=True, color=fg, size=11, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = align; cell.border = thin

def _estilo_row(ws, row, cols, bg=None, bold=False):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    font  = Font(bold=bold, name="Arial", size=10)
    align = Alignment(vertical="center", wrap_text=True)
    thin  = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"),  bottom=Side(style="thin"))
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font; cell.alignment = align; cell.border = thin
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

def _fmt_cur(ws, row, cols_idx):
    for c in cols_idx:
        ws.cell(row=row, column=c).number_format = "$#,##0.00"

def _titulo_sheet(ws, titulo, subtitulo=""):
    from openpyxl.styles import Font, Alignment, PatternFill
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = titulo
    c.font  = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill  = PatternFill("solid", fgColor="E91E8C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitulo:
        ws.merge_cells("A2:N2")
        c2 = ws["A2"]
        c2.value = subtitulo
        c2.font  = Font(italic=True, size=10, color="333333", name="Arial")
        c2.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18
        return 3
    return 2

def _query_rep(sql, params=None):
    conn = conectar_db()
    try:
        df = pd.read_sql_query(sql, conn, params=params)
    except Exception as e:
        conn.close()
        raise e
    finally:
        try: conn.close()
        except: pass
    return df

# ─── RIVIERA MAYA ────────────────────────────────────────────────────────────

def _excel_riviera(tipo_filtro="todos", valor_filtro=None, label="General", id_vend=None):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    if tipo_filtro == "bloqueos":
        where = "WHERE v.es_bloqueo = 1"
    elif tipo_filtro == "grupos":
        where = "WHERE v.es_grupo = 1"
    elif tipo_filtro == "vendedora" and valor_filtro:
        where = f"WHERE vd.nombre = '{valor_filtro}'"
    elif id_vend:
        where = f"WHERE v.vendedora_id = {id_vend}"
    else:
        where = ""

    df_v = _query_rep(f"""
        SELECT v.id, COALESCE(vd.nombre,'—') AS vendedora,
               v.cliente, v.celular_responsable AS celular, v.destino, v.tipo_habitacion,
               v.fecha_inicio, v.fecha_fin, v.noches, v.adultos, v.menores,
               v.precio_total, v.pagado, v.saldo, v.ganancia,
               COALESCE(v.comision_vendedora, ROUND(v.ganancia*0.10,2)) AS comision,
               v.estado,
               CASE v.comision_pagada WHEN 1 THEN 'Pagada' ELSE 'Pendiente' END AS estado_comision,
               v.fecha_registro
        FROM ventas v
        LEFT JOIN vendedoras vd ON v.vendedora_id = vd.id
        {where} ORDER BY v.id DESC
    """)
    df_a = _query_rep("""
        SELECT a.venta_id, a.fecha, a.monto, COALESCE(a.metodo_pago,'Efectivo') AS metodo
        FROM abonos a ORDER BY a.venta_id, a.fecha
    """)

    wb = Workbook()

    # Hoja 1 — Resumen
    ws1 = wb.active; ws1.title = "Resumen"
    r = _titulo_sheet(ws1, f"TURISMAR — Riviera Maya: {label}",
                      f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    totales = [
        ("Total ventas",          len(df_v)),
        ("Total vendido ($)",     float(df_v['precio_total'].sum()) if not df_v.empty else 0),
        ("Total cobrado ($)",     float(df_v['pagado'].sum()) if not df_v.empty else 0),
        ("Saldo por cobrar ($)",  float(df_v['saldo'].sum()) if not df_v.empty else 0),
        ("Ganancia total ($)",    float(df_v['ganancia'].sum()) if not df_v.empty else 0),
        ("Ventas ACTIVAS",        len(df_v[df_v['estado']=='ACTIVO']) if not df_v.empty else 0),
        ("Ventas LIQUIDADAS",     len(df_v[df_v['estado']=='LIQUIDADO']) if not df_v.empty else 0),
        ("Ventas CERRADAS",       len(df_v[df_v['estado']=='CERRADO']) if not df_v.empty else 0),
        ("Comisiones pendientes ($)", float(df_v[df_v['estado_comision']=='Pendiente']['comision'].sum()) if not df_v.empty else 0),
        ("Comisiones pagadas ($)",    float(df_v[df_v['estado_comision']=='Pagada']['comision'].sum()) if not df_v.empty else 0),
    ]
    ws1.cell(row=r,column=1).value="Indicador"; ws1.cell(row=r,column=2).value="Valor"
    _estilo_header(ws1,r,2,bg="29ABE2"); r+=1
    for i,(k,v_val) in enumerate(totales):
        ws1.cell(row=r,column=1).value=k; ws1.cell(row=r,column=2).value=v_val
        _estilo_row(ws1,r,2,bg="F0F9FF" if i%2==0 else "FFFFFF")
        if isinstance(v_val,float): _fmt_cur(ws1,r,[2])
        r+=1
    ws1.column_dimensions["A"].width=30; ws1.column_dimensions["B"].width=22

    # Hoja 2 — Clientes y Pagos
    ws2 = wb.create_sheet("Clientes y Pagos")
    r2 = _titulo_sheet(ws2, f"Clientes — {label}")
    hdrs=["ID","Vendedora","Cliente","Celular","Destino","Habitación","Salida","Regreso",
          "Noches","Adultos","Menores","Total","Pagado","Saldo","Ganancia","Comisión","Estado","Com.Estado"]
    for c,h in enumerate(hdrs,1): ws2.cell(row=r2,column=c).value=h
    _estilo_header(ws2,r2,len(hdrs),bg="E91E8C"); r2+=1
    for _,row in (df_v.iterrows() if not df_v.empty else []):
        vals=[row.get("id"),row.get("vendedora"),row.get("cliente"),row.get("celular"),
              row.get("destino"),row.get("tipo_habitacion"),row.get("fecha_inicio"),row.get("fecha_fin"),
              row.get("noches"),row.get("adultos"),row.get("menores"),
              row.get("precio_total"),row.get("pagado"),row.get("saldo"),
              row.get("ganancia"),row.get("comision"),row.get("estado"),row.get("estado_comision")]
        for c,val in enumerate(vals,1): ws2.cell(row=r2,column=c).value=val
        _estilo_row(ws2,r2,len(hdrs),bg="FFF0F7" if r2%2==0 else "FFFFFF")
        _fmt_cur(ws2,r2,[12,13,14,15,16]); r2+=1
    if not df_v.empty:
        n=len(df_v); ws2.cell(row=r2,column=1).value="TOTALES"
        for ci,col in [(12,"L"),(13,"M"),(14,"N"),(15,"O"),(16,"P")]:
            ws2.cell(row=r2,column=ci).value=f"=SUM({col}{r2-n}:{col}{r2-1})"
        _estilo_row(ws2,r2,len(hdrs),bg="29ABE2",bold=True); _fmt_cur(ws2,r2,[12,13,14,15,16])
    for c,w in enumerate([6,18,22,14,28,12,12,12,8,8,8,14,14,14,14,14,12,14],1):
        ws2.column_dimensions[get_column_letter(c)].width=w

    # Hoja 3 — Habitaciones
    ws3 = wb.create_sheet("Habitaciones")
    r3 = _titulo_sheet(ws3, f"Distribución Habitaciones — {label}")
    if not df_v.empty:
        df_v_hab = df_v.copy()
        df_v_hab["tipo_habitacion"] = df_v_hab["tipo_habitacion"].fillna("SIN TIPO").replace("", "SIN TIPO")
        dist=df_v_hab.groupby("tipo_habitacion").agg(
            cantidad=("id","count"),total=("precio_total","sum"),
            cobrado=("pagado","sum"),saldo=("saldo","sum"),ganancia=("ganancia","sum")
        ).reset_index()
        for c,h in enumerate(["Tipo","Cantidad","Total Vendido","Total Cobrado","Saldo","Ganancia"],1):
            ws3.cell(row=r3,column=c).value=h
        _estilo_header(ws3,r3,6,bg="1F4E79"); r3+=1
        for _,row in dist.iterrows():
            for c,val in enumerate([row["tipo_habitacion"],row["cantidad"],row["total"],
                                    row["cobrado"],row["saldo"],row["ganancia"]],1):
                ws3.cell(row=r3,column=c).value=val
            _estilo_row(ws3,r3,6,bg="EFF9FF" if r3%2==0 else "FFFFFF")
            _fmt_cur(ws3,r3,[3,4,5,6]); r3+=1
        n=len(dist); ws3.cell(row=r3,column=1).value="TOTAL"
        for ci,col in [(2,"B"),(3,"C"),(4,"D"),(5,"E"),(6,"F")]:
            ws3.cell(row=r3,column=ci).value=f"=SUM({col}{r3-n}:{col}{r3-1})"
        _estilo_row(ws3,r3,6,bg="29ABE2",bold=True); _fmt_cur(ws3,r3,[3,4,5,6])
    for c,w in enumerate([22,12,18,18,18,18],1):
        ws3.column_dimensions[get_column_letter(c)].width=w

    # Hoja 4 — Historial Pagos
    ws4 = wb.create_sheet("Historial Pagos")
    r4 = _titulo_sheet(ws4, f"Historial de Pagos — {label}")
    if not df_a.empty and not df_v.empty:
        ids_v = set(df_v["id"].tolist())
        df_a2 = df_a[df_a["venta_id"].isin(ids_v)].copy()
        if df_a2.empty:
            ws4.cell(row=r4,column=1).value = "Sin abonos registrados para estas ventas"
        else:
            mapa = df_v.set_index("id")[["cliente","destino","vendedora"]].to_dict("index")
            df_a2["cliente"]   = df_a2["venta_id"].map(lambda x: mapa.get(x,{}).get("cliente","—"))
            df_a2["destino"]   = df_a2["venta_id"].map(lambda x: mapa.get(x,{}).get("destino","—"))
            df_a2["vendedora"] = df_a2["venta_id"].map(lambda x: mapa.get(x,{}).get("vendedora","—"))
            for c,h in enumerate(["ID Venta","Cliente","Destino","Vendedora","Fecha","Monto","Método"],1):
                ws4.cell(row=r4,column=c).value = h
            _estilo_header(ws4,r4,7,bg="2D6A4F"); r4+=1
            for _,row in df_a2.iterrows():
                for c,val in enumerate([row["venta_id"],row["cliente"],row["destino"],
                                        row["vendedora"],row["fecha"],row["monto"],row["metodo"]],1):
                    ws4.cell(row=r4,column=c).value = val
                _estilo_row(ws4,r4,7,bg="F0FFF4" if r4%2==0 else "FFFFFF")
                _fmt_cur(ws4,r4,[6]); r4+=1
            n4 = len(df_a2)
            ws4.cell(row=r4,column=5).value = "TOTAL"
            ws4.cell(row=r4,column=6).value = f"=SUM(F{r4-n4}:F{r4-1})"
            _estilo_row(ws4,r4,7,bg="29ABE2",bold=True); _fmt_cur(ws4,r4,[6])
    else:
        ws4.cell(row=r4,column=1).value = "Sin abonos registrados"
    for c,w in enumerate([10,24,28,18,18,14,16],1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    return _excel_bytes_from_wb(wb)


# ─── NACIONALES ──────────────────────────────────────────────────────────────
# NOTA: clientes_nacionales usa habitaciones_doble/triple (no tipo_habitacion)
#       y los pasajeros están en pasajeros_nacionales

def _excel_nacionales_viaje(viaje_id, nombre_viaje, id_vend_filtro=None):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    filtro_vend = f"AND cn.vendedora_id = {id_vend_filtro}" if id_vend_filtro else ""

    df_v = _query_rep(f"""
        SELECT cn.id, COALESCE(vd.nombre,'—') AS vendedora,
               cn.nombre_cliente AS cliente, cn.celular_responsable AS celular,
               cn.adultos, cn.menores,
               cn.habitaciones_doble, cn.habitaciones_triple,
               cn.total_pagar, cn.total_abonado AS pagado, cn.saldo,
               COALESCE(cn.ganancia, 0) AS ganancia,
               cn.estado, cn.fecha_registro
        FROM clientes_nacionales cn
        LEFT JOIN vendedoras vd ON cn.vendedora_id = vd.id
        WHERE cn.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY cn.id
    """)
    df_pas = _query_rep(f"""
        SELECT pn.cliente_id, pn.nombre_completo, pn.tipo, pn.habitacion_asignada
        FROM pasajeros_nacionales pn
        JOIN clientes_nacionales cn ON pn.cliente_id = cn.id
        WHERE cn.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY pn.cliente_id, pn.tipo
    """)
    df_vj = _query_rep(f"""
        SELECT nombre_viaje, destino, fecha_salida, fecha_regreso,
               dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles,
               precio_persona_doble, precio_persona_triple, estado
        FROM viajes_nacionales WHERE id = {viaje_id}
    """)
    df_ab = _query_rep(f"""
        SELECT an.cliente_id, cn.nombre_cliente AS cliente,
               an.fecha, an.monto, COALESCE(an.metodo_pago,'Efectivo') AS metodo
        FROM abonos_nacionales an
        JOIN clientes_nacionales cn ON an.cliente_id = cn.id
        WHERE cn.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY an.fecha
    """)

    wb = Workbook()

    # Hoja 1 — Info General
    ws1 = wb.active; ws1.title = "Info General"
    r = _titulo_sheet(ws1, f"TURISMAR — {nombre_viaje}",
                      f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    if not df_vj.empty:
        rv = df_vj.iloc[0]
        campos = [
            ("Nombre del viaje",   rv.get("nombre_viaje","")),
            ("Destino",            rv.get("destino","")),
            ("Fecha salida",       rv.get("fecha_salida","")),
            ("Fecha regreso",      rv.get("fecha_regreso","")),
            ("Días / Noches",      f"{rv.get('dias','')} días / {rv.get('noches','')} noches"),
            ("Cupos totales",      rv.get("cupos_totales","")),
            ("Cupos vendidos",     rv.get("cupos_vendidos","")),
            ("Cupos disponibles",  rv.get("cupos_disponibles","")),
            ("Precio doble",       float(rv.get("precio_persona_doble") or 0)),
            ("Precio triple",      float(rv.get("precio_persona_triple") or 0)),
            ("Estado del viaje",   rv.get("estado","")),
            ("Total clientes",     len(df_v)),
            ("Total a pagar ($)",  float(df_v['total_pagar'].sum()) if not df_v.empty else 0),
            ("Total abonado ($)",  float(df_v['pagado'].sum()) if not df_v.empty else 0),
            ("Saldo pendiente ($)",float(df_v['saldo'].sum()) if not df_v.empty else 0),
            ("Ganancia total ($)", float(df_v['ganancia'].sum()) if not df_v.empty else 0),
        ]
        ws1.cell(row=r,column=1).value="Campo"; ws1.cell(row=r,column=2).value="Valor"
        _estilo_header(ws1,r,2,bg="29ABE2"); r+=1
        for i,(k,val) in enumerate(campos):
            ws1.cell(row=r,column=1).value=k; ws1.cell(row=r,column=2).value=val
            _estilo_row(ws1,r,2,bg="F0F9FF" if i%2==0 else "FFFFFF")
            if isinstance(val,float): _fmt_cur(ws1,r,[2])
            r+=1
    ws1.column_dimensions["A"].width=28; ws1.column_dimensions["B"].width=30

    # Hoja 2 — Clientes y Pagos
    ws2 = wb.create_sheet("Clientes y Pagos")
    r2 = _titulo_sheet(ws2, f"Clientes — {nombre_viaje}")
    hdrs=["ID","Vendedora","Cliente","Celular","Adultos","Menores",
          "Hab. Doble","Hab. Triple","Total","Abonado","Saldo","Ganancia","Estado"]
    for c,h in enumerate(hdrs,1): ws2.cell(row=r2,column=c).value=h
    _estilo_header(ws2,r2,len(hdrs),bg="E91E8C"); r2+=1
    for _,row in (df_v.iterrows() if not df_v.empty else []):
        vals=[row.get("id"),row.get("vendedora"),row.get("cliente"),row.get("celular"),
              row.get("adultos"),row.get("menores"),
              row.get("habitaciones_doble"),row.get("habitaciones_triple"),
              row.get("total_pagar"),row.get("pagado"),row.get("saldo"),
              row.get("ganancia"),row.get("estado")]
        for c,val in enumerate(vals,1): ws2.cell(row=r2,column=c).value=val
        _estilo_row(ws2,r2,len(hdrs),bg="FFF0F7" if r2%2==0 else "FFFFFF")
        _fmt_cur(ws2,r2,[9,10,11,12]); r2+=1
    if not df_v.empty:
        n=len(df_v); ws2.cell(row=r2,column=1).value="TOTALES"
        for ci,col in [(9,"I"),(10,"J"),(11,"K"),(12,"L")]:
            ws2.cell(row=r2,column=ci).value=f"=SUM({col}{r2-n}:{col}{r2-1})"
        _estilo_row(ws2,r2,len(hdrs),bg="29ABE2",bold=True); _fmt_cur(ws2,r2,[9,10,11,12])
    for c,w in enumerate([6,18,24,14,8,8,10,10,14,14,14,14,12],1):
        ws2.column_dimensions[get_column_letter(c)].width=w

    # Hoja 3 — Pasajeros (nombre real de cada uno)
    ws3 = wb.create_sheet("Pasajeros")
    r3 = _titulo_sheet(ws3, f"Lista de Pasajeros — {nombre_viaje}")
    if not df_pas.empty and not df_v.empty:
        mapa_cl = df_v.set_index("id")["cliente"].to_dict()
        df_pas["nombre_cliente"] = df_pas["cliente_id"].map(lambda x: mapa_cl.get(x,"—"))
        for c,h in enumerate(["ID Cliente","Nombre Cliente","Pasajero","Tipo","Habitación"],1):
            ws3.cell(row=r3,column=c).value=h
        _estilo_header(ws3,r3,5,bg="1F4E79"); r3+=1
        for _,row in df_pas.iterrows():
            for c,val in enumerate([row["cliente_id"],row["nombre_cliente"],
                                    row["nombre_completo"],row["tipo"],
                                    row["habitacion_asignada"]],1):
                ws3.cell(row=r3,column=c).value=val
            _estilo_row(ws3,r3,5,bg="EFF9FF" if r3%2==0 else "FFFFFF")
            r3+=1
    else:
        ws3.cell(row=r3,column=1).value="Sin pasajeros registrados"
    for c,w in enumerate([10,24,28,12,14],1):
        ws3.column_dimensions[get_column_letter(c)].width=w

    # Hoja 4 — Distribución Habitaciones
    ws4 = wb.create_sheet("Habitaciones")
    r4 = _titulo_sheet(ws4, f"Distribución Habitaciones — {nombre_viaje}")
    if not df_v.empty:
        total_dobles  = int(df_v["habitaciones_doble"].sum())
        total_triples = int(df_v["habitaciones_triple"].sum())
        total_pax     = int(df_v["adultos"].sum()) + int(df_v["menores"].sum())
        for c,h in enumerate(["Tipo Habitación","Cantidad","Pax promedio"],1):
            ws4.cell(row=r4,column=c).value=h
        _estilo_header(ws4,r4,3,bg="1F4E79"); r4+=1
        for tipo,cant in [("DOBLE",total_dobles),("TRIPLE",total_triples)]:
            ws4.cell(row=r4,column=1).value=tipo
            ws4.cell(row=r4,column=2).value=cant
            ws4.cell(row=r4,column=3).value=2 if tipo=="DOBLE" else 3
            _estilo_row(ws4,r4,3,bg="EFF9FF" if r4%2==0 else "FFFFFF"); r4+=1
        ws4.cell(row=r4,column=1).value="TOTAL HABITACIONES"
        ws4.cell(row=r4,column=2).value=total_dobles+total_triples
        ws4.cell(row=r4,column=3).value=f"{total_pax} pax"
        _estilo_row(ws4,r4,3,bg="29ABE2",bold=True)
    for c,w in enumerate([22,14,16],1):
        ws4.column_dimensions[get_column_letter(c)].width=w

    # Hoja 5 — Historial Pagos
    ws5 = wb.create_sheet("Historial Pagos")
    r5 = _titulo_sheet(ws5, f"Historial de Pagos — {nombre_viaje}")
    if not df_ab.empty:
        for c,h in enumerate(["ID Cliente","Cliente","Fecha","Monto","Método"],1):
            ws5.cell(row=r5,column=c).value=h
        _estilo_header(ws5,r5,5,bg="2D6A4F"); r5+=1
        for _,row in df_ab.iterrows():
            for c,val in enumerate([row["cliente_id"],row["cliente"],
                                    row["fecha"],row["monto"],row["metodo"]],1):
                ws5.cell(row=r5,column=c).value=val
            _estilo_row(ws5,r5,5,bg="F0FFF4" if r5%2==0 else "FFFFFF")
            _fmt_cur(ws5,r5,[4]); r5+=1
        n5=len(df_ab); ws5.cell(row=r5,column=3).value="TOTAL"
        ws5.cell(row=r5,column=4).value=f"=SUM(D{r5-n5}:D{r5-1})"
        _estilo_row(ws5,r5,5,bg="29ABE2",bold=True); _fmt_cur(ws5,r5,[4])
    else:
        ws5.cell(row=r5,column=1).value="Sin abonos registrados"
    for c,w in enumerate([10,26,18,14,16],1):
        ws5.column_dimensions[get_column_letter(c)].width=w

    return _excel_bytes_from_wb(wb)


# ─── INTERNACIONALES ─────────────────────────────────────────────────────────

def _excel_internacionales_viaje(viaje_id, nombre_viaje, id_vend_filtro=None):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    filtro_vend = f"AND ci.vendedora_id = {id_vend_filtro}" if id_vend_filtro else ""

    df_v = _query_rep(f"""
        SELECT ci.id, COALESCE(vd.nombre,'—') AS vendedora,
               ci.nombre_cliente AS cliente,
               ci.adultos, ci.menores,
               ci.habitaciones_doble, ci.habitaciones_triple,
               ci.total_usd AS total, ci.abonado_usd AS pagado,
               ci.saldo_usd AS saldo,
               COALESCE(ci.ganancia_usd, 0) AS ganancia,
               ci.estado, ci.fecha_registro
        FROM clientes_internacionales ci
        LEFT JOIN vendedoras vd ON ci.vendedora_id = vd.id
        WHERE ci.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY ci.id
    """)
    df_pas = _query_rep(f"""
        SELECT pi.cliente_id, pi.nombre_completo, pi.tipo, pi.habitacion_asignada
        FROM pasajeros_internacionales pi
        JOIN clientes_internacionales ci ON pi.cliente_id = ci.id
        WHERE ci.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY pi.cliente_id, pi.tipo
    """)
    df_vj = _query_rep(f"""SELECT id, destino AS nombre_viaje, destino, fecha_salida, fecha_regreso,
        dias, noches, cupos_totales, cupos_vendidos, cupos_disponibles, estado
        FROM viajes_internacionales WHERE id = {viaje_id}""")
    df_ab = _query_rep(f"""
        SELECT ai.cliente_id, ci.nombre_cliente AS cliente, ai.fecha,
               ai.monto_usd, COALESCE(ai.tipo_cambio, 0) AS tipo_cambio,
               COALESCE(ai.monto_mxn, 0) AS monto_mxn,
               COALESCE(ai.metodo_pago,'Efectivo') AS metodo
        FROM abonos_internacionales ai
        JOIN clientes_internacionales ci ON ai.cliente_id = ci.id
        WHERE ci.viaje_id = {viaje_id} {filtro_vend}
        ORDER BY ai.fecha
    """)

    wb = Workbook()

    # Hoja 1 — Info General
    ws1 = wb.active; ws1.title = "Info General"
    r = _titulo_sheet(ws1, f"TURISMAR — Internacional: {nombre_viaje}",
                      f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    if not df_vj.empty:
        rv = df_vj.iloc[0]
        campos=[("Viaje",rv.get("nombre_viaje","")),("Destino",rv.get("destino","")),
                ("Salida",rv.get("fecha_salida","")),("Regreso",rv.get("fecha_regreso","")),
                ("Días/Noches",f"{rv.get('dias','')} / {rv.get('noches','')}"),
                ("Estado",rv.get("estado","")),
                ("Total clientes",len(df_v)),
                ("Total USD",float(df_v['total'].sum()) if not df_v.empty else 0),
                ("Cobrado USD",float(df_v['pagado'].sum()) if not df_v.empty else 0),
                ("Saldo USD",float(df_v['saldo'].sum()) if not df_v.empty else 0)]
        ws1.cell(row=r,column=1).value="Campo"; ws1.cell(row=r,column=2).value="Valor"
        _estilo_header(ws1,r,2,bg="1F4E79"); r+=1
        for i,(k,val) in enumerate(campos):
            ws1.cell(row=r,column=1).value=k; ws1.cell(row=r,column=2).value=val
            _estilo_row(ws1,r,2,bg="EFF9FF" if i%2==0 else "FFFFFF")
            if isinstance(val,float): _fmt_cur(ws1,r,[2])
            r+=1
    ws1.column_dimensions["A"].width=25; ws1.column_dimensions["B"].width=30

    # Hoja 2 — Clientes y Pagos
    ws2 = wb.create_sheet("Clientes y Pagos")
    r2 = _titulo_sheet(ws2, f"Clientes — {nombre_viaje}")
    hdrs=["ID","Vendedora","Cliente","Adultos","Menores","Hab.Doble","Hab.Triple",
          "Total USD","Pagado USD","Saldo USD","Ganancia USD","Estado"]
    for c,h in enumerate(hdrs,1): ws2.cell(row=r2,column=c).value=h
    _estilo_header(ws2,r2,len(hdrs),bg="E91E8C"); r2+=1
    for _,row in (df_v.iterrows() if not df_v.empty else []):
        vals=[row.get("id"),row.get("vendedora"),row.get("cliente"),
              row.get("adultos"),row.get("menores"),
              row.get("habitaciones_doble"),row.get("habitaciones_triple"),
              row.get("total"),row.get("pagado"),row.get("saldo"),
              row.get("ganancia"),row.get("estado")]
        for c,val in enumerate(vals,1): ws2.cell(row=r2,column=c).value=val
        _estilo_row(ws2,r2,len(hdrs),bg="FFF0F7" if r2%2==0 else "FFFFFF")
        _fmt_cur(ws2,r2,[8,9,10,11]); r2+=1
    if not df_v.empty:
        n=len(df_v); ws2.cell(row=r2,column=1).value="TOTALES"
        for ci,col in [(8,"H"),(9,"I"),(10,"J"),(11,"K")]:
            ws2.cell(row=r2,column=ci).value=f"=SUM({col}{r2-n}:{col}{r2-1})"
        _estilo_row(ws2,r2,len(hdrs),bg="29ABE2",bold=True); _fmt_cur(ws2,r2,[8,9,10,11])
    for c,w in enumerate([6,18,26,8,8,10,10,14,14,14,14,12],1):
        ws2.column_dimensions[get_column_letter(c)].width=w

    # Hoja 3 — Pasajeros
    ws3 = wb.create_sheet("Pasajeros")
    r3 = _titulo_sheet(ws3, f"Pasajeros — {nombre_viaje}")
    if not df_pas.empty and not df_v.empty:
        mapa_cl = df_v.set_index("id")["cliente"].to_dict()
        df_pas["nombre_cliente"] = df_pas["cliente_id"].map(lambda x: mapa_cl.get(x,"—"))
        for c,h in enumerate(["ID Cliente","Nombre Cliente","Pasajero","Tipo","Habitación"],1):
            ws3.cell(row=r3,column=c).value=h
        _estilo_header(ws3,r3,5,bg="1F4E79"); r3+=1
        for _,row in df_pas.iterrows():
            for c,val in enumerate([row["cliente_id"],row["nombre_cliente"],
                                    row["nombre_completo"],row["tipo"],
                                    row.get("habitacion_asignada","")],1):
                ws3.cell(row=r3,column=c).value=val
            _estilo_row(ws3,r3,5,bg="EFF9FF" if r3%2==0 else "FFFFFF"); r3+=1
    else:
        ws3.cell(row=r3,column=1).value="Sin pasajeros registrados"
    for c,w in enumerate([10,24,28,12,14],1):
        ws3.column_dimensions[get_column_letter(c)].width=w

    # Hoja 4 — Distribución Habitaciones
    ws4 = wb.create_sheet("Habitaciones")
    r4 = _titulo_sheet(ws4, f"Habitaciones — {nombre_viaje}")
    if not df_v.empty:
        total_dobles  = int(df_v["habitaciones_doble"].sum())
        total_triples = int(df_v["habitaciones_triple"].sum())
        for c,h in enumerate(["Tipo","Cantidad"],1):
            ws4.cell(row=r4,column=c).value=h
        _estilo_header(ws4,r4,2,bg="1F4E79"); r4+=1
        for tipo,cant in [("DOBLE",total_dobles),("TRIPLE",total_triples)]:
            ws4.cell(row=r4,column=1).value=tipo
            ws4.cell(row=r4,column=2).value=cant
            _estilo_row(ws4,r4,2,bg="EFF9FF" if r4%2==0 else "FFFFFF"); r4+=1
        ws4.cell(row=r4,column=1).value="TOTAL"
        ws4.cell(row=r4,column=2).value=total_dobles+total_triples
        _estilo_row(ws4,r4,2,bg="29ABE2",bold=True)
    for c,w in enumerate([22,14],1):
        ws4.column_dimensions[get_column_letter(c)].width=w

    # Hoja 5 — Historial Pagos
    ws5 = wb.create_sheet("Historial Pagos")
    r5 = _titulo_sheet(ws5, f"Historial Pagos — {nombre_viaje}")
    if not df_ab.empty:
        for c,h in enumerate(["ID Cliente","Cliente","Fecha","Monto USD","Tipo Cambio","Monto MXN","Método"],1):
            ws5.cell(row=r5,column=c).value=h
        _estilo_header(ws5,r5,7,bg="2D6A4F"); r5+=1
        for _,row in df_ab.iterrows():
            for c,val in enumerate([row["cliente_id"],row["cliente"],row["fecha"],
                                    row["monto_usd"],row["tipo_cambio"],
                                    row["monto_mxn"],row["metodo"]],1):
                ws5.cell(row=r5,column=c).value=val
            _estilo_row(ws5,r5,7,bg="F0FFF4" if r5%2==0 else "FFFFFF")
            _fmt_cur(ws5,r5,[4,6]); r5+=1
        n5=len(df_ab); ws5.cell(row=r5,column=3).value="TOTAL"
        ws5.cell(row=r5,column=4).value=f"=SUM(D{r5-n5}:D{r5-1})"
        ws5.cell(row=r5,column=6).value=f"=SUM(F{r5-n5}:F{r5-1})"
        _estilo_row(ws5,r5,7,bg="29ABE2",bold=True); _fmt_cur(ws5,r5,[4,6])
    else:
        ws5.cell(row=r5,column=1).value="Sin abonos registrados"
    for c,w in enumerate([10,26,18,14,12,14,14],1):
        ws5.column_dimensions[get_column_letter(c)].width=w

    return _excel_bytes_from_wb(wb)


# ─── FINANCIERO ──────────────────────────────────────────────────────────────

def _excel_financiero(anio):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    MESES=["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    def mes_dict(df,col,col_mes="mes"):
        d={}
        if not df.empty:
            for _,row in df.iterrows():
                try: d[int(row[col_mes])]=float(row[col] or 0)
                except: pass
        return d
    df_rv=_query_rep(f"""SELECT strftime('%m',fecha_registro) AS mes, COUNT(*) AS ventas,
        SUM(precio_total) AS vendido, SUM(pagado) AS cobrado, SUM(ganancia) AS ganancia
        FROM ventas WHERE strftime('%Y',fecha_registro)='{anio}' GROUP BY mes ORDER BY mes""")
    df_nac=_query_rep(f"""SELECT strftime('%m',fecha_registro) AS mes, COUNT(*) AS ventas,
        SUM(total_pagar) AS vendido, SUM(total_abonado) AS cobrado
        FROM clientes_nacionales WHERE strftime('%Y',fecha_registro)='{anio}'
        GROUP BY mes ORDER BY mes""")
    df_int=_query_rep(f"""SELECT strftime('%m',fecha_registro) AS mes, COUNT(*) AS ventas,
        SUM(total_usd) AS vendido, SUM(abonado_usd) AS cobrado
        FROM clientes_internacionales WHERE strftime('%Y',fecha_registro)='{anio}'
        GROUP BY mes ORDER BY mes""")
    df_flujo=_query_rep(f"""SELECT strftime('%m',a.fecha) AS mes, SUM(a.monto) AS ingreso
        FROM abonos a WHERE strftime('%Y',a.fecha)='{anio}' GROUP BY mes ORDER BY mes""")
    df_ant=_query_rep(f"""SELECT strftime('%m',fecha_registro) AS mes,
        SUM(precio_total) AS vendido, SUM(pagado) AS cobrado, SUM(ganancia) AS ganancia
        FROM ventas WHERE strftime('%Y',fecha_registro)='{anio-1}'
        GROUP BY mes ORDER BY mes""")
    df_vend=_query_rep(f"""SELECT vd.nombre AS vendedora, COUNT(v.id) AS ventas,
        SUM(v.precio_total) AS vendido, SUM(v.pagado) AS cobrado,
        SUM(v.saldo) AS saldo, SUM(v.ganancia) AS ganancia,
        SUM(COALESCE(v.comision_vendedora,ROUND(v.ganancia*0.10,2))) AS comision_total,
        COUNT(CASE WHEN v.estado='LIQUIDADO' THEN 1 END) AS liquidadas,
        COUNT(CASE WHEN v.estado='CERRADO' THEN 1 END) AS cerradas
        FROM ventas v JOIN vendedoras vd ON v.vendedora_id=vd.id
        WHERE strftime('%Y',v.fecha_registro)='{anio}'
        GROUP BY vd.nombre ORDER BY vendido DESC""")
    rv_v=mes_dict(df_rv,"ventas"); rv_vnd=mes_dict(df_rv,"vendido")
    rv_cob=mes_dict(df_rv,"cobrado"); rv_gan=mes_dict(df_rv,"ganancia")
    nac_v=mes_dict(df_nac,"ventas"); nac_vnd=mes_dict(df_nac,"vendido"); nac_cob=mes_dict(df_nac,"cobrado")
    int_v=mes_dict(df_int,"ventas"); int_vnd=mes_dict(df_int,"vendido"); int_cob=mes_dict(df_int,"cobrado")
    flujo=mes_dict(df_flujo,"ingreso")
    ant_vnd=mes_dict(df_ant,"vendido"); ant_cob=mes_dict(df_ant,"cobrado"); ant_gan=mes_dict(df_ant,"ganancia")
    wb=Workbook()
    # Hoja 1
    ws1=wb.active; ws1.title="Resumen Anual"
    r=_titulo_sheet(ws1,f"TURISMAR — Reporte Financiero {anio}",
                    f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    hdrs=["Mes","Ventas RV","Vendido RV","Cobrado RV","Ganancia RV",
          "Ventas Nac","Vendido Nac","Cobrado Nac",
          "Ventas Int","Vendido Int USD","Cobrado Int USD","Flujo Caja"]
    for c,h in enumerate(hdrs,1): ws1.cell(row=r,column=c).value=h
    _estilo_header(ws1,r,len(hdrs),bg="1F4E79"); r+=1
    start_r=r
    for m in range(1,13):
        rd=[MESES[m-1],rv_v.get(m,0),rv_vnd.get(m,0),rv_cob.get(m,0),rv_gan.get(m,0),
            nac_v.get(m,0),nac_vnd.get(m,0),nac_cob.get(m,0),
            int_v.get(m,0),int_vnd.get(m,0),int_cob.get(m,0),flujo.get(m,0)]
        for c,val in enumerate(rd,1): ws1.cell(row=r,column=c).value=val
        _estilo_row(ws1,r,len(hdrs),bg="F0F9FF" if m%2==0 else "FFFFFF")
        _fmt_cur(ws1,r,[3,4,5,7,8,10,11,12]); r+=1
    ws1.cell(row=r,column=1).value=f"TOTAL {anio}"
    for ci in range(2,13):
        col_l=get_column_letter(ci)
        ws1.cell(row=r,column=ci).value=f"=SUM({col_l}{start_r}:{col_l}{r-1})"
    _estilo_row(ws1,r,len(hdrs),bg="E91E8C",bold=True); _fmt_cur(ws1,r,[3,4,5,7,8,10,11,12])
    for c,w in enumerate([10,10,16,16,16,10,16,16,10,16,16,16],1):
        ws1.column_dimensions[get_column_letter(c)].width=w
    # Hoja 2
    ws2=wb.create_sheet("Comparativa")
    r2=_titulo_sheet(ws2,f"Comparativa {anio-1} vs {anio}")
    hdrs2=["Mes",f"Vendido {anio-1}",f"Vendido {anio}","Var $","Var %",
           f"Cobrado {anio-1}",f"Cobrado {anio}","Var $","Var %",
           f"Ganancia {anio-1}",f"Ganancia {anio}","Var $","Var %"]
    for c,h in enumerate(hdrs2,1): ws2.cell(row=r2,column=c).value=h
    _estilo_header(ws2,r2,len(hdrs2),bg="2D6A4F"); r2+=1
    for m in range(1,13):
        a0v=ant_vnd.get(m,0); a1v=rv_vnd.get(m,0)
        a0c=ant_cob.get(m,0); a1c=rv_cob.get(m,0)
        a0g=ant_gan.get(m,0); a1g=rv_gan.get(m,0)
        rd=[MESES[m-1],a0v,a1v,a1v-a0v,(a1v-a0v)/a0v if a0v else 0,
            a0c,a1c,a1c-a0c,(a1c-a0c)/a0c if a0c else 0,
            a0g,a1g,a1g-a0g,(a1g-a0g)/a0g if a0g else 0]
        for c,val in enumerate(rd,1): ws2.cell(row=r2,column=c).value=val
        _estilo_row(ws2,r2,len(hdrs2),bg="F0FFF4" if m%2==0 else "FFFFFF")
        _fmt_cur(ws2,r2,[2,3,4,6,7,8,10,11,12])
        for ci in [5,9,13]: ws2.cell(row=r2,column=ci).number_format="0.0%"
        r2+=1
    for c,w in enumerate([10]+[14]*12,1):
        ws2.column_dimensions[get_column_letter(c)].width=w
    # Hoja 3
    ws3=wb.create_sheet("Por Vendedora")
    r3=_titulo_sheet(ws3,f"Rendimiento por Vendedora — {anio}")
    hdrs3=["Vendedora","Ventas","Vendido","Cobrado","Saldo","Ganancia",
           "Comisión Total","Liquidadas","Cerradas","% Liquidación"]
    for c,h in enumerate(hdrs3,1): ws3.cell(row=r3,column=c).value=h
    _estilo_header(ws3,r3,len(hdrs3),bg="E91E8C"); r3+=1
    for _,row in (df_vend.iterrows() if not df_vend.empty else []):
        tot=row.get("ventas",0)
        liq_pct=(float(row.get("liquidadas",0))+float(row.get("cerradas",0)))/float(tot) if tot>0 else 0
        for c,val in enumerate([row.get("vendedora"),row.get("ventas"),row.get("vendido"),
                                 row.get("cobrado"),row.get("saldo"),row.get("ganancia"),
                                 row.get("comision_total"),row.get("liquidadas"),
                                 row.get("cerradas"),liq_pct],1):
            ws3.cell(row=r3,column=c).value=val
        _estilo_row(ws3,r3,len(hdrs3),bg="FFF0F7" if r3%2==0 else "FFFFFF")
        _fmt_cur(ws3,r3,[3,4,5,6,7])
        ws3.cell(row=r3,column=10).number_format="0.0%"; r3+=1
    for c,w in enumerate([20,8,16,16,16,16,16,10,10,14],1):
        ws3.column_dimensions[get_column_letter(c)].width=w
    return _excel_bytes_from_wb(wb)


# ════════════════════════════════════════════════════════════════════════════
#  PÁGINA PRINCIPAL DE REPORTES — con control de roles
# ════════════════════════════════════════════════════════════════════════════

def pagina_reportes():
    """Reportes — Admin ve todo, vendedora solo sus propios reportes."""
    usuario  = st.session_state.usuario_actual
    es_admin = usuario.get("rol") == "ADMIN"
    id_vend  = usuario.get("id_vendedora")

    st.title("📊 Reportes y Análisis")

    # ── VENDEDORA ────────────────────────────────────────────────────────────
    if not es_admin:
        st.info(f"👋 Hola **{usuario['nombre']}** — descarga el reporte de tus ventas.")
        tabs_v = st.tabs(["🏖️ Mis Ventas Riviera","🎫 Mis Nacionales","🌎 Mis Internacionales"])

        with tabs_v[0]:
            st.markdown("Reporte de **tus** ventas Riviera Maya con resumen, clientes, habitaciones e historial.")
            if st.button("📥 Generar mi reporte Riviera", type="primary", key="v_rv"):
                with st.spinner("Generando..."):
                    try:
                        xls=_excel_riviera(id_vend=id_vend, label=usuario['nombre'])
                        st.download_button("⬇️ Descargar Excel", data=xls,
                            file_name=f"mis_ventas_riviera_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_v_rv")
                        st.success("✅ Listo")
                    except Exception as e:
                        st.error(f"❌ {e}")

        with tabs_v[1]:
            df_vj_v=_query_rep(f"""SELECT DISTINCT vj.id, vj.nombre_viaje, vj.destino, vj.fecha_salida
                FROM viajes_nacionales vj JOIN clientes_nacionales cn ON cn.viaje_id=vj.id
                WHERE cn.vendedora_id={id_vend} ORDER BY vj.fecha_salida DESC""")
            if df_vj_v.empty:
                st.info("No tienes clientes en viajes nacionales.")
            else:
                op={f"{r['nombre_viaje']} — {r['destino']} ({r['fecha_salida']})": r['id'] for _,r in df_vj_v.iterrows()}
                sel=st.selectbox("Viaje:", list(op.keys()), key="v_nac_sel")
                if st.button("📥 Generar reporte", type="primary", key="v_nac_btn"):
                    with st.spinner("Generando..."):
                        try:
                            xls=_excel_nacionales_viaje(op[sel], sel.split("—")[0].strip(), id_vend_filtro=id_vend)
                            st.download_button("⬇️ Descargar Excel", data=xls,
                                file_name=f"mis_nacionales_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_v_nac")
                            st.success("✅ Listo")
                        except Exception as e:
                            st.error(f"❌ {e}")

        with tabs_v[2]:
            df_vj_vi=_query_rep(f"""SELECT DISTINCT vi.id, vi.nombre_viaje, vi.destino, vi.fecha_salida
                FROM viajes_internacionales vi JOIN clientes_internacionales ci ON ci.viaje_id=vi.id
                WHERE ci.vendedora_id={id_vend} ORDER BY vi.fecha_salida DESC""")
            if df_vj_vi.empty:
                st.info("No tienes clientes en viajes internacionales.")
            else:
                op_i={f"{r['nombre_viaje']} — {r['destino']} ({r['fecha_salida']})": r['id'] for _,r in df_vj_vi.iterrows()}
                sel_i=st.selectbox("Viaje:", list(op_i.keys()), key="v_int_sel")
                if st.button("📥 Generar reporte", type="primary", key="v_int_btn"):
                    with st.spinner("Generando..."):
                        try:
                            xls=_excel_internacionales_viaje(op_i[sel_i], sel_i.split("—")[0].strip(), id_vend_filtro=id_vend)
                            st.download_button("⬇️ Descargar Excel", data=xls,
                                file_name=f"mis_internacionales_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_v_int")
                            st.success("✅ Listo")
                        except Exception as e:
                            st.error(f"❌ {e}")
        return

    # ── ADMIN ────────────────────────────────────────────────────────────────
    tabs = st.tabs(["🏖️ Riviera Maya","🎫 Nacionales","🌎 Internacionales","💰 Financiero"])

    with tabs[0]:
        st.subheader("🏖️ Reportes — Riviera Maya")
        subtabs=st.tabs(["📋 General","🔒 Bloqueos","👥 Grupos","🗂️ Todas","👩‍💼 Por Vendedora"])
        def _btn_rv(label, tipo, key, valor=None):
            st.caption("4 hojas: Resumen · Clientes y pagos · Habitaciones · Historial pagos")
            if st.button(f"📥 {label}", type="primary", key=key):
                with st.spinner("Generando..."):
                    try:
                        xls=_excel_riviera(tipo, valor_filtro=valor, label=label)
                        st.download_button("⬇️ Descargar Excel", data=xls,
                            file_name=f"riviera_{key}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{key}")
                        st.success("✅ Excel generado")
                    except Exception as e:
                        st.error(f"❌ {e}")
        with subtabs[0]: _btn_rv("Generar Excel General","todos","rv_gen")
        with subtabs[1]: _btn_rv("Generar Excel Bloqueos","bloqueos","rv_bloq")
        with subtabs[2]: _btn_rv("Generar Excel Grupos","grupos","rv_grp")
        with subtabs[3]: _btn_rv("Generar Excel — Todas las Ventas","todos","rv_all")
        with subtabs[4]:
            df_vends=_query_rep("SELECT nombre FROM vendedoras ORDER BY nombre")
            if not df_vends.empty:
                vsel=st.selectbox("Vendedora:", df_vends["nombre"].tolist(), key="rv_vend_sel")
                _btn_rv(f"Generar Excel — {vsel}","vendedora","rv_vend",valor=vsel)

    with tabs[1]:
        st.subheader("🎫 Reportes — Nacionales")
        subtabs_n=st.tabs(["📋 Por Viaje","🌐 Resumen Global"])
        with subtabs_n[0]:
            df_vj=_query_rep("SELECT id,nombre_viaje,destino,fecha_salida,estado,cupos_vendidos,cupos_totales FROM viajes_nacionales ORDER BY fecha_salida DESC")
            if df_vj.empty:
                st.info("Sin viajes nacionales.")
            else:
                op={f"{r['nombre_viaje']} — {r['destino']} ({r['fecha_salida']}) [{r['estado']}]": r['id'] for _,r in df_vj.iterrows()}
                sel=st.selectbox("Viaje:", list(op.keys()), key="nac_sel")
                rs=df_vj[df_vj['id']==op[sel]].iloc[0]
                st.caption(f"Cupos: {int(rs['cupos_vendidos'])}/{int(rs['cupos_totales'])} vendidos")
                st.caption("5 hojas: Info general · Clientes · Pasajeros · Habitaciones · Historial pagos")
                if st.button("📥 Generar Excel", type="primary", key="nac_btn"):
                    with st.spinner("Generando..."):
                        try:
                            xls=_excel_nacionales_viaje(op[sel], sel.split("—")[0].strip())
                            st.download_button("⬇️ Descargar Excel", data=xls,
                                file_name=f"nacional_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_nac")
                            st.success("✅ Excel generado con 5 hojas")
                        except Exception as e:
                            st.error(f"❌ {e}")
        with subtabs_n[1]:
            st.caption("Resumen de todos los viajes nacionales con ocupación y saldo.")
            if st.button("📥 Generar Excel Global", type="primary", key="nac_global"):
                with st.spinner("Generando..."):
                    try:
                        from openpyxl import Workbook
                        from openpyxl.utils import get_column_letter
                        df_g=_query_rep("""SELECT vj.nombre_viaje,vj.destino,vj.fecha_salida,vj.fecha_regreso,
                            vj.estado,vj.cupos_totales,vj.cupos_vendidos,
                            COUNT(cn.id) AS clientes,
                            SUM(cn.total_pagar) AS total_vendido,
                            SUM(cn.total_abonado) AS total_cobrado,
                            SUM(cn.saldo) AS saldo_pendiente
                            FROM viajes_nacionales vj
                            LEFT JOIN clientes_nacionales cn ON cn.viaje_id=vj.id
                            GROUP BY vj.id ORDER BY vj.fecha_salida DESC""")
                        wb2=Workbook(); ws=wb2.active; ws.title="Resumen Global"
                        r=_titulo_sheet(ws,"TURISMAR — Resumen Global Nacionales",
                                        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                        hdrs=["Viaje","Destino","Salida","Regreso","Estado",
                              "Cupos Tot","Cupos Vend","Clientes","Total Vendido","Total Cobrado","Saldo"]
                        for c,h in enumerate(hdrs,1): ws.cell(row=r,column=c).value=h
                        _estilo_header(ws,r,len(hdrs),bg="E91E8C"); r+=1
                        for _,row in (df_g.iterrows() if not df_g.empty else []):
                            for c,val in enumerate(row.tolist(),1): ws.cell(row=r,column=c).value=val
                            _estilo_row(ws,r,len(hdrs),bg="FFF0F7" if r%2==0 else "FFFFFF")
                            _fmt_cur(ws,r,[9,10,11]); r+=1
                        for c,w in enumerate([28,20,12,12,10,8,8,8,16,16,16],1):
                            ws.column_dimensions[get_column_letter(c)].width=w
                        xls2=_excel_bytes_from_wb(wb2)
                        st.download_button("⬇️ Descargar Excel", data=xls2,
                            file_name=f"nacionales_global_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_nac_g")
                        st.success("✅ Listo")
                    except Exception as e:
                        st.error(f"❌ {e}")

    with tabs[2]:
        st.subheader("🌎 Reportes — Internacionales")
        subtabs_i=st.tabs(["📋 Por Viaje","🌐 Resumen Global"])
        with subtabs_i[0]:
            df_vi=_query_rep("SELECT id, destino AS nombre_viaje, destino, fecha_salida, estado FROM viajes_internacionales ORDER BY fecha_salida DESC")
            if df_vi.empty:
                st.info("Sin viajes internacionales.")
            else:
                op_i={f"{r['nombre_viaje']} — {r['destino']} ({r['fecha_salida']}) [{r['estado']}]": r['id'] for _,r in df_vi.iterrows()}
                sel_i=st.selectbox("Viaje:", list(op_i.keys()), key="int_sel")
                st.caption("5 hojas: Info general · Clientes · Pasajeros · Habitaciones · Historial pagos")
                if st.button("📥 Generar Excel", type="primary", key="int_btn"):
                    with st.spinner("Generando..."):
                        try:
                            xls=_excel_internacionales_viaje(op_i[sel_i], sel_i.split("—")[0].strip())
                            st.download_button("⬇️ Descargar Excel", data=xls,
                                file_name=f"internacional_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_int")
                            st.success("✅ Excel generado con 5 hojas")
                        except Exception as e:
                            st.error(f"❌ {e}")
        with subtabs_i[1]:
            st.caption("Resumen de todos los viajes internacionales en USD.")
            if st.button("📥 Generar Excel Global", type="primary", key="int_global"):
                with st.spinner("Generando..."):
                    try:
                        from openpyxl import Workbook
                        from openpyxl.utils import get_column_letter
                        df_gi=_query_rep("""SELECT vi.destino AS nombre_viaje, vi.destino, vi.fecha_salida, vi.estado,
                            COUNT(ci.id) AS clientes,
                            SUM(ci.total_usd) AS total_usd,
                            SUM(ci.abonado_usd) AS cobrado_usd,
                            SUM(ci.saldo_usd) AS saldo_usd
                            FROM viajes_internacionales vi
                            LEFT JOIN clientes_internacionales ci ON ci.viaje_id=vi.id
                            GROUP BY vi.id ORDER BY vi.fecha_salida DESC""")
                        wb3=Workbook(); ws=wb3.active; ws.title="Resumen Global"
                        r=_titulo_sheet(ws,"TURISMAR — Resumen Global Internacionales",
                                        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
                        hdrs=["Viaje","Destino","Salida","Estado","Clientes","Total USD","Cobrado USD","Saldo USD"]
                        for c,h in enumerate(hdrs,1): ws.cell(row=r,column=c).value=h
                        _estilo_header(ws,r,len(hdrs),bg="1F4E79"); r+=1
                        for _,row in (df_gi.iterrows() if not df_gi.empty else []):
                            for c,val in enumerate(row.tolist(),1): ws.cell(row=r,column=c).value=val
                            _estilo_row(ws,r,len(hdrs),bg="EFF9FF" if r%2==0 else "FFFFFF")
                            _fmt_cur(ws,r,[6,7,8]); r+=1
                        for c,w in enumerate([28,20,12,10,8,14,14,14],1):
                            ws.column_dimensions[get_column_letter(c)].width=w
                        xls3=_excel_bytes_from_wb(wb3)
                        st.download_button("⬇️ Descargar Excel", data=xls3,
                            file_name=f"internacionales_global_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_int_g")
                        st.success("✅ Listo")
                    except Exception as e:
                        st.error(f"❌ {e}")

    with tabs[3]:
        st.subheader("💰 Reporte Financiero")
        col1,col2=st.columns([1,2])
        with col1:
            anio_sel=st.selectbox("Año:", list(range(datetime.now().year, 2022, -1)), key="fin_anio")
        with col2:
            st.info("3 hojas: Resumen anual · Comparativa vs año anterior · Por vendedora")
        MESES_ES=["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
        df_prev=_query_rep(f"""SELECT strftime('%m',fecha_registro) AS mes, COUNT(*) AS ventas,
            SUM(precio_total) AS vendido, SUM(pagado) AS cobrado, SUM(ganancia) AS ganancia
            FROM ventas WHERE strftime('%Y',fecha_registro)='{anio_sel}'
            GROUP BY mes ORDER BY mes""")
        if not df_prev.empty:
            df_prev["mes_nombre"]=df_prev["mes"].astype(int).apply(lambda x: MESES_ES[x-1])
            c1,c2,c3,c4=st.columns(4)
            c1.metric("Ventas Riviera",int(df_prev["ventas"].sum()))
            c2.metric("Total vendido",f"${df_prev['vendido'].sum():,.0f}")
            c3.metric("Total cobrado",f"${df_prev['cobrado'].sum():,.0f}")
            c4.metric("Ganancia",f"${df_prev['ganancia'].sum():,.0f}")
            st.divider()
            df_ch=df_prev[["mes_nombre","vendido","cobrado","ganancia"]].copy()
            df_ch.columns=["Mes","Vendido","Cobrado","Ganancia"]
            st.bar_chart(df_ch.set_index("Mes"))
        else:
            st.info(f"Sin datos de Riviera Maya para {anio_sel}.")
        st.divider()
        if st.button("📥 Generar Excel Financiero Completo", type="primary", key="fin_btn"):
            with st.spinner("Generando..."):
                try:
                    xls=_excel_financiero(anio_sel)
                    st.download_button("⬇️ Descargar Excel", data=xls,
                        file_name=f"financiero_turismar_{anio_sel}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_fin")
                    st.success("✅ Excel generado")
                except Exception as e:
                    st.error(f"❌ {e}")



def pagina_configuracion():
    """Página de configuración y comisiones — solo ADMIN"""
    st.title("⚙️ Configuración")

    tabs = st.tabs([
        "💰 Comisiones",
        "👩‍💼 Vendedoras",
        "🏨 Hoteles / Catálogos",
        "🏢 Mayoristas",
        "📦 Bloqueos",
        "👥 Grupos"
    ])

    # ══════════════════════════════════════════════════════
    # TAB 0 — COMISIONES
    # ══════════════════════════════════════════════════════
    with tabs[0]:
        st.subheader("💰 Gestión de Comisiones")
        st.info("💡 **Flujo:** Venta **ACTIVO** → se liquida → **LIQUIDADO** → se paga comisión → **CERRADO** (historial)")

        subtabs_com = st.tabs([
            "📋 Pendientes de Pago",
            "✅ Historial (Cerradas)",
            "⚙️ Configurar %"
        ])

        # ── Pendientes ────────────────────────────────────
        with subtabs_com[0]:
            st.markdown("#### 💸 Comisiones Pendientes de Pago")
            st.markdown("Ventas en estado **LIQUIDADO** cuya comisión aún no se ha pagado. Al pagar → pasan a **CERRADO**.")

            conn = conectar_db()

            # ── Riviera Maya: usa comision_vendedora ya calculada (10% ganancia) ──
            try:
                df_riv = pd.read_sql_query("""
                    SELECT
                        vd.nombre        AS vendedora,
                        v.id,
                        v.cliente,
                        v.destino,
                        v.ganancia,
                        COALESCE(v.comision_vendedora, ROUND(v.ganancia * 0.10, 2)) AS comision,
                        v.fecha_registro
                    FROM ventas v
                    JOIN vendedoras vd ON v.vendedora_id = vd.id
                    WHERE v.estado = 'LIQUIDADO'
                      AND COALESCE(v.comision_pagada, 0) = 0
                    ORDER BY vd.nombre, v.id
                """, conn)
            except Exception as e:
                df_riv = pd.DataFrame()

            # ── Nacionales: % configurable sobre ganancia ──
            try:
                df_nac = pd.read_sql_query("""
                    SELECT
                        vd.nombre           AS vendedora,
                        cn.id,
                        cn.nombre_cliente   AS cliente,
                        vj.nombre_viaje     AS destino,
                        COALESCE(cn.ganancia, cn.costo_neto, 0)  AS ganancia,
                        ROUND(COALESCE(cn.ganancia, cn.costo_neto, 0)
                              * COALESCE(cfg.porcentaje_comision, 10) / 100.0, 2) AS comision
                    FROM clientes_nacionales cn
                    JOIN vendedoras vd ON cn.vendedora_id = vd.id
                    JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                    LEFT JOIN config_comisiones cfg ON cfg.tipo = 'nacionales'
                    WHERE cn.estado = 'LIQUIDADO'
                      AND COALESCE(cn.comision_pagada, 0) = 0
                    ORDER BY vd.nombre, cn.id
                """, conn)
            except Exception as e:
                df_nac = pd.DataFrame()

            # ── Internacionales: % configurable sobre ganancia_usd ──
            try:
                df_int = pd.read_sql_query("""
                    SELECT
                        'Sin asignar'    AS vendedora,
                        ci.id,
                        ci.nombre_cliente AS cliente,
                        vi.destino,
                        COALESCE(ci.ganancia_usd, 0)  AS ganancia,
                        ROUND(COALESCE(ci.ganancia_usd, 0)
                              * COALESCE(cfg.porcentaje_comision, 10) / 100.0, 2) AS comision
                    FROM clientes_internacionales ci
                    JOIN viajes_internacionales vi ON ci.viaje_id = vi.id
                    LEFT JOIN config_comisiones cfg ON cfg.tipo = 'internacionales'
                    WHERE ci.estado = 'LIQUIDADO'
                      AND COALESCE(ci.comision_pagada, 0) = 0
                    ORDER BY ci.id
                """, conn)
            except Exception as e:
                df_int = pd.DataFrame()

            conn.close()

            # ── Resumen pivot por vendedora ───────────────────────────────────
            frames_sum = []
            if not df_riv.empty:
                tmp = df_riv[["vendedora","comision"]].copy(); tmp["tipo"] = "Riviera"
                frames_sum.append(tmp)
            if not df_nac.empty:
                tmp = df_nac[["vendedora","comision"]].copy(); tmp["tipo"] = "Nacional"
                frames_sum.append(tmp)
            if not df_int.empty:
                tmp = df_int[["vendedora","comision"]].copy(); tmp["tipo"] = "Internacional"
                frames_sum.append(tmp)

            if frames_sum:
                df_res = pd.concat(frames_sum, ignore_index=True)
                df_piv = df_res.groupby(["vendedora","tipo"])["comision"].sum().unstack(fill_value=0)
                df_piv["TOTAL"] = df_piv.sum(axis=1)
                df_piv = df_piv.reset_index()
                total_global = df_piv["TOTAL"].sum()
                col1, col2 = st.columns([3,1])
                col1.dataframe(
                    df_piv.style.format({c: "${:,.2f}" for c in df_piv.columns if c != "vendedora"}),
                    use_container_width=True, hide_index=True)
                col2.metric("💰 Total global pendiente", f"${total_global:,.2f}")
            else:
                st.success("✅ No hay comisiones pendientes de pago.")

            st.divider()

            # ── Selección de tipo y pago individual por venta ────────────────
            tipo_com = st.radio("Trabajar con:", ["🏖️ Riviera Maya", "🎫 Nacionales", "🌎 Internacionales"],
                                horizontal=True, key="com_tipo_radio")

            if "Riviera" in tipo_com:
                df_det  = df_riv.copy()
                tabla   = "ventas"
                moneda  = "$"
                nota_tc = "Comisión = 10% de la ganancia (fijo, según pagos.py)"
            elif "Nacionales" in tipo_com:
                df_det  = df_nac.copy()
                tabla   = "clientes_nacionales"
                moneda  = "$"
                nota_tc = "Comisión = % configurable sobre la ganancia"
            else:
                df_det  = df_int.copy()
                tabla   = "clientes_internacionales"
                moneda  = "USD $"
                nota_tc = "Comisión = % configurable sobre ganancia_usd"

            st.caption(f"ℹ️ {nota_tc}")

            if df_det.empty:
                st.info(f"No hay comisiones pendientes en {tipo_com}.")
            else:
                # Tabla detalle
                df_det_show = df_det.copy()
                df_det_show["ganancia"] = df_det_show["ganancia"].apply(lambda x: f"{moneda}{x:,.2f}")
                df_det_show["comision"] = df_det_show["comision"].apply(lambda x: f"{moneda}{x:,.2f}")
                st.dataframe(
                    df_det_show[["vendedora","id","cliente","destino","ganancia","comision"]],
                    use_container_width=True, hide_index=True)

                st.divider()
                st.markdown("##### 💸 Pagar comisión — selecciona venta(s)")

                # Seleccionar vendedora y ver sus ventas pendientes
                vendedoras_disp = sorted(df_det["vendedora"].unique().tolist())
                col1, col2 = st.columns(2)
                with col1:
                    vend_sel_com = st.selectbox("Vendedora:", vendedoras_disp, key=f"com_vend_{tipo_com}")
                
                df_vend_com = df_det[df_det["vendedora"] == vend_sel_com].copy()

                # Checkboxes para seleccionar qué ventas pagar
                st.markdown(f"**Ventas liquidadas de {vend_sel_com}:**")
                ids_seleccionados = []
                monto_seleccionado = 0.0

                for _, row_com in df_vend_com.iterrows():
                    check = st.checkbox(
                        f"ID {int(row_com['id'])} | {row_com['cliente']} — {row_com['destino']} | Comisión: {moneda}{row_com['comision']:,.2f}",
                        value=True,
                        key=f"com_chk_{tipo_com}_{int(row_com['id'])}"
                    )
                    if check:
                        ids_seleccionados.append(int(row_com["id"]))
                        monto_seleccionado += row_com["comision"]

                if ids_seleccionados:
                    col1, col2, col3 = st.columns(3)
                    col1.metric("Ventas seleccionadas", len(ids_seleccionados))
                    col2.metric("Total a pagar", f"{moneda}{monto_seleccionado:,.2f}")
                    col3.metric("Vendedora", vend_sel_com)

                    col1, col2 = st.columns(2)
                    with col1:
                        metodo_com = st.selectbox("Método de pago:",
                            ["Efectivo","Transferencia","Tarjeta de débito","Otro"],
                            key=f"com_metodo_{tipo_com}")
                    with col2:
                        nota_com = st.text_input("Nota (opcional):", key=f"com_nota_{tipo_com}")

                    if st.button(
                        f"💸 Pagar {moneda}{monto_seleccionado:,.2f} a {vend_sel_com} → ventas pasan a CERRADO",
                        type="primary", use_container_width=True, key=f"com_btn_pagar_{tipo_com}"
                    ):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            fecha_pago_com = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                            vend_id_com = 0
                            try:
                                _cur = conn.cursor()
                                _cur.execute("SELECT id FROM vendedoras WHERE nombre = ?", (vend_sel_com,))
                                _row = _cur.fetchone()
                                if _row: vend_id_com = _row[0]
                            except: pass

                            for vid in ids_seleccionados:
                                comision_vid = float(df_vend_com[df_vend_com["id"]==vid]["comision"].values[0])

                                # Marcar comision_pagada = 1 y pasar a CERRADO
                                if "Riviera" in tipo_com:
                                    cursor.execute("""
                                        UPDATE ventas
                                        SET comision_pagada = 1,
                                            estado = 'CERRADO',
                                            fecha_pago_comision = ?
                                        WHERE id = ?
                                    """, (fecha_pago_com, vid))
                                else:
                                    try:
                                        cursor.execute(f"""
                                            UPDATE {tabla}
                                            SET comision_pagada = 1,
                                                estado = 'CERRADO'
                                            WHERE id = ?
                                        """, (vid,))
                                    except:
                                        cursor.execute(f"""
                                            UPDATE {tabla}
                                            SET comision_pagada = 1
                                            WHERE id = ?
                                        """, (vid,))

                                # Registrar en historial
                                cursor.execute("""CREATE TABLE IF NOT EXISTS historial_comisiones (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    vendedora TEXT, vendedora_id INTEGER DEFAULT 0,
                                    fecha TEXT, tipo TEXT, referencia_id INTEGER,
                                    monto REAL, metodo_pago TEXT, nota TEXT)""")
                                try: cursor.execute("ALTER TABLE historial_comisiones ADD COLUMN vendedora_id INTEGER DEFAULT 0")
                                except: pass
                                cursor.execute("""INSERT INTO historial_comisiones
                                    (tipo, referencia_id, vendedora, vendedora_id, monto, metodo_pago, nota, fecha)
                                    VALUES (?,?,?,?,?,?,?,?)""",
                                    (tipo_com.replace("🏖️ ","").replace("🎫 ","").replace("🌎 ",""),
                                     vid, vend_sel_com, vend_id_com, comision_vid,
                                     metodo_com, nota_com, fecha_pago_com))

                            conn.commit()
                            conn.close()
                            st.success(f"✅ Comisión de {moneda}{monto_seleccionado:,.2f} pagada a **{vend_sel_com}**. Las ventas ahora están en **CERRADO**.")
                            st.balloons()
                            st.rerun()

                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                            st.caption("💡 Asegúrate de que existan: columna `comision_pagada` (INTEGER) y `fecha_pago_comision` (TEXT) en `ventas`, y tabla `historial_comisiones`.")

        # ── Historial pagadas ─────────────────────────────
        with subtabs_com[1]:
            st.markdown("#### ✅ Historial de Comisiones Pagadas")

            conn = conectar_db()
            try:
                df_hist = pd.read_sql_query("""
                    SELECT fecha, vendedora, tipo, referencia_id,
                           monto, metodo_pago, nota
                    FROM historial_comisiones
                    ORDER BY fecha DESC
                """, conn)
            except:
                df_hist = pd.DataFrame()
            conn.close()

            if df_hist.empty:
                st.info("No hay comisiones pagadas registradas aún.")
            else:
                # Filtros
                col1, col2 = st.columns(2)
                with col1:
                    filtro_vend_h = st.selectbox("Filtrar por vendedora:",
                        ["Todas"] + sorted(df_hist["vendedora"].unique().tolist()),
                        key="hist_vend")
                with col2:
                    filtro_tipo_h = st.selectbox("Filtrar por tipo:",
                        ["Todos"] + sorted(df_hist["tipo"].unique().tolist()),
                        key="hist_tipo")

                df_hf = df_hist.copy()
                if filtro_vend_h != "Todas":
                    df_hf = df_hf[df_hf["vendedora"] == filtro_vend_h]
                if filtro_tipo_h != "Todos":
                    df_hf = df_hf[df_hf["tipo"] == filtro_tipo_h]

                total_hist = df_hf["monto"].sum()
                col1, col2 = st.columns(2)
                col1.metric("Registros", len(df_hf))
                col2.metric("Total pagado", f"${total_hist:,.2f}")

                df_hf_show = df_hf.copy()
                df_hf_show["monto"] = df_hf_show["monto"].apply(lambda x: f"${x:,.2f}")
                df_hf_show.columns = ["Fecha","Vendedora","Tipo","ID Ref.","Monto","Método","Nota"]
                st.dataframe(df_hf_show, use_container_width=True, hide_index=True)

                csv_hist = df_hf_show.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Descargar historial",
                    csv_hist, file_name=f"historial_comisiones_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv")

        # ── Configurar porcentaje ─────────────────────────
        with subtabs_com[2]:
            st.markdown("#### ⚙️ Porcentaje de Comisión por Tipo")
            st.info("💡 Este porcentaje se aplica **sobre la ganancia** de cada venta liquidada.")

            conn = conectar_db()
            try:
                df_cfg = pd.read_sql_query(
                    "SELECT tipo, porcentaje_comision FROM config_comisiones", conn)
                cfg_dict = dict(zip(df_cfg["tipo"], df_cfg["porcentaje_comision"]))
            except:
                cfg_dict = {}
            conn.close()

            col1, col2, col3 = st.columns(3)
            with col1:
                pct_riv = st.number_input("Riviera Maya (%)",
                    min_value=1, max_value=100,
                    value=int(cfg_dict.get("riviera", 30)),
                    key="cfg_pct_riv")
            with col2:
                pct_nac = st.number_input("Nacionales (%)",
                    min_value=1, max_value=100,
                    value=int(cfg_dict.get("nacionales", 30)),
                    key="cfg_pct_nac")
            with col3:
                pct_int = st.number_input("Internacionales (%)",
                    min_value=1, max_value=100,
                    value=int(cfg_dict.get("internacionales", 30)),
                    key="cfg_pct_int")

            st.markdown("")
            st.markdown("**Ejemplo con 30%:** Si una venta tuvo $1,000 de ganancia → comisión = $300")

            if st.button("💾 Guardar Porcentajes", type="primary", key="cfg_guardar_pct"):
                try:
                    conn = conectar_db()
                    cursor = conn.cursor()
                    for tipo, pct in [("riviera", pct_riv), ("nacionales", pct_nac), ("internacionales", pct_int)]:
                        cursor.execute("""
                            INSERT INTO config_comisiones (tipo, porcentaje_comision)
                            VALUES (?, ?)
                            ON CONFLICT(tipo) DO UPDATE SET porcentaje_comision = excluded.porcentaje_comision
                        """, (tipo, pct))
                    conn.commit()
                    conn.close()
                    st.success("✅ Porcentajes guardados correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Error: {e}")
                    st.caption("💡 Asegúrate de que exista la tabla `config_comisiones` con columnas `tipo` (UNIQUE) y `porcentaje_comision`.")

    # ══════════════════════════════════════════════════════
    # TAB 1 — VENDEDORAS
    # ══════════════════════════════════════════════════════
    with tabs[1]:
        st.subheader("👩‍💼 Gestión de Vendedoras")
        subtabs_vend = st.tabs(["➕ Nueva Vendedora", "✏️ Editar / Activar"])

        with subtabs_vend[0]:
            st.markdown("#### Registrar Nueva Vendedora")
            nv_nombre = st.text_input("Nombre completo *", key="nv_nombre")
            if st.button("💾 Registrar Vendedora", type="primary", key="nv_guardar"):
                if not nv_nombre.strip():
                    st.error("❌ El nombre es obligatorio.")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute(
                            "INSERT INTO vendedoras (nombre, activa, fecha_registro) VALUES (?, 1, ?)",
                            (nv_nombre.strip(), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                        )
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Vendedora **{nv_nombre}** registrada.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

        with subtabs_vend[1]:
            st.markdown("#### Editar o cambiar estado de vendedora")
            conn = conectar_db()
            try:
                df_vends = pd.read_sql_query(
                    "SELECT id, nombre, activa FROM vendedoras ORDER BY nombre", conn)
            except:
                df_vends = pd.DataFrame()
            conn.close()

            if df_vends.empty:
                st.info("No hay vendedoras registradas.")
            else:
                vend_opts = {
                    f"ID {row['id']} | {row['nombre']} ({'ACTIVA' if row['activa'] else 'INACTIVA'})": row
                    for _, row in df_vends.iterrows()
                }
                vend_edit_label = st.selectbox("Selecciona vendedora:", list(vend_opts.keys()), key="vend_edit_sel")
                vend_edit = vend_opts[vend_edit_label]

                col1, col2 = st.columns(2)
                with col1:
                    nuevo_nombre_v = st.text_input("Nombre", value=vend_edit["nombre"], key="vend_edit_nom")
                with col2:
                    nuevo_estado_v = st.selectbox("Estado",
                        ["ACTIVA", "INACTIVA"],
                        index=0 if vend_edit["activa"] else 1,
                        key="vend_edit_est")

                if st.button("💾 Guardar Cambios", type="primary", key="vend_edit_guardar"):
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute(
                            "UPDATE vendedoras SET nombre = ?, activa = ? WHERE id = ?",
                            (nuevo_nombre_v.strip(), 1 if nuevo_estado_v == "ACTIVA" else 0, int(vend_edit["id"]))
                        )
                        conn.commit()
                        conn.close()
                        st.success("✅ Vendedora actualizada.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

    # ══════════════════════════════════════════════════════
    # TAB 2 — HOTELES / CATÁLOGOS
    # ══════════════════════════════════════════════════════
    with tabs[2]:
        st.subheader("🏨 Catálogo de Hoteles")
        subtabs_hot = st.tabs(["➕ Nuevo Hotel", "📋 Ver Hoteles"])

        with subtabs_hot[0]:
            st.markdown("#### ➕ Registrar Nuevo Hotel")
            col1, col2 = st.columns(2)
            with col1:
                nh_nombre = st.text_input("Nombre del hotel *",
                    placeholder="Ej: Barceló Maya Grand", key="nh_nombre")
            with col2:
                nh_estrellas = st.selectbox("Estrellas", [3, 4, 5], index=1, key="nh_estrellas")

            nh_direccion = st.text_input("Dirección",
                placeholder="Ej: Carr. Cancún-Tulum Km. 266,3, Xpu Ha, Q.R.",
                key="nh_direccion")
            nh_telefono = st.text_input("Teléfono",
                placeholder="Ej: 984 875 1500", key="nh_telefono")

            if st.button("💾 Agregar Hotel", type="primary",
                         use_container_width=True, key="nh_guardar"):
                if not nh_nombre.strip():
                    st.error("❌ El nombre es obligatorio.")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        try:
                            cursor.execute(
                                """INSERT INTO hoteles
                                   (nombre, direccion, telefono, estrellas, activo, veces_usado)
                                   VALUES (?, ?, ?, ?, 1, 0)""",
                                (nh_nombre.strip(),
                                 nh_direccion.strip() or None,
                                 nh_telefono.strip() or None,
                                 nh_estrellas)
                            )
                        except Exception:
                            cursor.execute(
                                "INSERT INTO hoteles (nombre, activo, veces_usado) VALUES (?, 1, 0)",
                                (nh_nombre.strip(),)
                            )
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Hotel **{nh_nombre}** agregado.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

        with subtabs_hot[1]:
            conn = conectar_db()
            try:
                df_hoteles = pd.read_sql_query(
                    """SELECT id, nombre,
                              COALESCE(estrellas, 4) as estrellas,
                              COALESCE(direccion, '—') as direccion,
                              COALESCE(telefono, '—') as telefono,
                              activo, veces_usado
                       FROM hoteles ORDER BY nombre""", conn)
            except Exception:
                df_hoteles = pd.DataFrame()
            conn.close()

            if df_hoteles.empty:
                st.info("No hay hoteles en el catálogo.")
            else:
                df_hoteles["activo"]    = df_hoteles["activo"].apply(lambda x: "✅ Activo" if x else "❌ Inactivo")
                df_hoteles["estrellas"] = df_hoteles["estrellas"].apply(lambda x: "★" * int(x) if x else "")
                df_hoteles.columns = ["ID","Hotel","Estrellas","Dirección","Teléfono","Estado","Veces usado"]
                st.dataframe(df_hoteles, use_container_width=True, hide_index=True)

    # ══════════════════════════════════════════════════════
    # TAB 3 — MAYORISTAS / OPERADORES
    # ══════════════════════════════════════════════════════
    with tabs[3]:
        st.subheader("🏢 Mayoristas / Operadores")
        st.info("💡 Los mayoristas aparecen en el selector **Operador** al registrar una Venta General en Riviera Maya.")

        subtabs_may = st.tabs(["➕ Nuevo Mayorista", "📋 Ver / Editar"])

        # ── Nuevo Mayorista ───────────────────────────────
        with subtabs_may[0]:
            st.markdown("#### Registrar Nuevo Mayorista")
            col1, col2 = st.columns(2)
            with col1:
                may_nombre = st.text_input("Nombre del mayorista / operador *",
                    placeholder="Ej: Amstar, Despegar, Apple Leisure", key="may_nombre")
            with col2:
                may_contacto = st.text_input("Contacto / teléfono (opcional)", key="may_contacto")
            may_notas = st.text_area("Notas (opcional)", height=80, key="may_notas")

            if st.button("💾 Agregar Mayorista", type="primary", use_container_width=True, key="may_guardar"):
                if not may_nombre.strip():
                    st.error("❌ El nombre es obligatorio.")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        # Intentar con columnas extra, si no existen usar solo nombre y activo
                        try:
                            cursor.execute(
                                "INSERT INTO operadores (nombre, contacto, notas, activo) VALUES (?, ?, ?, 1)",
                                (may_nombre.strip(), may_contacto.strip() or None, may_notas.strip() or None)
                            )
                        except Exception:
                            cursor.execute(
                                "INSERT INTO operadores (nombre, activo) VALUES (?, 1)",
                                (may_nombre.strip(),)
                            )
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Mayorista **{may_nombre}** agregado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

        # ── Ver / Editar Mayoristas ───────────────────────
        with subtabs_may[1]:
            st.markdown("#### 📋 Mayoristas Registrados")
            conn = conectar_db()
            try:
                df_may = pd.read_sql_query(
                    "SELECT id, nombre, activo FROM operadores ORDER BY nombre", conn)
            except Exception:
                df_may = pd.DataFrame()
            conn.close()

            if df_may.empty:
                st.info("No hay mayoristas registrados. Agrégalos en la pestaña ➕ Nuevo Mayorista.")
            else:
                # Métricas
                activos_may = df_may[df_may['activo'] == 1]
                col1, col2 = st.columns(2)
                col1.metric("Total mayoristas", len(df_may))
                col2.metric("Activos", len(activos_may))

                st.divider()

                # Tabla con estado legible
                df_may_show = df_may.copy()
                df_may_show['activo'] = df_may_show['activo'].apply(
                    lambda x: "✅ Activo" if x else "❌ Inactivo")
                df_may_show.columns = ["ID", "Mayorista / Operador", "Estado"]
                st.dataframe(df_may_show, use_container_width=True, hide_index=True)

                st.divider()
                st.markdown("##### ✏️ Editar un mayorista")

                may_opts = {
                    f"ID {row['id']} | {row['nombre']} ({'Activo' if row['activo'] == 1 else 'Inactivo'})": row
                    for _, row in df_may.iterrows()
                }
                may_sel_label = st.selectbox("Selecciona:", list(may_opts.keys()), key="may_ed_sel")
                may_sel = may_opts[may_sel_label]

                col1, col2 = st.columns(2)
                with col1:
                    may_ed_nombre = st.text_input("Nombre", value=may_sel['nombre'], key="may_ed_nombre")
                with col2:
                    may_ed_estado = st.selectbox("Estado",
                        ["ACTIVO", "INACTIVO"],
                        index=0 if may_sel['activo'] == 1 else 1,
                        key="may_ed_estado")

                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("💾 Guardar cambios", type="primary", key="may_ed_guardar"):
                        if not may_ed_nombre.strip():
                            st.error("❌ El nombre no puede estar vacío.")
                        else:
                            try:
                                conn = conectar_db()
                                cursor = conn.cursor()
                                cursor.execute(
                                    "UPDATE operadores SET nombre = ?, activo = ? WHERE id = ?",
                                    (may_ed_nombre.strip(),
                                     1 if may_ed_estado == "ACTIVO" else 0,
                                     int(may_sel['id']))
                                )
                                conn.commit()
                                conn.close()
                                st.success(f"✅ Mayorista actualizado → **{may_ed_nombre}** ({may_ed_estado})")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error: {e}")
                with col_btn2:
                    if st.button("🗑️ Eliminar mayorista", key="may_ed_borrar"):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            cursor.execute("DELETE FROM operadores WHERE id = ?", (int(may_sel['id']),))
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Mayorista **{may_sel['nombre']}** eliminado.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error al eliminar: {e}")

    # ══════════════════════════════════════════════════════
    # TAB 4 — BLOQUEOS
    # ══════════════════════════════════════════════════════
    with tabs[4]:
        st.subheader("📦 Gestión de Bloqueos")
        st.info("💡 Un **bloqueo** es un lote de habitaciones pre-compradas al mayorista. Al registrar una venta de tipo Bloqueo, se descuenta automáticamente del inventario.")

        subtabs_bl = st.tabs(["➕ Nuevo Bloqueo", "📋 Ver Bloqueos", "✏️ Editar Bloqueo"])

        # ── Nuevo Bloqueo ─────────────────────────────────
        with subtabs_bl[0]:
            st.markdown("#### Registrar Nuevo Bloqueo")

            col1, col2 = st.columns(2)
            with col1:
                hoteles_cat = obtener_hoteles()
                if hoteles_cat:
                    bl_hotel = st.selectbox("Hotel *", hoteles_cat, key="bl_hotel")
                else:
                    bl_hotel = st.text_input("Hotel *", key="bl_hotel_txt")
                bl_operador    = st.text_input("Operador mayorista (opcional)", key="bl_operador")
                bl_responsable = st.text_input("Celular del responsable", key="bl_responsable")
            with col2:
                bl_fecha_ini = st.date_input("Fecha de inicio *", key="bl_fi")
                bl_fecha_fin = st.date_input("Fecha de fin *",    key="bl_ff")
                bl_hab       = st.number_input("Habitaciones compradas *", min_value=1, value=10, key="bl_hab")

            if bl_fecha_fin > bl_fecha_ini:
                bl_noches = (bl_fecha_fin - bl_fecha_ini).days
                st.success(f"✅ {bl_noches} noches")
            else:
                bl_noches = 0
                st.error("❌ Fecha de fin debe ser posterior a la de inicio")

            st.divider()
            st.markdown("##### 💵 Precios por noche — Adultos")
            col1, col2, col3 = st.columns(3)
            with col1:
                bl_p_doble   = st.number_input("Doble/noche *",    min_value=0.0, step=50.0, format="%.2f", key="bl_pd")
            with col2:
                bl_p_triple  = st.number_input("Triple/noche *",   min_value=0.0, step=50.0, format="%.2f", key="bl_pt")
            with col3:
                bl_p_cuad    = st.number_input("Cuádruple/noche *",min_value=0.0, step=50.0, format="%.2f", key="bl_pc")

            st.markdown("##### 👶 Precios por noche — Menores")
            col1, col2, col3 = st.columns(3)
            with col1:
                bl_mn_doble  = st.number_input("Menor Doble/noche",   min_value=0.0, step=50.0, format="%.2f", key="bl_md")
            with col2:
                bl_mn_triple = st.number_input("Menor Triple/noche",  min_value=0.0, step=50.0, format="%.2f", key="bl_mt")
            with col3:
                bl_mn_cuad   = st.number_input("Menor Cuádruple/noche",min_value=0.0, step=50.0, format="%.2f", key="bl_mc")

            bl_costo_real = st.number_input("Costo real por habitación (referencia interna)", min_value=0.0, step=100.0, format="%.2f", key="bl_costo")

            # Preview rápido
            if bl_noches > 0 and bl_p_doble > 0:
                st.markdown("**📊 Vista previa (habitación doble × estancia completa):**")
                col1, col2, col3 = st.columns(3)
                col1.metric("Total doble", f"${bl_p_doble * bl_noches:,.2f}")
                col2.metric("Total triple", f"${bl_p_triple * bl_noches:,.2f}")
                col3.metric("Total cuádruple", f"${bl_p_cuad * bl_noches:,.2f}")

            if st.button("💾 Registrar Bloqueo", type="primary", use_container_width=True, key="bl_guardar"):
                errores_bl = []
                if not bl_hotel: errores_bl.append("El hotel es obligatorio.")
                if bl_noches <= 0: errores_bl.append("Las fechas no son válidas.")
                if bl_p_doble <= 0: errores_bl.append("El precio doble es obligatorio.")

                if errores_bl:
                    for e in errores_bl: st.error(f"❌ {e}")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute("""
                            INSERT INTO bloqueos (
                                hotel, operador, celular_responsable, fecha_inicio, fecha_fin, noches,
                                habitaciones_totales, habitaciones_vendidas, habitaciones_disponibles,
                                precio_noche_doble, precio_noche_triple, precio_noche_cuadruple,
                                precio_menor_doble, precio_menor_triple, precio_menor_cuadruple,
                                costo_real, estado, fecha_registro
                            ) VALUES (?,?,?,?,?,?,?,0,?,?,?,?,?,?,?,?,'ACTIVO',?)
                        """, (
                            bl_hotel, bl_operador or None, bl_responsable,
                            bl_fecha_ini.strftime("%d-%m-%Y"), bl_fecha_fin.strftime("%d-%m-%Y"), bl_noches,
                            bl_hab, bl_hab,
                            bl_p_doble, bl_p_triple, bl_p_cuad,
                            bl_mn_doble, bl_mn_triple, bl_mn_cuad,
                            bl_costo_real, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ))
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Bloqueo en **{bl_hotel}** registrado con {bl_hab} habitaciones.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

        # ── Ver Bloqueos ──────────────────────────────────
        with subtabs_bl[1]:
            st.markdown("#### 📋 Todos los Bloqueos")
            conn = conectar_db()
            try:
                df_bl = pd.read_sql_query("""
                    SELECT id, hotel, fecha_inicio, fecha_fin, noches,
                           habitaciones_totales, habitaciones_vendidas, habitaciones_disponibles,
                           precio_noche_doble, precio_noche_triple, precio_noche_cuadruple,
                           costo_real, estado
                    FROM bloqueos ORDER BY fecha_inicio DESC
                """, conn)
            except:
                df_bl = pd.DataFrame()
            conn.close()

            if df_bl.empty:
                st.info("No hay bloqueos registrados aún.")
            else:
                # Métricas resumen
                col1, col2, col3, col4 = st.columns(4)
                activos_bl = df_bl[df_bl['estado'] == 'ACTIVO']
                col1.metric("Total bloqueos", len(df_bl))
                col2.metric("Activos", len(activos_bl))
                col3.metric("Hab. disponibles", int(df_bl['habitaciones_disponibles'].sum()))
                col4.metric("Hab. vendidas",    int(df_bl['habitaciones_vendidas'].sum()))

                st.divider()
                filtro_est_bl = st.selectbox("Filtrar por estado:",
                    ["Todos"] + sorted(df_bl['estado'].unique().tolist()), key="bl_filtro_est")
                df_blf = df_bl if filtro_est_bl == "Todos" else df_bl[df_bl['estado'] == filtro_est_bl]

                df_bl_show = df_blf.copy()
                df_bl_show['precio_noche_doble']   = df_bl_show['precio_noche_doble'].apply(lambda x: f"${x:,.2f}")
                df_bl_show['precio_noche_triple']  = df_bl_show['precio_noche_triple'].apply(lambda x: f"${x:,.2f}")
                df_bl_show['precio_noche_cuadruple']= df_bl_show['precio_noche_cuadruple'].apply(lambda x: f"${x:,.2f}")
                df_bl_show['costo_real']           = df_bl_show['costo_real'].apply(lambda x: f"${x:,.2f}")
                df_bl_show.columns = ['ID','Hotel','Inicio','Fin','Noches',
                                       'Total','Vendidas','Disponibles',
                                       'P. Doble/n','P. Triple/n','P. Cuád/n','Costo/hab','Estado']
                st.dataframe(df_bl_show, use_container_width=True, hide_index=True)

                # Cambiar estado
                st.divider()
                st.markdown("##### 🔄 Cambiar estado de un bloqueo")
                bl_ids = df_blf['id'].tolist()
                if bl_ids:
                    id_bl_cambio = st.selectbox("ID del bloqueo:", bl_ids, key="bl_cambio_id")
                    row_bl = df_blf[df_blf['id'] == id_bl_cambio].iloc[0]
                    nuevo_est_bl = "CERRADO" if row_bl['estado'] == "ACTIVO" else "ACTIVO"
                    if st.button(f"{'🔒 Cerrar' if nuevo_est_bl == 'CERRADO' else '🔓 Reactivar'} bloqueo ID {id_bl_cambio}",
                                 key="bl_cambio_btn"):
                        try:
                            conn = conectar_db()
                            cursor = conn.cursor()
                            cursor.execute("UPDATE bloqueos SET estado = ? WHERE id = ?", (nuevo_est_bl, int(id_bl_cambio)))
                            conn.commit()
                            conn.close()
                            st.success(f"✅ Bloqueo ID {id_bl_cambio} → {nuevo_est_bl}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {e}")

        # ── Editar Bloqueo ────────────────────────────────
        with subtabs_bl[2]:
            st.markdown("#### ✏️ Editar Bloqueo")
            conn = conectar_db()
            try:
                df_bl_ed = pd.read_sql_query(
                    "SELECT id, hotel, habitaciones_totales, habitaciones_disponibles, habitaciones_vendidas, "
                    "precio_noche_doble, precio_noche_triple, precio_noche_cuadruple, costo_real, estado "
                    "FROM bloqueos ORDER BY id DESC", conn)
            except:
                df_bl_ed = pd.DataFrame()
            conn.close()

            if df_bl_ed.empty:
                st.info("No hay bloqueos para editar.")
            else:
                bl_ed_opts = {
                    f"ID {row['id']} | {row['hotel']} | {row['estado']}": row
                    for _, row in df_bl_ed.iterrows()
                }
                bl_ed_label = st.selectbox("Selecciona bloqueo:", list(bl_ed_opts.keys()), key="bl_ed_sel")
                bl_ed = bl_ed_opts[bl_ed_label]

                col1, col2 = st.columns(2)
                with col1:
                    hoteles_cat2 = obtener_hoteles()
                    if hoteles_cat2 and bl_ed['hotel'] in hoteles_cat2:
                        bl_ed_hotel = st.selectbox("Hotel", hoteles_cat2,
                            index=hoteles_cat2.index(bl_ed['hotel']), key="bl_ed_hotel")
                    else:
                        bl_ed_hotel = st.text_input("Hotel", value=bl_ed['hotel'], key="bl_ed_hotel_txt")
                    bl_ed_hab    = st.number_input("Habitaciones totales",
                        min_value=int(bl_ed['habitaciones_vendidas']), value=int(bl_ed['habitaciones_totales']),
                        key="bl_ed_hab")
                with col2:
                    bl_ed_pd  = st.number_input("Precio doble/noche",    min_value=0.0, value=float(bl_ed['precio_noche_doble']),   step=50.0, format="%.2f", key="bl_ed_pd")
                    bl_ed_pt  = st.number_input("Precio triple/noche",   min_value=0.0, value=float(bl_ed['precio_noche_triple']),  step=50.0, format="%.2f", key="bl_ed_pt")
                    bl_ed_pc  = st.number_input("Precio cuádruple/noche",min_value=0.0, value=float(bl_ed['precio_noche_cuadruple']),step=50.0, format="%.2f", key="bl_ed_pc")
                    bl_ed_costo = st.number_input("Costo real/hab", min_value=0.0, value=float(bl_ed['costo_real']), step=100.0, format="%.2f", key="bl_ed_costo")

                nuevas_disp_bl = bl_ed_hab - int(bl_ed['habitaciones_vendidas'])
                st.caption(f"Nuevas disponibles calculadas: {nuevas_disp_bl}")

                if st.button("💾 Guardar Cambios en Bloqueo", type="primary", key="bl_ed_guardar"):
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE bloqueos
                            SET hotel = ?, habitaciones_totales = ?, habitaciones_disponibles = ?,
                                precio_noche_doble = ?, precio_noche_triple = ?, precio_noche_cuadruple = ?,
                                costo_real = ?
                            WHERE id = ?
                        """, (bl_ed_hotel, bl_ed_hab, max(0, nuevas_disp_bl),
                              bl_ed_pd, bl_ed_pt, bl_ed_pc, bl_ed_costo, int(bl_ed['id'])))
                        conn.commit()
                        conn.close()
                        st.success("✅ Bloqueo actualizado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

    # ══════════════════════════════════════════════════════
    # TAB 5 — GRUPOS
    # ══════════════════════════════════════════════════════
    with tabs[5]:
        st.subheader("👥 Gestión de Grupos")
        st.info("💡 Un **grupo** es una compra en bloque para un conjunto de personas (ej. viaje de empresa, excursión familiar). Al registrar una venta de tipo Grupo, se descuenta automáticamente del inventario de habitaciones.")

        subtabs_gr = st.tabs(["➕ Nuevo Grupo", "📋 Ver Grupos", "✏️ Editar / Cerrar"])

        # ── Nuevo Grupo ───────────────────────────────────
        with subtabs_gr[0]:
            st.markdown("#### Registrar Nuevo Grupo")

            col1, col2 = st.columns(2)
            with col1:
                gr_nombre    = st.text_input("Nombre del grupo *", placeholder="Ej: Familia García Verano 2025", key="gr_nombre")
                hoteles_cat3 = obtener_hoteles()
                if hoteles_cat3:
                    gr_hotel = st.selectbox("Hotel *", hoteles_cat3, key="gr_hotel")
                else:
                    gr_hotel = st.text_input("Hotel *", key="gr_hotel_txt")
                gr_operador  = st.text_input("Operador mayorista (opcional)", key="gr_operador")
            with col2:
                gr_fecha_ini  = st.date_input("Fecha de inicio *", key="gr_fi")
                gr_fecha_fin  = st.date_input("Fecha de fin *",    key="gr_ff")
                gr_hab        = st.number_input("Habitaciones totales *", min_value=1, value=5, key="gr_hab")

            if gr_fecha_fin > gr_fecha_ini:
                gr_noches = (gr_fecha_fin - gr_fecha_ini).days
                st.success(f"✅ {gr_noches} noches")
            else:
                gr_noches = 0
                st.error("❌ Fecha de fin debe ser posterior a la de inicio")

            st.divider()
            st.markdown("##### 💵 Precios por noche — Adultos")
            col1, col2, col3 = st.columns(3)
            with col1:
                gr_p_doble   = st.number_input("Doble/noche *",    min_value=0.0, step=50.0, format="%.2f", key="gr_pd")
            with col2:
                gr_p_triple  = st.number_input("Triple/noche",     min_value=0.0, step=50.0, format="%.2f", key="gr_pt")
            with col3:
                gr_p_cuad    = st.number_input("Cuádruple/noche",  min_value=0.0, step=50.0, format="%.2f", key="gr_pc")

            st.markdown("##### 👶 Precios por noche — Menores")
            col1, col2, col3 = st.columns(3)
            with col1:
                gr_mn_doble  = st.number_input("Menor Doble/noche",    min_value=0.0, step=50.0, format="%.2f", key="gr_md")
            with col2:
                gr_mn_triple = st.number_input("Menor Triple/noche",   min_value=0.0, step=50.0, format="%.2f", key="gr_mt")
            with col3:
                gr_mn_cuad   = st.number_input("Menor Cuádruple/noche",min_value=0.0, step=50.0, format="%.2f", key="gr_mc")

            gr_costo_real = st.number_input("Costo real total del grupo ($)", min_value=0.0, step=500.0, format="%.2f", key="gr_costo")

            st.divider()
            st.markdown("##### 👤 Responsable del grupo")
            col1, col2 = st.columns(2)
            with col1:
                gr_responsable = st.text_input("Nombre del responsable", value="Mamá", key="gr_resp")
            with col2:
                gr_celular_resp = st.text_input("Celular del responsable", key="gr_cel_resp")

            # Preview
            if gr_noches > 0 and gr_p_doble > 0:
                st.markdown("**📊 Vista previa (habitación doble × estancia completa):**")
                col1, col2, col3 = st.columns(3)
                col1.metric("Total doble", f"${gr_p_doble * gr_noches:,.2f}")
                col2.metric("Total triple", f"${gr_p_triple * gr_noches:,.2f}")
                col3.metric("Total cuádruple", f"${gr_p_cuad * gr_noches:,.2f}")

            if st.button("💾 Registrar Grupo", type="primary", use_container_width=True, key="gr_guardar"):
                errores_gr = []
                if not gr_nombre.strip(): errores_gr.append("El nombre del grupo es obligatorio.")
                if not gr_hotel:          errores_gr.append("El hotel es obligatorio.")
                if gr_noches <= 0:        errores_gr.append("Las fechas no son válidas.")
                if gr_p_doble <= 0:       errores_gr.append("El precio doble es obligatorio.")

                if errores_gr:
                    for e in errores_gr: st.error(f"❌ {e}")
                else:
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute("""
                            INSERT INTO grupos (
                                nombre_grupo, operador, hotel, fecha_inicio, fecha_fin, noches,
                                habitaciones_totales, habitaciones_vendidas, habitaciones_disponibles,
                                precio_noche_doble, precio_noche_triple, precio_noche_cuadruple,
                                precio_menor_doble, precio_menor_triple, precio_menor_cuadruple,
                                costo_real, responsable, celular_responsable, estado, fecha_registro
                            ) VALUES (?,?,?,?,?,?,?,0,?,?,?,?,?,?,?,?,?,?,'ACTIVO',?)
                        """, (
                            gr_nombre.strip(), gr_operador or None, gr_hotel,
                            gr_fecha_ini.strftime("%d-%m-%Y"), gr_fecha_fin.strftime("%d-%m-%Y"), gr_noches,
                            gr_hab, gr_hab,
                            gr_p_doble, gr_p_triple, gr_p_cuad,
                            gr_mn_doble, gr_mn_triple, gr_mn_cuad,
                            gr_costo_real, gr_responsable, gr_celular_resp,
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ))
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Grupo **{gr_nombre}** registrado con {gr_hab} habitaciones.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
                        st.caption("💡 Asegúrate de que exista la tabla `grupos` en la base de datos.")

        # ── Ver Grupos ────────────────────────────────────
        with subtabs_gr[1]:
            st.markdown("#### 📋 Todos los Grupos")
            conn = conectar_db()
            try:
                df_gr = pd.read_sql_query("""
                    SELECT id, nombre_grupo, operador, hotel,
                           fecha_inicio, fecha_fin, noches,
                           habitaciones_totales, habitaciones_vendidas, habitaciones_disponibles,
                           precio_noche_doble, costo_real, responsable, estado
                    FROM grupos ORDER BY fecha_inicio DESC
                """, conn)
            except:
                df_gr = pd.DataFrame()
            conn.close()

            if df_gr.empty:
                st.info("No hay grupos registrados aún.")
            else:
                col1, col2, col3, col4 = st.columns(4)
                activos_gr = df_gr[df_gr['estado'] == 'ACTIVO']
                col1.metric("Total grupos",        len(df_gr))
                col2.metric("Activos",             len(activos_gr))
                col3.metric("Hab. disponibles",    int(df_gr['habitaciones_disponibles'].sum()))
                col4.metric("Hab. vendidas",        int(df_gr['habitaciones_vendidas'].sum()))

                st.divider()
                filtro_est_gr = st.selectbox("Filtrar por estado:",
                    ["Todos"] + sorted(df_gr['estado'].unique().tolist()), key="gr_filtro_est")
                df_grf = df_gr if filtro_est_gr == "Todos" else df_gr[df_gr['estado'] == filtro_est_gr]

                for _, row_gr in df_grf.iterrows():
                    ocupacion_gr = (row_gr['habitaciones_vendidas'] / row_gr['habitaciones_totales'] * 100) \
                                   if row_gr['habitaciones_totales'] > 0 else 0
                    estado_icon_gr = "✅" if row_gr['estado'] == "ACTIVO" else ("🔒" if row_gr['estado'] == "CERRADO" else "📭")
                    with st.expander(
                        f"{estado_icon_gr} ID {int(row_gr['id'])} | {row_gr['nombre_grupo']} — {row_gr['hotel']} | {row_gr['habitaciones_disponibles']} disp.",
                        expanded=False):
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("Hotel",       row_gr['hotel'])
                        c2.metric("Noches",      row_gr['noches'])
                        c3.metric("Hab. Vendidas",f"{int(row_gr['habitaciones_vendidas'])}/{int(row_gr['habitaciones_totales'])}")
                        c4.metric("Ocupación",   f"{ocupacion_gr:.0f}%")

                        c1b, c2b, c3b = st.columns(3)
                        c1b.write(f"**Fechas:** {row_gr['fecha_inicio']} → {row_gr['fecha_fin']}")
                        c2b.write(f"**Operador:** {row_gr['operador'] or '—'}")
                        c3b.write(f"**Responsable:** {row_gr['responsable']}")

                        c1c, c2c, c3c = st.columns(3)
                        c1c.metric("Precio doble/noche",   f"${row_gr['precio_noche_doble']:,.2f}")
                        c2c.metric("Costo real total",     f"${row_gr['costo_real']:,.2f}")
                        c3c.metric("Estado",               row_gr['estado'])

                        if row_gr['habitaciones_totales'] > 0:
                            st.progress(min(ocupacion_gr / 100, 1.0))

                csv_gr = df_grf.to_csv(index=False).encode('utf-8')
                st.download_button("📥 Descargar CSV de grupos", csv_gr,
                    file_name=f"grupos_{datetime.now().strftime('%Y%m%d')}.csv", mime="text/csv", key="gr_csv")

        # ── Editar / Cerrar Grupo ─────────────────────────
        with subtabs_gr[2]:
            st.markdown("#### ✏️ Editar o Cerrar un Grupo")
            conn = conectar_db()
            try:
                df_gr_ed = pd.read_sql_query("""
                    SELECT id, nombre_grupo, hotel, habitaciones_totales, habitaciones_vendidas,
                           habitaciones_disponibles, precio_noche_doble, precio_noche_triple,
                           precio_noche_cuadruple, costo_real, estado
                    FROM grupos ORDER BY id DESC
                """, conn)
            except:
                df_gr_ed = pd.DataFrame()
            conn.close()

            if df_gr_ed.empty:
                st.info("No hay grupos para editar.")
            else:
                gr_ed_opts = {
                    f"ID {row['id']} | {row['nombre_grupo']} | {row['hotel']} | {row['estado']}": row
                    for _, row in df_gr_ed.iterrows()
                }
                gr_ed_label = st.selectbox("Selecciona grupo:", list(gr_ed_opts.keys()), key="gr_ed_sel")
                gr_ed = gr_ed_opts[gr_ed_label]

                col1, col2 = st.columns(2)
                with col1:
                    gr_ed_nombre = st.text_input("Nombre del grupo", value=gr_ed['nombre_grupo'], key="gr_ed_nom")
                    hoteles_cat4 = obtener_hoteles()
                    if hoteles_cat4 and gr_ed['hotel'] in hoteles_cat4:
                        gr_ed_hotel = st.selectbox("Hotel", hoteles_cat4,
                            index=hoteles_cat4.index(gr_ed['hotel']), key="gr_ed_hotel")
                    else:
                        gr_ed_hotel = st.text_input("Hotel", value=gr_ed['hotel'], key="gr_ed_hotel_txt")
                    gr_ed_hab  = st.number_input("Habitaciones totales",
                        min_value=int(gr_ed['habitaciones_vendidas']), value=int(gr_ed['habitaciones_totales']),
                        key="gr_ed_hab")
                    gr_ed_estado = st.selectbox("Estado", ["ACTIVO","CERRADO","AGOTADO"],
                        index=["ACTIVO","CERRADO","AGOTADO"].index(gr_ed['estado'])
                              if gr_ed['estado'] in ["ACTIVO","CERRADO","AGOTADO"] else 0,
                        key="gr_ed_estado")
                with col2:
                    gr_ed_pd    = st.number_input("Precio doble/noche",    min_value=0.0, value=float(gr_ed['precio_noche_doble']),   step=50.0, format="%.2f", key="gr_ed_pd")
                    gr_ed_pt    = st.number_input("Precio triple/noche",   min_value=0.0, value=float(gr_ed['precio_noche_triple']),  step=50.0, format="%.2f", key="gr_ed_pt")
                    gr_ed_pc    = st.number_input("Precio cuádruple/noche",min_value=0.0, value=float(gr_ed['precio_noche_cuadruple']),step=50.0, format="%.2f", key="gr_ed_pc")
                    gr_ed_costo = st.number_input("Costo real total",      min_value=0.0, value=float(gr_ed['costo_real']),           step=500.0, format="%.2f", key="gr_ed_costo")

                nuevas_disp_gr = max(0, gr_ed_hab - int(gr_ed['habitaciones_vendidas']))
                st.caption(f"Habitaciones disponibles resultantes: {nuevas_disp_gr}")

                if st.button("💾 Guardar Cambios en Grupo", type="primary", key="gr_ed_guardar"):
                    try:
                        conn = conectar_db()
                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE grupos
                            SET nombre_grupo = ?, hotel = ?,
                                habitaciones_totales = ?, habitaciones_disponibles = ?,
                                precio_noche_doble = ?, precio_noche_triple = ?, precio_noche_cuadruple = ?,
                                costo_real = ?, estado = ?
                            WHERE id = ?
                        """, (gr_ed_nombre.strip(), gr_ed_hotel, gr_ed_hab, nuevas_disp_gr,
                              gr_ed_pd, gr_ed_pt, gr_ed_pc, gr_ed_costo, gr_ed_estado,
                              int(gr_ed['id'])))
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Grupo **{gr_ed_nombre}** actualizado → {gr_ed_estado}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")


def pagina_mi_historial():
    """Historial personal de ventas y comisiones — solo VENDEDORA"""
    usuario = st.session_state.usuario_actual
    id_vend = usuario["id_vendedora"]

    st.title(f"📂 Mi Historial — {usuario['nombre']}")

    tabs = st.tabs([
        "🏖️ Ventas Riviera",
        "🎫 Viajes Nacionales",
        "🌎 Viajes Internacionales",
        "💰 Mis Comisiones"
    ])

    # ── TAB 0: Riviera ────────────────────────────────────────────────────
    with tabs[0]:
        st.subheader("🏖️ Mis Ventas — Riviera Maya")
        conn = conectar_db()
        try:
            df_rv = pd.read_sql_query(f"""
                SELECT v.id, v.cliente, v.destino, v.tipo_habitacion,
                       v.fecha_inicio, v.fecha_fin, v.noches,
                       v.adultos, v.menores,
                       v.precio_total, v.pagado, v.saldo,
                       v.ganancia,
                       COALESCE(v.comision_vendedora, ROUND(v.ganancia*0.10,2)) AS comision,
                       v.estado,
                       CASE v.comision_pagada WHEN 1 THEN '✅ Pagada' ELSE '⏳ Pendiente' END AS estado_comision,
                       v.fecha_registro
                FROM ventas v
                LEFT JOIN vendedoras vd ON v.vendedora_id = vd.id
                WHERE v.vendedora_id = {id_vend}
                ORDER BY v.id DESC
            """, conn)
        except Exception as e:
            df_rv = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_rv.empty:
            st.info("No tienes ventas de Riviera Maya registradas.")
        else:
            # Métricas resumen
            col1, col2, col3, col4, col5 = st.columns(5)
            col1.metric("Total ventas",    len(df_rv))
            col2.metric("Total vendido",   f"${df_rv['precio_total'].sum():,.2f}")
            col3.metric("Total cobrado",   f"${df_rv['pagado'].sum():,.2f}")
            col4.metric("Saldo pendiente", f"${df_rv['saldo'].sum():,.2f}")
            col5.metric("Mis ganancias",   f"${df_rv['ganancia'].sum():,.2f}")

            st.divider()

            # Filtro por estado
            col1, col2 = st.columns(2)
            with col1:
                filtro_rv = st.selectbox("Filtrar por estado:",
                    ["Todos"] + sorted(df_rv["estado"].unique().tolist()), key="mh_rv_estado")
            with col2:
                buscar_rv = st.text_input("🔍 Buscar cliente o destino", key="mh_rv_buscar")

            df_rvf = df_rv.copy()
            if filtro_rv != "Todos":
                df_rvf = df_rvf[df_rvf["estado"] == filtro_rv]
            if buscar_rv:
                mask = (df_rvf["cliente"].str.contains(buscar_rv, case=False, na=False) |
                        df_rvf["destino"].str.contains(buscar_rv, case=False, na=False))
                df_rvf = df_rvf[mask]

            df_rv_show = df_rvf.copy()
            for col in ["precio_total","pagado","saldo","ganancia","comision"]:
                df_rv_show[col] = df_rv_show[col].apply(lambda x: f"${x:,.2f}")
            df_rv_show.columns = ["ID","Cliente","Destino","Habitación","Salida","Regreso","Noches",
                                   "Adultos","Menores","Total","Pagado","Saldo","Ganancia",
                                   "Mi Comisión","Estado","Estado Comisión","Registro"]
            st.dataframe(df_rv_show, use_container_width=True, hide_index=True)

            # Detalle expandible
            st.divider()
            ids_rv = df_rvf["id"].tolist()
            if ids_rv:
                id_rv_sel = st.selectbox("Ver detalle:", ids_rv, key="mh_rv_det")
                row_rv = df_rvf[df_rvf["id"] == id_rv_sel].iloc[0]
                with st.expander(f"📄 {row_rv['cliente']} | {row_rv['destino']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Cliente:** {row_rv['cliente']}")
                    c1.write(f"**Salida:** {row_rv['fecha_inicio']} → {row_rv['fecha_fin']}")
                    c1.write(f"**Noches:** {row_rv['noches']}")
                    c2.metric("Total",   f"${row_rv['precio_total']:,.2f}")
                    c2.metric("Pagado",  f"${row_rv['pagado']:,.2f}")
                    c3.metric("Saldo",   f"${row_rv['saldo']:,.2f}")
                    c3.metric("Mi comisión", f"${row_rv['comision']:,.2f} ({row_rv['estado_comision']})")

                    conn2 = conectar_db()
                    try:
                        df_pas_rv = pd.read_sql_query(
                            "SELECT nombre, tipo FROM pasajeros WHERE venta_id = ?",
                            conn2, params=(id_rv_sel,))
                        df_ab_rv = pd.read_sql_query(
                            "SELECT fecha, monto FROM abonos WHERE venta_id = ? ORDER BY fecha",
                            conn2, params=(id_rv_sel,))
                    except:
                        df_pas_rv = pd.DataFrame()
                        df_ab_rv = pd.DataFrame()
                    conn2.close()

                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**👥 Pasajeros:**")
                        if not df_pas_rv.empty:
                            st.dataframe(df_pas_rv, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin pasajeros")
                    with c2:
                        st.markdown("**💳 Abonos:**")
                        if not df_ab_rv.empty:
                            df_ab_rv["monto"] = df_ab_rv["monto"].apply(lambda x: f"${x:,.2f}")
                            st.dataframe(df_ab_rv, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin abonos")

            csv_rv = df_rv_show.to_csv(index=False).encode("utf-8")
            st.download_button("📥 Descargar CSV", csv_rv,
                file_name=f"mis_ventas_riviera_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # ── TAB 1: Nacionales ─────────────────────────────────────────────────
    with tabs[1]:
        st.subheader("🎫 Mis Clientes — Viajes Nacionales")
        conn = conectar_db()
        try:
            df_nv = pd.read_sql_query(f"""
                SELECT cn.id, cn.nombre_cliente, vj.nombre_viaje, vj.destino,
                       vj.fecha_salida, vj.fecha_regreso,
                       cn.adultos, cn.menores,
                       cn.habitaciones_doble, cn.habitaciones_triple,
                       cn.total_pagar, cn.total_abonado, cn.saldo,
                       COALESCE(cn.ganancia, cn.costo_neto, 0) AS ganancia,
                       cn.estado,
                       CASE cn.comision_pagada WHEN 1 THEN '✅ Pagada' ELSE '⏳ Pendiente' END AS estado_comision,
                       cn.fecha_registro
                FROM clientes_nacionales cn
                JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                WHERE cn.vendedora_id = {id_vend}
                ORDER BY cn.id DESC
            """, conn)
        except Exception as e:
            df_nv = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_nv.empty:
            st.info("No tienes clientes de viajes nacionales registrados.")
        else:
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Clientes",      len(df_nv))
            col2.metric("Total cobrar",  f"${df_nv['total_pagar'].sum():,.2f}")
            col3.metric("Total cobrado", f"${df_nv['total_abonado'].sum():,.2f}")
            col4.metric("Saldo pend.",   f"${df_nv['saldo'].sum():,.2f}")

            st.divider()

            col1, col2 = st.columns(2)
            with col1:
                filtro_nv = st.selectbox("Filtrar por estado:",
                    ["Todos"] + sorted(df_nv["estado"].unique().tolist()), key="mh_nv_estado")
            with col2:
                buscar_nv = st.text_input("🔍 Buscar cliente o viaje", key="mh_nv_buscar")

            df_nvf = df_nv.copy()
            if filtro_nv != "Todos":
                df_nvf = df_nvf[df_nvf["estado"] == filtro_nv]
            if buscar_nv:
                mask = (df_nvf["nombre_cliente"].str.contains(buscar_nv, case=False, na=False) |
                        df_nvf["nombre_viaje"].str.contains(buscar_nv, case=False, na=False))
                df_nvf = df_nvf[mask]

            df_nv_show = df_nvf.copy()
            for col in ["total_pagar","total_abonado","saldo","ganancia"]:
                df_nv_show[col] = df_nv_show[col].apply(lambda x: f"${x:,.2f}")
            df_nv_show.columns = ["ID","Cliente","Viaje","Destino","Salida","Regreso",
                                   "Adultos","Menores","Hab.Doble","Hab.Triple",
                                   "Total","Abonado","Saldo","Ganancia",
                                   "Estado","Estado Comisión","Registro"]
            st.dataframe(df_nv_show, use_container_width=True, hide_index=True)

            # Detalle
            st.divider()
            ids_nv = df_nvf["id"].tolist()
            if ids_nv:
                id_nv_sel = st.selectbox("Ver detalle:", ids_nv, key="mh_nv_det")
                row_nv = df_nvf[df_nvf["id"] == id_nv_sel].iloc[0]
                with st.expander(f"📄 {row_nv['nombre_cliente']} | {row_nv['nombre_viaje']}", expanded=True):
                    c1, c2, c3 = st.columns(3)
                    c1.write(f"**Cliente:** {row_nv['nombre_cliente']}")
                    c1.write(f"**Viaje:** {row_nv['nombre_viaje']}")
                    c1.write(f"**Salida:** {row_nv['fecha_salida']} → {row_nv['fecha_regreso']}")
                    c2.metric("Total",   f"${row_nv['total_pagar']:,.2f}")
                    c2.metric("Abonado", f"${row_nv['total_abonado']:,.2f}")
                    c3.metric("Saldo",   f"${row_nv['saldo']:,.2f}")
                    c3.write(f"**Comisión:** {row_nv['estado_comision']}")

                    conn2 = conectar_db()
                    try:
                        df_pas_nv = pd.read_sql_query(
                            "SELECT nombre_completo, tipo, habitacion_asignada FROM pasajeros_nacionales WHERE cliente_id = ?",
                            conn2, params=(id_nv_sel,))
                        df_ab_nv = pd.read_sql_query(
                            "SELECT fecha, monto FROM abonos_nacionales WHERE cliente_id = ? ORDER BY fecha",
                            conn2, params=(id_nv_sel,))
                    except:
                        df_pas_nv = pd.DataFrame()
                        df_ab_nv = pd.DataFrame()
                    conn2.close()

                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**🧳 Pasajeros:**")
                        if not df_pas_nv.empty:
                            df_pas_nv.columns = ["Nombre","Tipo","Habitación"]
                            st.dataframe(df_pas_nv, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin pasajeros")
                    with c2:
                        st.markdown("**💳 Abonos:**")
                        if not df_ab_nv.empty:
                            df_ab_nv["monto"] = df_ab_nv["monto"].apply(lambda x: f"${x:,.2f}")
                            st.dataframe(df_ab_nv, hide_index=True, use_container_width=True)
                        else:
                            st.caption("Sin abonos")

            csv_nv = df_nv_show.to_csv(index=False).encode("utf-8")
            st.download_button("📥 Descargar CSV", csv_nv,
                file_name=f"mis_clientes_nacionales_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")

    # ── TAB 2: Internacionales ────────────────────────────────────────────
    with tabs[2]:
        st.subheader("🌎 Mis Clientes — Viajes Internacionales")
        st.info("ℹ️ Los viajes internacionales no tienen vendedora asignada directamente. Muestra todos los clientes para tu referencia.")
        conn = conectar_db()
        try:
            df_iv = pd.read_sql_query("""
                SELECT ci.id, ci.nombre_cliente, vi.destino,
                       vi.fecha_salida, vi.fecha_regreso,
                       ci.adultos, ci.menores,
                       ci.total_usd, ci.abonado_usd, ci.saldo_usd,
                       ci.ganancia_usd,
                       ci.estado,
                       CASE ci.comision_pagada WHEN 1 THEN '✅ Pagada' ELSE '⏳ Pendiente' END AS estado_comision
                FROM clientes_internacionales ci
                JOIN viajes_internacionales vi ON ci.viaje_id = vi.id
                WHERE ci.estado != 'CERRADO'
                ORDER BY ci.id DESC
            """, conn)
        except Exception as e:
            df_iv = pd.DataFrame()
            st.caption(f"⚠️ {e}")
        conn.close()

        if df_iv.empty:
            st.info("No hay clientes internacionales activos.")
        else:
            col1, col2, col3 = st.columns(3)
            col1.metric("Clientes",    len(df_iv))
            col2.metric("Total USD",   f"${df_iv['total_usd'].sum():,.2f}")
            col3.metric("Saldo USD",   f"${df_iv['saldo_usd'].sum():,.2f}")

            df_iv_show = df_iv.copy()
            for col in ["total_usd","abonado_usd","saldo_usd","ganancia_usd"]:
                df_iv_show[col] = df_iv_show[col].apply(lambda x: f"${x:,.2f} USD")
            df_iv_show.columns = ["ID","Cliente","Destino","Salida","Regreso",
                                   "Adultos","Menores","Total","Abonado","Saldo",
                                   "Ganancia","Estado","Estado Comisión"]
            st.dataframe(df_iv_show, use_container_width=True, hide_index=True)

    # ── TAB 3: Mis Comisiones ─────────────────────────────────────────────
    with tabs[3]:
        st.subheader("💰 Mis Comisiones")

        conn = conectar_db()

        # Comisiones pagadas (historial)
        try:
            df_com_pag = pd.read_sql_query(f"""
                SELECT h.fecha, h.tipo, h.referencia_id, h.monto,
                       h.metodo_pago, h.nota
                FROM historial_comisiones h
                WHERE h.vendedora_id = {id_vend}
                   OR h.vendedora = '{usuario["nombre"]}'
                ORDER BY h.fecha DESC
            """, conn)
        except:
            df_com_pag = pd.DataFrame()

        # Comisiones pendientes Riviera
        try:
            df_com_pen_rv = pd.read_sql_query(f"""
                SELECT v.id, v.cliente, v.destino, v.ganancia,
                       COALESCE(v.comision_vendedora, ROUND(v.ganancia*0.10,2)) AS comision
                FROM ventas v
                WHERE v.usuario_id = {id_vend}
                  AND v.estado = 'LIQUIDADO'
                  AND COALESCE(v.comision_pagada, 0) = 0
            """, conn)
        except:
            df_com_pen_rv = pd.DataFrame()

        # Comisiones pendientes Nacionales
        try:
            df_com_pen_nac = pd.read_sql_query(f"""
                SELECT cn.id, cn.nombre_cliente AS cliente,
                       vj.nombre_viaje AS destino,
                       COALESCE(cn.ganancia, 0) AS ganancia,
                       ROUND(COALESCE(cn.ganancia,0) * COALESCE(cfg.porcentaje_comision,10) / 100.0, 2) AS comision
                FROM clientes_nacionales cn
                JOIN viajes_nacionales vj ON cn.viaje_id = vj.id
                LEFT JOIN config_comisiones cfg ON cfg.tipo = 'nacionales'
                WHERE cn.vendedora_id = {id_vend}
                  AND cn.estado = 'LIQUIDADO'
                  AND COALESCE(cn.comision_pagada, 0) = 0
            """, conn)
        except:
            df_com_pen_nac = pd.DataFrame()

        conn.close()

        # ── Resumen de comisiones ─────────────────────────────────────────
        total_pagado_com  = df_com_pag["monto"].sum() if not df_com_pag.empty else 0
        total_pen_rv      = df_com_pen_rv["comision"].sum() if not df_com_pen_rv.empty else 0
        total_pen_nac     = df_com_pen_nac["comision"].sum() if not df_com_pen_nac.empty else 0
        total_pendiente   = total_pen_rv + total_pen_nac

        col1, col2, col3 = st.columns(3)
        col1.metric("💵 Comisiones cobradas",  f"${total_pagado_com:,.2f}",
                    help="Total histórico de comisiones ya pagadas")
        col2.metric("⏳ Comisiones pendientes", f"${total_pendiente:,.2f}",
                    help="Ventas liquidadas cuya comisión aún no se ha pagado")
        col3.metric("📊 Total acumulado",       f"${total_pagado_com + total_pendiente:,.2f}")

        st.divider()

        # ── Pendientes detalle ────────────────────────────────────────────
        st.markdown("#### ⏳ Comisiones Pendientes de Cobro")
        if df_com_pen_rv.empty and df_com_pen_nac.empty:
            st.success("✅ No tienes comisiones pendientes. ¡Todo cobrado!")
        else:
            if not df_com_pen_rv.empty:
                st.markdown("**🏖️ Riviera Maya** (10% de ganancia)")
                df_pen_rv_show = df_com_pen_rv.copy()
                df_pen_rv_show["ganancia"] = df_pen_rv_show["ganancia"].apply(lambda x: f"${x:,.2f}")
                df_pen_rv_show["comision"] = df_pen_rv_show["comision"].apply(lambda x: f"${x:,.2f}")
                df_pen_rv_show.columns = ["ID","Cliente","Destino","Ganancia","Mi Comisión"]
                st.dataframe(df_pen_rv_show, use_container_width=True, hide_index=True)

            if not df_com_pen_nac.empty:
                st.markdown("**🎫 Nacionales**")
                df_pen_nac_show = df_com_pen_nac.copy()
                df_pen_nac_show["ganancia"] = df_pen_nac_show["ganancia"].apply(lambda x: f"${x:,.2f}")
                df_pen_nac_show["comision"] = df_pen_nac_show["comision"].apply(lambda x: f"${x:,.2f}")
                df_pen_nac_show.columns = ["ID","Cliente","Destino","Ganancia","Mi Comisión"]
                st.dataframe(df_pen_nac_show, use_container_width=True, hide_index=True)

        st.divider()

        # ── Historial pagadas ─────────────────────────────────────────────
        st.markdown("#### ✅ Historial de Comisiones Cobradas")
        if df_com_pag.empty:
            st.info("Aún no tienes comisiones cobradas en el historial.")
        else:
            df_com_pag_show = df_com_pag.copy()
            df_com_pag_show["monto"] = df_com_pag_show["monto"].apply(lambda x: f"${x:,.2f}")
            df_com_pag_show.columns = ["Fecha","Tipo","ID Ref.","Monto","Método","Nota"]
            st.dataframe(df_com_pag_show, use_container_width=True, hide_index=True)

            csv_com = df_com_pag_show.to_csv(index=False).encode("utf-8")
            st.download_button("📥 Descargar historial",
                csv_com, file_name=f"mis_comisiones_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv")




def pagina_otros():
    """Página de Otros servicios: Pasaportes, Visas, Vuelos"""
    st.title("🗂️ Otros Servicios")

    usuario  = st.session_state.usuario_actual
    es_admin = usuario.get("rol") == "ADMIN"
    id_vend  = usuario.get("id_vendedora", 0) or 0

    tab_pas, tab_vis, tab_vue = st.tabs(["🛂 Pasaportes", "🌎 Visas", "✈️ Vuelos"])

    # ══════════════════════════════════════════════════════════════
    # TAB 1 — PASAPORTES
    # ══════════════════════════════════════════════════════════════
    with tab_pas:
        st.subheader("🛂 Gestión de Pasaportes")

        with st.expander("➕ Registrar nuevo trámite de pasaporte", expanded=False):
            conn = conectar_db()
            vendedoras_df = pd.read_sql_query("SELECT id, nombre FROM vendedoras WHERE activa=1 ORDER BY nombre", conn)
            conn.close()

            c1, c2 = st.columns(2)
            with c1:
                pas_cliente  = st.text_input("👤 Nombre del cliente *", key="pas_cliente")
                pas_celular  = st.text_input("📞 Celular", key="pas_celular")
                pas_tipo     = st.selectbox("Tipo", ["Nuevo", "Renovación"], key="pas_tipo")
                pas_estado   = st.selectbox("Estado", ["En trámite", "Cita agendada", "Listo", "Entregado"], key="pas_estado")
            with c2:
                pas_cita     = st.date_input("📅 Fecha de cita", value=None, key="pas_cita")
                pas_entrega  = st.date_input("📅 Fecha estimada entrega", value=None, key="pas_entrega")
                if es_admin and not vendedoras_df.empty:
                    pas_vend_idx = st.selectbox("👩‍💼 Vendedora", vendedoras_df['nombre'].tolist(), key="pas_vend")
                    pas_vend_id  = int(vendedoras_df[vendedoras_df['nombre']==pas_vend_idx]['id'].values[0])
                else:
                    pas_vend_id = id_vend

            c3, c4 = st.columns(2)
            with c3:
                pas_costo    = st.number_input("💰 Costo oficial ($)", min_value=0.0, step=50.0, key="pas_costo")
            with c4:
                pas_servicio = st.number_input("🛎️ Cargo por servicio ($)", min_value=0.0, step=50.0, key="pas_servicio")

            pas_total = pas_costo + pas_servicio
            st.markdown(f"**Total a cobrar al cliente: ${pas_total:,.2f}**")
            pas_notas = st.text_area("📝 Notas", key="pas_notas", height=80)

            if st.button("💾 Guardar Pasaporte", type="primary", key="btn_guardar_pas"):
                if not pas_cliente.strip():
                    st.error("El nombre del cliente es obligatorio.")
                else:
                    conn = conectar_db()
                    conn.execute("""
                        INSERT INTO pasaportes
                        (vendedora_id, cliente, celular, tipo, fecha_cita, fecha_entrega_est,
                         costo_oficial, cargo_servicio, total, pagado, saldo, estado, notas, fecha_registro)
                        VALUES (?,?,?,?,?,?,?,?,?,0,?,?,?,?)
                    """, (
                        pas_vend_id, pas_cliente.strip(), pas_celular.strip(), pas_tipo,
                        str(pas_cita) if pas_cita else None,
                        str(pas_entrega) if pas_entrega else None,
                        pas_costo, pas_servicio, pas_total, pas_total,
                        pas_estado, pas_notas.strip(),
                        datetime.now().strftime("%Y-%m-%d %H:%M")
                    ))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Pasaporte de {pas_cliente} registrado correctamente.")
                    st.rerun()

        st.divider()

        conn = conectar_db()
        filtro_pas = "" if es_admin else f"WHERE p.vendedora_id = {id_vend}"
        try:
            df_pas = pd.read_sql_query(f"""
                SELECT p.id, p.cliente, p.celular, p.tipo, p.fecha_cita,
                       p.fecha_entrega_est, p.costo_oficial, p.cargo_servicio,
                       p.total, p.pagado, p.saldo, p.estado, p.notas,
                       COALESCE(v.nombre,'—') AS vendedora, p.fecha_registro
                FROM pasaportes p
                LEFT JOIN vendedoras v ON p.vendedora_id = v.id
                {filtro_pas}
                ORDER BY p.fecha_registro DESC
            """, conn)
        except:
            df_pas = pd.DataFrame()
        conn.close()

        if df_pas.empty:
            st.info("No hay trámites de pasaporte registrados aún.")
        else:
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                filtro_est_pas = st.selectbox("Filtrar por estado", ["Todos","En trámite","Cita agendada","Listo","Entregado"], key="filtro_est_pas")
            with col_f2:
                buscar_pas = st.text_input("🔍 Buscar cliente", key="buscar_pas")

            df_pas_show = df_pas.copy()
            if filtro_est_pas != "Todos":
                df_pas_show = df_pas_show[df_pas_show['estado'] == filtro_est_pas]
            if buscar_pas:
                df_pas_show = df_pas_show[df_pas_show['cliente'].str.contains(buscar_pas, case=False, na=False)]

            for _, row in df_pas_show.iterrows():
                estado_color = {"En trámite":"#E65100","Cita agendada":"#0077B6","Listo":"#2E7D32","Entregado":"#555"}.get(row['estado'],"#555")
                saldo_v = float(row['saldo']) if row['saldo'] else 0
                with st.expander(f"🛂 {row['cliente']} — {row['tipo']} — {row['estado']} | Saldo: ${saldo_v:,.2f}"):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.markdown(f"**👤 Cliente:** {row['cliente']}")
                        st.markdown(f"**📞 Celular:** {row['celular'] or '—'}")
                        st.markdown(f"**📋 Tipo:** {row['tipo']}")
                        st.markdown(f"**👩‍💼 Vendedora:** {row['vendedora']}")
                    with c2:
                        st.markdown(f"**📅 Fecha cita:** {row['fecha_cita'] or '—'}")
                        st.markdown(f"**📅 Entrega est.:** {row['fecha_entrega_est'] or '—'}")
                        st.markdown(f"<span style='color:{estado_color};font-weight:700;'>Estado: {row['estado']}</span>", unsafe_allow_html=True)
                    with c3:
                        st.markdown(f"**💰 Costo oficial:** ${float(row['costo_oficial']):,.2f}")
                        st.markdown(f"**🛎️ Servicio:** ${float(row['cargo_servicio']):,.2f}")
                        st.markdown(f"**Total: ${float(row['total']):,.2f}**")
                        st.markdown(f"**✅ Pagado:** ${float(row['pagado']):,.2f}")
                        st.markdown(f"**⚠️ Saldo:** ${saldo_v:,.2f}")

                    if row['notas']:
                        st.caption(f"📝 {row['notas']}")

                    col_est, col_abn = st.columns(2)
                    with col_est:
                        nuevo_est = st.selectbox("Cambiar estado", ["En trámite","Cita agendada","Listo","Entregado"],
                            index=["En trámite","Cita agendada","Listo","Entregado"].index(row['estado']) if row['estado'] in ["En trámite","Cita agendada","Listo","Entregado"] else 0,
                            key=f"est_pas_{row['id']}")
                        if st.button("💾 Actualizar estado", key=f"upd_pas_{row['id']}"):
                            conn = conectar_db()
                            conn.execute("UPDATE pasaportes SET estado=? WHERE id=?", (nuevo_est, row['id']))
                            conn.commit()
                            conn.close()
                            st.rerun()
                    with col_abn:
                        if saldo_v > 0:
                            monto_abn = st.number_input("💵 Registrar abono ($)", min_value=0.0, max_value=saldo_v, step=50.0, key=f"abn_pas_{row['id']}")
                            metodo_abn = st.selectbox("Método", ["Efectivo","Transferencia","Tarjeta"], key=f"met_pas_{row['id']}")
                            if st.button("➕ Abonar", key=f"btn_abn_pas_{row['id']}"):
                                if monto_abn > 0:
                                    conn = conectar_db()
                                    nuevo_pagado = float(row['pagado']) + monto_abn
                                    nuevo_saldo  = float(row['total']) - nuevo_pagado
                                    nuevo_estado = "Entregado" if nuevo_saldo <= 0 else row['estado']
                                    conn.execute("INSERT INTO abonos_pasaportes (pasaporte_id, monto, fecha, metodo_pago) VALUES (?,?,?,?)",
                                        (row['id'], monto_abn, datetime.now().strftime("%Y-%m-%d"), metodo_abn))
                                    conn.execute("UPDATE pasaportes SET pagado=?, saldo=?, estado=? WHERE id=?",
                                        (nuevo_pagado, nuevo_saldo, nuevo_estado, row['id']))
                                    conn.commit()
                                    conn.close()
                                    st.success(f"✅ Abono de ${monto_abn:,.2f} registrado.")
                                    st.rerun()

    # ══════════════════════════════════════════════════════════════
    # TAB 2 — VISAS
    # ══════════════════════════════════════════════════════════════
    with tab_vis:
        st.subheader("🌎 Gestión de Visas")

        with st.expander("➕ Registrar nuevo trámite de visa", expanded=False):
            conn = conectar_db()
            vendedoras_df = pd.read_sql_query("SELECT id, nombre FROM vendedoras WHERE activa=1 ORDER BY nombre", conn)
            conn.close()

            c1, c2 = st.columns(2)
            with c1:
                vis_cliente   = st.text_input("👤 Nombre del cliente *", key="vis_cliente")
                vis_celular   = st.text_input("📞 Celular", key="vis_celular")
                vis_pais      = st.text_input("🌍 País destino", key="vis_pais", placeholder="Ej: Estados Unidos")
                vis_tipo      = st.selectbox("Tipo de visa", ["Turista","Negocios","Estudiante","Trabajo","Tránsito","Otra"], key="vis_tipo")
            with c2:
                vis_familiar  = st.checkbox("👨‍👩‍👧 Trámite familiar", key="vis_familiar")
                vis_integ     = st.number_input("# Integrantes", min_value=1, value=1, step=1, key="vis_integ") if vis_familiar else 1
                vis_cita      = st.date_input("📅 Fecha de cita", value=None, key="vis_cita")
                vis_entrega   = st.date_input("📅 Fecha estimada entrega", value=None, key="vis_entrega")
                vis_estado    = st.selectbox("Estado", ["En trámite","Cita agendada","Aprobada","Rechazada","Entregada"], key="vis_estado")
                if es_admin and not vendedoras_df.empty:
                    vis_vend_idx = st.selectbox("👩‍💼 Vendedora", vendedoras_df['nombre'].tolist(), key="vis_vend")
                    vis_vend_id  = int(vendedoras_df[vendedoras_df['nombre']==vis_vend_idx]['id'].values[0])
                else:
                    vis_vend_id = id_vend

            c3, c4 = st.columns(2)
            with c3:
                vis_costo    = st.number_input("💰 Costo oficial ($)", min_value=0.0, step=50.0, key="vis_costo")
            with c4:
                vis_servicio = st.number_input("🛎️ Cargo por servicio ($)", min_value=0.0, step=50.0, key="vis_servicio")

            vis_total = vis_costo + vis_servicio
            st.markdown(f"**Total a cobrar al cliente: ${vis_total:,.2f}**")
            vis_notas = st.text_area("📝 Notas", key="vis_notas", height=80)

            if st.button("💾 Guardar Visa", type="primary", key="btn_guardar_vis"):
                if not vis_cliente.strip():
                    st.error("El nombre del cliente es obligatorio.")
                else:
                    conn = conectar_db()
                    conn.execute("""
                        INSERT INTO visas
                        (vendedora_id, cliente, celular, pais_destino, tipo_visa, es_familiar,
                         num_integrantes, fecha_cita, fecha_entrega_est, costo_oficial, cargo_servicio,
                         total, pagado, saldo, estado, notas, fecha_registro)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,0,?,?,?,?)
                    """, (
                        vis_vend_id, vis_cliente.strip(), vis_celular.strip(),
                        vis_pais.strip(), vis_tipo, 1 if vis_familiar else 0, int(vis_integ),
                        str(vis_cita) if vis_cita else None,
                        str(vis_entrega) if vis_entrega else None,
                        vis_costo, vis_servicio, vis_total, vis_total,
                        vis_estado, vis_notas.strip(),
                        datetime.now().strftime("%Y-%m-%d %H:%M")
                    ))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Visa de {vis_cliente} registrada correctamente.")
                    st.rerun()

        st.divider()

        conn = conectar_db()
        filtro_vis = "" if es_admin else f"WHERE vi.vendedora_id = {id_vend}"
        try:
            df_vis = pd.read_sql_query(f"""
                SELECT vi.id, vi.cliente, vi.celular, vi.pais_destino, vi.tipo_visa,
                       vi.es_familiar, vi.num_integrantes, vi.fecha_cita, vi.fecha_entrega_est,
                       vi.costo_oficial, vi.cargo_servicio, vi.total, vi.pagado, vi.saldo,
                       vi.estado, vi.notas, COALESCE(v.nombre,'—') AS vendedora
                FROM visas vi
                LEFT JOIN vendedoras v ON vi.vendedora_id = v.id
                {filtro_vis}
                ORDER BY vi.fecha_registro DESC
            """, conn)
        except:
            df_vis = pd.DataFrame()
        conn.close()

        if df_vis.empty:
            st.info("No hay trámites de visa registrados aún.")
        else:
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                filtro_est_vis = st.selectbox("Filtrar por estado", ["Todos","En trámite","Cita agendada","Aprobada","Rechazada","Entregada"], key="filtro_est_vis")
            with col_f2:
                buscar_vis = st.text_input("🔍 Buscar cliente", key="buscar_vis")

            df_vis_show = df_vis.copy()
            if filtro_est_vis != "Todos":
                df_vis_show = df_vis_show[df_vis_show['estado'] == filtro_est_vis]
            if buscar_vis:
                df_vis_show = df_vis_show[df_vis_show['cliente'].str.contains(buscar_vis, case=False, na=False)]

            for _, row in df_vis_show.iterrows():
                estado_color = {"En trámite":"#E65100","Cita agendada":"#0077B6","Aprobada":"#2E7D32","Rechazada":"#CC0000","Entregada":"#555"}.get(row['estado'],"#555")
                saldo_v = float(row['saldo']) if row['saldo'] else 0
                fam_txt = f"Familiar ({int(row['num_integrantes'])} integrantes)" if row['es_familiar'] else "Personal"
                with st.expander(f"🌎 {row['cliente']} — {row['pais_destino'] or '—'} — {row['tipo_visa']} — {row['estado']}"):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.markdown(f"**👤 Cliente:** {row['cliente']}")
                        st.markdown(f"**📞 Celular:** {row['celular'] or '—'}")
                        st.markdown(f"**🌍 País:** {row['pais_destino'] or '—'}")
                        st.markdown(f"**📋 Tipo:** {row['tipo_visa']} — {fam_txt}")
                        st.markdown(f"**👩‍💼 Vendedora:** {row['vendedora']}")
                    with c2:
                        st.markdown(f"**📅 Fecha cita:** {row['fecha_cita'] or '—'}")
                        st.markdown(f"**📅 Entrega est.:** {row['fecha_entrega_est'] or '—'}")
                        st.markdown(f"<span style='color:{estado_color};font-weight:700;'>Estado: {row['estado']}</span>", unsafe_allow_html=True)
                    with c3:
                        st.markdown(f"**💰 Costo oficial:** ${float(row['costo_oficial']):,.2f}")
                        st.markdown(f"**🛎️ Servicio:** ${float(row['cargo_servicio']):,.2f}")
                        st.markdown(f"**Total: ${float(row['total']):,.2f}**")
                        st.markdown(f"**✅ Pagado:** ${float(row['pagado']):,.2f}")
                        st.markdown(f"**⚠️ Saldo:** ${saldo_v:,.2f}")

                    if row['notas']:
                        st.caption(f"📝 {row['notas']}")

                    col_est, col_abn = st.columns(2)
                    with col_est:
                        estados_vis = ["En trámite","Cita agendada","Aprobada","Rechazada","Entregada"]
                        nuevo_est = st.selectbox("Cambiar estado", estados_vis,
                            index=estados_vis.index(row['estado']) if row['estado'] in estados_vis else 0,
                            key=f"est_vis_{row['id']}")
                        if st.button("💾 Actualizar estado", key=f"upd_vis_{row['id']}"):
                            conn = conectar_db()
                            conn.execute("UPDATE visas SET estado=? WHERE id=?", (nuevo_est, row['id']))
                            conn.commit()
                            conn.close()
                            st.rerun()
                    with col_abn:
                        if saldo_v > 0:
                            monto_abn = st.number_input("💵 Registrar abono ($)", min_value=0.0, max_value=saldo_v, step=50.0, key=f"abn_vis_{row['id']}")
                            metodo_abn = st.selectbox("Método", ["Efectivo","Transferencia","Tarjeta"], key=f"met_vis_{row['id']}")
                            if st.button("➕ Abonar", key=f"btn_abn_vis_{row['id']}"):
                                if monto_abn > 0:
                                    conn = conectar_db()
                                    nuevo_pagado = float(row['pagado']) + monto_abn
                                    nuevo_saldo  = float(row['total']) - nuevo_pagado
                                    nuevo_estado = "Entregada" if nuevo_saldo <= 0 else row['estado']
                                    conn.execute("INSERT INTO abonos_visas (visa_id, monto, fecha, metodo_pago) VALUES (?,?,?,?)",
                                        (row['id'], monto_abn, datetime.now().strftime("%Y-%m-%d"), metodo_abn))
                                    conn.execute("UPDATE visas SET pagado=?, saldo=?, estado=? WHERE id=?",
                                        (nuevo_pagado, nuevo_saldo, nuevo_estado, row['id']))
                                    conn.commit()
                                    conn.close()
                                    st.success(f"✅ Abono de ${monto_abn:,.2f} registrado.")
                                    st.rerun()

    # ══════════════════════════════════════════════════════════════
    # TAB 3 — VUELOS
    # ══════════════════════════════════════════════════════════════
    with tab_vue:
        st.subheader("✈️ Gestión de Vuelos")

        with st.expander("➕ Registrar nuevo vuelo", expanded=False):
            conn = conectar_db()
            vendedoras_df = pd.read_sql_query("SELECT id, nombre FROM vendedoras WHERE activa=1 ORDER BY nombre", conn)
            conn.close()

            c1, c2 = st.columns(2)
            with c1:
                vue_pasajero  = st.text_input("👤 Nombre del pasajero *", key="vue_pasajero")
                vue_celular   = st.text_input("📞 Celular", key="vue_celular")
                vue_tipo      = st.selectbox("Tipo de vuelo", ["Nacional","Internacional"], key="vue_tipo")
                vue_aerolinea = st.selectbox("✈️ Aerolínea", ["Volaris","Aeroméxico","Viva Aerobus","Interjet","American Airlines","United","Delta","Copa Airlines","Otra"], key="vue_aerolinea")
                vue_num_pas   = st.number_input("👥 Número de pasajeros", min_value=1, value=1, step=1, key="vue_num_pas")
            with c2:
                vue_origen    = st.text_input("🛫 Origen", key="vue_origen", placeholder="Ej: Mérida (MID)")
                vue_destino   = st.text_input("🛬 Destino", key="vue_destino", placeholder="Ej: Cancún (CUN)")
                vue_fecha     = st.date_input("📅 Fecha del vuelo", value=None, key="vue_fecha")
                vue_hora      = st.text_input("🕐 Hora del vuelo", key="vue_hora", placeholder="Ej: 14:30")
                vue_estado    = st.selectbox("Estado", ["Cotizado","Confirmado","Emitido","Cancelado"], key="vue_estado")
                if es_admin and not vendedoras_df.empty:
                    vue_vend_idx = st.selectbox("👩‍💼 Vendedora", vendedoras_df['nombre'].tolist(), key="vue_vend")
                    vue_vend_id  = int(vendedoras_df[vendedoras_df['nombre']==vue_vend_idx]['id'].values[0])
                else:
                    vue_vend_id = id_vend

            c3, c4 = st.columns(2)
            with c3:
                vue_costo    = st.number_input("💰 Costo de compra ($)", min_value=0.0, step=100.0, key="vue_costo",
                                               help="Lo que pagó la agencia en la plataforma")
            with c4:
                vue_servicio = st.number_input("🛎️ Cargo por servicio ($)", min_value=0.0, step=50.0, key="vue_servicio")

            vue_total = vue_costo + vue_servicio
            st.markdown(f"**Total a cobrar al cliente: ${vue_total:,.2f}**")
            vue_notas = st.text_area("📝 Notas (# confirmación, equipaje, etc.)", key="vue_notas", height=80)

            if st.button("💾 Guardar Vuelo", type="primary", key="btn_guardar_vue"):
                if not vue_pasajero.strip():
                    st.error("El nombre del pasajero es obligatorio.")
                else:
                    conn = conectar_db()
                    conn.execute("""
                        INSERT INTO vuelos
                        (vendedora_id, pasajero, celular, tipo, aerolinea, origen, destino,
                         fecha_vuelo, hora_vuelo, num_pasajeros, costo_compra, cargo_servicio,
                         total, pagado, saldo, estado, notas, fecha_registro)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,0,?,?,?,?)
                    """, (
                        vue_vend_id, vue_pasajero.strip(), vue_celular.strip(),
                        vue_tipo, vue_aerolinea, vue_origen.strip(), vue_destino.strip(),
                        str(vue_fecha) if vue_fecha else None, vue_hora.strip(),
                        int(vue_num_pas), vue_costo, vue_servicio, vue_total, vue_total,
                        vue_estado, vue_notas.strip(),
                        datetime.now().strftime("%Y-%m-%d %H:%M")
                    ))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Vuelo de {vue_pasajero} registrado correctamente.")
                    st.rerun()

        st.divider()

        conn = conectar_db()
        filtro_vue = "" if es_admin else f"WHERE vu.vendedora_id = {id_vend}"
        try:
            df_vue = pd.read_sql_query(f"""
                SELECT vu.id, vu.pasajero, vu.celular, vu.tipo, vu.aerolinea,
                       vu.origen, vu.destino, vu.fecha_vuelo, vu.hora_vuelo,
                       vu.num_pasajeros, vu.costo_compra, vu.cargo_servicio,
                       vu.total, vu.pagado, vu.saldo, vu.estado, vu.notas,
                       COALESCE(v.nombre,'—') AS vendedora
                FROM vuelos vu
                LEFT JOIN vendedoras v ON vu.vendedora_id = v.id
                {filtro_vue}
                ORDER BY vu.fecha_vuelo DESC
            """, conn)
        except:
            df_vue = pd.DataFrame()
        conn.close()

        if df_vue.empty:
            st.info("No hay vuelos registrados aún.")
        else:
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                filtro_est_vue = st.selectbox("Estado", ["Todos","Cotizado","Confirmado","Emitido","Cancelado"], key="filtro_est_vue")
            with col_f2:
                filtro_tipo_vue = st.selectbox("Tipo", ["Todos","Nacional","Internacional"], key="filtro_tipo_vue")
            with col_f3:
                buscar_vue = st.text_input("🔍 Buscar pasajero", key="buscar_vue")

            df_vue_show = df_vue.copy()
            if filtro_est_vue != "Todos":
                df_vue_show = df_vue_show[df_vue_show['estado'] == filtro_est_vue]
            if filtro_tipo_vue != "Todos":
                df_vue_show = df_vue_show[df_vue_show['tipo'] == filtro_tipo_vue]
            if buscar_vue:
                df_vue_show = df_vue_show[df_vue_show['pasajero'].str.contains(buscar_vue, case=False, na=False)]

            for _, row in df_vue_show.iterrows():
                estado_color = {"Cotizado":"#E65100","Confirmado":"#0077B6","Emitido":"#2E7D32","Cancelado":"#CC0000"}.get(row['estado'],"#555")
                saldo_v = float(row['saldo']) if row['saldo'] else 0
                with st.expander(f"✈️ {row['pasajero']} — {row['origen'] or '—'} → {row['destino'] or '—'} | {row['aerolinea']} | {row['fecha_vuelo'] or '—'} | {row['estado']}"):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.markdown(f"**👤 Pasajero:** {row['pasajero']}")
                        st.markdown(f"**📞 Celular:** {row['celular'] or '—'}")
                        st.markdown(f"**✈️ Aerolínea:** {row['aerolinea']}")
                        st.markdown(f"**🏷️ Tipo:** {row['tipo']}")
                        st.markdown(f"**👥 Pasajeros:** {int(row['num_pasajeros']) if row['num_pasajeros'] else 1}")
                        st.markdown(f"**👩‍💼 Vendedora:** {row['vendedora']}")
                    with c2:
                        st.markdown(f"**🛫 Origen:** {row['origen'] or '—'}")
                        st.markdown(f"**🛬 Destino:** {row['destino'] or '—'}")
                        st.markdown(f"**📅 Fecha:** {row['fecha_vuelo'] or '—'}")
                        st.markdown(f"**🕐 Hora:** {row['hora_vuelo'] or '—'}")
                        st.markdown(f"<span style='color:{estado_color};font-weight:700;'>Estado: {row['estado']}</span>", unsafe_allow_html=True)
                    with c3:
                        st.markdown(f"**💰 Costo compra:** ${float(row['costo_compra']):,.2f}")
                        st.markdown(f"**🛎️ Servicio:** ${float(row['cargo_servicio']):,.2f}")
                        st.markdown(f"**Total: ${float(row['total']):,.2f}**")
                        st.markdown(f"**✅ Pagado:** ${float(row['pagado']):,.2f}")
                        st.markdown(f"**⚠️ Saldo:** ${saldo_v:,.2f}")

                    if row['notas']:
                        st.caption(f"📝 {row['notas']}")

                    col_est, col_abn = st.columns(2)
                    with col_est:
                        estados_vue = ["Cotizado","Confirmado","Emitido","Cancelado"]
                        nuevo_est = st.selectbox("Cambiar estado", estados_vue,
                            index=estados_vue.index(row['estado']) if row['estado'] in estados_vue else 0,
                            key=f"est_vue_{row['id']}")
                        if st.button("💾 Actualizar estado", key=f"upd_vue_{row['id']}"):
                            conn = conectar_db()
                            conn.execute("UPDATE vuelos SET estado=? WHERE id=?", (nuevo_est, row['id']))
                            conn.commit()
                            conn.close()
                            st.rerun()
                    with col_abn:
                        if saldo_v > 0:
                            monto_abn = st.number_input("💵 Registrar abono ($)", min_value=0.0, max_value=saldo_v, step=50.0, key=f"abn_vue_{row['id']}")
                            metodo_abn = st.selectbox("Método", ["Efectivo","Transferencia","Tarjeta"], key=f"met_vue_{row['id']}")
                            if st.button("➕ Abonar", key=f"btn_abn_vue_{row['id']}"):
                                if monto_abn > 0:
                                    conn = conectar_db()
                                    nuevo_pagado = float(row['pagado']) + monto_abn
                                    nuevo_saldo  = float(row['total']) - nuevo_pagado
                                    nuevo_estado = "Emitido" if nuevo_saldo <= 0 else row['estado']
                                    conn.execute("INSERT INTO abonos_vuelos (vuelo_id, monto, fecha, metodo_pago) VALUES (?,?,?,?)",
                                        (row['id'], monto_abn, datetime.now().strftime("%Y-%m-%d"), metodo_abn))
                                    conn.execute("UPDATE vuelos SET pagado=?, saldo=?, estado=? WHERE id=?",
                                        (nuevo_pagado, nuevo_saldo, nuevo_estado, row['id']))
                                    conn.commit()
                                    conn.close()
                                    st.success(f"✅ Abono de ${monto_abn:,.2f} registrado.")
                                    st.rerun()


# Menú lateral
def menu_lateral():
    """Menú de navegación lateral"""
    with st.sidebar:
        if os.path.exists(LOGO_PATH):
            st.image(LOGO_PATH, use_container_width=True)
        else:
            st.markdown("<div style='color:#90caf9;font-size:1.1rem;font-weight:700;padding:8px 0;'>✈️ Turismar</div>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        usuario = st.session_state.usuario_actual
        st.markdown(f"**👤 {usuario['nombre']}**")
        st.markdown(f"*{usuario['rol']}*")
        
        st.markdown("---")
        
        # Opciones del menú
        opciones = {
            "🏠 Dashboard": "dashboard",
            "🏖️ Riviera Maya": "riviera",
            "🎫 Viajes Nacionales": "nacionales",
            "🌎 Viajes Internacionales": "internacionales",
            "💸 Transferencias": "transferencias",
            "🗂️ Otros": "otros",
            "📊 Reportes": "reportes"
        }
        
        if usuario["rol"] == "ADMIN":
            opciones["💰 Comisiones / Config"] = "config"
        if usuario["rol"] == "VENDEDORA":
            opciones["📂 Mi Historial"] = "mi_historial"
        
        for label, key in opciones.items():
            if st.button(label, use_container_width=True):
                st.session_state.pagina_actual = key
        
        st.markdown("---")
        
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.usuario_actual = None
            st.rerun()

# Aplicación principal
def main():
    """Función principal de la aplicación"""

    # Inicializar base de datos (crea tablas si no existen)
    inicializar_base_datos()

    # Inicializar session state
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    
    if 'pagina_actual' not in st.session_state:
        st.session_state.pagina_actual = "dashboard"
    
    # Login
    if not st.session_state.logged_in:
        # Logo centrado en login
        col_logo1, col_logo2, col_logo3 = st.columns([1, 1, 1])
        with col_logo2:
            if os.path.exists(LOGO_PATH):
                st.image(LOGO_PATH, use_container_width=True)
            else:
                st.markdown("<h1 style='text-align:center;'>✈️</h1>", unsafe_allow_html=True)
        st.markdown("<h1 style='text-align: center;'>Sistema Agencia de Viajes</h1>", unsafe_allow_html=True)
        st.markdown("<h3 style='text-align: center; color:#0077B6;'>Riviera Maya</h3>", unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            st.markdown("### 🔐 Iniciar Sesión")
            
            usuario = st.text_input("Usuario", placeholder="Ingresa tu usuario")
            password = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            
            if st.button("Ingresar", type="primary", use_container_width=True):
                if usuario and password:
                    user_data = verificar_login(usuario, password)
                    if user_data:
                        st.session_state.logged_in = True
                        st.session_state.usuario_actual = user_data
                        st.rerun()
                    else:
                        st.error("❌ Usuario o contraseña incorrectos")
                else:
                    st.warning("⚠️ Por favor ingresa usuario y contraseña")
            
            st.info("💡 **Usuario de prueba:**\n\nUsuario: `admin`\n\nContraseña: `admin123`")
    
    else:
        # Menú lateral
        menu_lateral()
        
        # Contenido principal según la página seleccionada
        if st.session_state.pagina_actual == "dashboard":
            mostrar_dashboard()
        elif st.session_state.pagina_actual == "riviera":
            pagina_ventas_riviera()
        elif st.session_state.pagina_actual == "nacionales":
            pagina_viajes_nacionales()
        elif st.session_state.pagina_actual == "internacionales":
            pagina_viajes_internacionales()
        elif st.session_state.pagina_actual == "transferencias":
            if TRANSFERENCIAS_DISPONIBLE:
                mostrar_pagina_transferencias()
            else:
                st.error("⚠️ Módulo de transferencias no disponible")
        elif st.session_state.pagina_actual == "otros":
            pagina_otros()
        elif st.session_state.pagina_actual == "reportes":
            pagina_reportes()
        elif st.session_state.pagina_actual == "config":
            pagina_configuracion()
        elif st.session_state.pagina_actual == "mi_historial":
            pagina_mi_historial()

if __name__ == "__main__":
    main()
