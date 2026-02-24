from database import conectar
from datetime import datetime
import os

def reporte_viaje_internacional_consola():
    """Muestra reporte detallado de un viaje internacional en consola"""
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Mostrar viajes disponibles
    cursor.execute("""
        SELECT id, destino, fecha_salida, fecha_regreso, estado
        FROM viajes_internacionales
        ORDER BY fecha_salida DESC
    """)
    
    viajes = cursor.fetchall()
    
    if not viajes:
        print("\n❌ No hay viajes internacionales registrados.")
        conexion.close()
        return
    
    print("\n" + "="*70)
    print("🌎 VIAJES INTERNACIONALES DISPONIBLES")
    print("="*70)
    
    for v in viajes:
        estado_icon = "✅" if v[4] == "ACTIVO" else "🔒"
        print(f"{estado_icon} {v[0]}. {v[1]} ({v[2]} al {v[3]}) - {v[4]}")
    
    try:
        viaje_id = int(input("\nSelecciona ID del viaje para ver reporte: "))
    except ValueError:
        print("❌ ID inválido.")
        conexion.close()
        return
    
    # Obtener información del viaje
    cursor.execute("""
        SELECT destino, fecha_salida, fecha_regreso, dias, noches,
               cupos_totales, cupos_vendidos, cupos_disponibles,
               precio_adulto_doble_usd, precio_adulto_triple_usd,
               precio_menor_doble_usd, precio_menor_triple_usd,
               porcentaje_ganancia, estado
        FROM viajes_internacionales
        WHERE id = ?
    """, (viaje_id,))
    
    viaje = cursor.fetchone()
    
    if not viaje:
        print("❌ Viaje no encontrado.")
        conexion.close()
        return
    
    # ===== ENCABEZADO =====
    print("\n" + "="*70)
    print(f"🌎 REPORTE DETALLADO - {viaje[0].upper()}")
    print("="*70)
    
    print(f"\n📅 INFORMACIÓN DEL VIAJE:")
    print(f"   Fechas: {viaje[1]} al {viaje[2]}")
    print(f"   Duración: {viaje[3]} días / {viaje[4]} noches")
    print(f"   Estado: {viaje[13]}")
    
    print(f"\n👥 CUPOS:")
    print(f"   Total: {viaje[5]} personas")
    print(f"   Vendidos: {viaje[6]} ({(viaje[6]/viaje[5]*100):.1f}%)")
    print(f"   Disponibles: {viaje[7]}")
    
    print(f"\n💵 PRECIOS BASE (USD):")
    print(f"   Adulto doble: ${viaje[8]:,.2f}")
    print(f"   Adulto triple: ${viaje[9]:,.2f}")
    print(f"   Menor doble: ${viaje[10]:,.2f}")
    print(f"   Menor triple: ${viaje[11]:,.2f}")
    print(f"   Margen de ganancia: {viaje[12]:.1f}%")
    
    # ===== CLIENTES Y PASAJEROS =====
    cursor.execute("""
        SELECT c.id, c.nombre_cliente, c.adultos, c.menores,
               c.habitaciones_doble, c.habitaciones_triple,
               c.total_usd, c.abonado_usd, c.saldo_usd, 
               c.ganancia_usd, c.estado
        FROM clientes_internacionales c
        WHERE c.viaje_id = ?
        ORDER BY c.nombre_cliente
    """, (viaje_id,))
    
    clientes = cursor.fetchall()
    
    if not clientes:
        print("\n\n⚠️  No hay clientes registrados en este viaje.")
        conexion.close()
        return
    
    print(f"\n\n{'='*70}")
    print(f"👥 CLIENTES REGISTRADOS ({len(clientes)})")
    print("="*70)
    
    # Totales
    total_vendido = 0
    total_abonado = 0
    total_saldo = 0
    total_ganancia = 0
    total_hab_dobles = 0
    total_hab_triples = 0
    
    for cliente in clientes:
        cliente_id = cliente[0]
        estado_icon = "✅" if cliente[10] == "LIQUIDADO" else "⏳"
        
        print(f"\n{estado_icon} {cliente[1]}")
        print(f"   Pasajeros: {cliente[2]} adultos + {cliente[3]} menores = {cliente[2] + cliente[3]} total")
        print(f"   Habitaciones: {cliente[4]} doble(s) + {cliente[5]} triple(s)")
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre_completo, tipo, habitacion_asignada
            FROM pasajeros_internacionales
            WHERE cliente_id = ?
            ORDER BY habitacion_asignada, tipo DESC
        """, (cliente_id,))
        
        pasajeros = cursor.fetchall()
        
        if pasajeros:
            print(f"\n   📋 DISTRIBUCIÓN DE HABITACIONES:")
            
            # Agrupar por habitación
            habitaciones = {}
            for p in pasajeros:
                hab = p[2] if p[2] else "Sin asignar"
                if hab not in habitaciones:
                    habitaciones[hab] = []
                habitaciones[hab].append(f"{p[0]} ({p[1]})")
            
            for hab, personas in sorted(habitaciones.items()):
                print(f"      {hab}:")
                for persona in personas:
                    print(f"         • {persona}")
        
        print(f"\n   💵 PAGOS (USD):")
        print(f"      Total: ${cliente[6]:,.2f}")
        print(f"      Abonado: ${cliente[7]:,.2f}")
        print(f"      Saldo: ${cliente[8]:,.2f}")
        print(f"      Ganancia: ${cliente[9]:,.2f}")
        
        # Historial de abonos
        cursor.execute("""
            SELECT fecha, moneda, monto_original, tipo_cambio, monto_usd
            FROM abonos_internacionales
            WHERE cliente_id = ?
            ORDER BY fecha
        """, (cliente_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            print(f"\n   📅 Historial de pagos:")
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
                if abono[1] == "USD":
                    print(f"      {fecha}: ${abono[4]:,.2f} USD")
                else:
                    print(f"      {fecha}: ${abono[2]:,.2f} MXN (TC: {abono[3]:.2f}) = ${abono[4]:,.2f} USD")
        
        print("-" * 70)
        
        # Acumular totales
        total_vendido += cliente[6]
        total_abonado += cliente[7]
        total_saldo += cliente[8]
        total_ganancia += cliente[9]
        total_hab_dobles += cliente[4]
        total_hab_triples += cliente[5]
    
    # ===== RESUMEN GENERAL =====
    print(f"\n{'='*70}")
    print("📊 RESUMEN GENERAL")
    print("="*70)
    
    print(f"\n🏨 HABITACIONES:")
    print(f"   Dobles: {total_hab_dobles}")
    print(f"   Triples: {total_hab_triples}")
    print(f"   Total: {total_hab_dobles + total_hab_triples}")
    
    print(f"\n💰 FINANCIERO (USD):")
    print(f"   Total vendido: ${total_vendido:,.2f}")
    print(f"   Total abonado: ${total_abonado:,.2f}")
    print(f"   Total saldo: ${total_saldo:,.2f}")
    porcentaje_cobrado = (total_abonado / total_vendido * 100) if total_vendido > 0 else 0
    print(f"   % Cobrado: {porcentaje_cobrado:.1f}%")
    
    print(f"\n📈 GANANCIA:")
    print(f"   Ganancia total: ${total_ganancia:,.2f}")
    if total_vendido > 0:
        margen = (total_ganancia / total_vendido * 100)
        print(f"   Margen promedio: {margen:.1f}%")
    
    # ===== ALERTAS =====
    print(f"\n⚠️  ALERTAS:")
    alertas = []
    
    # Clientes con saldo alto
    cursor.execute("""
        SELECT COUNT(*) FROM clientes_internacionales
        WHERE viaje_id = ? AND saldo_usd > 1000
    """, (viaje_id,))
    
    clientes_saldo_alto = cursor.fetchone()[0]
    if clientes_saldo_alto > 0:
        alertas.append(f"   • {clientes_saldo_alto} cliente(s) con saldo > $1,000 USD")
    
    # % Cobrado bajo
    if porcentaje_cobrado < 50 and len(clientes) > 0:
        alertas.append(f"   • ⚠️  Solo {porcentaje_cobrado:.0f}% cobrado del total")
    
    # Cupos disponibles
    if viaje[7] > 0 and viaje[13] == "ACTIVO":
        alertas.append(f"   • {viaje[7]} cupos aún disponibles para venta")
    
    if alertas:
        for alerta in alertas:
            print(alerta)
    else:
        print("   ✅ Sin alertas")
    
    print("\n" + "="*70)
    
    conexion.close()


def exportar_reporte_excel():
    """Exporta reporte de viaje internacional a Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("\n❌ Error: Se requiere openpyxl para exportar a Excel.")
        print("Instala con: pip install openpyxl")
        return
    
    conexion = conectar()
    cursor = conexion.cursor()
    
    # Mostrar viajes disponibles
    cursor.execute("""
        SELECT id, destino, fecha_salida, fecha_regreso, estado
        FROM viajes_internacionales
        ORDER BY fecha_salida DESC
    """)
    
    viajes = cursor.fetchall()
    
    if not viajes:
        print("\n❌ No hay viajes internacionales registrados.")
        conexion.close()
        return
    
    print("\n" + "="*70)
    print("🌎 VIAJES INTERNACIONALES DISPONIBLES")
    print("="*70)
    
    for v in viajes:
        estado_icon = "✅" if v[4] == "ACTIVO" else "🔒"
        print(f"{estado_icon} {v[0]}. {v[1]} ({v[2]} al {v[3]}) - {v[4]}")
    
    try:
        viaje_id = int(input("\nSelecciona ID del viaje para exportar: "))
    except ValueError:
        print("❌ ID inválido.")
        conexion.close()
        return
    
    # Obtener información del viaje
    cursor.execute("""
        SELECT destino, fecha_salida, fecha_regreso, dias, noches,
               cupos_totales, cupos_vendidos, cupos_disponibles,
               precio_adulto_doble_usd, precio_adulto_triple_usd,
               precio_menor_doble_usd, precio_menor_triple_usd,
               porcentaje_ganancia, estado
        FROM viajes_internacionales
        WHERE id = ?
    """, (viaje_id,))
    
    viaje = cursor.fetchone()
    
    if not viaje:
        print("❌ Viaje no encontrado.")
        conexion.close()
        return
    
    destino = viaje[0]
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: INFORMACIÓN GENERAL =====
    ws_info = wb.create_sheet("Información General")
    
    ws_info['A1'] = f"REPORTE - {destino.upper()}"
    ws_info['A1'].font = Font(bold=True, size=16)
    ws_info.merge_cells('A1:D1')
    
    row = 3
    ws_info[f'A{row}'] = "INFORMACIÓN DEL VIAJE"
    ws_info[f'A{row}'].font = title_font
    
    row += 1
    info_data = [
        ["Destino:", destino],
        ["Fecha salida:", viaje[1]],
        ["Fecha regreso:", viaje[2]],
        ["Duración:", f"{viaje[3]} días / {viaje[4]} noches"],
        ["Estado:", viaje[13]],
        ["", ""],
        ["CUPOS", ""],
        ["Total:", viaje[5]],
        ["Vendidos:", viaje[6]],
        ["Disponibles:", viaje[7]],
        ["% Ocupación:", f"{(viaje[6]/viaje[5]*100):.1f}%"],
        ["", ""],
        ["PRECIOS BASE (USD)", ""],
        ["Adulto doble:", f"${viaje[8]:,.2f}"],
        ["Adulto triple:", f"${viaje[9]:,.2f}"],
        ["Menor doble:", f"${viaje[10]:,.2f}"],
        ["Menor triple:", f"${viaje[11]:,.2f}"],
        ["Margen ganancia:", f"{viaje[12]:.1f}%"]
    ]
    
    for data in info_data:
        ws_info[f'A{row}'] = data[0]
        ws_info[f'B{row}'] = data[1]
        if data[0] in ["CUPOS", "PRECIOS BASE (USD)", "INFORMACIÓN DEL VIAJE"]:
            ws_info[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Ajustar anchos
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 20
    
    # ===== HOJA 2: CLIENTES Y PAGOS =====
    ws_clientes = wb.create_sheet("Clientes y Pagos")
    
    ws_clientes['A1'] = "CLIENTES Y ESTADO DE PAGOS"
    ws_clientes['A1'].font = Font(bold=True, size=14)
    ws_clientes.merge_cells('A1:J1')
    
    # Encabezados
    headers = ["ID", "Cliente", "Adultos", "Menores", "Total Pax", "Dobles", "Triples", 
               "Total USD", "Abonado USD", "Saldo USD", "Estado"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws_clientes.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Obtener clientes
    cursor.execute("""
        SELECT c.id, c.nombre_cliente, c.adultos, c.menores,
               c.habitaciones_doble, c.habitaciones_triple,
               c.total_usd, c.abonado_usd, c.saldo_usd, c.estado
        FROM clientes_internacionales c
        WHERE c.viaje_id = ?
        ORDER BY c.nombre_cliente
    """, (viaje_id,))
    
    clientes = cursor.fetchall()
    
    row = 4
    total_vendido = 0
    total_abonado = 0
    total_saldo = 0
    
    for cliente in clientes:
        ws_clientes.cell(row=row, column=1, value=cliente[0])
        ws_clientes.cell(row=row, column=2, value=cliente[1])
        ws_clientes.cell(row=row, column=3, value=cliente[2])
        ws_clientes.cell(row=row, column=4, value=cliente[3])
        ws_clientes.cell(row=row, column=5, value=cliente[2] + cliente[3])
        ws_clientes.cell(row=row, column=6, value=cliente[4])
        ws_clientes.cell(row=row, column=7, value=cliente[5])
        ws_clientes.cell(row=row, column=8, value=cliente[6])
        ws_clientes.cell(row=row, column=9, value=cliente[7])
        ws_clientes.cell(row=row, column=10, value=cliente[8])
        ws_clientes.cell(row=row, column=11, value=cliente[9])
        
        # Formato moneda
        for col in [8, 9, 10]:
            ws_clientes.cell(row=row, column=col).number_format = '$#,##0.00'
        
        # Borders
        for col in range(1, 12):
            ws_clientes.cell(row=row, column=col).border = border
        
        total_vendido += cliente[6]
        total_abonado += cliente[7]
        total_saldo += cliente[8]
        
        row += 1
    
    # Totales
    row += 1
    ws_clientes.cell(row=row, column=7, value="TOTALES:").font = Font(bold=True)
    ws_clientes.cell(row=row, column=8, value=total_vendido).font = Font(bold=True)
    ws_clientes.cell(row=row, column=9, value=total_abonado).font = Font(bold=True)
    ws_clientes.cell(row=row, column=10, value=total_saldo).font = Font(bold=True)
    
    for col in [8, 9, 10]:
        ws_clientes.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_clientes.column_dimensions['A'].width = 8
    ws_clientes.column_dimensions['B'].width = 30
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws_clientes.column_dimensions[col].width = 10
    for col in ['H', 'I', 'J']:
        ws_clientes.column_dimensions[col].width = 15
    ws_clientes.column_dimensions['K'].width = 12
    
    # ===== HOJA 3: DISTRIBUCIÓN DE HABITACIONES =====
    ws_hab = wb.create_sheet("Habitaciones")
    
    ws_hab['A1'] = "DISTRIBUCIÓN DE HABITACIONES Y PASAJEROS"
    ws_hab['A1'].font = Font(bold=True, size=14)
    ws_hab.merge_cells('A1:E1')
    
    # Encabezados
    headers_hab = ["Cliente", "Habitación", "Pasajero", "Tipo", "Total Pax"]
    
    for col, header in enumerate(headers_hab, start=1):
        cell = ws_hab.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_cliente = cliente[1]
        total_pax = cliente[2] + cliente[3]
        
        # Obtener pasajeros
        cursor.execute("""
            SELECT nombre_completo, tipo, habitacion_asignada
            FROM pasajeros_internacionales
            WHERE cliente_id = ?
            ORDER BY habitacion_asignada, tipo DESC
        """, (cliente_id,))
        
        pasajeros = cursor.fetchall()
        
        primera_fila = True
        for pasajero in pasajeros:
            ws_hab.cell(row=row, column=1, value=nombre_cliente if primera_fila else "")
            ws_hab.cell(row=row, column=2, value=pasajero[2] if pasajero[2] else "Sin asignar")
            ws_hab.cell(row=row, column=3, value=pasajero[0])
            ws_hab.cell(row=row, column=4, value=pasajero[1])
            ws_hab.cell(row=row, column=5, value=total_pax if primera_fila else "")
            
            for col in range(1, 6):
                ws_hab.cell(row=row, column=col).border = border
            
            primera_fila = False
            row += 1
        
        # Línea separadora
        row += 1
    
    # Ajustar anchos
    ws_hab.column_dimensions['A'].width = 30
    ws_hab.column_dimensions['B'].width = 20
    ws_hab.column_dimensions['C'].width = 35
    ws_hab.column_dimensions['D'].width = 12
    ws_hab.column_dimensions['E'].width = 12
    
    # ===== HOJA 4: HISTORIAL DE PAGOS =====
    ws_pagos = wb.create_sheet("Historial de Pagos")
    
    ws_pagos['A1'] = "HISTORIAL DETALLADO DE PAGOS"
    ws_pagos['A1'].font = Font(bold=True, size=14)
    ws_pagos.merge_cells('A1:F1')
    
    # Encabezados
    headers_pagos = ["Cliente", "Fecha", "Moneda", "Monto Original", "Tipo Cambio", "Monto USD"]
    
    for col, header in enumerate(headers_pagos, start=1):
        cell = ws_pagos.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row = 4
    
    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_cliente = cliente[1]
        
        # Obtener abonos
        cursor.execute("""
            SELECT fecha, moneda, monto_original, tipo_cambio, monto_usd
            FROM abonos_internacionales
            WHERE cliente_id = ?
            ORDER BY fecha
        """, (cliente_id,))
        
        abonos = cursor.fetchall()
        
        if abonos:
            primera_fila = True
            for abono in abonos:
                fecha = datetime.strptime(abono[0], "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M")
                
                ws_pagos.cell(row=row, column=1, value=nombre_cliente if primera_fila else "")
                ws_pagos.cell(row=row, column=2, value=fecha)
                ws_pagos.cell(row=row, column=3, value=abono[1])
                ws_pagos.cell(row=row, column=4, value=abono[2])
                ws_pagos.cell(row=row, column=5, value=abono[3])
                ws_pagos.cell(row=row, column=6, value=abono[4])
                
                # Formato
                ws_pagos.cell(row=row, column=4).number_format = '$#,##0.00'
                ws_pagos.cell(row=row, column=5).number_format = '#,##0.00'
                ws_pagos.cell(row=row, column=6).number_format = '$#,##0.00'
                
                for col in range(1, 7):
                    ws_pagos.cell(row=row, column=col).border = border
                
                primera_fila = False
                row += 1
    
    # Ajustar anchos
    ws_pagos.column_dimensions['A'].width = 30
    ws_pagos.column_dimensions['B'].width = 20
    ws_pagos.column_dimensions['C'].width = 10
    ws_pagos.column_dimensions['D'].width = 18
    ws_pagos.column_dimensions['E'].width = 15
    ws_pagos.column_dimensions['F'].width = 18
    
    # Guardar archivo
    destino_limpio = destino.replace(" ", "_").replace("/", "-")
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"Reporte_{destino_limpio}_{fecha_actual}.xlsx"
    
    # Guardar en el directorio actual (Windows compatible)
    ruta_salida = nombre_archivo
    wb.save(ruta_salida)
    
    print(f"\n✅ Reporte exportado exitosamente:")
    print(f"📁 {nombre_archivo}")
    
    # Mostrar resumen
    print(f"\n📊 RESUMEN:")
    print(f"   Clientes: {len(clientes)}")
    print(f"   Total vendido: ${total_vendido:,.2f} USD")
    print(f"   Total abonado: ${total_abonado:,.2f} USD")
    print(f"   Saldo pendiente: ${total_saldo:,.2f} USD")
    
    conexion.close()
    
    return ruta_salida


def menu_reportes_internacionales():
    """Menú principal de reportes internacionales"""
    while True:
        print("\n" + "="*70)
        print("📊 REPORTES - VIAJES INTERNACIONALES")
        print("="*70)
        print("\n1. Ver reporte en consola")
        print("2. Exportar reporte a Excel")
        print("3. Volver")
        
        opcion = input("\nSelecciona una opción: ").strip()
        
        if opcion == "1":
            reporte_viaje_internacional_consola()
        elif opcion == "2":
            archivo = exportar_reporte_excel()
            if archivo:
                print(f"\n💡 El archivo se guardó en: {archivo}")
        elif opcion == "3":
            break
        else:
            print("❌ Opción inválida.")


if __name__ == "__main__":
    menu_reportes_internacionales()
